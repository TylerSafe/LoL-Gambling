# program written by Tyler Safe to facilitate league of legends kill total gambling
# program scrapes data from given stats pages and performs calculations to give meaningful output, utlising a GUI
# written Jan 2022

import requests
from bs4 import BeautifulSoup
from PyQt5 import QtCore, QtGui, QtWidgets
from PyQt5.QtWidgets import QApplication, QMainWindow
from PyQt5.QtGui import QMovie
import sys
from openpyxl import load_workbook
from datetime import date

class Ui_MainWindow(object):
    def setupUi(self, MainWindow):
        MainWindow.setObjectName("MainWindow")
        MainWindow.resize(1301, 857)
        self.centralwidget = QtWidgets.QWidget(MainWindow)
        self.centralwidget.setObjectName("centralwidget")
        self.stackedWidget = QtWidgets.QStackedWidget(self.centralwidget)
        self.stackedWidget.setGeometry(QtCore.QRect(0, -20, 1471, 891))
        self.stackedWidget.setObjectName("stackedWidget")
        self.page = QtWidgets.QWidget()
        self.page.setObjectName("page")
        self.label = QtWidgets.QLabel(self.page)
        self.label.setGeometry(QtCore.QRect(460, 50, 361, 61))
        self.label.setObjectName("label")
        self.pushButton = QtWidgets.QPushButton(self.page)
        self.pushButton.setGeometry(QtCore.QRect(260, 150, 111, 51))
        self.pushButton.setObjectName("pushButton")
        self.lcs = QtWidgets.QPushButton(self.page)
        self.lcs.setGeometry(QtCore.QRect(260, 240, 111, 51))
        self.lcs.setObjectName("lcs")
        self.lec = QtWidgets.QPushButton(self.page)
        self.lec.setGeometry(QtCore.QRect(260, 330, 111, 51))
        self.lec.setObjectName("lec")
        self.vcs = QtWidgets.QPushButton(self.page)
        self.vcs.setGeometry(QtCore.QRect(480, 240, 111, 51))
        self.vcs.setObjectName("vcs")
        self.lec_6 = QtWidgets.QPushButton(self.page)
        self.lec_6.setGeometry(QtCore.QRect(700, 240, 111, 51))
        self.lec_6.setObjectName("lec_6")
        self.lec_7 = QtWidgets.QPushButton(self.page)
        self.lec_7.setGeometry(QtCore.QRect(920, 150, 111, 51))
        self.lec_7.setObjectName("lec_7")
        self.lec_8 = QtWidgets.QPushButton(self.page)
        self.lec_8.setGeometry(QtCore.QRect(700, 150, 111, 51))
        self.lec_8.setObjectName("lec_8")
        self.lec_9 = QtWidgets.QPushButton(self.page)
        self.lec_9.setGeometry(QtCore.QRect(480, 330, 111, 51))
        self.lec_9.setObjectName("lec_9")
        self.lpl = QtWidgets.QPushButton(self.page)
        self.lpl.setGeometry(QtCore.QRect(480, 150, 111, 51))
        self.lpl.setObjectName("lpl")
        self.lec_11 = QtWidgets.QPushButton(self.page)
        self.lec_11.setGeometry(QtCore.QRect(700, 330, 111, 51))
        self.lec_11.setObjectName("lec_11")
        self.lec_12 = QtWidgets.QPushButton(self.page)
        self.lec_12.setGeometry(QtCore.QRect(920, 240, 111, 51))
        self.lec_12.setObjectName("lec_12")
        self.lec_13 = QtWidgets.QPushButton(self.page)
        self.lec_13.setGeometry(QtCore.QRect(920, 330, 111, 51))
        self.lec_13.setObjectName("lec_13")
        self.upcoming_table = QtWidgets.QTableWidget(self.page)
        self.upcoming_table.setGeometry(QtCore.QRect(60, 480, 561, 311))
        self.upcoming_table.setSizeAdjustPolicy(QtWidgets.QAbstractScrollArea.AdjustToContentsOnFirstShow)
        self.upcoming_table.setObjectName("upcoming_table")
        self.upcoming_table.setColumnCount(2)
        self.upcoming_table.setRowCount(0)
        item = QtWidgets.QTableWidgetItem()
        self.upcoming_table.setHorizontalHeaderItem(0, item)
        item = QtWidgets.QTableWidgetItem()
        self.upcoming_table.setHorizontalHeaderItem(1, item)
        self.label_77 = QtWidgets.QLabel(self.page)
        self.label_77.setGeometry(QtCore.QRect(270, 430, 221, 31))
        self.label_77.setObjectName("label_77")
        # added animated gif
        self.label_13 = QtWidgets.QLabel(self.page)
        self.label_13.setGeometry(QtCore.QRect(720, 490, 521, 291))
        self.label_13.setText("")
        self.label_13.setObjectName("label_13")
        self.movie = QMovie("C:\\Users\\Legen\\Documents\\League Program\\league_image.webp")
        self.label_13.setMovie(self.movie)
        self.movie.start()
        # keep code
        self.stackedWidget.addWidget(self.page)
        self.page_2 = QtWidgets.QWidget()
        self.page_2.setObjectName("page_2")
        self.label_2 = QtWidgets.QLabel(self.page_2)
        self.label_2.setGeometry(QtCore.QRect(440, 20, 361, 61))
        self.label_2.setObjectName("label_2")
        self.tableWidget = QtWidgets.QTableWidget(self.page_2)
        self.tableWidget.setGeometry(QtCore.QRect(100, 80, 831, 351))
        self.tableWidget.setObjectName("tableWidget")
        self.tableWidget.setColumnCount(8)
        self.tableWidget.setRowCount(0)
        item = QtWidgets.QTableWidgetItem()
        self.tableWidget.setHorizontalHeaderItem(0, item)
        item = QtWidgets.QTableWidgetItem()
        self.tableWidget.setHorizontalHeaderItem(1, item)
        item = QtWidgets.QTableWidgetItem()
        self.tableWidget.setHorizontalHeaderItem(2, item)
        item = QtWidgets.QTableWidgetItem()
        self.tableWidget.setHorizontalHeaderItem(3, item)
        item = QtWidgets.QTableWidgetItem()
        self.tableWidget.setHorizontalHeaderItem(4, item)
        item = QtWidgets.QTableWidgetItem()
        self.tableWidget.setHorizontalHeaderItem(5, item)
        item = QtWidgets.QTableWidgetItem()
        self.tableWidget.setHorizontalHeaderItem(6, item)
        item = QtWidgets.QTableWidgetItem()
        self.tableWidget.setHorizontalHeaderItem(7, item)
        self.update_lck = QtWidgets.QPushButton(self.page_2)
        self.update_lck.setGeometry(QtCore.QRect(830, 440, 101, 41))
        self.update_lck.setObjectName("update_lck")
        self.back = QtWidgets.QPushButton(self.page_2)
        self.back.setGeometry(QtCore.QRect(10, 30, 91, 41))
        self.back.setObjectName("back")
        self.ladder_1 = QtWidgets.QTableWidget(self.page_2)
        self.ladder_1.setGeometry(QtCore.QRect(1000, 80, 231, 351))
        self.ladder_1.setObjectName("ladder_1")
        self.ladder_1.setColumnCount(2)
        self.ladder_1.setRowCount(0)
        item = QtWidgets.QTableWidgetItem()
        self.ladder_1.setHorizontalHeaderItem(0, item)
        item = QtWidgets.QTableWidgetItem()
        self.ladder_1.setHorizontalHeaderItem(1, item)
        self.label_4 = QtWidgets.QLabel(self.page_2)
        self.label_4.setGeometry(QtCore.QRect(1080, 20, 361, 61))
        self.label_4.setObjectName("label_4")
        self.lck_matches = QtWidgets.QTableWidget(self.page_2)
        self.lck_matches.setGeometry(QtCore.QRect(90, 500, 1141, 281))
        self.lck_matches.setObjectName("lck_matches")
        self.lck_matches.setColumnCount(9)
        self.lck_matches.setRowCount(0)
        item = QtWidgets.QTableWidgetItem()
        self.lck_matches.setHorizontalHeaderItem(0, item)
        item = QtWidgets.QTableWidgetItem()
        self.lck_matches.setHorizontalHeaderItem(1, item)
        item = QtWidgets.QTableWidgetItem()
        self.lck_matches.setHorizontalHeaderItem(2, item)
        item = QtWidgets.QTableWidgetItem()
        self.lck_matches.setHorizontalHeaderItem(3, item)
        item = QtWidgets.QTableWidgetItem()
        self.lck_matches.setHorizontalHeaderItem(4, item)
        item = QtWidgets.QTableWidgetItem()
        self.lck_matches.setHorizontalHeaderItem(5, item)
        item = QtWidgets.QTableWidgetItem()
        self.lck_matches.setHorizontalHeaderItem(6, item)
        item = QtWidgets.QTableWidgetItem()
        self.lck_matches.setHorizontalHeaderItem(7, item)
        item = QtWidgets.QTableWidgetItem()
        self.lck_matches.setHorizontalHeaderItem(8, item)
        self.label_8 = QtWidgets.QLabel(self.page_2)
        self.label_8.setGeometry(QtCore.QRect(440, 800, 61, 21))
        self.label_8.setObjectName("label_8")
        self.label_9 = QtWidgets.QLabel(self.page_2)
        self.label_9.setGeometry(QtCore.QRect(620, 800, 51, 21))
        self.label_9.setObjectName("label_9")
        self.lck_line = QtWidgets.QTextEdit(self.page_2)
        self.lck_line.setGeometry(QtCore.QRect(480, 800, 71, 21))
        self.lck_line.setObjectName("lck_line")
        self.lck_odds = QtWidgets.QTextEdit(self.page_2)
        self.lck_odds.setGeometry(QtCore.QRect(670, 800, 71, 21))
        self.lck_odds.setObjectName("lck_odds")
        self.lck_calculate = QtWidgets.QPushButton(self.page_2)
        self.lck_calculate.setGeometry(QtCore.QRect(840, 790, 141, 41))
        self.lck_calculate.setObjectName("lck_calculate")
        self.stackedWidget.addWidget(self.page_2)
        self.page_3 = QtWidgets.QWidget()
        self.page_3.setObjectName("page_3")
        self.tableWidget_2 = QtWidgets.QTableWidget(self.page_3)
        self.tableWidget_2.setGeometry(QtCore.QRect(100, 80, 831, 351))
        self.tableWidget_2.setObjectName("tableWidget_2")
        self.tableWidget_2.setColumnCount(8)
        self.tableWidget_2.setRowCount(0)
        item = QtWidgets.QTableWidgetItem()
        self.tableWidget_2.setHorizontalHeaderItem(0, item)
        item = QtWidgets.QTableWidgetItem()
        self.tableWidget_2.setHorizontalHeaderItem(1, item)
        item = QtWidgets.QTableWidgetItem()
        self.tableWidget_2.setHorizontalHeaderItem(2, item)
        item = QtWidgets.QTableWidgetItem()
        self.tableWidget_2.setHorizontalHeaderItem(3, item)
        item = QtWidgets.QTableWidgetItem()
        self.tableWidget_2.setHorizontalHeaderItem(4, item)
        item = QtWidgets.QTableWidgetItem()
        self.tableWidget_2.setHorizontalHeaderItem(5, item)
        item = QtWidgets.QTableWidgetItem()
        self.tableWidget_2.setHorizontalHeaderItem(6, item)
        item = QtWidgets.QTableWidgetItem()
        self.tableWidget_2.setHorizontalHeaderItem(7, item)
        self.label_3 = QtWidgets.QLabel(self.page_3)
        self.label_3.setGeometry(QtCore.QRect(440, 20, 361, 61))
        self.label_3.setObjectName("label_3")
        self.back_2 = QtWidgets.QPushButton(self.page_3)
        self.back_2.setGeometry(QtCore.QRect(10, 30, 91, 41))
        self.back_2.setObjectName("back_2")
        self.update_lcs = QtWidgets.QPushButton(self.page_3)
        self.update_lcs.setGeometry(QtCore.QRect(830, 440, 101, 41))
        self.update_lcs.setStyleSheet("")
        self.update_lcs.setObjectName("update_lcs")
        self.ladder_2 = QtWidgets.QTableWidget(self.page_3)
        self.ladder_2.setGeometry(QtCore.QRect(1000, 80, 231, 351))
        self.ladder_2.setObjectName("ladder_2")
        self.ladder_2.setColumnCount(2)
        self.ladder_2.setRowCount(0)
        item = QtWidgets.QTableWidgetItem()
        self.ladder_2.setHorizontalHeaderItem(0, item)
        item = QtWidgets.QTableWidgetItem()
        self.ladder_2.setHorizontalHeaderItem(1, item)
        self.label_5 = QtWidgets.QLabel(self.page_3)
        self.label_5.setGeometry(QtCore.QRect(1080, 20, 361, 61))
        self.label_5.setObjectName("label_5")
        self.lcs_matches = QtWidgets.QTableWidget(self.page_3)
        self.lcs_matches.setGeometry(QtCore.QRect(90, 500, 1141, 281))
        self.lcs_matches.setObjectName("lcs_matches")
        self.lcs_matches.setColumnCount(9)
        self.lcs_matches.setRowCount(0)
        item = QtWidgets.QTableWidgetItem()
        self.lcs_matches.setHorizontalHeaderItem(0, item)
        item = QtWidgets.QTableWidgetItem()
        self.lcs_matches.setHorizontalHeaderItem(1, item)
        item = QtWidgets.QTableWidgetItem()
        self.lcs_matches.setHorizontalHeaderItem(2, item)
        item = QtWidgets.QTableWidgetItem()
        self.lcs_matches.setHorizontalHeaderItem(3, item)
        item = QtWidgets.QTableWidgetItem()
        self.lcs_matches.setHorizontalHeaderItem(4, item)
        item = QtWidgets.QTableWidgetItem()
        self.lcs_matches.setHorizontalHeaderItem(5, item)
        item = QtWidgets.QTableWidgetItem()
        self.lcs_matches.setHorizontalHeaderItem(6, item)
        item = QtWidgets.QTableWidgetItem()
        self.lcs_matches.setHorizontalHeaderItem(7, item)
        item = QtWidgets.QTableWidgetItem()
        self.lcs_matches.setHorizontalHeaderItem(8, item)
        self.lcs_odds = QtWidgets.QTextEdit(self.page_3)
        self.lcs_odds.setGeometry(QtCore.QRect(650, 800, 71, 21))
        self.lcs_odds.setObjectName("lcs_odds")
        self.lcs_calculate = QtWidgets.QPushButton(self.page_3)
        self.lcs_calculate.setGeometry(QtCore.QRect(820, 790, 141, 41))
        self.lcs_calculate.setObjectName("lcs_calculate")
        self.label_10 = QtWidgets.QLabel(self.page_3)
        self.label_10.setGeometry(QtCore.QRect(600, 800, 51, 21))
        self.label_10.setObjectName("label_10")
        self.label_14 = QtWidgets.QLabel(self.page_3)
        self.label_14.setGeometry(QtCore.QRect(420, 800, 61, 21))
        self.label_14.setObjectName("label_14")
        self.lcs_line = QtWidgets.QTextEdit(self.page_3)
        self.lcs_line.setGeometry(QtCore.QRect(460, 800, 71, 21))
        self.lcs_line.setObjectName("lcs_line")
        self.stackedWidget.addWidget(self.page_3)
        self.page_4 = QtWidgets.QWidget()
        self.page_4.setObjectName("page_4")
        self.tableWidget_3 = QtWidgets.QTableWidget(self.page_4)
        self.tableWidget_3.setGeometry(QtCore.QRect(100, 80, 831, 351))
        self.tableWidget_3.setObjectName("tableWidget_3")
        self.tableWidget_3.setColumnCount(8)
        self.tableWidget_3.setRowCount(0)
        item = QtWidgets.QTableWidgetItem()
        self.tableWidget_3.setHorizontalHeaderItem(0, item)
        item = QtWidgets.QTableWidgetItem()
        self.tableWidget_3.setHorizontalHeaderItem(1, item)
        item = QtWidgets.QTableWidgetItem()
        self.tableWidget_3.setHorizontalHeaderItem(2, item)
        item = QtWidgets.QTableWidgetItem()
        self.tableWidget_3.setHorizontalHeaderItem(3, item)
        item = QtWidgets.QTableWidgetItem()
        self.tableWidget_3.setHorizontalHeaderItem(4, item)
        item = QtWidgets.QTableWidgetItem()
        self.tableWidget_3.setHorizontalHeaderItem(5, item)
        item = QtWidgets.QTableWidgetItem()
        self.tableWidget_3.setHorizontalHeaderItem(6, item)
        item = QtWidgets.QTableWidgetItem()
        self.tableWidget_3.setHorizontalHeaderItem(7, item)
        self.label_6 = QtWidgets.QLabel(self.page_4)
        self.label_6.setGeometry(QtCore.QRect(430, 20, 361, 61))
        self.label_6.setObjectName("label_6")
        self.back_3 = QtWidgets.QPushButton(self.page_4)
        self.back_3.setGeometry(QtCore.QRect(10, 30, 91, 41))
        self.back_3.setObjectName("back_3")
        self.update_lec = QtWidgets.QPushButton(self.page_4)
        self.update_lec.setGeometry(QtCore.QRect(830, 440, 101, 41))
        self.update_lec.setObjectName("update_lec")
        self.label_7 = QtWidgets.QLabel(self.page_4)
        self.label_7.setGeometry(QtCore.QRect(1080, 20, 361, 61))
        self.label_7.setObjectName("label_7")
        self.ladder_3 = QtWidgets.QTableWidget(self.page_4)
        self.ladder_3.setGeometry(QtCore.QRect(1000, 80, 231, 351))
        self.ladder_3.setObjectName("ladder_3")
        self.ladder_3.setColumnCount(2)
        self.ladder_3.setRowCount(0)
        item = QtWidgets.QTableWidgetItem()
        self.ladder_3.setHorizontalHeaderItem(0, item)
        item = QtWidgets.QTableWidgetItem()
        self.ladder_3.setHorizontalHeaderItem(1, item)
        self.lec_matches = QtWidgets.QTableWidget(self.page_4)
        self.lec_matches.setGeometry(QtCore.QRect(90, 500, 1141, 281))
        self.lec_matches.setObjectName("lec_matches")
        self.lec_matches.setColumnCount(9)
        self.lec_matches.setRowCount(0)
        item = QtWidgets.QTableWidgetItem()
        self.lec_matches.setHorizontalHeaderItem(0, item)
        item = QtWidgets.QTableWidgetItem()
        self.lec_matches.setHorizontalHeaderItem(1, item)
        item = QtWidgets.QTableWidgetItem()
        self.lec_matches.setHorizontalHeaderItem(2, item)
        item = QtWidgets.QTableWidgetItem()
        self.lec_matches.setHorizontalHeaderItem(3, item)
        item = QtWidgets.QTableWidgetItem()
        self.lec_matches.setHorizontalHeaderItem(4, item)
        item = QtWidgets.QTableWidgetItem()
        self.lec_matches.setHorizontalHeaderItem(5, item)
        item = QtWidgets.QTableWidgetItem()
        self.lec_matches.setHorizontalHeaderItem(6, item)
        item = QtWidgets.QTableWidgetItem()
        self.lec_matches.setHorizontalHeaderItem(7, item)
        item = QtWidgets.QTableWidgetItem()
        self.lec_matches.setHorizontalHeaderItem(8, item)
        self.lec_odds = QtWidgets.QTextEdit(self.page_4)
        self.lec_odds.setGeometry(QtCore.QRect(640, 800, 71, 21))
        self.lec_odds.setObjectName("lec_odds")
        self.lec_calculate = QtWidgets.QPushButton(self.page_4)
        self.lec_calculate.setGeometry(QtCore.QRect(810, 790, 141, 41))
        self.lec_calculate.setObjectName("lec_calculate")
        self.label_11 = QtWidgets.QLabel(self.page_4)
        self.label_11.setGeometry(QtCore.QRect(590, 800, 51, 21))
        self.label_11.setObjectName("label_11")
        self.label_12 = QtWidgets.QLabel(self.page_4)
        self.label_12.setGeometry(QtCore.QRect(410, 800, 61, 21))
        self.label_12.setObjectName("label_12")
        self.lec_line = QtWidgets.QTextEdit(self.page_4)
        self.lec_line.setGeometry(QtCore.QRect(450, 800, 71, 21))
        self.lec_line.setObjectName("lec_line")
        self.stackedWidget.addWidget(self.page_4)
        self.page_5 = QtWidgets.QWidget()
        self.page_5.setObjectName("page_5")
        self.label_53 = QtWidgets.QLabel(self.page_5)
        self.label_53.setGeometry(QtCore.QRect(440, 20, 361, 61))
        self.label_53.setObjectName("label_53")
        self.label_54 = QtWidgets.QLabel(self.page_5)
        self.label_54.setGeometry(QtCore.QRect(620, 800, 51, 21))
        self.label_54.setObjectName("label_54")
        self.lpl_stats_table = QtWidgets.QTableWidget(self.page_5)
        self.lpl_stats_table.setGeometry(QtCore.QRect(100, 80, 831, 351))
        self.lpl_stats_table.setObjectName("lpl_stats_table")
        self.lpl_stats_table.setColumnCount(8)
        self.lpl_stats_table.setRowCount(0)
        item = QtWidgets.QTableWidgetItem()
        self.lpl_stats_table.setHorizontalHeaderItem(0, item)
        item = QtWidgets.QTableWidgetItem()
        self.lpl_stats_table.setHorizontalHeaderItem(1, item)
        item = QtWidgets.QTableWidgetItem()
        self.lpl_stats_table.setHorizontalHeaderItem(2, item)
        item = QtWidgets.QTableWidgetItem()
        self.lpl_stats_table.setHorizontalHeaderItem(3, item)
        item = QtWidgets.QTableWidgetItem()
        self.lpl_stats_table.setHorizontalHeaderItem(4, item)
        item = QtWidgets.QTableWidgetItem()
        self.lpl_stats_table.setHorizontalHeaderItem(5, item)
        item = QtWidgets.QTableWidgetItem()
        self.lpl_stats_table.setHorizontalHeaderItem(6, item)
        item = QtWidgets.QTableWidgetItem()
        self.lpl_stats_table.setHorizontalHeaderItem(7, item)
        self.label_55 = QtWidgets.QLabel(self.page_5)
        self.label_55.setGeometry(QtCore.QRect(1080, 20, 361, 61))
        self.label_55.setObjectName("label_55")
        self.lpl_odds = QtWidgets.QTextEdit(self.page_5)
        self.lpl_odds.setGeometry(QtCore.QRect(670, 800, 71, 21))
        self.lpl_odds.setObjectName("lpl_odds")
        self.ladder_lpl = QtWidgets.QTableWidget(self.page_5)
        self.ladder_lpl.setGeometry(QtCore.QRect(1000, 80, 231, 351))
        self.ladder_lpl.setObjectName("ladder_lpl")
        self.ladder_lpl.setColumnCount(2)
        self.ladder_lpl.setRowCount(0)
        item = QtWidgets.QTableWidgetItem()
        self.ladder_lpl.setHorizontalHeaderItem(0, item)
        item = QtWidgets.QTableWidgetItem()
        self.ladder_lpl.setHorizontalHeaderItem(1, item)
        self.lpl_matches = QtWidgets.QTableWidget(self.page_5)
        self.lpl_matches.setGeometry(QtCore.QRect(90, 500, 1141, 281))
        self.lpl_matches.setObjectName("lpl_matches")
        self.lpl_matches.setColumnCount(9)
        self.lpl_matches.setRowCount(0)
        item = QtWidgets.QTableWidgetItem()
        self.lpl_matches.setHorizontalHeaderItem(0, item)
        item = QtWidgets.QTableWidgetItem()
        self.lpl_matches.setHorizontalHeaderItem(1, item)
        item = QtWidgets.QTableWidgetItem()
        self.lpl_matches.setHorizontalHeaderItem(2, item)
        item = QtWidgets.QTableWidgetItem()
        self.lpl_matches.setHorizontalHeaderItem(3, item)
        item = QtWidgets.QTableWidgetItem()
        self.lpl_matches.setHorizontalHeaderItem(4, item)
        item = QtWidgets.QTableWidgetItem()
        self.lpl_matches.setHorizontalHeaderItem(5, item)
        item = QtWidgets.QTableWidgetItem()
        self.lpl_matches.setHorizontalHeaderItem(6, item)
        item = QtWidgets.QTableWidgetItem()
        self.lpl_matches.setHorizontalHeaderItem(7, item)
        item = QtWidgets.QTableWidgetItem()
        self.lpl_matches.setHorizontalHeaderItem(8, item)
        self.lpl_calculate = QtWidgets.QPushButton(self.page_5)
        self.lpl_calculate.setGeometry(QtCore.QRect(840, 790, 141, 41))
        self.lpl_calculate.setObjectName("lpl_calculate")
        self.label_56 = QtWidgets.QLabel(self.page_5)
        self.label_56.setGeometry(QtCore.QRect(440, 800, 61, 21))
        self.label_56.setObjectName("label_56")
        self.lpl_line = QtWidgets.QTextEdit(self.page_5)
        self.lpl_line.setGeometry(QtCore.QRect(480, 800, 71, 21))
        self.lpl_line.setObjectName("lpl_line")
        self.back_13 = QtWidgets.QPushButton(self.page_5)
        self.back_13.setGeometry(QtCore.QRect(10, 30, 91, 41))
        self.back_13.setObjectName("back_13")
        self.update_lpl = QtWidgets.QPushButton(self.page_5)
        self.update_lpl.setGeometry(QtCore.QRect(830, 440, 101, 41))
        self.update_lpl.setObjectName("update_lpl")
        self.stackedWidget.addWidget(self.page_5)
        self.page_vcs = QtWidgets.QWidget()
        self.page_vcs.setObjectName("page_vcs")
        self.update_vcs = QtWidgets.QPushButton(self.page_vcs)
        self.update_vcs.setGeometry(QtCore.QRect(830, 440, 101, 41))
        self.update_vcs.setObjectName("update_vcs")
        self.vcs_calculate = QtWidgets.QPushButton(self.page_vcs)
        self.vcs_calculate.setGeometry(QtCore.QRect(840, 790, 141, 41))
        self.vcs_calculate.setObjectName("vcs_calculate")
        self.label_60 = QtWidgets.QLabel(self.page_vcs)
        self.label_60.setGeometry(QtCore.QRect(440, 800, 61, 21))
        self.label_60.setObjectName("label_60")
        self.vcs_line = QtWidgets.QTextEdit(self.page_vcs)
        self.vcs_line.setGeometry(QtCore.QRect(480, 800, 71, 21))
        self.vcs_line.setObjectName("vcs_line")
        self.back_14 = QtWidgets.QPushButton(self.page_vcs)
        self.back_14.setGeometry(QtCore.QRect(10, 30, 91, 41))
        self.back_14.setObjectName("back_14")
        self.label_57 = QtWidgets.QLabel(self.page_vcs)
        self.label_57.setGeometry(QtCore.QRect(440, 20, 361, 61))
        self.label_57.setObjectName("label_57")
        self.vcs_odds = QtWidgets.QTextEdit(self.page_vcs)
        self.vcs_odds.setGeometry(QtCore.QRect(670, 800, 71, 21))
        self.vcs_odds.setObjectName("vcs_odds")
        self.label_58 = QtWidgets.QLabel(self.page_vcs)
        self.label_58.setGeometry(QtCore.QRect(620, 800, 51, 21))
        self.label_58.setObjectName("label_58")
        self.ladder_vcs = QtWidgets.QTableWidget(self.page_vcs)
        self.ladder_vcs.setGeometry(QtCore.QRect(1000, 80, 231, 351))
        self.ladder_vcs.setObjectName("ladder_vcs")
        self.ladder_vcs.setColumnCount(2)
        self.ladder_vcs.setRowCount(0)
        item = QtWidgets.QTableWidgetItem()
        self.ladder_vcs.setHorizontalHeaderItem(0, item)
        item = QtWidgets.QTableWidgetItem()
        self.ladder_vcs.setHorizontalHeaderItem(1, item)
        self.vcs_stats_table = QtWidgets.QTableWidget(self.page_vcs)
        self.vcs_stats_table.setGeometry(QtCore.QRect(100, 80, 831, 351))
        self.vcs_stats_table.setObjectName("vcs_stats_table")
        self.vcs_stats_table.setColumnCount(8)
        self.vcs_stats_table.setRowCount(0)
        item = QtWidgets.QTableWidgetItem()
        self.vcs_stats_table.setHorizontalHeaderItem(0, item)
        item = QtWidgets.QTableWidgetItem()
        self.vcs_stats_table.setHorizontalHeaderItem(1, item)
        item = QtWidgets.QTableWidgetItem()
        self.vcs_stats_table.setHorizontalHeaderItem(2, item)
        item = QtWidgets.QTableWidgetItem()
        self.vcs_stats_table.setHorizontalHeaderItem(3, item)
        item = QtWidgets.QTableWidgetItem()
        self.vcs_stats_table.setHorizontalHeaderItem(4, item)
        item = QtWidgets.QTableWidgetItem()
        self.vcs_stats_table.setHorizontalHeaderItem(5, item)
        item = QtWidgets.QTableWidgetItem()
        self.vcs_stats_table.setHorizontalHeaderItem(6, item)
        item = QtWidgets.QTableWidgetItem()
        self.vcs_stats_table.setHorizontalHeaderItem(7, item)
        self.label_59 = QtWidgets.QLabel(self.page_vcs)
        self.label_59.setGeometry(QtCore.QRect(1080, 20, 361, 61))
        self.label_59.setObjectName("label_59")
        self.vcs_matches = QtWidgets.QTableWidget(self.page_vcs)
        self.vcs_matches.setGeometry(QtCore.QRect(90, 500, 1141, 281))
        self.vcs_matches.setObjectName("vcs_matches")
        self.vcs_matches.setColumnCount(9)
        self.vcs_matches.setRowCount(0)
        item = QtWidgets.QTableWidgetItem()
        self.vcs_matches.setHorizontalHeaderItem(0, item)
        item = QtWidgets.QTableWidgetItem()
        self.vcs_matches.setHorizontalHeaderItem(1, item)
        item = QtWidgets.QTableWidgetItem()
        self.vcs_matches.setHorizontalHeaderItem(2, item)
        item = QtWidgets.QTableWidgetItem()
        self.vcs_matches.setHorizontalHeaderItem(3, item)
        item = QtWidgets.QTableWidgetItem()
        self.vcs_matches.setHorizontalHeaderItem(4, item)
        item = QtWidgets.QTableWidgetItem()
        self.vcs_matches.setHorizontalHeaderItem(5, item)
        item = QtWidgets.QTableWidgetItem()
        self.vcs_matches.setHorizontalHeaderItem(6, item)
        item = QtWidgets.QTableWidgetItem()
        self.vcs_matches.setHorizontalHeaderItem(7, item)
        item = QtWidgets.QTableWidgetItem()
        self.vcs_matches.setHorizontalHeaderItem(8, item)
        self.stackedWidget.addWidget(self.page_vcs)
        self.page_21 = QtWidgets.QWidget()
        self.page_21.setObjectName("page_21")
        self.update_lck_7 = QtWidgets.QPushButton(self.page_21)
        self.update_lck_7.setGeometry(QtCore.QRect(830, 440, 101, 41))
        self.update_lck_7.setObjectName("update_lck_7")
        self.ladder_15 = QtWidgets.QTableWidget(self.page_21)
        self.ladder_15.setGeometry(QtCore.QRect(1000, 80, 231, 351))
        self.ladder_15.setObjectName("ladder_15")
        self.ladder_15.setColumnCount(2)
        self.ladder_15.setRowCount(0)
        item = QtWidgets.QTableWidgetItem()
        self.ladder_15.setHorizontalHeaderItem(0, item)
        item = QtWidgets.QTableWidgetItem()
        self.ladder_15.setHorizontalHeaderItem(1, item)
        self.label_62 = QtWidgets.QLabel(self.page_21)
        self.label_62.setGeometry(QtCore.QRect(620, 800, 91, 21))
        self.label_62.setObjectName("label_62")
        self.tableWidget_15 = QtWidgets.QTableWidget(self.page_21)
        self.tableWidget_15.setGeometry(QtCore.QRect(100, 80, 831, 351))
        self.tableWidget_15.setObjectName("tableWidget_15")
        self.tableWidget_15.setColumnCount(8)
        self.tableWidget_15.setRowCount(0)
        item = QtWidgets.QTableWidgetItem()
        self.tableWidget_15.setHorizontalHeaderItem(0, item)
        item = QtWidgets.QTableWidgetItem()
        self.tableWidget_15.setHorizontalHeaderItem(1, item)
        item = QtWidgets.QTableWidgetItem()
        self.tableWidget_15.setHorizontalHeaderItem(2, item)
        item = QtWidgets.QTableWidgetItem()
        self.tableWidget_15.setHorizontalHeaderItem(3, item)
        item = QtWidgets.QTableWidgetItem()
        self.tableWidget_15.setHorizontalHeaderItem(4, item)
        item = QtWidgets.QTableWidgetItem()
        self.tableWidget_15.setHorizontalHeaderItem(5, item)
        item = QtWidgets.QTableWidgetItem()
        self.tableWidget_15.setHorizontalHeaderItem(6, item)
        item = QtWidgets.QTableWidgetItem()
        self.tableWidget_15.setHorizontalHeaderItem(7, item)
        self.lck_odds_7 = QtWidgets.QTextEdit(self.page_21)
        self.lck_odds_7.setGeometry(QtCore.QRect(670, 800, 71, 21))
        self.lck_odds_7.setObjectName("lck_odds_7")
        self.back_15 = QtWidgets.QPushButton(self.page_21)
        self.back_15.setGeometry(QtCore.QRect(10, 30, 91, 41))
        self.back_15.setObjectName("back_15")
        self.label_61 = QtWidgets.QLabel(self.page_21)
        self.label_61.setGeometry(QtCore.QRect(440, 20, 361, 61))
        self.label_61.setObjectName("label_61")
        self.lck_matches_7 = QtWidgets.QTableWidget(self.page_21)
        self.lck_matches_7.setGeometry(QtCore.QRect(90, 500, 1141, 281))
        self.lck_matches_7.setObjectName("lck_matches_7")
        self.lck_matches_7.setColumnCount(9)
        self.lck_matches_7.setRowCount(0)
        item = QtWidgets.QTableWidgetItem()
        self.lck_matches_7.setHorizontalHeaderItem(0, item)
        item = QtWidgets.QTableWidgetItem()
        self.lck_matches_7.setHorizontalHeaderItem(1, item)
        item = QtWidgets.QTableWidgetItem()
        self.lck_matches_7.setHorizontalHeaderItem(2, item)
        item = QtWidgets.QTableWidgetItem()
        self.lck_matches_7.setHorizontalHeaderItem(3, item)
        item = QtWidgets.QTableWidgetItem()
        self.lck_matches_7.setHorizontalHeaderItem(4, item)
        item = QtWidgets.QTableWidgetItem()
        self.lck_matches_7.setHorizontalHeaderItem(5, item)
        item = QtWidgets.QTableWidgetItem()
        self.lck_matches_7.setHorizontalHeaderItem(6, item)
        item = QtWidgets.QTableWidgetItem()
        self.lck_matches_7.setHorizontalHeaderItem(7, item)
        item = QtWidgets.QTableWidgetItem()
        self.lck_matches_7.setHorizontalHeaderItem(8, item)
        self.label_63 = QtWidgets.QLabel(self.page_21)
        self.label_63.setGeometry(QtCore.QRect(1080, 20, 361, 61))
        self.label_63.setObjectName("label_63")
        self.label_64 = QtWidgets.QLabel(self.page_21)
        self.label_64.setGeometry(QtCore.QRect(440, 800, 61, 21))
        self.label_64.setObjectName("label_64")
        self.lck_line_7 = QtWidgets.QTextEdit(self.page_21)
        self.lck_line_7.setGeometry(QtCore.QRect(480, 800, 71, 21))
        self.lck_line_7.setObjectName("lck_line_7")
        self.lck_calculate_7 = QtWidgets.QPushButton(self.page_21)
        self.lck_calculate_7.setGeometry(QtCore.QRect(840, 790, 141, 41))
        self.lck_calculate_7.setObjectName("lck_calculate_7")
        self.stackedWidget.addWidget(self.page_21)
        self.page_22 = QtWidgets.QWidget()
        self.page_22.setObjectName("page_22")
        self.update_lck_8 = QtWidgets.QPushButton(self.page_22)
        self.update_lck_8.setGeometry(QtCore.QRect(830, 440, 101, 41))
        self.update_lck_8.setObjectName("update_lck_8")
        self.label_65 = QtWidgets.QLabel(self.page_22)
        self.label_65.setGeometry(QtCore.QRect(440, 20, 361, 61))
        self.label_65.setObjectName("label_65")
        self.label_66 = QtWidgets.QLabel(self.page_22)
        self.label_66.setGeometry(QtCore.QRect(620, 800, 91, 21))
        self.label_66.setObjectName("label_66")
        self.back_16 = QtWidgets.QPushButton(self.page_22)
        self.back_16.setGeometry(QtCore.QRect(10, 30, 91, 41))
        self.back_16.setObjectName("back_16")
        self.tableWidget_16 = QtWidgets.QTableWidget(self.page_22)
        self.tableWidget_16.setGeometry(QtCore.QRect(100, 80, 831, 351))
        self.tableWidget_16.setObjectName("tableWidget_16")
        self.tableWidget_16.setColumnCount(8)
        self.tableWidget_16.setRowCount(0)
        item = QtWidgets.QTableWidgetItem()
        self.tableWidget_16.setHorizontalHeaderItem(0, item)
        item = QtWidgets.QTableWidgetItem()
        self.tableWidget_16.setHorizontalHeaderItem(1, item)
        item = QtWidgets.QTableWidgetItem()
        self.tableWidget_16.setHorizontalHeaderItem(2, item)
        item = QtWidgets.QTableWidgetItem()
        self.tableWidget_16.setHorizontalHeaderItem(3, item)
        item = QtWidgets.QTableWidgetItem()
        self.tableWidget_16.setHorizontalHeaderItem(4, item)
        item = QtWidgets.QTableWidgetItem()
        self.tableWidget_16.setHorizontalHeaderItem(5, item)
        item = QtWidgets.QTableWidgetItem()
        self.tableWidget_16.setHorizontalHeaderItem(6, item)
        item = QtWidgets.QTableWidgetItem()
        self.tableWidget_16.setHorizontalHeaderItem(7, item)
        self.lck_calculate_8 = QtWidgets.QPushButton(self.page_22)
        self.lck_calculate_8.setGeometry(QtCore.QRect(840, 790, 141, 41))
        self.lck_calculate_8.setObjectName("lck_calculate_8")
        self.lck_odds_8 = QtWidgets.QTextEdit(self.page_22)
        self.lck_odds_8.setGeometry(QtCore.QRect(670, 800, 71, 21))
        self.lck_odds_8.setObjectName("lck_odds_8")
        self.lck_line_8 = QtWidgets.QTextEdit(self.page_22)
        self.lck_line_8.setGeometry(QtCore.QRect(480, 800, 71, 21))
        self.lck_line_8.setObjectName("lck_line_8")
        self.ladder_16 = QtWidgets.QTableWidget(self.page_22)
        self.ladder_16.setGeometry(QtCore.QRect(1000, 80, 231, 351))
        self.ladder_16.setObjectName("ladder_16")
        self.ladder_16.setColumnCount(2)
        self.ladder_16.setRowCount(0)
        item = QtWidgets.QTableWidgetItem()
        self.ladder_16.setHorizontalHeaderItem(0, item)
        item = QtWidgets.QTableWidgetItem()
        self.ladder_16.setHorizontalHeaderItem(1, item)
        self.label_67 = QtWidgets.QLabel(self.page_22)
        self.label_67.setGeometry(QtCore.QRect(1080, 20, 361, 61))
        self.label_67.setObjectName("label_67")
        self.label_68 = QtWidgets.QLabel(self.page_22)
        self.label_68.setGeometry(QtCore.QRect(440, 800, 61, 21))
        self.label_68.setObjectName("label_68")
        self.lck_matches_8 = QtWidgets.QTableWidget(self.page_22)
        self.lck_matches_8.setGeometry(QtCore.QRect(90, 500, 1141, 281))
        self.lck_matches_8.setObjectName("lck_matches_8")
        self.lck_matches_8.setColumnCount(9)
        self.lck_matches_8.setRowCount(0)
        item = QtWidgets.QTableWidgetItem()
        self.lck_matches_8.setHorizontalHeaderItem(0, item)
        item = QtWidgets.QTableWidgetItem()
        self.lck_matches_8.setHorizontalHeaderItem(1, item)
        item = QtWidgets.QTableWidgetItem()
        self.lck_matches_8.setHorizontalHeaderItem(2, item)
        item = QtWidgets.QTableWidgetItem()
        self.lck_matches_8.setHorizontalHeaderItem(3, item)
        item = QtWidgets.QTableWidgetItem()
        self.lck_matches_8.setHorizontalHeaderItem(4, item)
        item = QtWidgets.QTableWidgetItem()
        self.lck_matches_8.setHorizontalHeaderItem(5, item)
        item = QtWidgets.QTableWidgetItem()
        self.lck_matches_8.setHorizontalHeaderItem(6, item)
        item = QtWidgets.QTableWidgetItem()
        self.lck_matches_8.setHorizontalHeaderItem(7, item)
        item = QtWidgets.QTableWidgetItem()
        self.lck_matches_8.setHorizontalHeaderItem(8, item)
        self.stackedWidget.addWidget(self.page_22)
        self.page_23 = QtWidgets.QWidget()
        self.page_23.setObjectName("page_23")
        self.update_lck_9 = QtWidgets.QPushButton(self.page_23)
        self.update_lck_9.setGeometry(QtCore.QRect(830, 440, 101, 41))
        self.update_lck_9.setObjectName("update_lck_9")
        self.ladder_17 = QtWidgets.QTableWidget(self.page_23)
        self.ladder_17.setGeometry(QtCore.QRect(1000, 80, 231, 351))
        self.ladder_17.setObjectName("ladder_17")
        self.ladder_17.setColumnCount(2)
        self.ladder_17.setRowCount(0)
        item = QtWidgets.QTableWidgetItem()
        self.ladder_17.setHorizontalHeaderItem(0, item)
        item = QtWidgets.QTableWidgetItem()
        self.ladder_17.setHorizontalHeaderItem(1, item)
        self.back_17 = QtWidgets.QPushButton(self.page_23)
        self.back_17.setGeometry(QtCore.QRect(10, 30, 91, 41))
        self.back_17.setObjectName("back_17")
        self.lck_line_9 = QtWidgets.QTextEdit(self.page_23)
        self.lck_line_9.setGeometry(QtCore.QRect(480, 800, 71, 21))
        self.lck_line_9.setObjectName("lck_line_9")
        self.lck_odds_9 = QtWidgets.QTextEdit(self.page_23)
        self.lck_odds_9.setGeometry(QtCore.QRect(670, 800, 71, 21))
        self.lck_odds_9.setObjectName("lck_odds_9")
        self.label_69 = QtWidgets.QLabel(self.page_23)
        self.label_69.setGeometry(QtCore.QRect(440, 20, 361, 61))
        self.label_69.setObjectName("label_69")
        self.tableWidget_17 = QtWidgets.QTableWidget(self.page_23)
        self.tableWidget_17.setGeometry(QtCore.QRect(100, 80, 831, 351))
        self.tableWidget_17.setObjectName("tableWidget_17")
        self.tableWidget_17.setColumnCount(8)
        self.tableWidget_17.setRowCount(0)
        item = QtWidgets.QTableWidgetItem()
        self.tableWidget_17.setHorizontalHeaderItem(0, item)
        item = QtWidgets.QTableWidgetItem()
        self.tableWidget_17.setHorizontalHeaderItem(1, item)
        item = QtWidgets.QTableWidgetItem()
        self.tableWidget_17.setHorizontalHeaderItem(2, item)
        item = QtWidgets.QTableWidgetItem()
        self.tableWidget_17.setHorizontalHeaderItem(3, item)
        item = QtWidgets.QTableWidgetItem()
        self.tableWidget_17.setHorizontalHeaderItem(4, item)
        item = QtWidgets.QTableWidgetItem()
        self.tableWidget_17.setHorizontalHeaderItem(5, item)
        item = QtWidgets.QTableWidgetItem()
        self.tableWidget_17.setHorizontalHeaderItem(6, item)
        item = QtWidgets.QTableWidgetItem()
        self.tableWidget_17.setHorizontalHeaderItem(7, item)
        self.label_70 = QtWidgets.QLabel(self.page_23)
        self.label_70.setGeometry(QtCore.QRect(620, 800, 91, 21))
        self.label_70.setObjectName("label_70")
        self.label_71 = QtWidgets.QLabel(self.page_23)
        self.label_71.setGeometry(QtCore.QRect(1080, 20, 361, 61))
        self.label_71.setObjectName("label_71")
        self.label_72 = QtWidgets.QLabel(self.page_23)
        self.label_72.setGeometry(QtCore.QRect(440, 800, 61, 21))
        self.label_72.setObjectName("label_72")
        self.lck_matches_9 = QtWidgets.QTableWidget(self.page_23)
        self.lck_matches_9.setGeometry(QtCore.QRect(90, 500, 1141, 281))
        self.lck_matches_9.setObjectName("lck_matches_9")
        self.lck_matches_9.setColumnCount(9)
        self.lck_matches_9.setRowCount(0)
        item = QtWidgets.QTableWidgetItem()
        self.lck_matches_9.setHorizontalHeaderItem(0, item)
        item = QtWidgets.QTableWidgetItem()
        self.lck_matches_9.setHorizontalHeaderItem(1, item)
        item = QtWidgets.QTableWidgetItem()
        self.lck_matches_9.setHorizontalHeaderItem(2, item)
        item = QtWidgets.QTableWidgetItem()
        self.lck_matches_9.setHorizontalHeaderItem(3, item)
        item = QtWidgets.QTableWidgetItem()
        self.lck_matches_9.setHorizontalHeaderItem(4, item)
        item = QtWidgets.QTableWidgetItem()
        self.lck_matches_9.setHorizontalHeaderItem(5, item)
        item = QtWidgets.QTableWidgetItem()
        self.lck_matches_9.setHorizontalHeaderItem(6, item)
        item = QtWidgets.QTableWidgetItem()
        self.lck_matches_9.setHorizontalHeaderItem(7, item)
        item = QtWidgets.QTableWidgetItem()
        self.lck_matches_9.setHorizontalHeaderItem(8, item)
        self.lck_calculate_9 = QtWidgets.QPushButton(self.page_23)
        self.lck_calculate_9.setGeometry(QtCore.QRect(840, 790, 141, 41))
        self.lck_calculate_9.setObjectName("lck_calculate_9")
        self.stackedWidget.addWidget(self.page_23)
        self.page_24 = QtWidgets.QWidget()
        self.page_24.setObjectName("page_24")
        self.update_lck_10 = QtWidgets.QPushButton(self.page_24)
        self.update_lck_10.setGeometry(QtCore.QRect(830, 440, 101, 41))
        self.update_lck_10.setObjectName("update_lck_10")
        self.label_76 = QtWidgets.QLabel(self.page_24)
        self.label_76.setGeometry(QtCore.QRect(440, 800, 61, 21))
        self.label_76.setObjectName("label_76")
        self.ladder_18 = QtWidgets.QTableWidget(self.page_24)
        self.ladder_18.setGeometry(QtCore.QRect(1000, 80, 231, 351))
        self.ladder_18.setObjectName("ladder_18")
        self.ladder_18.setColumnCount(2)
        self.ladder_18.setRowCount(0)
        item = QtWidgets.QTableWidgetItem()
        self.ladder_18.setHorizontalHeaderItem(0, item)
        item = QtWidgets.QTableWidgetItem()
        self.ladder_18.setHorizontalHeaderItem(1, item)
        self.back_18 = QtWidgets.QPushButton(self.page_24)
        self.back_18.setGeometry(QtCore.QRect(10, 30, 91, 41))
        self.back_18.setObjectName("back_18")
        self.label_73 = QtWidgets.QLabel(self.page_24)
        self.label_73.setGeometry(QtCore.QRect(440, 20, 361, 61))
        self.label_73.setObjectName("label_73")
        self.lck_calculate_10 = QtWidgets.QPushButton(self.page_24)
        self.lck_calculate_10.setGeometry(QtCore.QRect(840, 790, 141, 41))
        self.lck_calculate_10.setObjectName("lck_calculate_10")
        self.lck_line_10 = QtWidgets.QTextEdit(self.page_24)
        self.lck_line_10.setGeometry(QtCore.QRect(480, 800, 71, 21))
        self.lck_line_10.setObjectName("lck_line_10")
        self.tableWidget_18 = QtWidgets.QTableWidget(self.page_24)
        self.tableWidget_18.setGeometry(QtCore.QRect(100, 80, 831, 351))
        self.tableWidget_18.setObjectName("tableWidget_18")
        self.tableWidget_18.setColumnCount(8)
        self.tableWidget_18.setRowCount(0)
        item = QtWidgets.QTableWidgetItem()
        self.tableWidget_18.setHorizontalHeaderItem(0, item)
        item = QtWidgets.QTableWidgetItem()
        self.tableWidget_18.setHorizontalHeaderItem(1, item)
        item = QtWidgets.QTableWidgetItem()
        self.tableWidget_18.setHorizontalHeaderItem(2, item)
        item = QtWidgets.QTableWidgetItem()
        self.tableWidget_18.setHorizontalHeaderItem(3, item)
        item = QtWidgets.QTableWidgetItem()
        self.tableWidget_18.setHorizontalHeaderItem(4, item)
        item = QtWidgets.QTableWidgetItem()
        self.tableWidget_18.setHorizontalHeaderItem(5, item)
        item = QtWidgets.QTableWidgetItem()
        self.tableWidget_18.setHorizontalHeaderItem(6, item)
        item = QtWidgets.QTableWidgetItem()
        self.tableWidget_18.setHorizontalHeaderItem(7, item)
        self.label_74 = QtWidgets.QLabel(self.page_24)
        self.label_74.setGeometry(QtCore.QRect(620, 800, 91, 21))
        self.label_74.setObjectName("label_74")
        self.lck_odds_10 = QtWidgets.QTextEdit(self.page_24)
        self.lck_odds_10.setGeometry(QtCore.QRect(670, 800, 71, 21))
        self.lck_odds_10.setObjectName("lck_odds_10")
        self.lck_matches_10 = QtWidgets.QTableWidget(self.page_24)
        self.lck_matches_10.setGeometry(QtCore.QRect(90, 500, 1141, 281))
        self.lck_matches_10.setObjectName("lck_matches_10")
        self.lck_matches_10.setColumnCount(9)
        self.lck_matches_10.setRowCount(0)
        item = QtWidgets.QTableWidgetItem()
        self.lck_matches_10.setHorizontalHeaderItem(0, item)
        item = QtWidgets.QTableWidgetItem()
        self.lck_matches_10.setHorizontalHeaderItem(1, item)
        item = QtWidgets.QTableWidgetItem()
        self.lck_matches_10.setHorizontalHeaderItem(2, item)
        item = QtWidgets.QTableWidgetItem()
        self.lck_matches_10.setHorizontalHeaderItem(3, item)
        item = QtWidgets.QTableWidgetItem()
        self.lck_matches_10.setHorizontalHeaderItem(4, item)
        item = QtWidgets.QTableWidgetItem()
        self.lck_matches_10.setHorizontalHeaderItem(5, item)
        item = QtWidgets.QTableWidgetItem()
        self.lck_matches_10.setHorizontalHeaderItem(6, item)
        item = QtWidgets.QTableWidgetItem()
        self.lck_matches_10.setHorizontalHeaderItem(7, item)
        item = QtWidgets.QTableWidgetItem()
        self.lck_matches_10.setHorizontalHeaderItem(8, item)
        self.label_75 = QtWidgets.QLabel(self.page_24)
        self.label_75.setGeometry(QtCore.QRect(1080, 20, 361, 61))
        self.label_75.setObjectName("label_75")
        self.stackedWidget.addWidget(self.page_24)
        MainWindow.setCentralWidget(self.centralwidget)
        self.menubar = QtWidgets.QMenuBar(MainWindow)
        self.menubar.setGeometry(QtCore.QRect(0, 0, 1301, 21))
        self.menubar.setObjectName("menubar")
        MainWindow.setMenuBar(self.menubar)
        self.statusbar = QtWidgets.QStatusBar(MainWindow)
        self.statusbar.setObjectName("statusbar")
        MainWindow.setStatusBar(self.statusbar)

        # button links to created functions
        # load data associated with the selected page
        self.pushButton.clicked.connect(lambda: self.get_lck())
        self.lcs.clicked.connect(lambda: self.get_lcs())
        self.lec.clicked.connect(lambda: self.get_lec())
        self.vcs.clicked.connect(lambda: self.get_vcs())
        self.lpl.clicked.connect(lambda: self.get_lpl())
        
        # update data in current table based on button clicked
        self.update_lck.clicked.connect(lambda: lck_data())
        self.update_lcs.clicked.connect(lambda: lcs_data())
        self.update_lec.clicked.connect(lambda: lec_data())
        self.update_vcs.clicked.connect(lambda: vcs_data())
        self.update_lpl.clicked.connect(lambda: lpl_data())
        
        # navigate back to main page from other pages
        self.back.clicked.connect(lambda: self.stackedWidget.setCurrentWidget(self.page))
        self.back_2.clicked.connect(lambda: self.stackedWidget.setCurrentWidget(self.page))
        self.back_3.clicked.connect(lambda: self.stackedWidget.setCurrentWidget(self.page))
        self.back_13.clicked.connect(lambda: self.stackedWidget.setCurrentWidget(self.page))
        self.back_14.clicked.connect(lambda: self.stackedWidget.setCurrentWidget(self.page))

        # alter the line/odds used in the calculation and refresh the upcoming games table
        self.lck_calculate.clicked.connect(lambda: self.calc_lck())
        self.lec_calculate.clicked.connect(lambda: self.calc_lec())
        self.lcs_calculate.clicked.connect(lambda: self.calc_lcs())
        self.vcs_calculate.clicked.connect(lambda: self.calc_vcs())
        self.lpl_calculate.clicked.connect(lambda: self.calc_lpl())

        # scrape the upcoming games and insert them into a table upon start up
        next_games = UpcomingGames()
        comp, match = next_games.next_games()
        self.insert_data(comp, 0, self.upcoming_table)
        self.insert_data(match, 1, self.upcoming_table)
        # resize the table for ease of use
        header = self.upcoming_table.horizontalHeader()       
        header.setSectionResizeMode(0, QtWidgets.QHeaderView.ResizeToContents)
        header.setSectionResizeMode(1, QtWidgets.QHeaderView.Stretch)

        self.retranslateUi(MainWindow)
        QtCore.QMetaObject.connectSlotsByName(MainWindow)

    def calc_lck(self):
        line = float(self.lck_line.toPlainText())
        odds = float(self.lck_odds.toPlainText())
        self.load_data('lck', self.tableWidget, self.page_2,'https://lol.fandom.com/wiki/LCK/2022_Season/Spring_Season', self.ladder_1, 'https://www.rivalry.com/esports/league-of-legends-betting/3254-champions-korea', self.lck_matches, line, odds)

    def calc_lcs(self):
        line = float(self.lcs_line.toPlainText())
        odds = float(self.lcs_odds.toPlainText())
        self.load_data('lcs', self.tableWidget_2, self.page_3,'https://lol.fandom.com/wiki/LCS/2022_Season/Spring_Season', self.ladder_2, 'https://www.rivalry.com/esports/league-of-legends-betting/3713-lcs-north-america', self.lcs_matches, line, odds)
    
    def calc_lec(self):
        line = float(self.lec_line.toPlainText())
        odds = float(self.lec_odds.toPlainText())
        self.load_data('lec', self.tableWidget_3, self.page_4,'https://lol.fandom.com/wiki/LEC/2022_Season/Spring_Season', self.ladder_3, 'https://www.rivalry.com/esports/league-of-legends-betting/3282-european-championship', self.lec_matches, line, odds)

    def calc_vcs(self):
        line = float(self.vcs_line.toPlainText())
        odds = float(self.vcs_odds.toPlainText())
        self.load_data('vcs', self.vcs_stats_table, self.page_vcs,'https://lol.fandom.com/wiki/VCS/2022_Season/Spring_Season', self.ladder_vcs, 'https://www.rivalry.com/esports/league-of-legends-betting/3296-vcs', self.vcs_matches, line, odds)

    def calc_lpl(self):
        line = float(self.lpl_line.toPlainText())
        odds = float(self.lpl_odds.toPlainText())
        self.load_data('lpl', self.lpl_stats_table, self.page_5,'https://lol.fandom.com/wiki/LPL/2022_Season/Spring_Season', self.ladder_lpl, 'https://www.rivalry.com/esports/league-of-legends-betting/2762-china-lpl', self.lpl_matches, line, odds)

    def get_lck(self):
        self.load_data('lck', self.tableWidget, self.page_2,'https://lol.fandom.com/wiki/LCK/2022_Season/Spring_Season', self.ladder_1, 'https://www.rivalry.com/esports/league-of-legends-betting/3254-champions-korea', self.lck_matches, 22.5)

    def get_lcs(self):
        self.load_data('lcs', self.tableWidget_2, self.page_3, 'https://lol.fandom.com/wiki/LCS/2022_Season/Spring_Season', self.ladder_2, 'https://www.rivalry.com/esports/league-of-legends-betting/3713-lcs-north-america', self.lcs_matches, 24.5)

    def get_lec(self):
        self.load_data('lec', self.tableWidget_3, self.page_4, 'https://lol.fandom.com/wiki/LEC/2022_Season/Spring_Season', self.ladder_3, 'https://www.rivalry.com/esports/league-of-legends-betting/3282-european-championship', self.lec_matches, 25.5)

    def get_vcs(self):
        self.load_data('vcs', self.vcs_stats_table, self.page_vcs,'https://lol.fandom.com/wiki/VCS/2022_Season/Spring_Season', self.ladder_vcs, 'https://www.rivalry.com/esports/league-of-legends-betting/3296-vcs', self.vcs_matches, 31.5)

    def get_lpl(self):
        self.load_data('lpl', self.lpl_stats_table, self.page_5,'https://lol.fandom.com/wiki/LPL/2022_Season/Spring_Season', self.ladder_lpl, 'https://www.rivalry.com/esports/league-of-legends-betting/2762-china-lpl', self.lpl_matches, 25.5)

    def retranslateUi(self, MainWindow):
        _translate = QtCore.QCoreApplication.translate
        MainWindow.setWindowTitle(_translate("MainWindow", "MainWindow"))
        self.label.setText(_translate("MainWindow", "<html><head/><body><p><span style=\" font-size:16pt;\">League of Legends Kill Total Gambling</span></p></body></html>"))
        self.pushButton.setText(_translate("MainWindow", "LCK"))
        self.lcs.setText(_translate("MainWindow", "LCS"))
        self.lec.setText(_translate("MainWindow", "LEC"))
        self.vcs.setText(_translate("MainWindow", "VCS"))
        self.lec_6.setText(_translate("MainWindow", "x"))
        self.lec_7.setText(_translate("MainWindow", "x"))
        self.lec_8.setText(_translate("MainWindow", "x"))
        self.lec_9.setText(_translate("MainWindow", "x"))
        self.lpl.setText(_translate("MainWindow", "LPL"))
        self.lec_11.setText(_translate("MainWindow", "x"))
        self.lec_12.setText(_translate("MainWindow", "x"))
        self.lec_13.setText(_translate("MainWindow", "x"))
        item = self.upcoming_table.horizontalHeaderItem(0)
        item.setText(_translate("MainWindow", "League"))
        item = self.upcoming_table.horizontalHeaderItem(1)
        item.setText(_translate("MainWindow", "Match"))
        self.label_77.setText(_translate("MainWindow", "<html><head/><body><p><span style=\" font-size:14pt;\">Upcoming Games</span></p></body></html>"))
        self.label_2.setText(_translate("MainWindow", "<html><head/><body><p><span style=\" font-size:16pt;\">LCK Statistics</span></p></body></html>"))
        item = self.tableWidget.horizontalHeaderItem(0)
        item.setText(_translate("MainWindow", "Team Name"))
        item = self.tableWidget.horizontalHeaderItem(1)
        item.setText(_translate("MainWindow", "Matches"))
        item = self.tableWidget.horizontalHeaderItem(2)
        item.setText(_translate("MainWindow", "Avg Game 1"))
        item = self.tableWidget.horizontalHeaderItem(3)
        item.setText(_translate("MainWindow", "% Game 1"))
        item = self.tableWidget.horizontalHeaderItem(4)
        item.setText(_translate("MainWindow", "Avg Game 2"))
        item = self.tableWidget.horizontalHeaderItem(5)
        item.setText(_translate("MainWindow", "% Game 2"))
        item = self.tableWidget.horizontalHeaderItem(6)
        item.setText(_translate("MainWindow", "Avg Total"))
        item = self.tableWidget.horizontalHeaderItem(7)
        item.setText(_translate("MainWindow", "% Total"))
        self.update_lck.setText(_translate("MainWindow", "Update"))
        self.back.setText(_translate("MainWindow", "Back"))
        item = self.ladder_1.horizontalHeaderItem(0)
        item.setText(_translate("MainWindow", "Team"))
        item = self.ladder_1.horizontalHeaderItem(1)
        item.setText(_translate("MainWindow", "Record"))
        self.label_4.setText(_translate("MainWindow", "<html><head/><body><p><span style=\" font-size:16pt;\">Ladder</span></p></body></html>"))
        item = self.lck_matches.horizontalHeaderItem(0)
        item.setText(_translate("MainWindow", "Match"))
        item = self.lck_matches.horizontalHeaderItem(1)
        item.setText(_translate("MainWindow", "Date"))
        item = self.lck_matches.horizontalHeaderItem(2)
        item.setText(_translate("MainWindow", "Line"))
        item = self.lck_matches.horizontalHeaderItem(3)
        item.setText(_translate("MainWindow", "G1 Avg"))
        item = self.lck_matches.horizontalHeaderItem(4)
        item.setText(_translate("MainWindow", "G1 %"))
        item = self.lck_matches.horizontalHeaderItem(5)
        item.setText(_translate("MainWindow", "G1 Value"))
        item = self.lck_matches.horizontalHeaderItem(6)
        item.setText(_translate("MainWindow", "G2 Avg"))
        item = self.lck_matches.horizontalHeaderItem(7)
        item.setText(_translate("MainWindow", "G2 %"))
        item = self.lck_matches.horizontalHeaderItem(8)
        item.setText(_translate("MainWindow", "G2 Value"))
        self.label_8.setText(_translate("MainWindow", "<html><head/><body><p><span style=\" font-size:12pt;\">Line:</span></p></body></html>"))
        self.label_9.setText(_translate("MainWindow", "<html><head/><body><p><span style=\" font-size:12pt;\">Odds:</span></p></body></html>"))
        self.lck_calculate.setText(_translate("MainWindow", "Calculate"))
        item = self.tableWidget_2.horizontalHeaderItem(0)
        item.setText(_translate("MainWindow", "Team Name"))
        item = self.tableWidget_2.horizontalHeaderItem(1)
        item.setText(_translate("MainWindow", "Matches"))
        item = self.tableWidget_2.horizontalHeaderItem(2)
        item.setText(_translate("MainWindow", "Avg Game 1"))
        item = self.tableWidget_2.horizontalHeaderItem(3)
        item.setText(_translate("MainWindow", "% Game 1"))
        item = self.tableWidget_2.horizontalHeaderItem(4)
        item.setText(_translate("MainWindow", "Avg Game 2"))
        item = self.tableWidget_2.horizontalHeaderItem(5)
        item.setText(_translate("MainWindow", "% Game 2"))
        item = self.tableWidget_2.horizontalHeaderItem(6)
        item.setText(_translate("MainWindow", "Avg Total"))
        item = self.tableWidget_2.horizontalHeaderItem(7)
        item.setText(_translate("MainWindow", "% Total"))
        self.label_3.setText(_translate("MainWindow", "<html><head/><body><p><span style=\" font-size:16pt;\">LCS Statistics</span></p></body></html>"))
        self.back_2.setText(_translate("MainWindow", "Back"))
        self.update_lcs.setText(_translate("MainWindow", "Update"))
        item = self.ladder_2.horizontalHeaderItem(0)
        item.setText(_translate("MainWindow", "Team"))
        item = self.ladder_2.horizontalHeaderItem(1)
        item.setText(_translate("MainWindow", "Record"))
        self.label_5.setText(_translate("MainWindow", "<html><head/><body><p><span style=\" font-size:16pt;\">Ladder</span></p></body></html>"))
        item = self.lcs_matches.horizontalHeaderItem(0)
        item.setText(_translate("MainWindow", "Match"))
        item = self.lcs_matches.horizontalHeaderItem(1)
        item.setText(_translate("MainWindow", "Date"))
        item = self.lcs_matches.horizontalHeaderItem(2)
        item.setText(_translate("MainWindow", "Line"))
        item = self.lcs_matches.horizontalHeaderItem(3)
        item.setText(_translate("MainWindow", "G1 Avg"))
        item = self.lcs_matches.horizontalHeaderItem(4)
        item.setText(_translate("MainWindow", "G1 %"))
        item = self.lcs_matches.horizontalHeaderItem(5)
        item.setText(_translate("MainWindow", "G1 Value"))
        item = self.lcs_matches.horizontalHeaderItem(6)
        item.setText(_translate("MainWindow", "G2 Avg"))
        item = self.lcs_matches.horizontalHeaderItem(7)
        item.setText(_translate("MainWindow", "G2 %"))
        item = self.lcs_matches.horizontalHeaderItem(8)
        item.setText(_translate("MainWindow", "G2 Value"))
        self.lcs_calculate.setText(_translate("MainWindow", "Calculate"))
        self.label_10.setText(_translate("MainWindow", "<html><head/><body><p><span style=\" font-size:12pt;\">Odds:</span></p></body></html>"))
        self.label_14.setText(_translate("MainWindow", "<html><head/><body><p><span style=\" font-size:12pt;\">Line:</span></p></body></html>"))
        item = self.tableWidget_3.horizontalHeaderItem(0)
        item.setText(_translate("MainWindow", "Team Name"))
        item = self.tableWidget_3.horizontalHeaderItem(1)
        item.setText(_translate("MainWindow", "Matches"))
        item = self.tableWidget_3.horizontalHeaderItem(2)
        item.setText(_translate("MainWindow", "Avg Game 1"))
        item = self.tableWidget_3.horizontalHeaderItem(3)
        item.setText(_translate("MainWindow", "% Game 1"))
        item = self.tableWidget_3.horizontalHeaderItem(4)
        item.setText(_translate("MainWindow", "Avg Game 2"))
        item = self.tableWidget_3.horizontalHeaderItem(5)
        item.setText(_translate("MainWindow", "% Game 2"))
        item = self.tableWidget_3.horizontalHeaderItem(6)
        item.setText(_translate("MainWindow", "Avg Total"))
        item = self.tableWidget_3.horizontalHeaderItem(7)
        item.setText(_translate("MainWindow", "% Total"))
        self.label_6.setText(_translate("MainWindow", "<html><head/><body><p><span style=\" font-size:16pt;\">LEC Statistics</span></p></body></html>"))
        self.back_3.setText(_translate("MainWindow", "Back"))
        self.update_lec.setText(_translate("MainWindow", "Update"))
        self.label_7.setText(_translate("MainWindow", "<html><head/><body><p><span style=\" font-size:16pt;\">Ladder</span></p></body></html>"))
        item = self.ladder_3.horizontalHeaderItem(0)
        item.setText(_translate("MainWindow", "Team"))
        item = self.ladder_3.horizontalHeaderItem(1)
        item.setText(_translate("MainWindow", "Record"))
        item = self.lec_matches.horizontalHeaderItem(0)
        item.setText(_translate("MainWindow", "Match"))
        item = self.lec_matches.horizontalHeaderItem(1)
        item.setText(_translate("MainWindow", "Date"))
        item = self.lec_matches.horizontalHeaderItem(2)
        item.setText(_translate("MainWindow", "Line"))
        item = self.lec_matches.horizontalHeaderItem(3)
        item.setText(_translate("MainWindow", "G1 Avg"))
        item = self.lec_matches.horizontalHeaderItem(4)
        item.setText(_translate("MainWindow", "G1 %"))
        item = self.lec_matches.horizontalHeaderItem(5)
        item.setText(_translate("MainWindow", "G1 Value"))
        item = self.lec_matches.horizontalHeaderItem(6)
        item.setText(_translate("MainWindow", "G2 Avg"))
        item = self.lec_matches.horizontalHeaderItem(7)
        item.setText(_translate("MainWindow", "G2 %"))
        item = self.lec_matches.horizontalHeaderItem(8)
        item.setText(_translate("MainWindow", "G2 Value"))
        self.lec_calculate.setText(_translate("MainWindow", "Calculate"))
        self.label_11.setText(_translate("MainWindow", "<html><head/><body><p><span style=\" font-size:12pt;\">Odds:</span></p></body></html>"))
        self.label_12.setText(_translate("MainWindow", "<html><head/><body><p><span style=\" font-size:12pt;\">Line:</span></p></body></html>"))
        self.label_53.setText(_translate("MainWindow", "<html><head/><body><p><span style=\" font-size:16pt;\">LPL Statistics</span></p></body></html>"))
        self.label_54.setText(_translate("MainWindow", "<html><head/><body><p><span style=\" font-size:12pt;\">Odds:</span></p></body></html>"))
        item = self.lpl_stats_table.horizontalHeaderItem(0)
        item.setText(_translate("MainWindow", "Team Name"))
        item = self.lpl_stats_table.horizontalHeaderItem(1)
        item.setText(_translate("MainWindow", "Matches"))
        item = self.lpl_stats_table.horizontalHeaderItem(2)
        item.setText(_translate("MainWindow", "Avg Game 1"))
        item = self.lpl_stats_table.horizontalHeaderItem(3)
        item.setText(_translate("MainWindow", "% Game 1"))
        item = self.lpl_stats_table.horizontalHeaderItem(4)
        item.setText(_translate("MainWindow", "Avg Game 2"))
        item = self.lpl_stats_table.horizontalHeaderItem(5)
        item.setText(_translate("MainWindow", "% Game 2"))
        item = self.lpl_stats_table.horizontalHeaderItem(6)
        item.setText(_translate("MainWindow", "Avg Total"))
        item = self.lpl_stats_table.horizontalHeaderItem(7)
        item.setText(_translate("MainWindow", "% Total"))
        self.label_55.setText(_translate("MainWindow", "<html><head/><body><p><span style=\" font-size:16pt;\">Ladder</span></p></body></html>"))
        item = self.ladder_lpl.horizontalHeaderItem(0)
        item.setText(_translate("MainWindow", "Team"))
        item = self.ladder_lpl.horizontalHeaderItem(1)
        item.setText(_translate("MainWindow", "Record"))
        item = self.lpl_matches.horizontalHeaderItem(0)
        item.setText(_translate("MainWindow", "Match"))
        item = self.lpl_matches.horizontalHeaderItem(1)
        item.setText(_translate("MainWindow", "Date"))
        item = self.lpl_matches.horizontalHeaderItem(2)
        item.setText(_translate("MainWindow", "Line"))
        item = self.lpl_matches.horizontalHeaderItem(3)
        item.setText(_translate("MainWindow", "G1 Avg"))
        item = self.lpl_matches.horizontalHeaderItem(4)
        item.setText(_translate("MainWindow", "G1 %"))
        item = self.lpl_matches.horizontalHeaderItem(5)
        item.setText(_translate("MainWindow", "G1 Value"))
        item = self.lpl_matches.horizontalHeaderItem(6)
        item.setText(_translate("MainWindow", "G2 Avg"))
        item = self.lpl_matches.horizontalHeaderItem(7)
        item.setText(_translate("MainWindow", "G2 %"))
        item = self.lpl_matches.horizontalHeaderItem(8)
        item.setText(_translate("MainWindow", "G2 Value"))
        self.lpl_calculate.setText(_translate("MainWindow", "Calculate"))
        self.label_56.setText(_translate("MainWindow", "<html><head/><body><p><span style=\" font-size:12pt;\">Line:</span></p></body></html>"))
        self.back_13.setText(_translate("MainWindow", "Back"))
        self.update_lpl.setText(_translate("MainWindow", "Update"))
        self.update_vcs.setText(_translate("MainWindow", "Update"))
        self.vcs_calculate.setText(_translate("MainWindow", "Calculate"))
        self.label_60.setText(_translate("MainWindow", "<html><head/><body><p><span style=\" font-size:12pt;\">Line:</span></p></body></html>"))
        self.back_14.setText(_translate("MainWindow", "Back"))
        self.label_57.setText(_translate("MainWindow", "<html><head/><body><p><span style=\" font-size:16pt;\">VCS Statistics</span></p></body></html>"))
        self.label_58.setText(_translate("MainWindow", "<html><head/><body><p><span style=\" font-size:12pt;\">Odds:</span></p></body></html>"))
        item = self.ladder_vcs.horizontalHeaderItem(0)
        item.setText(_translate("MainWindow", "Team"))
        item = self.ladder_vcs.horizontalHeaderItem(1)
        item.setText(_translate("MainWindow", "Record"))
        item = self.vcs_stats_table.horizontalHeaderItem(0)
        item.setText(_translate("MainWindow", "Team Name"))
        item = self.vcs_stats_table.horizontalHeaderItem(1)
        item.setText(_translate("MainWindow", "Matches"))
        item = self.vcs_stats_table.horizontalHeaderItem(2)
        item.setText(_translate("MainWindow", "Avg Game 1"))
        item = self.vcs_stats_table.horizontalHeaderItem(3)
        item.setText(_translate("MainWindow", "% Game 1"))
        item = self.vcs_stats_table.horizontalHeaderItem(4)
        item.setText(_translate("MainWindow", "Avg Game 2"))
        item = self.vcs_stats_table.horizontalHeaderItem(5)
        item.setText(_translate("MainWindow", "% Game 2"))
        item = self.vcs_stats_table.horizontalHeaderItem(6)
        item.setText(_translate("MainWindow", "Avg Total"))
        item = self.vcs_stats_table.horizontalHeaderItem(7)
        item.setText(_translate("MainWindow", "% Total"))
        self.label_59.setText(_translate("MainWindow", "<html><head/><body><p><span style=\" font-size:16pt;\">Ladder</span></p></body></html>"))
        item = self.vcs_matches.horizontalHeaderItem(0)
        item.setText(_translate("MainWindow", "Match"))
        item = self.vcs_matches.horizontalHeaderItem(1)
        item.setText(_translate("MainWindow", "Date"))
        item = self.vcs_matches.horizontalHeaderItem(2)
        item.setText(_translate("MainWindow", "Line"))
        item = self.vcs_matches.horizontalHeaderItem(3)
        item.setText(_translate("MainWindow", "G1 Avg"))
        item = self.vcs_matches.horizontalHeaderItem(4)
        item.setText(_translate("MainWindow", "G1 %"))
        item = self.vcs_matches.horizontalHeaderItem(5)
        item.setText(_translate("MainWindow", "G1 Value"))
        item = self.vcs_matches.horizontalHeaderItem(6)
        item.setText(_translate("MainWindow", "G2 Avg"))
        item = self.vcs_matches.horizontalHeaderItem(7)
        item.setText(_translate("MainWindow", "G2 %"))
        item = self.vcs_matches.horizontalHeaderItem(8)
        item.setText(_translate("MainWindow", "G2 Value"))
        self.update_lck_7.setText(_translate("MainWindow", "Update"))
        item = self.ladder_15.horizontalHeaderItem(0)
        item.setText(_translate("MainWindow", "Team"))
        item = self.ladder_15.horizontalHeaderItem(1)
        item.setText(_translate("MainWindow", "Record"))
        self.label_62.setText(_translate("MainWindow", "<html><head/><body><p><span style=\" font-size:12pt;\">Odds:</span></p></body></html>"))
        item = self.tableWidget_15.horizontalHeaderItem(0)
        item.setText(_translate("MainWindow", "Team Name"))
        item = self.tableWidget_15.horizontalHeaderItem(1)
        item.setText(_translate("MainWindow", "Matches"))
        item = self.tableWidget_15.horizontalHeaderItem(2)
        item.setText(_translate("MainWindow", "Avg Game 1"))
        item = self.tableWidget_15.horizontalHeaderItem(3)
        item.setText(_translate("MainWindow", "% Game 1"))
        item = self.tableWidget_15.horizontalHeaderItem(4)
        item.setText(_translate("MainWindow", "Avg Game 2"))
        item = self.tableWidget_15.horizontalHeaderItem(5)
        item.setText(_translate("MainWindow", "% Game 2"))
        item = self.tableWidget_15.horizontalHeaderItem(6)
        item.setText(_translate("MainWindow", "Avg Total"))
        item = self.tableWidget_15.horizontalHeaderItem(7)
        item.setText(_translate("MainWindow", "% Total"))
        self.back_15.setText(_translate("MainWindow", "Back"))
        self.label_61.setText(_translate("MainWindow", "<html><head/><body><p><span style=\" font-size:16pt;\">LCK Statistics</span></p></body></html>"))
        item = self.lck_matches_7.horizontalHeaderItem(0)
        item.setText(_translate("MainWindow", "Match"))
        item = self.lck_matches_7.horizontalHeaderItem(1)
        item.setText(_translate("MainWindow", "Date"))
        item = self.lck_matches_7.horizontalHeaderItem(2)
        item.setText(_translate("MainWindow", "Line"))
        item = self.lck_matches_7.horizontalHeaderItem(3)
        item.setText(_translate("MainWindow", "G1 Avg"))
        item = self.lck_matches_7.horizontalHeaderItem(4)
        item.setText(_translate("MainWindow", "G1 %"))
        item = self.lck_matches_7.horizontalHeaderItem(5)
        item.setText(_translate("MainWindow", "G1 Value"))
        item = self.lck_matches_7.horizontalHeaderItem(6)
        item.setText(_translate("MainWindow", "G2 Avg"))
        item = self.lck_matches_7.horizontalHeaderItem(7)
        item.setText(_translate("MainWindow", "G2 %"))
        item = self.lck_matches_7.horizontalHeaderItem(8)
        item.setText(_translate("MainWindow", "G2 Value"))
        self.label_63.setText(_translate("MainWindow", "<html><head/><body><p><span style=\" font-size:16pt;\">Ladder</span></p></body></html>"))
        self.label_64.setText(_translate("MainWindow", "<html><head/><body><p><span style=\" font-size:12pt;\">Line:</span></p></body></html>"))
        self.lck_calculate_7.setText(_translate("MainWindow", "Calculate"))
        self.update_lck_8.setText(_translate("MainWindow", "Update"))
        self.label_65.setText(_translate("MainWindow", "<html><head/><body><p><span style=\" font-size:16pt;\">LCK Statistics</span></p></body></html>"))
        self.label_66.setText(_translate("MainWindow", "<html><head/><body><p><span style=\" font-size:12pt;\">Odds:</span></p></body></html>"))
        self.back_16.setText(_translate("MainWindow", "Back"))
        item = self.tableWidget_16.horizontalHeaderItem(0)
        item.setText(_translate("MainWindow", "Team Name"))
        item = self.tableWidget_16.horizontalHeaderItem(1)
        item.setText(_translate("MainWindow", "Matches"))
        item = self.tableWidget_16.horizontalHeaderItem(2)
        item.setText(_translate("MainWindow", "Avg Game 1"))
        item = self.tableWidget_16.horizontalHeaderItem(3)
        item.setText(_translate("MainWindow", "% Game 1"))
        item = self.tableWidget_16.horizontalHeaderItem(4)
        item.setText(_translate("MainWindow", "Avg Game 2"))
        item = self.tableWidget_16.horizontalHeaderItem(5)
        item.setText(_translate("MainWindow", "% Game 2"))
        item = self.tableWidget_16.horizontalHeaderItem(6)
        item.setText(_translate("MainWindow", "Avg Total"))
        item = self.tableWidget_16.horizontalHeaderItem(7)
        item.setText(_translate("MainWindow", "% Total"))
        self.lck_calculate_8.setText(_translate("MainWindow", "Calculate"))
        item = self.ladder_16.horizontalHeaderItem(0)
        item.setText(_translate("MainWindow", "Team"))
        item = self.ladder_16.horizontalHeaderItem(1)
        item.setText(_translate("MainWindow", "Record"))
        self.label_67.setText(_translate("MainWindow", "<html><head/><body><p><span style=\" font-size:16pt;\">Ladder</span></p></body></html>"))
        self.label_68.setText(_translate("MainWindow", "<html><head/><body><p><span style=\" font-size:12pt;\">Line:</span></p></body></html>"))
        item = self.lck_matches_8.horizontalHeaderItem(0)
        item.setText(_translate("MainWindow", "Match"))
        item = self.lck_matches_8.horizontalHeaderItem(1)
        item.setText(_translate("MainWindow", "Date"))
        item = self.lck_matches_8.horizontalHeaderItem(2)
        item.setText(_translate("MainWindow", "Line"))
        item = self.lck_matches_8.horizontalHeaderItem(3)
        item.setText(_translate("MainWindow", "G1 Avg"))
        item = self.lck_matches_8.horizontalHeaderItem(4)
        item.setText(_translate("MainWindow", "G1 %"))
        item = self.lck_matches_8.horizontalHeaderItem(5)
        item.setText(_translate("MainWindow", "G1 Value"))
        item = self.lck_matches_8.horizontalHeaderItem(6)
        item.setText(_translate("MainWindow", "G2 Avg"))
        item = self.lck_matches_8.horizontalHeaderItem(7)
        item.setText(_translate("MainWindow", "G2 %"))
        item = self.lck_matches_8.horizontalHeaderItem(8)
        item.setText(_translate("MainWindow", "G2 Value"))
        self.update_lck_9.setText(_translate("MainWindow", "Update"))
        item = self.ladder_17.horizontalHeaderItem(0)
        item.setText(_translate("MainWindow", "Team"))
        item = self.ladder_17.horizontalHeaderItem(1)
        item.setText(_translate("MainWindow", "Record"))
        self.back_17.setText(_translate("MainWindow", "Back"))
        self.label_69.setText(_translate("MainWindow", "<html><head/><body><p><span style=\" font-size:16pt;\">LCK Statistics</span></p></body></html>"))
        item = self.tableWidget_17.horizontalHeaderItem(0)
        item.setText(_translate("MainWindow", "Team Name"))
        item = self.tableWidget_17.horizontalHeaderItem(1)
        item.setText(_translate("MainWindow", "Matches"))
        item = self.tableWidget_17.horizontalHeaderItem(2)
        item.setText(_translate("MainWindow", "Avg Game 1"))
        item = self.tableWidget_17.horizontalHeaderItem(3)
        item.setText(_translate("MainWindow", "% Game 1"))
        item = self.tableWidget_17.horizontalHeaderItem(4)
        item.setText(_translate("MainWindow", "Avg Game 2"))
        item = self.tableWidget_17.horizontalHeaderItem(5)
        item.setText(_translate("MainWindow", "% Game 2"))
        item = self.tableWidget_17.horizontalHeaderItem(6)
        item.setText(_translate("MainWindow", "Avg Total"))
        item = self.tableWidget_17.horizontalHeaderItem(7)
        item.setText(_translate("MainWindow", "% Total"))
        self.label_70.setText(_translate("MainWindow", "<html><head/><body><p><span style=\" font-size:12pt;\">Odds:</span></p></body></html>"))
        self.label_71.setText(_translate("MainWindow", "<html><head/><body><p><span style=\" font-size:16pt;\">Ladder</span></p></body></html>"))
        self.label_72.setText(_translate("MainWindow", "<html><head/><body><p><span style=\" font-size:12pt;\">Line:</span></p></body></html>"))
        item = self.lck_matches_9.horizontalHeaderItem(0)
        item.setText(_translate("MainWindow", "Match"))
        item = self.lck_matches_9.horizontalHeaderItem(1)
        item.setText(_translate("MainWindow", "Date"))
        item = self.lck_matches_9.horizontalHeaderItem(2)
        item.setText(_translate("MainWindow", "Line"))
        item = self.lck_matches_9.horizontalHeaderItem(3)
        item.setText(_translate("MainWindow", "G1 Avg"))
        item = self.lck_matches_9.horizontalHeaderItem(4)
        item.setText(_translate("MainWindow", "G1 %"))
        item = self.lck_matches_9.horizontalHeaderItem(5)
        item.setText(_translate("MainWindow", "G1 Value"))
        item = self.lck_matches_9.horizontalHeaderItem(6)
        item.setText(_translate("MainWindow", "G2 Avg"))
        item = self.lck_matches_9.horizontalHeaderItem(7)
        item.setText(_translate("MainWindow", "G2 %"))
        item = self.lck_matches_9.horizontalHeaderItem(8)
        item.setText(_translate("MainWindow", "G2 Value"))
        self.lck_calculate_9.setText(_translate("MainWindow", "Calculate"))
        self.update_lck_10.setText(_translate("MainWindow", "Update"))
        self.label_76.setText(_translate("MainWindow", "<html><head/><body><p><span style=\" font-size:12pt;\">Line:</span></p></body></html>"))
        item = self.ladder_18.horizontalHeaderItem(0)
        item.setText(_translate("MainWindow", "Team"))
        item = self.ladder_18.horizontalHeaderItem(1)
        item.setText(_translate("MainWindow", "Record"))
        self.back_18.setText(_translate("MainWindow", "Back"))
        self.label_73.setText(_translate("MainWindow", "<html><head/><body><p><span style=\" font-size:16pt;\">LCK Statistics</span></p></body></html>"))
        self.lck_calculate_10.setText(_translate("MainWindow", "Calculate"))
        item = self.tableWidget_18.horizontalHeaderItem(0)
        item.setText(_translate("MainWindow", "Team Name"))
        item = self.tableWidget_18.horizontalHeaderItem(1)
        item.setText(_translate("MainWindow", "Matches"))
        item = self.tableWidget_18.horizontalHeaderItem(2)
        item.setText(_translate("MainWindow", "Avg Game 1"))
        item = self.tableWidget_18.horizontalHeaderItem(3)
        item.setText(_translate("MainWindow", "% Game 1"))
        item = self.tableWidget_18.horizontalHeaderItem(4)
        item.setText(_translate("MainWindow", "Avg Game 2"))
        item = self.tableWidget_18.horizontalHeaderItem(5)
        item.setText(_translate("MainWindow", "% Game 2"))
        item = self.tableWidget_18.horizontalHeaderItem(6)
        item.setText(_translate("MainWindow", "Avg Total"))
        item = self.tableWidget_18.horizontalHeaderItem(7)
        item.setText(_translate("MainWindow", "% Total"))
        self.label_74.setText(_translate("MainWindow", "<html><head/><body><p><span style=\" font-size:12pt;\">Odds:</span></p></body></html>"))
        item = self.lck_matches_10.horizontalHeaderItem(0)
        item.setText(_translate("MainWindow", "Match"))
        item = self.lck_matches_10.horizontalHeaderItem(1)
        item.setText(_translate("MainWindow", "Date"))
        item = self.lck_matches_10.horizontalHeaderItem(2)
        item.setText(_translate("MainWindow", "Line"))
        item = self.lck_matches_10.horizontalHeaderItem(3)
        item.setText(_translate("MainWindow", "G1 Avg"))
        item = self.lck_matches_10.horizontalHeaderItem(4)
        item.setText(_translate("MainWindow", "G1 %"))
        item = self.lck_matches_10.horizontalHeaderItem(5)
        item.setText(_translate("MainWindow", "G1 Value"))
        item = self.lck_matches_10.horizontalHeaderItem(6)
        item.setText(_translate("MainWindow", "G2 Avg"))
        item = self.lck_matches_10.horizontalHeaderItem(7)
        item.setText(_translate("MainWindow", "G2 %"))
        item = self.lck_matches_10.horizontalHeaderItem(8)
        item.setText(_translate("MainWindow", "G2 Value"))
        self.label_75.setText(_translate("MainWindow", "<html><head/><body><p><span style=\" font-size:16pt;\">Ladder</span></p></body></html>"))


    # navigate to page and fill it with data relating to the given leagues statistics, rankings and upcomning games
    def load_data(self, competition, table, page, url, ladder_table, url_2, upcoming_table, line = 23.5, odds = 1.83):   
        # change pages if it has not been done
        try:
            self.stackedWidget.setCurrentWidget(page)
        except:
            pass

        # load the ladder upon opening/updating the page
        league = Ladder()
        ladder_details = league.ladder(url, competition)       
        self.insert_data(ladder_details[0], 0, ladder_table)
        self.insert_data(ladder_details[1], 1, ladder_table)

        # load data for the upcoming games for the specified league
        upcoming = UpcomingGames()
        games = upcoming.games(url_2)
       
        # open and retrieve all statistics data from excel sgeet for relevant league
        wb = load_workbook('C:\\Users\\Legen\\Documents\\League Program\\data\\' + competition + '_data.xlsx')
        ws = wb['Sheet1']

        name = ws['A']
        matches = ws['B']
        a_g1 = ws['C']
        a_g2 = ws['E']
        a_all = ws['G']
        p_g1 = ws['D']
        p_g2 = ws['F']
        p_all = ws['H']

        name_data = [name[x].value for x in range(len(name))]
        match_data = [str(matches[x].value) for x in range(len(name))]
        a_g1_data = [a_g1[x].value for x in range(len(name))]
        a_g2_data = [a_g2[x].value for x in range(len(name))]
        a_all_data = [a_all[x].value for x in range(len(name))]
        p_g1_data = [p_g1[x].value for x in range(len(name))]
        p_g2_data = [p_g2[x].value for x in range(len(name))]
        p_all_data = [p_all[x].value for x in range(len(name))]

        wb.close()

        # use function to insert data into the statistics table
        self.insert_data(name_data, 0, table)
        self.insert_data(match_data, 1, table)
        self.insert_data(a_g1_data, 2, table)
        self.insert_data(p_g1_data, 3, table)
        self.insert_data(a_g2_data, 4, table)
        self.insert_data(p_g2_data, 5, table)
        self.insert_data(a_all_data, 6, table)
        self.insert_data(p_all_data, 7, table)

        avg_g1 = []
        avg_g2 = []
        perc_g1 = []
        perc_g2 = []

        # sometimes data has match line attached, remove it before comparisons
        for i in range(len(games)):
            if games[i][-5:] == ' +1.5' or games[i][-5:] == ' -1.5':
                length = len(games[i]) - 5
                games[i] = games[i][:length]
        
        # fix name discrepency in data pulled from different locations
        try:
            if competition == 'lck':
                name_data[0] = 'DragonX'
                name_data[8] = 'Nongshim RedForce'
            elif competition == 'lec':
                name_data[3] = 'G2'
                name_data[8] = 'BDS'
                name_data[5] = 'Misfits'
                name_data[9] = 'Vitality'
            elif competition == 'lcs':
                name_data[2] = 'Counter Logic'
                name_data[8] = 'Liquid'
                name_data[9] = 'Team SoloMid'
            elif competition == 'vcs':
                name_data[1] = 'Cerberus'
                name_data[2] = 'GAM'
                name_data[3] = 'Luxury'
                name_data[5] = 'SBTC'
                name_data[6] = 'Flash'
                name_data[7] = 'Secret'
            elif competition == 'lpl':
                name_data[1] = 'Bilibili'
                name_data[2] = 'Edward Gaming'
                name_data[4] = 'Invictus'
                name_data[6] = 'LGD'
                name_data[7] = 'LNG'
                name_data[10] = 'Royal Never Give Up'
                name_data[13] = 'ThunderTalk'
                name_data[16] = 'Weibo'

        except:
            pass

        matches_2 = []
        combined_avg_g1 = []
        combined_avg_g2 = []
        combined_perc_g1 = []
        combined_perc_g2 = []
        g1_value = []
        g2_value = []
        line_list = []
        
        try:
            # get the stats for each team in the upcoming matches and store in a list
            for i in range(0, len(games), 1):
                no_matches = 0
                for j in range(0, len(name_data), 1):
                    if games[i] == name_data[j]:
                        avg_g1.append(float(a_g1_data[j]))
                        perc_g1.append(float(p_g1_data[j]))
                        avg_g2.append(float(a_g2_data[j]))
                        perc_g2.append(float(p_g2_data[j]))
                    else:
                        # if no matches are found there is a discrepency in the names, append a large number to avoid list index out of range 
                        # but make also make it clear that the stats are not real and need to be reviewed manually
                        no_matches += 1
                        if no_matches == len(name_data):
                            avg_g1.append(999)
                            perc_g1.append(999)
                            avg_g2.append(999)
                            perc_g2.append(999)

            # turn the games into 'team 1 vs team 2' and combine data to reflect the match stats rather than individual teams
            for i in range(0, len(games) - 1, 2):
                team_1 = games[i]
                team_2 = games[i + 1]
                match = team_1 + ' vs ' + team_2
                matches_2.append(match) 
                combined_avg_g1.append(round(((avg_g1[i] + avg_g1[i + 1]) / 2), 2))
                combined_perc_g1.append(round(((perc_g1[i] + perc_g1[i + 1]) / 2), 2))
                try:    
                    combined_avg_g2.append(round(((avg_g2[i] + avg_g2[i + 1]) / 2), 2))
                    combined_perc_g2.append(round(((perc_g2[i] + perc_g2[i + 1]) / 2), 2))   
                except:
                    pass   
        except:
            print('No data in excel doc')

        # calculate the value of bets based on provided data
        g1_value = upcoming.calculate_value(combined_perc_g1, combined_avg_g1, line, odds)
        g2_value = upcoming.calculate_value(combined_perc_g2, combined_avg_g2, line, odds)

        # convert data in lists to strings for insertion into table
        a_g1 = [str(x) for x in combined_avg_g1]
        a_g2 = [str(x) for x in combined_avg_g2]
        p_g1 = [str(x) for x in combined_perc_g1]
        p_g2 = [str(x) for x in combined_perc_g2]
        v_g1 = [str(x) for x in g1_value]
        v_g2 = [str(x) for x in g2_value]
   
        # create list of given line to insert into table
        for i in range(len(matches_2)):
            line_list.append(str(line))

        # insert all data into the correct columns of the upcoming table
        self.insert_data(matches_2, 0, upcoming_table)
        self.insert_data(line_list, 2, upcoming_table)
        self.insert_data(a_g1, 3, upcoming_table)
        self.insert_data(p_g1, 4, upcoming_table)
        self.insert_data(v_g1, 5, upcoming_table)
        self.insert_data(a_g2, 6, upcoming_table)
        self.insert_data(p_g2, 7, upcoming_table)
        self.insert_data(v_g2, 8, upcoming_table)

    # insert given data into table
    def insert_data(self, data, col, table):    
        row = 0        
        table.setRowCount(len(data))       
        for i in data:
            table.setItem(row, col, QtWidgets.QTableWidgetItem(i))
            row += 1

class TableData:        
    # required initial information to establish a class
    def __init__(self, kill_average, teams, competition):
        # class variables
        self.kill_average = kill_average
        self.teams = teams
        self.competition = competition
        self.kills_g1 = []
        self.kills_g2 = []
        self.kills_all = []
        self.games_g1 = []
        self.games_g2 = []
        self.games_all = []
        self.percent_g1 = []
        self.percent_g2 = []
        self.percent_all = []
        
        # create lists of the correct length based on the amount of teams in the league
        for team in teams:
            self.kills_g1.append(0)
            self.kills_g2.append(0)
            self.kills_all.append(0)
            self.games_g1.append(0)
            self.games_g2.append(0)
            self.games_all.append(0)
            self.percent_g1.append(0)
            self.percent_g2.append(0)
            self.percent_all.append(0)

    # get team, kill and game data from the given url and store it in the appropriate list
    def kill_data(self, url):
        # get information from website
        result = requests.get(url)
        soup = BeautifulSoup(result.text,'html.parser')

        # declare variables
        max_games = 20
        team_list = []
        game_kills = []
        blue_side = []
        red_side = []
        game_total = []

        try:
            # find all games on the page to get data from
            for i in range(max_games):
                game = soup.find_all('table', class_ = 'sb')[i]

                # check number of games played on blue side
                for team in game.find_all('th', class_ = 'side-blue'):
                    game_number = team.text.strip()
                    try:
                        int(game_number)
                    except:
                        game_number = None
                    
                    if game_number != None:
                        blue_side.append(game_number)

                # check number of games played on red side
                for team in game.find_all('th', class_ = 'side-red'):
                    game_number = team.text.strip()
                    try:
                        int(game_number)
                    except:
                        game_number = None
                    
                    if game_number != None:
                        red_side.append(game_number)
                
                # get the team names associated with each game
                for team in game.find_all('span', class_ = 'teamname'):
                    team_list.append(team.text.strip())

                # get the amount of kills for game one
                for team in game.find_all('div', class_ = 'sb-header-Kills'):
                    game_kills.append(team.text.strip())
        except:
            pass

        # convert list elements from strings to ints to operate on
        blue_side = list(map(int, blue_side))
        red_side = list(map(int, red_side))
        game_kills = list(map(int, game_kills))
        # combine red and blue to determine the game number 
        game_total = [a + b for a, b in zip(red_side, blue_side)]

        # double the size of the game number list to match game kills and team names size
        for i in range(len(game_total)):
            game_total.insert((i * 2 + 1), game_total[i * 2])

        # insert kill data for each team into overall record for game 1 and 2 as well as overall
        for i in range(len(team_list)):
            for j in range(len(self.teams)):
                if self.teams[j] == team_list[i]:              

                    # convert the team totals to match totals based on the teams position in the list, even number is team 1, odd is team 2
                    total_kills = 0
                    if i % 2 == 0:
                        total_kills = game_kills[i] + game_kills[i + 1]
                    else:
                        total_kills = game_kills[i] + game_kills[i - 1]
                    
                    self.games_all[j] += 1
                    self.kills_all[j] += total_kills

                    # use a 1 to indicate kills went over the total and 0 to indicate under
                    if total_kills > self.kill_average:
                        self.percent_all[j] += 1

                    if game_total[i] == 1:
                        self.games_g1[j] += 1
                        self.kills_g1[j] += total_kills
                        # use a 1 to indicate kills went over the total and 0 to indicate under
                        if total_kills > self.kill_average:
                            self.percent_g1[j] += 1
                    elif game_total[i] == 2:
                        self.games_g2[j] += 1
                        self.kills_g2[j] += total_kills
                        # use a 1 to indicate kills went over the total and 0 to indicate under
                        if total_kills > self.kill_average:
                            self.percent_g2[j] += 1

    # calculate the average kills over the season
    def calculate_average(self, kills, games):
        average = []
        for i in range(len(games)):
            if games[i] != 0:
                average.append(str(round(kills[i]/games[i], 2)))
            else:
                average.append(0)
        # if there has been greater than 10 games only use the 10 most recent
        #else:
        #    games.reverse()
        #    kills.reverse()
        #    for i in range(len(games)):
        #        if games[i] != 0:
        #            average.append(str(round(kills[i]/games[i], 2)))
        #        else:
        #            average.append(0)
        return average

    # calculate the percentage of games that have gone over the predetermined line
    def calculate_percentage(self, percent, games):
        percentage = []
        for i in range(len(games)):
            if games[i] != 0:
                percentage.append(str(round((percent[i]/games[i]) * 100)))
            else:
                percentage.append(0)

        return percentage

    # perform all calculations on the retrieved data
    def calculate_all(self):
        average_g1 = self.calculate_average(self.kills_g1, self.games_g1)
        average_g2 = self.calculate_average(self.kills_g2, self.games_g2)
        average_all = self.calculate_average(self.kills_all, self.games_all)
        percentage_g1 = self.calculate_percentage(self.percent_g1, self.games_g1)
        percentage_g2 = self.calculate_percentage(self.percent_g2, self.games_g2)
        percentage_all = self.calculate_percentage(self.percent_all, self.games_all)

        self.save_data(average_g1, average_g2, average_all, percentage_g1, percentage_g2, percentage_all)

    # save the data to an excel doc to be later retrieved
    def save_data(self, average_g1, average_g2, average_all, percentage_g1, percentage_g2, percentage_all):
        # open correct excel doc
        wb = load_workbook('C:\\Users\\Legen\\Documents\\League Program\\data\\' + self.competition + '_data.xlsx')
        ws = wb['Sheet1']

        # save all required data to the excel doc
        for i in range(len(self.teams)):
            ws.cell(column = 1, row = i + 1, value = self.teams[i])
            ws.cell(column = 2, row = i + 1, value = self.games_g1[i])
            ws.cell(column = 3, row = i + 1, value = average_g1[i])
            ws.cell(column = 4, row = i + 1, value = percentage_g1[i])
            ws.cell(column = 5, row = i + 1, value = average_g2[i])
            ws.cell(column = 6, row = i + 1, value = percentage_g2[i])
            ws.cell(column = 7, row = i + 1, value = average_all[i])
            ws.cell(column = 8, row = i + 1, value = percentage_all[i])
                     
        # save and close to excel doc
        wb.save('C:\\Users\\Legen\\Documents\\League Program\\data\\' + self.competition + '_data.xlsx')
        wb.close()
        
# creates a ladder based on the current standings of the given league
class Ladder:
    
    # get all information associated with the ladder for the specific league
    def ladder(self, url, competition):
        result = requests.get(url)
        soup = BeautifulSoup(result.text,'html.parser')
        table = soup.find('table', class_ = 'wikitable2 standings')
        team_names = []
        
        # no class/id available so have to specify specific td in the html to retrieve teams records
        #try:    
        name = table.find_all('span', class_ = 'teamname')

        for team in name:
            team_names.append(team.text)
        
        number_teams = len(team_names)
        
        if competition == 'lck':
            # scrape data from chosen table and get information on team names for LCK
            team_records = self.record_position(table, 5, 8, number_teams)
        
        elif competition == 'lcs':
            team_records = self.record_position(table, 5, 5, number_teams)
        
        elif competition == 'lec':
            # scrape data from chosen table and get information on team names for LCK
            team_records = self.record_position(table, 5, 5, number_teams)
            # scrape data from chosen table and get information on team names for LEC
            #name = table.find_all('a', class_ = 'catlink-teams tWACM tWAFM tWAN')
            #for team in name:
            #    team_names.append(team.text)
            #team_records = self.record_position(table, 5, 5)
        elif competition == 'vcs':
            team_records = self.record_position(table, 5, 8, number_teams)
        elif competition == 'lpl':
            team_records = self.record_position(table, 7, 8, number_teams)

        #except:
        #    print('Error! The HTML tds do not line up')

        return team_names, team_records

    # trim off extra data gathered from scraping the team records for the correct number of teams in the league
    def record_position(self, table, starting_value, increments, number_teams):
        team_records = []
        for i in range(number_teams):
            team_records.append(self.trim_data(str(table.find_all('td')[starting_value + (increments * i)])))
        return team_records
    
    # trim extra characters from the records so they can be displayed neatly
    def trim_data(self, data):
        size = len(data)
        trimmed_data = data[4:size - 5]
        return trimmed_data

class UpcomingGames:
        
    # 
    def games(self, url):
        result = requests.get(url)
        soup = BeautifulSoup(result.text,'lxml')
        games = []

        all_games = soup.find_all('div', class_ = 'outcome-name')
        for team in all_games:
            name = team.text
            games.append(name)        

        return games

    # apply the model calculations and return a list that represents the value in each game
    def calculate_value(self, percentage, average, line, odds):
        game_value = []
        for i in range(0, len(percentage), 1):
            system_odds_g1 = 50 + abs(((percentage[i] - 50) + (average[i] - line) * 2) / 2)
            g1_odds = (1 / odds) * 100
            game_value.append(round(system_odds_g1 - g1_odds, 2))
        return game_value

    
    def next_games(self):
        # declare variables and scrape the upcoming games
        games = self.games('https://www.rivalry.com/au/esports/league-of-legends-betting')
        matches = []

        # change the list of teams into a list of matches
        for i in range(0, len(games) - 1, 2):
            team_1 = games[i]
            team_2 = games[i + 1]
            match = team_1 + ' vs ' + team_2
            matches.append(match)

        result = requests.get('https://www.rivalry.com/au/esports/league-of-legends-betting')
        soup = BeautifulSoup(result.text,'lxml')
        league = []

        # scrape data for the competition that each match takes place in
        all_leagues = soup.find_all('div', class_ = 'text-league-of-legends-shade dark:text-league-of-legends-tint text-[11px]')
        for comp in all_leagues:
            league.append(comp.text)

        return league, matches

        
# provides necessary data used in calculation and creation of table
def lck_data():
    # create object of league class for the LCK
    lck_teams = ['DRX', 'DWG KIA', 'Fredit BRION', 'Gen.G', 'Hanwha Life', 'KT Rolster', 'Kwangdong Freecs', 'Liiv SANDBOX', 'NS RedForce', 'T1']
    competition = 'lck'
    lck = TableData(22.5, lck_teams, competition)
    
    # call function to scrape data and perform calculations
    lck.kill_data('https://lol.fandom.com/wiki/LCK/2022_Season/Spring_Season/Scoreboards')
    lck.kill_data('https://lol.fandom.com/wiki/LCK/2022_Season/Spring_Season/Scoreboards/Week_1_(2)')
    lck.kill_data('https://lol.fandom.com/wiki/LCK/2022_Season/Spring_Season/Scoreboards/Week_2')
    lck.kill_data('https://lol.fandom.com/wiki/LCK/2022_Season/Spring_Season/Scoreboards/Week_2_(2)')
    lck.kill_data('https://lol.fandom.com/wiki/LCK/2022_Season/Spring_Season/Scoreboards/Week_3')
    lck.kill_data('https://lol.fandom.com/wiki/LCK/2022_Season/Spring_Season/Scoreboards/Week_3_(2)')
    lck.kill_data('https://lol.fandom.com/wiki/LCK/2022_Season/Spring_Season/Scoreboards/Week_4')
    lck.kill_data('https://lol.fandom.com/wiki/LCK/2022_Season/Spring_Season/Scoreboards/Week_4_(2)')
    lck.kill_data('https://lol.fandom.com/wiki/LCK/2022_Season/Spring_Season/Scoreboards/Week_5')
    lck.kill_data('https://lol.fandom.com/wiki/LCK/2022_Season/Spring_Season/Scoreboards/Week_5_(2)')
    lck.kill_data('https://lol.fandom.com/wiki/LCK/2022_Season/Spring_Season/Scoreboards/Week_6')
    lck.kill_data('https://lol.fandom.com/wiki/LCK/2022_Season/Spring_Season/Scoreboards/Week_6_(2)')
    lck.kill_data('https://lol.fandom.com/wiki/LCK/2022_Season/Spring_Season/Scoreboards/Week_7')
    lck.kill_data('https://lol.fandom.com/wiki/LCK/2022_Season/Spring_Season/Scoreboards/Week_7_(2)')
    lck.kill_data('https://lol.fandom.com/wiki/LCK/2022_Season/Spring_Season/Scoreboards/Week_8')
    lck.kill_data('https://lol.fandom.com/wiki/LCK/2022_Season/Spring_Season/Scoreboards/Week_8_(2)')
    lck.kill_data('https://lol.fandom.com/wiki/LCK/2022_Season/Spring_Season/Scoreboards/Week_9')
    lck.kill_data('https://lol.fandom.com/wiki/LCK/2022_Season/Spring_Season/Scoreboards/Week_9_(2)')

    # update the displayed table with the new data
    lck.calculate_all()
    ui.get_lck()

# provides necessary data used in calculation and creation of table
# yes these have a lot of links but these are the only sources for the required data, one of the reasons an edge still exists
def lcs_data():
    # create object of league class for the LCS
    lcs_teams = ['100 Thieves', 'Cloud9', 'CLG', 'Dignitas', 'Evil Geniuses', 'FlyQuest', 'Golden Guardians', 'Immortals', 'Team Liquid', 'TSM']
    competition = 'lcs'
    lcs = TableData(24.5, lcs_teams, competition)
    
    # call function to scrape data and perform calculations
    lcs.kill_data('https://lol.fandom.com/wiki/LCS/2022_Season/Spring_Season/Scoreboards')
    lcs.kill_data('https://lol.fandom.com/wiki/LCS/2022_Season/Spring_Season/Scoreboards/Week_2')
    lcs.kill_data('https://lol.fandom.com/wiki/LCS/2022_Season/Spring_Season/Scoreboards/Week_3')
    lcs.kill_data('https://lol.fandom.com/wiki/LCS/2022_Season/Spring_Season/Scoreboards/Week_4')
    lcs.kill_data('https://lol.fandom.com/wiki/LCS/2022_Season/Spring_Season/Scoreboards/Week_5')
    lcs.kill_data('https://lol.fandom.com/wiki/LCS/2022_Season/Spring_Season/Scoreboards/Week_6')
    lcs.kill_data('https://lol.fandom.com/wiki/LCS/2022_Season/Spring_Season/Scoreboards/Week_7')
    lcs.kill_data('https://lol.fandom.com/wiki/LCS/2022_Season/Spring_Season/Scoreboards/Week_8')
    
    # update the displayed table with the new data
    lcs.calculate_all()
    ui.get_lcs()

# provides necessary data used in calculation and creation of table
def lec_data():
    # create object of league class for the LCS
    lec_teams = ['Astralis', 'Excel', 'Fnatic', 'G2 Esports', 'MAD Lions', 'Misfits Gaming', 'Rogue', 'SK Gaming', 'Team BDS', 'Team Vitality']
    competition = 'lec'
    lec = TableData(25.5, lec_teams, competition)
    
    # call function to scrape data and perform calculations
    lec.kill_data('https://lol.fandom.com/wiki/LEC/2022_Season/Spring_Season/Scoreboards')
    lec.kill_data('https://lol.fandom.com/wiki/LEC/2022_Season/Spring_Season/Scoreboards/Week_2')
    lec.kill_data('https://lol.fandom.com/wiki/LEC/2022_Season/Spring_Season/Scoreboards/Week_3')
    lec.kill_data('https://lol.fandom.com/wiki/LEC/2022_Season/Spring_Season/Scoreboards/Week_4')
    lec.kill_data('https://lol.fandom.com/wiki/LEC/2022_Season/Spring_Season/Scoreboards/Week_5')
    lec.kill_data('https://lol.fandom.com/wiki/LEC/2022_Season/Spring_Season/Scoreboards/Week_6')
    lec.kill_data('https://lol.fandom.com/wiki/LEC/2022_Season/Spring_Season/Scoreboards/Week_7')
    lec.kill_data('https://lol.fandom.com/wiki/LEC/2022_Season/Spring_Season/Scoreboards/Week_8')
    
    # update the displayed table with the new data
    lec.calculate_all()
    ui.get_lec()

# provides necessary data used in calculation and creation of table
def vcs_data():
    # create object of league class for the LCS
    vcs_teams = ['Burst The Sky', 'CERBERUS', 'GAM Esports', 'Luxury Esports', 'Saigon Buffalo', 'SBTC Esports', 'Team Flash', 'Team Secret']
    competition = 'vcs'
    vcs = TableData(31.5, vcs_teams, competition)
    
    # call function to scrape data and perform calculations
    vcs.kill_data('https://lol.fandom.com/wiki/VCS/2022_Season/Spring_Season/Scoreboards')
    vcs.kill_data('https://lol.fandom.com/wiki/VCS/2022_Season/Spring_Season/Scoreboards/Week_2')
    vcs.kill_data('https://lol.fandom.com/wiki/VCS/2022_Season/Spring_Season/Scoreboards/Week_3')
    vcs.kill_data('https://lol.fandom.com/wiki/VCS/2022_Season/Spring_Season/Scoreboards/Week_4')
    vcs.kill_data('https://lol.fandom.com/wiki/VCS/2022_Season/Spring_Season/Scoreboards/Week_5')
    vcs.kill_data('https://lol.fandom.com/wiki/VCS/2022_Season/Spring_Season/Scoreboards/Week_6')
    vcs.kill_data('https://lol.fandom.com/wiki/VCS/2022_Season/Spring_Season/Scoreboards/Week_7')
    vcs.kill_data('https://lol.fandom.com/wiki/VCS/2022_Season/Spring_Season/Scoreboards/Week_8')
    vcs.kill_data('https://lol.fandom.com/wiki/VCS/2022_Season/Spring_Season/Scoreboards/Week_9')
    
    # update the displayed table with the new data
    vcs.calculate_all()
    ui.get_vcs()

def lpl_data():
    # create object of league class for the LCS
    lpl_teams = ["Anyone's Legend", 'Bilibili Gaming', 'EDward Gaming', 'FunPlus Phoenix', 'Invictus Gaming', 'JD Gaming', 'LGD Gaming', 'LNG Esports', 'Oh My God', 'Rare Atom', 'RNG', 'Team WE', 'Top Esports', 'TT Gaming', 'Ultra Prime', 'Victory Five', 'Weibo Gaming']
    competition = 'lpl'
    lpl = TableData(25.5, lpl_teams, competition)
    
    # call function to scrape data and perform calculations
    lpl.kill_data('https://lol.fandom.com/wiki/LPL/2022_Season/Spring_Season/Scoreboards')
    lpl.kill_data('https://lol.fandom.com/wiki/LPL/2022_Season/Spring_Season/Scoreboards/Week_1_(2)')
    lpl.kill_data('https://lol.fandom.com/wiki/LPL/2022_Season/Spring_Season/Scoreboards/Week_1_(3)')
    lpl.kill_data('https://lol.fandom.com/wiki/LPL/2022_Season/Spring_Season/Scoreboards/Week_2')
    lpl.kill_data('https://lol.fandom.com/wiki/LPL/2022_Season/Spring_Season/Scoreboards/Week_2_(2)')
    lpl.kill_data('https://lol.fandom.com/wiki/LPL/2022_Season/Spring_Season/Scoreboards/Week_2_(3)')
    lpl.kill_data('https://lol.fandom.com/wiki/LPL/2022_Season/Spring_Season/Scoreboards/Week_3')
    lpl.kill_data('https://lol.fandom.com/wiki/LPL/2022_Season/Spring_Season/Scoreboards/Week_3_(2)')
    lpl.kill_data('https://lol.fandom.com/wiki/LPL/2022_Season/Spring_Season/Scoreboards/Week_4')
    lpl.kill_data('https://lol.fandom.com/wiki/LPL/2022_Season/Spring_Season/Scoreboards/Week_4_(2)')
    lpl.kill_data('https://lol.fandom.com/wiki/LPL/2022_Season/Spring_Season/Scoreboards/Week_5')
    lpl.kill_data('https://lol.fandom.com/wiki/LPL/2022_Season/Spring_Season/Scoreboards/Week_5_(2)')
    lpl.kill_data('https://lol.fandom.com/wiki/LPL/2022_Season/Spring_Season/Scoreboards/Week_5_(3)')
    lpl.kill_data('https://lol.fandom.com/wiki/LPL/2022_Season/Spring_Season/Scoreboards/Week_6')
    lpl.kill_data('https://lol.fandom.com/wiki/LPL/2022_Season/Spring_Season/Scoreboards/Week_6_(2)')
    lpl.kill_data('https://lol.fandom.com/wiki/LPL/2022_Season/Spring_Season/Scoreboards/Week_6_(3)')
    lpl.kill_data('https://lol.fandom.com/wiki/LPL/2022_Season/Spring_Season/Scoreboards/Week_7')
    lpl.kill_data('https://lol.fandom.com/wiki/LPL/2022_Season/Spring_Season/Scoreboards/Week_7_(2)')
    lpl.kill_data('https://lol.fandom.com/wiki/LPL/2022_Season/Spring_Season/Scoreboards/Week_7_(3)')
    lpl.kill_data('https://lol.fandom.com/wiki/LPL/2022_Season/Spring_Season/Scoreboards/Week_8')
    lpl.kill_data('https://lol.fandom.com/wiki/LPL/2022_Season/Spring_Season/Scoreboards/Week_8_(2)')
    lpl.kill_data('https://lol.fandom.com/wiki/LPL/2022_Season/Spring_Season/Scoreboards/Week_8_(3)')
    lpl.kill_data('https://lol.fandom.com/wiki/LPL/2022_Season/Spring_Season/Scoreboards/Week_9')
    lpl.kill_data('https://lol.fandom.com/wiki/LPL/2022_Season/Spring_Season/Scoreboards/Week_9_(2)')
    lpl.kill_data('https://lol.fandom.com/wiki/LPL/2022_Season/Spring_Season/Scoreboards/Week_9_(3)')
    
    # update the displayed table with the new data
    lpl.calculate_all()
    ui.get_lpl()

# alter style of the UI
stylesheet = """
QToolTip
{
     border: 1px solid black;
     background-color: #ffa02f;
     padding: 1px;
     border-radius: 3px;
     opacity: 100;
}

QWidget
{
    color: #b1b1b1;
    background-color: #323232;
}

QTreeView, QListView
{
    background-color: silver;
    margin-left: 5px;
}

QWidget:item:hover
{
    background-color: QLinearGradient( x1: 0, y1: 0, x2: 0, y2: 1, stop: 0 #ffa02f, stop: 1 #ca0619);
    color: #000000;
}

QWidget:item:selected
{
    background-color: QLinearGradient( x1: 0, y1: 0, x2: 0, y2: 1, stop: 0 #ffa02f, stop: 1 #d7801a);
}

QMenuBar::item
{
    background: transparent;
}

QMenuBar::item:selected
{
    background: transparent;
    border: 1px solid #ffaa00;
}

QMenuBar::item:pressed
{
    background: #444;
    border: 1px solid #000;
    background-color: QLinearGradient(
        x1:0, y1:0,
        x2:0, y2:1,
        stop:1 #212121,
        stop:0.4 #343434/*,
        stop:0.2 #343434,
        stop:0.1 #ffaa00*/
    );
    margin-bottom:-1px;
    padding-bottom:1px;
}

QMenu
{
    border: 1px solid #000;
}

QMenu::item
{
    padding: 2px 20px 2px 20px;
}

QMenu::item:selected
{
    color: #000000;
}

QWidget:disabled
{
    color: #808080;
    background-color: #323232;
}

QAbstractItemView
{
    background-color: QLinearGradient( x1: 0, y1: 0, x2: 0, y2: 1, stop: 0 #4d4d4d, stop: 0.1 #646464, stop: 1 #5d5d5d);
}

QWidget:focus
{
    /*border: 1px solid darkgray;*/
}

QLineEdit
{
    background-color: QLinearGradient( x1: 0, y1: 0, x2: 0, y2: 1, stop: 0 #4d4d4d, stop: 0 #646464, stop: 1 #5d5d5d);
    padding: 1px;
    border-style: solid;
    border: 1px solid #1e1e1e;
    border-radius: 5;
}

QPushButton
{
    color: #b1b1b1;
    background-color: QLinearGradient( x1: 0, y1: 0, x2: 0, y2: 1, stop: 0 #565656, stop: 0.1 #525252, stop: 0.5 #4e4e4e, stop: 0.9 #4a4a4a, stop: 1 #464646);
    border-width: 1px;
    border-color: #1e1e1e;
    border-style: solid;
    border-radius: 6;
    padding: 3px;
    font-size: 12px;
    padding-left: 5px;
    padding-right: 5px;
    min-width: 40px;
}

QPushButton:pressed
{
    background-color: QLinearGradient( x1: 0, y1: 0, x2: 0, y2: 1, stop: 0 #2d2d2d, stop: 0.1 #2b2b2b, stop: 0.5 #292929, stop: 0.9 #282828, stop: 1 #252525);
}

QComboBox
{
    selection-background-color: #ffaa00;
    background-color: QLinearGradient( x1: 0, y1: 0, x2: 0, y2: 1, stop: 0 #565656, stop: 0.1 #525252, stop: 0.5 #4e4e4e, stop: 0.9 #4a4a4a, stop: 1 #464646);
    border-style: solid;
    border: 1px solid #1e1e1e;
    border-radius: 5;
}

QComboBox:hover,QPushButton:hover
{
    border: 2px solid QLinearGradient( x1: 0, y1: 0, x2: 0, y2: 1, stop: 0 #ffa02f, stop: 1 #d7801a);
}


QComboBox:on
{
    padding-top: 3px;
    padding-left: 4px;
    background-color: QLinearGradient( x1: 0, y1: 0, x2: 0, y2: 1, stop: 0 #2d2d2d, stop: 0.1 #2b2b2b, stop: 0.5 #292929, stop: 0.9 #282828, stop: 1 #252525);
    selection-background-color: #ffaa00;
}

QComboBox QAbstractItemView
{
    border: 2px solid darkgray;
    selection-background-color: QLinearGradient( x1: 0, y1: 0, x2: 0, y2: 1, stop: 0 #ffa02f, stop: 1 #d7801a);
}

QComboBox::drop-down
{
     subcontrol-origin: padding;
     subcontrol-position: top right;
     width: 15px;

     border-left-width: 0px;
     border-left-color: darkgray;
     border-left-style: solid; /* just a single line */
     border-top-right-radius: 3px; /* same radius as the QComboBox */
     border-bottom-right-radius: 3px;
 }

QComboBox::down-arrow
{
     image: url(:/dark_orange/img/down_arrow.png);
}

QGroupBox
{
    border: 1px solid darkgray;
    margin-top: 10px;
}

QGroupBox:focus
{
    border: 1px solid darkgray;
}

QTextEdit:focus
{
    border: 1px solid darkgray;
}

QScrollBar:horizontal {
     border: 1px solid #222222;
     background: QLinearGradient( x1: 0, y1: 0, x2: 0, y2: 1, stop: 0.0 #121212, stop: 0.2 #282828, stop: 1 #484848);
     height: 7px;
     margin: 0px 16px 0 16px;
}

QScrollBar::handle:horizontal
{
      background: QLinearGradient( x1: 0, y1: 0, x2: 1, y2: 0, stop: 0 #ffa02f, stop: 0.5 #d7801a, stop: 1 #ffa02f);
      min-height: 20px;
      border-radius: 2px;
}

QScrollBar::add-line:horizontal {
      border: 1px solid #1b1b19;
      border-radius: 2px;
      background: QLinearGradient( x1: 0, y1: 0, x2: 1, y2: 0, stop: 0 #ffa02f, stop: 1 #d7801a);
      width: 14px;
      subcontrol-position: right;
      subcontrol-origin: margin;
}

QScrollBar::sub-line:horizontal {
      border: 1px solid #1b1b19;
      border-radius: 2px;
      background: QLinearGradient( x1: 0, y1: 0, x2: 1, y2: 0, stop: 0 #ffa02f, stop: 1 #d7801a);
      width: 14px;
     subcontrol-position: left;
     subcontrol-origin: margin;
}

QScrollBar::right-arrow:horizontal, QScrollBar::left-arrow:horizontal
{
      border: 1px solid black;
      width: 1px;
      height: 1px;
      background: white;
}

QScrollBar::add-page:horizontal, QScrollBar::sub-page:horizontal
{
      background: none;
}

QScrollBar:vertical
{
      background: QLinearGradient( x1: 0, y1: 0, x2: 1, y2: 0, stop: 0.0 #121212, stop: 0.2 #282828, stop: 1 #484848);
      width: 7px;
      margin: 16px 0 16px 0;
      border: 1px solid #222222;
}

QScrollBar::handle:vertical
{
      background: QLinearGradient( x1: 0, y1: 0, x2: 0, y2: 1, stop: 0 #ffa02f, stop: 0.5 #d7801a, stop: 1 #ffa02f);
      min-height: 20px;
      border-radius: 2px;
}

QScrollBar::add-line:vertical
{
      border: 1px solid #1b1b19;
      border-radius: 2px;
      background: QLinearGradient( x1: 0, y1: 0, x2: 0, y2: 1, stop: 0 #ffa02f, stop: 1 #d7801a);
      height: 14px;
      subcontrol-position: bottom;
      subcontrol-origin: margin;
}

QScrollBar::sub-line:vertical
{
      border: 1px solid #1b1b19;
      border-radius: 2px;
      background: QLinearGradient( x1: 0, y1: 0, x2: 0, y2: 1, stop: 0 #d7801a, stop: 1 #ffa02f);
      height: 14px;
      subcontrol-position: top;
      subcontrol-origin: margin;
}

QScrollBar::up-arrow:vertical, QScrollBar::down-arrow:vertical
{
      border: 1px solid black;
      width: 1px;
      height: 1px;
      background: white;
}


QScrollBar::add-page:vertical, QScrollBar::sub-page:vertical
{
      background: none;
}

QTextEdit
{
    background-color: #242424;
}

QPlainTextEdit
{
    background-color: #242424;
}

QHeaderView::section
{
    background-color: QLinearGradient(x1:0, y1:0, x2:0, y2:1, stop:0 #616161, stop: 0.5 #505050, stop: 0.6 #434343, stop:1 #656565);
    color: white;
    padding-left: 4px;
    border: 1px solid #6c6c6c;
}

QCheckBox:disabled
{
color: #414141;
}

QDockWidget::title
{
    text-align: center;
    spacing: 3px; /* spacing between items in the tool bar */
    background-color: QLinearGradient(x1:0, y1:0, x2:0, y2:1, stop:0 #323232, stop: 0.5 #242424, stop:1 #323232);
}

QDockWidget::close-button, QDockWidget::float-button
{
    text-align: center;
    spacing: 1px; /* spacing between items in the tool bar */
    background-color: QLinearGradient(x1:0, y1:0, x2:0, y2:1, stop:0 #323232, stop: 0.5 #242424, stop:1 #323232);
}

QDockWidget::close-button:hover, QDockWidget::float-button:hover
{
    background: #242424;
}

QDockWidget::close-button:pressed, QDockWidget::float-button:pressed
{
    padding: 1px -1px -1px 1px;
}

QMainWindow::separator
{
    background-color: QLinearGradient(x1:0, y1:0, x2:0, y2:1, stop:0 #161616, stop: 0.5 #151515, stop: 0.6 #212121, stop:1 #343434);
    color: white;
    padding-left: 4px;
    border: 1px solid #4c4c4c;
    spacing: 3px; /* spacing between items in the tool bar */
}

QMainWindow::separator:hover
{

    background-color: QLinearGradient(x1:0, y1:0, x2:0, y2:1, stop:0 #d7801a, stop:0.5 #b56c17 stop:1 #ffa02f);
    color: white;
    padding-left: 4px;
    border: 1px solid #6c6c6c;
    spacing: 3px; /* spacing between items in the tool bar */
}

QToolBar::handle
{
     spacing: 3px; /* spacing between items in the tool bar */
     background: url(:/dark_orange/img/handle.png);
}

QMenu::separator
{
    height: 2px;
    background-color: QLinearGradient(x1:0, y1:0, x2:0, y2:1, stop:0 #161616, stop: 0.5 #151515, stop: 0.6 #212121, stop:1 #343434);
    color: white;
    padding-left: 4px;
    margin-left: 10px;
    margin-right: 5px;
}

QProgressBar
{
    border: 2px solid grey;
    border-radius: 5px;
    text-align: center;
}

QProgressBar::chunk
{
    background-color: #d7801a;
    width: 2.15px;
    margin: 0.5px;
}

QTabBar::tab {
    color: #b1b1b1;
    border: 1px solid #444;
    border-bottom-style: none;
    background-color: #323232;
    padding-left: 10px;
    padding-right: 10px;
    padding-top: 3px;
    padding-bottom: 2px;
    margin-right: -1px;
}

QTabWidget::pane {
    border: 1px solid #444;
    top: 1px;
}

QTabBar::tab:last
{
    margin-right: 0; /* the last selected tab has nothing to overlap with on the right */
    border-top-right-radius: 3px;
}

QTabBar::tab:first:!selected
{
 margin-left: 0px; /* the last selected tab has nothing to overlap with on the right */


    border-top-left-radius: 3px;
}

QTabBar::tab:!selected
{
    color: #b1b1b1;
    border-bottom-style: solid;
    margin-top: 3px;
    background-color: QLinearGradient(x1:0, y1:0, x2:0, y2:1, stop:1 #212121, stop:.4 #343434);
}

QTabBar::tab:selected
{
    border-top-left-radius: 3px;
    border-top-right-radius: 3px;
    margin-bottom: 0px;
}

QTabBar::tab:!selected:hover
{
    /*border-top: 2px solid #ffaa00;
    padding-bottom: 3px;*/
    border-top-left-radius: 3px;
    border-top-right-radius: 3px;
    background-color: QLinearGradient(x1:0, y1:0, x2:0, y2:1, stop:1 #212121, stop:0.4 #343434, stop:0.2 #343434, stop:0.1 #ffaa00);
}

QRadioButton::indicator:checked, QRadioButton::indicator:unchecked{
    color: #b1b1b1;
    background-color: #323232;
    border: 1px solid #b1b1b1;
    border-radius: 6px;
}

QRadioButton::indicator:checked
{
    background-color: qradialgradient(
        cx: 0.5, cy: 0.5,
        fx: 0.5, fy: 0.5,
        radius: 1.0,
        stop: 0.25 #ffaa00,
        stop: 0.3 #323232
    );
}

QCheckBox::indicator{
    color: #b1b1b1;
    background-color: #323232;
    border: 1px solid #b1b1b1;
    width: 9px;
    height: 9px;
}

QRadioButton::indicator
{
    border-radius: 6px;
}

QRadioButton::indicator:hover, QCheckBox::indicator:hover
{
    border: 1px solid #ffaa00;
}

QCheckBox::indicator:checked
{
    image:url(:/dark_orange/img/checkbox.png);
}

QCheckBox::indicator:disabled, QRadioButton::indicator:disabled
{
    border: 1px solid #444;
}


QSlider::groove:horizontal {
    border: 1px solid #3A3939;
    height: 8px;
    background: #201F1F;
    margin: 2px 0;
    border-radius: 2px;
}

QSlider::handle:horizontal {
    background: QLinearGradient( x1: 0, y1: 0, x2: 0, y2: 1,
      stop: 0.0 silver, stop: 0.2 #a8a8a8, stop: 1 #727272);
    border: 1px solid #3A3939;
    width: 14px;
    height: 14px;
    margin: -4px 0;
    border-radius: 2px;
}

QSlider::groove:vertical {
    border: 1px solid #3A3939;
    width: 8px;
    background: #201F1F;
    margin: 0 0px;
    border-radius: 2px;
}

QSlider::handle:vertical {
    background: QLinearGradient( x1: 0, y1: 0, x2: 0, y2: 1, stop: 0.0 silver,
      stop: 0.2 #a8a8a8, stop: 1 #727272);
    border: 1px solid #3A3939;
    width: 14px;
    height: 14px;
    margin: 0 -4px;
    border-radius: 2px;
}

QAbstractSpinBox {
    padding-top: 2px;
    padding-bottom: 2px;
    border: 1px solid darkgray;

    border-radius: 2px;
    min-width: 50px;
}
"""

if __name__ == "__main__":
    # create UI
    app = QApplication(sys.argv)
    app.setStyleSheet(stylesheet)
    MainWindow = QtWidgets.QMainWindow()
    ui = Ui_MainWindow()
    ui.setupUi(MainWindow)
    MainWindow.show()
    sys.exit(app.exec_())
