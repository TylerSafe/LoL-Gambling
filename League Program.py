# program written by Tyler Safe to facilitate league of legends kill total gambling
# program scrapes data from given stats pages and performs calculations to give meaningful output, utlising a GUI
# written Jan 2022

import requests
from bs4 import BeautifulSoup
from PyQt5 import QtCore, QtGui, QtWidgets
from PyQt5.QtWidgets import QApplication, QMainWindow
from PyQt5.QtGui import QMovie
import sys
from openpyxl import load_workbook
from datetime import date

class Ui_MainWindow(object):
    # auto generated code using PyQt5 to create the UI
    def setupUi(self, MainWindow):
        MainWindow.setObjectName("MainWindow")
        MainWindow.resize(1301, 857)
        self.centralwidget = QtWidgets.QWidget(MainWindow)
        self.centralwidget.setObjectName("centralwidget")
        self.stackedWidget = QtWidgets.QStackedWidget(self.centralwidget)
        self.stackedWidget.setGeometry(QtCore.QRect(0, -20, 1471, 891))
        self.stackedWidget.setObjectName("stackedWidget")
        self.page = QtWidgets.QWidget()
        self.page.setObjectName("page")
        self.label = QtWidgets.QLabel(self.page)
        self.label.setGeometry(QtCore.QRect(460, 50, 361, 61))
        self.label.setObjectName("label")
        self.pushButton = QtWidgets.QPushButton(self.page)
        self.pushButton.setGeometry(QtCore.QRect(40, 150, 111, 51))
        self.pushButton.setObjectName("pushButton")
        self.lcs = QtWidgets.QPushButton(self.page)
        self.lcs.setGeometry(QtCore.QRect(40, 240, 111, 51))
        self.lcs.setObjectName("lcs")
        self.lec = QtWidgets.QPushButton(self.page)
        self.lec.setGeometry(QtCore.QRect(40, 330, 111, 51))
        self.lec.setObjectName("lec")
        self.vcs = QtWidgets.QPushButton(self.page)
        self.vcs.setGeometry(QtCore.QRect(200, 240, 111, 51))
        self.vcs.setObjectName("vcs")
        self.cblol = QtWidgets.QPushButton(self.page)
        self.cblol.setGeometry(QtCore.QRect(360, 240, 111, 51))
        self.cblol.setObjectName("cblol")
        self.ul = QtWidgets.QPushButton(self.page)
        self.ul.setGeometry(QtCore.QRect(520, 150, 111, 51))
        self.ul.setObjectName("ul")
        self.lla = QtWidgets.QPushButton(self.page)
        self.lla.setGeometry(QtCore.QRect(360, 150, 111, 51))
        self.lla.setObjectName("lla")
        self.na_acad = QtWidgets.QPushButton(self.page)
        self.na_acad.setGeometry(QtCore.QRect(200, 330, 111, 51))
        self.na_acad.setObjectName("na_acad")
        self.lpl = QtWidgets.QPushButton(self.page)
        self.lpl.setGeometry(QtCore.QRect(200, 150, 111, 51))
        self.lpl.setObjectName("lpl")
        self.lco = QtWidgets.QPushButton(self.page)
        self.lco.setGeometry(QtCore.QRect(360, 330, 111, 51))
        self.lco.setObjectName("lco")
        self.pld = QtWidgets.QPushButton(self.page)
        self.pld.setGeometry(QtCore.QRect(520, 240, 111, 51))
        self.pld.setObjectName("pld")
        self.pgn = QtWidgets.QPushButton(self.page)
        self.pgn.setGeometry(QtCore.QRect(520, 330, 111, 51))
        self.pgn.setObjectName("pgn")
        self.upcoming_table = QtWidgets.QTableWidget(self.page)
        self.upcoming_table.setGeometry(QtCore.QRect(60, 480, 561, 311))
        self.upcoming_table.setSizeAdjustPolicy(QtWidgets.QAbstractScrollArea.AdjustToContentsOnFirstShow)
        self.upcoming_table.setObjectName("upcoming_table")
        self.upcoming_table.setColumnCount(2)
        self.upcoming_table.setRowCount(0)
        item = QtWidgets.QTableWidgetItem()
        self.upcoming_table.setHorizontalHeaderItem(0, item)
        item = QtWidgets.QTableWidgetItem()
        self.upcoming_table.setHorizontalHeaderItem(1, item)
        self.label_77 = QtWidgets.QLabel(self.page)
        self.label_77.setGeometry(QtCore.QRect(270, 430, 221, 31))
        self.label_77.setObjectName("label_77")
        # added animated gif
        self.label_13 = QtWidgets.QLabel(self.page)
        self.label_13.setGeometry(QtCore.QRect(720, 490, 521, 291))
        self.label_13.setText("")
        self.label_13.setObjectName("label_13")
        self.movie = QMovie("C:\\Users\\Legen\\Documents\\League Program\\league_image.webp")
        self.label_13.setMovie(self.movie)
        self.movie.start()
        # keep code
        self.ljl = QtWidgets.QPushButton(self.page)
        self.ljl.setGeometry(QtCore.QRect(680, 240, 111, 51))
        self.ljl.setObjectName("ljl")
        self.lvp = QtWidgets.QPushButton(self.page)
        self.lvp.setGeometry(QtCore.QRect(680, 150, 111, 51))
        self.lvp.setObjectName("lvp")
        self.nlc = QtWidgets.QPushButton(self.page)
        self.nlc.setGeometry(QtCore.QRect(680, 330, 111, 51))
        self.nlc.setObjectName("nlc")
        self.lfl = QtWidgets.QPushButton(self.page)
        self.lfl.setGeometry(QtCore.QRect(840, 150, 111, 51))
        self.lfl.setObjectName("lfl")
        self.lplol = QtWidgets.QPushButton(self.page)
        self.lplol.setGeometry(QtCore.QRect(840, 240, 111, 51))
        self.lplol.setObjectName("lplol")
        self.hpm = QtWidgets.QPushButton(self.page)
        self.hpm.setGeometry(QtCore.QRect(840, 330, 111, 51))
        self.hpm.setObjectName("hpm")
        self.lck_chal = QtWidgets.QPushButton(self.page)
        self.lck_chal.setGeometry(QtCore.QRect(1000, 150, 111, 51))
        self.lck_chal.setObjectName("lck_chal")
        self.tcl = QtWidgets.QPushButton(self.page)
        self.tcl.setGeometry(QtCore.QRect(1000, 240, 111, 51))
        self.tcl.setObjectName("tcl")
        self.pcs = QtWidgets.QPushButton(self.page)
        self.pcs.setGeometry(QtCore.QRect(1000, 330, 111, 51))
        self.pcs.setObjectName("pcs")
        self.none = QtWidgets.QPushButton(self.page)
        self.none.setGeometry(QtCore.QRect(1160, 330, 111, 51))
        self.none.setText("")
        self.none.setObjectName("none")
        self.lcl = QtWidgets.QPushButton(self.page)
        self.lcl.setGeometry(QtCore.QRect(1160, 240, 111, 51))
        self.lcl.setText("")
        self.lcl.setObjectName("lcl")
        self.ebl = QtWidgets.QPushButton(self.page)
        self.ebl.setGeometry(QtCore.QRect(1160, 150, 111, 51))
        self.ebl.setObjectName("ebl")
        self.stackedWidget.addWidget(self.page)
        self.page_lvp = QtWidgets.QWidget()
        self.page_lvp.setObjectName("page_lvp")
        self.label_102 = QtWidgets.QLabel(self.page_lvp)
        self.label_102.setGeometry(QtCore.QRect(440, 800, 61, 21))
        self.label_102.setObjectName("label_102")
        self.lvp_matches = QtWidgets.QTableWidget(self.page_lvp)
        self.lvp_matches.setGeometry(QtCore.QRect(90, 500, 1141, 281))
        self.lvp_matches.setObjectName("lvp_matches")
        self.lvp_matches.setColumnCount(9)
        self.lvp_matches.setRowCount(0)
        item = QtWidgets.QTableWidgetItem()
        self.lvp_matches.setHorizontalHeaderItem(0, item)
        item = QtWidgets.QTableWidgetItem()
        self.lvp_matches.setHorizontalHeaderItem(1, item)
        item = QtWidgets.QTableWidgetItem()
        self.lvp_matches.setHorizontalHeaderItem(2, item)
        item = QtWidgets.QTableWidgetItem()
        self.lvp_matches.setHorizontalHeaderItem(3, item)
        item = QtWidgets.QTableWidgetItem()
        self.lvp_matches.setHorizontalHeaderItem(4, item)
        item = QtWidgets.QTableWidgetItem()
        self.lvp_matches.setHorizontalHeaderItem(5, item)
        item = QtWidgets.QTableWidgetItem()
        self.lvp_matches.setHorizontalHeaderItem(6, item)
        item = QtWidgets.QTableWidgetItem()
        self.lvp_matches.setHorizontalHeaderItem(7, item)
        item = QtWidgets.QTableWidgetItem()
        self.lvp_matches.setHorizontalHeaderItem(8, item)
        self.label_103 = QtWidgets.QLabel(self.page_lvp)
        self.label_103.setGeometry(QtCore.QRect(620, 800, 51, 21))
        self.label_103.setObjectName("label_103")
        self.lvp_calculate = QtWidgets.QPushButton(self.page_lvp)
        self.lvp_calculate.setGeometry(QtCore.QRect(840, 790, 141, 41))
        self.lvp_calculate.setObjectName("lvp_calculate")
        self.lvp_stats_table = QtWidgets.QTableWidget(self.page_lvp)
        self.lvp_stats_table.setGeometry(QtCore.QRect(100, 80, 831, 351))
        self.lvp_stats_table.setObjectName("lvp_stats_table")
        self.lvp_stats_table.setColumnCount(8)
        self.lvp_stats_table.setRowCount(0)
        item = QtWidgets.QTableWidgetItem()
        self.lvp_stats_table.setHorizontalHeaderItem(0, item)
        item = QtWidgets.QTableWidgetItem()
        self.lvp_stats_table.setHorizontalHeaderItem(1, item)
        item = QtWidgets.QTableWidgetItem()
        self.lvp_stats_table.setHorizontalHeaderItem(2, item)
        item = QtWidgets.QTableWidgetItem()
        self.lvp_stats_table.setHorizontalHeaderItem(3, item)
        item = QtWidgets.QTableWidgetItem()
        self.lvp_stats_table.setHorizontalHeaderItem(4, item)
        item = QtWidgets.QTableWidgetItem()
        self.lvp_stats_table.setHorizontalHeaderItem(5, item)
        item = QtWidgets.QTableWidgetItem()
        self.lvp_stats_table.setHorizontalHeaderItem(6, item)
        item = QtWidgets.QTableWidgetItem()
        self.lvp_stats_table.setHorizontalHeaderItem(7, item)
        self.label_104 = QtWidgets.QLabel(self.page_lvp)
        self.label_104.setGeometry(QtCore.QRect(1080, 20, 361, 61))
        self.label_104.setObjectName("label_104")
        self.back_25 = QtWidgets.QPushButton(self.page_lvp)
        self.back_25.setGeometry(QtCore.QRect(10, 30, 91, 41))
        self.back_25.setObjectName("back_25")
        self.ladder_lvp = QtWidgets.QTableWidget(self.page_lvp)
        self.ladder_lvp.setGeometry(QtCore.QRect(1000, 80, 231, 351))
        self.ladder_lvp.setObjectName("ladder_lvp")
        self.ladder_lvp.setColumnCount(2)
        self.ladder_lvp.setRowCount(0)
        item = QtWidgets.QTableWidgetItem()
        self.ladder_lvp.setHorizontalHeaderItem(0, item)
        item = QtWidgets.QTableWidgetItem()
        self.ladder_lvp.setHorizontalHeaderItem(1, item)
        self.update_lvp = QtWidgets.QPushButton(self.page_lvp)
        self.update_lvp.setGeometry(QtCore.QRect(830, 440, 101, 41))
        self.update_lvp.setObjectName("update_lvp")
        self.lvp_line = QtWidgets.QTextEdit(self.page_lvp)
        self.lvp_line.setGeometry(QtCore.QRect(480, 800, 71, 21))
        self.lvp_line.setObjectName("lvp_line")
        self.label_105 = QtWidgets.QLabel(self.page_lvp)
        self.label_105.setGeometry(QtCore.QRect(440, 20, 361, 61))
        self.label_105.setObjectName("label_105")
        self.lvp_odds = QtWidgets.QTextEdit(self.page_lvp)
        self.lvp_odds.setGeometry(QtCore.QRect(670, 800, 71, 21))
        self.lvp_odds.setObjectName("lvp_odds")
        self.stackedWidget.addWidget(self.page_lvp)
        self.page_ljl = QtWidgets.QWidget()
        self.page_ljl.setObjectName("page_ljl")
        self.ljl_odds = QtWidgets.QTextEdit(self.page_ljl)
        self.ljl_odds.setGeometry(QtCore.QRect(670, 800, 71, 21))
        self.ljl_odds.setObjectName("ljl_odds")
        self.label_108 = QtWidgets.QLabel(self.page_ljl)
        self.label_108.setGeometry(QtCore.QRect(1080, 20, 361, 61))
        self.label_108.setObjectName("label_108")
        self.label_106 = QtWidgets.QLabel(self.page_ljl)
        self.label_106.setGeometry(QtCore.QRect(440, 800, 61, 21))
        self.label_106.setObjectName("label_106")
        self.ljl_matches = QtWidgets.QTableWidget(self.page_ljl)
        self.ljl_matches.setGeometry(QtCore.QRect(90, 500, 1141, 281))
        self.ljl_matches.setObjectName("ljl_matches")
        self.ljl_matches.setColumnCount(9)
        self.ljl_matches.setRowCount(0)
        item = QtWidgets.QTableWidgetItem()
        self.ljl_matches.setHorizontalHeaderItem(0, item)
        item = QtWidgets.QTableWidgetItem()
        self.ljl_matches.setHorizontalHeaderItem(1, item)
        item = QtWidgets.QTableWidgetItem()
        self.ljl_matches.setHorizontalHeaderItem(2, item)
        item = QtWidgets.QTableWidgetItem()
        self.ljl_matches.setHorizontalHeaderItem(3, item)
        item = QtWidgets.QTableWidgetItem()
        self.ljl_matches.setHorizontalHeaderItem(4, item)
        item = QtWidgets.QTableWidgetItem()
        self.ljl_matches.setHorizontalHeaderItem(5, item)
        item = QtWidgets.QTableWidgetItem()
        self.ljl_matches.setHorizontalHeaderItem(6, item)
        item = QtWidgets.QTableWidgetItem()
        self.ljl_matches.setHorizontalHeaderItem(7, item)
        item = QtWidgets.QTableWidgetItem()
        self.ljl_matches.setHorizontalHeaderItem(8, item)
        self.ladder_ljl = QtWidgets.QTableWidget(self.page_ljl)
        self.ladder_ljl.setGeometry(QtCore.QRect(1000, 80, 231, 351))
        self.ladder_ljl.setObjectName("ladder_ljl")
        self.ladder_ljl.setColumnCount(2)
        self.ladder_ljl.setRowCount(0)
        item = QtWidgets.QTableWidgetItem()
        self.ladder_ljl.setHorizontalHeaderItem(0, item)
        item = QtWidgets.QTableWidgetItem()
        self.ladder_ljl.setHorizontalHeaderItem(1, item)
        self.update_ljl = QtWidgets.QPushButton(self.page_ljl)
        self.update_ljl.setGeometry(QtCore.QRect(830, 440, 101, 41))
        self.update_ljl.setObjectName("update_ljl")
        self.label_109 = QtWidgets.QLabel(self.page_ljl)
        self.label_109.setGeometry(QtCore.QRect(440, 20, 361, 61))
        self.label_109.setObjectName("label_109")
        self.ljl_calculate = QtWidgets.QPushButton(self.page_ljl)
        self.ljl_calculate.setGeometry(QtCore.QRect(840, 790, 141, 41))
        self.ljl_calculate.setObjectName("ljl_calculate")
        self.label_107 = QtWidgets.QLabel(self.page_ljl)
        self.label_107.setGeometry(QtCore.QRect(620, 800, 51, 21))
        self.label_107.setObjectName("label_107")
        self.ljl_line = QtWidgets.QTextEdit(self.page_ljl)
        self.ljl_line.setGeometry(QtCore.QRect(480, 800, 71, 21))
        self.ljl_line.setObjectName("ljl_line")
        self.ljl_stats_table = QtWidgets.QTableWidget(self.page_ljl)
        self.ljl_stats_table.setGeometry(QtCore.QRect(100, 80, 831, 351))
        self.ljl_stats_table.setObjectName("ljl_stats_table")
        self.ljl_stats_table.setColumnCount(8)
        self.ljl_stats_table.setRowCount(0)
        item = QtWidgets.QTableWidgetItem()
        self.ljl_stats_table.setHorizontalHeaderItem(0, item)
        item = QtWidgets.QTableWidgetItem()
        self.ljl_stats_table.setHorizontalHeaderItem(1, item)
        item = QtWidgets.QTableWidgetItem()
        self.ljl_stats_table.setHorizontalHeaderItem(2, item)
        item = QtWidgets.QTableWidgetItem()
        self.ljl_stats_table.setHorizontalHeaderItem(3, item)
        item = QtWidgets.QTableWidgetItem()
        self.ljl_stats_table.setHorizontalHeaderItem(4, item)
        item = QtWidgets.QTableWidgetItem()
        self.ljl_stats_table.setHorizontalHeaderItem(5, item)
        item = QtWidgets.QTableWidgetItem()
        self.ljl_stats_table.setHorizontalHeaderItem(6, item)
        item = QtWidgets.QTableWidgetItem()
        self.ljl_stats_table.setHorizontalHeaderItem(7, item)
        self.back_26 = QtWidgets.QPushButton(self.page_ljl)
        self.back_26.setGeometry(QtCore.QRect(10, 30, 91, 41))
        self.back_26.setObjectName("back_26")
        self.stackedWidget.addWidget(self.page_ljl)
        self.page_nlc = QtWidgets.QWidget()
        self.page_nlc.setObjectName("page_nlc")
        self.label_110 = QtWidgets.QLabel(self.page_nlc)
        self.label_110.setGeometry(QtCore.QRect(440, 800, 61, 21))
        self.label_110.setObjectName("label_110")
        self.nlc_matches = QtWidgets.QTableWidget(self.page_nlc)
        self.nlc_matches.setGeometry(QtCore.QRect(90, 500, 1141, 281))
        self.nlc_matches.setObjectName("nlc_matches")
        self.nlc_matches.setColumnCount(9)
        self.nlc_matches.setRowCount(0)
        item = QtWidgets.QTableWidgetItem()
        self.nlc_matches.setHorizontalHeaderItem(0, item)
        item = QtWidgets.QTableWidgetItem()
        self.nlc_matches.setHorizontalHeaderItem(1, item)
        item = QtWidgets.QTableWidgetItem()
        self.nlc_matches.setHorizontalHeaderItem(2, item)
        item = QtWidgets.QTableWidgetItem()
        self.nlc_matches.setHorizontalHeaderItem(3, item)
        item = QtWidgets.QTableWidgetItem()
        self.nlc_matches.setHorizontalHeaderItem(4, item)
        item = QtWidgets.QTableWidgetItem()
        self.nlc_matches.setHorizontalHeaderItem(5, item)
        item = QtWidgets.QTableWidgetItem()
        self.nlc_matches.setHorizontalHeaderItem(6, item)
        item = QtWidgets.QTableWidgetItem()
        self.nlc_matches.setHorizontalHeaderItem(7, item)
        item = QtWidgets.QTableWidgetItem()
        self.nlc_matches.setHorizontalHeaderItem(8, item)
        self.label_111 = QtWidgets.QLabel(self.page_nlc)
        self.label_111.setGeometry(QtCore.QRect(620, 800, 51, 21))
        self.label_111.setObjectName("label_111")
        self.nlc_calculate = QtWidgets.QPushButton(self.page_nlc)
        self.nlc_calculate.setGeometry(QtCore.QRect(840, 790, 141, 41))
        self.nlc_calculate.setObjectName("nlc_calculate")
        self.nlc_stats_table = QtWidgets.QTableWidget(self.page_nlc)
        self.nlc_stats_table.setGeometry(QtCore.QRect(100, 80, 831, 351))
        self.nlc_stats_table.setObjectName("nlc_stats_table")
        self.nlc_stats_table.setColumnCount(8)
        self.nlc_stats_table.setRowCount(0)
        item = QtWidgets.QTableWidgetItem()
        self.nlc_stats_table.setHorizontalHeaderItem(0, item)
        item = QtWidgets.QTableWidgetItem()
        self.nlc_stats_table.setHorizontalHeaderItem(1, item)
        item = QtWidgets.QTableWidgetItem()
        self.nlc_stats_table.setHorizontalHeaderItem(2, item)
        item = QtWidgets.QTableWidgetItem()
        self.nlc_stats_table.setHorizontalHeaderItem(3, item)
        item = QtWidgets.QTableWidgetItem()
        self.nlc_stats_table.setHorizontalHeaderItem(4, item)
        item = QtWidgets.QTableWidgetItem()
        self.nlc_stats_table.setHorizontalHeaderItem(5, item)
        item = QtWidgets.QTableWidgetItem()
        self.nlc_stats_table.setHorizontalHeaderItem(6, item)
        item = QtWidgets.QTableWidgetItem()
        self.nlc_stats_table.setHorizontalHeaderItem(7, item)
        self.label_112 = QtWidgets.QLabel(self.page_nlc)
        self.label_112.setGeometry(QtCore.QRect(1080, 20, 361, 61))
        self.label_112.setObjectName("label_112")
        self.back_27 = QtWidgets.QPushButton(self.page_nlc)
        self.back_27.setGeometry(QtCore.QRect(10, 30, 91, 41))
        self.back_27.setObjectName("back_27")
        self.ladder_nlc = QtWidgets.QTableWidget(self.page_nlc)
        self.ladder_nlc.setGeometry(QtCore.QRect(1000, 80, 231, 351))
        self.ladder_nlc.setObjectName("ladder_nlc")
        self.ladder_nlc.setColumnCount(2)
        self.ladder_nlc.setRowCount(0)
        item = QtWidgets.QTableWidgetItem()
        self.ladder_nlc.setHorizontalHeaderItem(0, item)
        item = QtWidgets.QTableWidgetItem()
        self.ladder_nlc.setHorizontalHeaderItem(1, item)
        self.update_nlc = QtWidgets.QPushButton(self.page_nlc)
        self.update_nlc.setGeometry(QtCore.QRect(830, 440, 101, 41))
        self.update_nlc.setObjectName("update_nlc")
        self.nlc_line = QtWidgets.QTextEdit(self.page_nlc)
        self.nlc_line.setGeometry(QtCore.QRect(480, 800, 71, 21))
        self.nlc_line.setObjectName("nlc_line")
        self.label_113 = QtWidgets.QLabel(self.page_nlc)
        self.label_113.setGeometry(QtCore.QRect(440, 20, 361, 61))
        self.label_113.setObjectName("label_113")
        self.nlc_odds = QtWidgets.QTextEdit(self.page_nlc)
        self.nlc_odds.setGeometry(QtCore.QRect(670, 800, 71, 21))
        self.nlc_odds.setObjectName("nlc_odds")
        self.stackedWidget.addWidget(self.page_nlc)
        self.page_ul = QtWidgets.QWidget()
        self.page_ul.setObjectName("page_ul")
        self.update_ul = QtWidgets.QPushButton(self.page_ul)
        self.update_ul.setGeometry(QtCore.QRect(830, 440, 101, 41))
        self.update_ul.setObjectName("update_ul")
        self.back_19 = QtWidgets.QPushButton(self.page_ul)
        self.back_19.setGeometry(QtCore.QRect(10, 30, 91, 41))
        self.back_19.setObjectName("back_19")
        self.label_78 = QtWidgets.QLabel(self.page_ul)
        self.label_78.setGeometry(QtCore.QRect(620, 800, 51, 21))
        self.label_78.setObjectName("label_78")
        self.ul_calculate = QtWidgets.QPushButton(self.page_ul)
        self.ul_calculate.setGeometry(QtCore.QRect(840, 790, 141, 41))
        self.ul_calculate.setObjectName("ul_calculate")
        self.ul_matches = QtWidgets.QTableWidget(self.page_ul)
        self.ul_matches.setGeometry(QtCore.QRect(90, 500, 1141, 281))
        self.ul_matches.setObjectName("ul_matches")
        self.ul_matches.setColumnCount(9)
        self.ul_matches.setRowCount(0)
        item = QtWidgets.QTableWidgetItem()
        self.ul_matches.setHorizontalHeaderItem(0, item)
        item = QtWidgets.QTableWidgetItem()
        self.ul_matches.setHorizontalHeaderItem(1, item)
        item = QtWidgets.QTableWidgetItem()
        self.ul_matches.setHorizontalHeaderItem(2, item)
        item = QtWidgets.QTableWidgetItem()
        self.ul_matches.setHorizontalHeaderItem(3, item)
        item = QtWidgets.QTableWidgetItem()
        self.ul_matches.setHorizontalHeaderItem(4, item)
        item = QtWidgets.QTableWidgetItem()
        self.ul_matches.setHorizontalHeaderItem(5, item)
        item = QtWidgets.QTableWidgetItem()
        self.ul_matches.setHorizontalHeaderItem(6, item)
        item = QtWidgets.QTableWidgetItem()
        self.ul_matches.setHorizontalHeaderItem(7, item)
        item = QtWidgets.QTableWidgetItem()
        self.ul_matches.setHorizontalHeaderItem(8, item)
        self.label_79 = QtWidgets.QLabel(self.page_ul)
        self.label_79.setGeometry(QtCore.QRect(440, 20, 361, 61))
        self.label_79.setObjectName("label_79")
        self.label_80 = QtWidgets.QLabel(self.page_ul)
        self.label_80.setGeometry(QtCore.QRect(440, 800, 61, 21))
        self.label_80.setObjectName("label_80")
        self.ul_stats_table = QtWidgets.QTableWidget(self.page_ul)
        self.ul_stats_table.setGeometry(QtCore.QRect(100, 80, 831, 351))
        self.ul_stats_table.setObjectName("ul_stats_table")
        self.ul_stats_table.setColumnCount(8)
        self.ul_stats_table.setRowCount(0)
        item = QtWidgets.QTableWidgetItem()
        self.ul_stats_table.setHorizontalHeaderItem(0, item)
        item = QtWidgets.QTableWidgetItem()
        self.ul_stats_table.setHorizontalHeaderItem(1, item)
        item = QtWidgets.QTableWidgetItem()
        self.ul_stats_table.setHorizontalHeaderItem(2, item)
        item = QtWidgets.QTableWidgetItem()
        self.ul_stats_table.setHorizontalHeaderItem(3, item)
        item = QtWidgets.QTableWidgetItem()
        self.ul_stats_table.setHorizontalHeaderItem(4, item)
        item = QtWidgets.QTableWidgetItem()
        self.ul_stats_table.setHorizontalHeaderItem(5, item)
        item = QtWidgets.QTableWidgetItem()
        self.ul_stats_table.setHorizontalHeaderItem(6, item)
        item = QtWidgets.QTableWidgetItem()
        self.ul_stats_table.setHorizontalHeaderItem(7, item)
        self.ul_line = QtWidgets.QTextEdit(self.page_ul)
        self.ul_line.setGeometry(QtCore.QRect(480, 800, 71, 21))
        self.ul_line.setObjectName("ul_line")
        self.label_81 = QtWidgets.QLabel(self.page_ul)
        self.label_81.setGeometry(QtCore.QRect(1080, 20, 361, 61))
        self.label_81.setObjectName("label_81")
        self.ladder_ul = QtWidgets.QTableWidget(self.page_ul)
        self.ladder_ul.setGeometry(QtCore.QRect(1000, 80, 231, 351))
        self.ladder_ul.setObjectName("ladder_ul")
        self.ladder_ul.setColumnCount(2)
        self.ladder_ul.setRowCount(0)
        item = QtWidgets.QTableWidgetItem()
        self.ladder_ul.setHorizontalHeaderItem(0, item)
        item = QtWidgets.QTableWidgetItem()
        self.ladder_ul.setHorizontalHeaderItem(1, item)
        self.ul_odds = QtWidgets.QTextEdit(self.page_ul)
        self.ul_odds.setGeometry(QtCore.QRect(670, 800, 71, 21))
        self.ul_odds.setObjectName("ul_odds")
        self.stackedWidget.addWidget(self.page_ul)
        self.page_pld = QtWidgets.QWidget()
        self.page_pld.setObjectName("page_pld")
        self.update_pld = QtWidgets.QPushButton(self.page_pld)
        self.update_pld.setGeometry(QtCore.QRect(830, 440, 101, 41))
        self.update_pld.setObjectName("update_pld")
        self.back_20 = QtWidgets.QPushButton(self.page_pld)
        self.back_20.setGeometry(QtCore.QRect(10, 30, 91, 41))
        self.back_20.setObjectName("back_20")
        self.label_82 = QtWidgets.QLabel(self.page_pld)
        self.label_82.setGeometry(QtCore.QRect(620, 800, 51, 21))
        self.label_82.setObjectName("label_82")
        self.pld_calculate = QtWidgets.QPushButton(self.page_pld)
        self.pld_calculate.setGeometry(QtCore.QRect(840, 790, 141, 41))
        self.pld_calculate.setObjectName("pld_calculate")
        self.pld_matches = QtWidgets.QTableWidget(self.page_pld)
        self.pld_matches.setGeometry(QtCore.QRect(90, 500, 1141, 281))
        self.pld_matches.setObjectName("pld_matches")
        self.pld_matches.setColumnCount(9)
        self.pld_matches.setRowCount(0)
        item = QtWidgets.QTableWidgetItem()
        self.pld_matches.setHorizontalHeaderItem(0, item)
        item = QtWidgets.QTableWidgetItem()
        self.pld_matches.setHorizontalHeaderItem(1, item)
        item = QtWidgets.QTableWidgetItem()
        self.pld_matches.setHorizontalHeaderItem(2, item)
        item = QtWidgets.QTableWidgetItem()
        self.pld_matches.setHorizontalHeaderItem(3, item)
        item = QtWidgets.QTableWidgetItem()
        self.pld_matches.setHorizontalHeaderItem(4, item)
        item = QtWidgets.QTableWidgetItem()
        self.pld_matches.setHorizontalHeaderItem(5, item)
        item = QtWidgets.QTableWidgetItem()
        self.pld_matches.setHorizontalHeaderItem(6, item)
        item = QtWidgets.QTableWidgetItem()
        self.pld_matches.setHorizontalHeaderItem(7, item)
        item = QtWidgets.QTableWidgetItem()
        self.pld_matches.setHorizontalHeaderItem(8, item)
        self.label_83 = QtWidgets.QLabel(self.page_pld)
        self.label_83.setGeometry(QtCore.QRect(440, 20, 361, 61))
        self.label_83.setObjectName("label_83")
        self.label_84 = QtWidgets.QLabel(self.page_pld)
        self.label_84.setGeometry(QtCore.QRect(440, 800, 61, 21))
        self.label_84.setObjectName("label_84")
        self.pld_stats_table = QtWidgets.QTableWidget(self.page_pld)
        self.pld_stats_table.setGeometry(QtCore.QRect(100, 80, 831, 351))
        self.pld_stats_table.setObjectName("pld_stats_table")
        self.pld_stats_table.setColumnCount(8)
        self.pld_stats_table.setRowCount(0)
        item = QtWidgets.QTableWidgetItem()
        self.pld_stats_table.setHorizontalHeaderItem(0, item)
        item = QtWidgets.QTableWidgetItem()
        self.pld_stats_table.setHorizontalHeaderItem(1, item)
        item = QtWidgets.QTableWidgetItem()
        self.pld_stats_table.setHorizontalHeaderItem(2, item)
        item = QtWidgets.QTableWidgetItem()
        self.pld_stats_table.setHorizontalHeaderItem(3, item)
        item = QtWidgets.QTableWidgetItem()
        self.pld_stats_table.setHorizontalHeaderItem(4, item)
        item = QtWidgets.QTableWidgetItem()
        self.pld_stats_table.setHorizontalHeaderItem(5, item)
        item = QtWidgets.QTableWidgetItem()
        self.pld_stats_table.setHorizontalHeaderItem(6, item)
        item = QtWidgets.QTableWidgetItem()
        self.pld_stats_table.setHorizontalHeaderItem(7, item)
        self.pld_line = QtWidgets.QTextEdit(self.page_pld)
        self.pld_line.setGeometry(QtCore.QRect(480, 800, 71, 21))
        self.pld_line.setObjectName("pld_line")
        self.label_85 = QtWidgets.QLabel(self.page_pld)
        self.label_85.setGeometry(QtCore.QRect(1080, 20, 361, 61))
        self.label_85.setObjectName("label_85")
        self.ladder_pld = QtWidgets.QTableWidget(self.page_pld)
        self.ladder_pld.setGeometry(QtCore.QRect(1000, 80, 231, 351))
        self.ladder_pld.setObjectName("ladder_pld")
        self.ladder_pld.setColumnCount(2)
        self.ladder_pld.setRowCount(0)
        item = QtWidgets.QTableWidgetItem()
        self.ladder_pld.setHorizontalHeaderItem(0, item)
        item = QtWidgets.QTableWidgetItem()
        self.ladder_pld.setHorizontalHeaderItem(1, item)
        self.pld_odds = QtWidgets.QTextEdit(self.page_pld)
        self.pld_odds.setGeometry(QtCore.QRect(670, 800, 71, 21))
        self.pld_odds.setObjectName("pld_odds")
        self.stackedWidget.addWidget(self.page_pld)
        self.page_pgn = QtWidgets.QWidget()
        self.page_pgn.setObjectName("page_pgn")
        self.pgn_odds = QtWidgets.QTextEdit(self.page_pgn)
        self.pgn_odds.setGeometry(QtCore.QRect(670, 800, 71, 21))
        self.pgn_odds.setObjectName("pgn_odds")
        self.label_88 = QtWidgets.QLabel(self.page_pgn)
        self.label_88.setGeometry(QtCore.QRect(440, 800, 61, 21))
        self.label_88.setObjectName("label_88")
        self.label_89 = QtWidgets.QLabel(self.page_pgn)
        self.label_89.setGeometry(QtCore.QRect(1080, 20, 361, 61))
        self.label_89.setObjectName("label_89")
        self.pgn_calculate = QtWidgets.QPushButton(self.page_pgn)
        self.pgn_calculate.setGeometry(QtCore.QRect(840, 790, 141, 41))
        self.pgn_calculate.setObjectName("pgn_calculate")
        self.pgn_line = QtWidgets.QTextEdit(self.page_pgn)
        self.pgn_line.setGeometry(QtCore.QRect(480, 800, 71, 21))
        self.pgn_line.setObjectName("pgn_line")
        self.back_21 = QtWidgets.QPushButton(self.page_pgn)
        self.back_21.setGeometry(QtCore.QRect(10, 30, 91, 41))
        self.back_21.setObjectName("back_21")
        self.label_86 = QtWidgets.QLabel(self.page_pgn)
        self.label_86.setGeometry(QtCore.QRect(620, 800, 51, 21))
        self.label_86.setObjectName("label_86")
        self.pgn_stats_table = QtWidgets.QTableWidget(self.page_pgn)
        self.pgn_stats_table.setGeometry(QtCore.QRect(100, 80, 831, 351))
        self.pgn_stats_table.setObjectName("pgn_stats_table")
        self.pgn_stats_table.setColumnCount(8)
        self.pgn_stats_table.setRowCount(0)
        item = QtWidgets.QTableWidgetItem()
        self.pgn_stats_table.setHorizontalHeaderItem(0, item)
        item = QtWidgets.QTableWidgetItem()
        self.pgn_stats_table.setHorizontalHeaderItem(1, item)
        item = QtWidgets.QTableWidgetItem()
        self.pgn_stats_table.setHorizontalHeaderItem(2, item)
        item = QtWidgets.QTableWidgetItem()
        self.pgn_stats_table.setHorizontalHeaderItem(3, item)
        item = QtWidgets.QTableWidgetItem()
        self.pgn_stats_table.setHorizontalHeaderItem(4, item)
        item = QtWidgets.QTableWidgetItem()
        self.pgn_stats_table.setHorizontalHeaderItem(5, item)
        item = QtWidgets.QTableWidgetItem()
        self.pgn_stats_table.setHorizontalHeaderItem(6, item)
        item = QtWidgets.QTableWidgetItem()
        self.pgn_stats_table.setHorizontalHeaderItem(7, item)
        self.update_pgn = QtWidgets.QPushButton(self.page_pgn)
        self.update_pgn.setGeometry(QtCore.QRect(830, 440, 101, 41))
        self.update_pgn.setObjectName("update_pgn")
        self.ladder_pgn = QtWidgets.QTableWidget(self.page_pgn)
        self.ladder_pgn.setGeometry(QtCore.QRect(1000, 80, 231, 351))
        self.ladder_pgn.setObjectName("ladder_pgn")
        self.ladder_pgn.setColumnCount(2)
        self.ladder_pgn.setRowCount(0)
        item = QtWidgets.QTableWidgetItem()
        self.ladder_pgn.setHorizontalHeaderItem(0, item)
        item = QtWidgets.QTableWidgetItem()
        self.ladder_pgn.setHorizontalHeaderItem(1, item)
        self.pgn_matches = QtWidgets.QTableWidget(self.page_pgn)
        self.pgn_matches.setGeometry(QtCore.QRect(90, 500, 1141, 281))
        self.pgn_matches.setObjectName("pgn_matches")
        self.pgn_matches.setColumnCount(9)
        self.pgn_matches.setRowCount(0)
        item = QtWidgets.QTableWidgetItem()
        self.pgn_matches.setHorizontalHeaderItem(0, item)
        item = QtWidgets.QTableWidgetItem()
        self.pgn_matches.setHorizontalHeaderItem(1, item)
        item = QtWidgets.QTableWidgetItem()
        self.pgn_matches.setHorizontalHeaderItem(2, item)
        item = QtWidgets.QTableWidgetItem()
        self.pgn_matches.setHorizontalHeaderItem(3, item)
        item = QtWidgets.QTableWidgetItem()
        self.pgn_matches.setHorizontalHeaderItem(4, item)
        item = QtWidgets.QTableWidgetItem()
        self.pgn_matches.setHorizontalHeaderItem(5, item)
        item = QtWidgets.QTableWidgetItem()
        self.pgn_matches.setHorizontalHeaderItem(6, item)
        item = QtWidgets.QTableWidgetItem()
        self.pgn_matches.setHorizontalHeaderItem(7, item)
        item = QtWidgets.QTableWidgetItem()
        self.pgn_matches.setHorizontalHeaderItem(8, item)
        self.label_87 = QtWidgets.QLabel(self.page_pgn)
        self.label_87.setGeometry(QtCore.QRect(440, 20, 361, 61))
        self.label_87.setObjectName("label_87")
        self.stackedWidget.addWidget(self.page_pgn)
        self.page_lfl = QtWidgets.QWidget()
        self.page_lfl.setObjectName("page_lfl")
        self.update_lfl = QtWidgets.QPushButton(self.page_lfl)
        self.update_lfl.setGeometry(QtCore.QRect(830, 440, 101, 41))
        self.update_lfl.setObjectName("update_lfl")
        self.back_22 = QtWidgets.QPushButton(self.page_lfl)
        self.back_22.setGeometry(QtCore.QRect(10, 30, 91, 41))
        self.back_22.setObjectName("back_22")
        self.label_90 = QtWidgets.QLabel(self.page_lfl)
        self.label_90.setGeometry(QtCore.QRect(620, 800, 51, 21))
        self.label_90.setObjectName("label_90")
        self.lfl_calculate = QtWidgets.QPushButton(self.page_lfl)
        self.lfl_calculate.setGeometry(QtCore.QRect(840, 790, 141, 41))
        self.lfl_calculate.setObjectName("lfl_calculate")
        self.lfl_matches = QtWidgets.QTableWidget(self.page_lfl)
        self.lfl_matches.setGeometry(QtCore.QRect(90, 500, 1141, 281))
        self.lfl_matches.setObjectName("lfl_matches")
        self.lfl_matches.setColumnCount(9)
        self.lfl_matches.setRowCount(0)
        item = QtWidgets.QTableWidgetItem()
        self.lfl_matches.setHorizontalHeaderItem(0, item)
        item = QtWidgets.QTableWidgetItem()
        self.lfl_matches.setHorizontalHeaderItem(1, item)
        item = QtWidgets.QTableWidgetItem()
        self.lfl_matches.setHorizontalHeaderItem(2, item)
        item = QtWidgets.QTableWidgetItem()
        self.lfl_matches.setHorizontalHeaderItem(3, item)
        item = QtWidgets.QTableWidgetItem()
        self.lfl_matches.setHorizontalHeaderItem(4, item)
        item = QtWidgets.QTableWidgetItem()
        self.lfl_matches.setHorizontalHeaderItem(5, item)
        item = QtWidgets.QTableWidgetItem()
        self.lfl_matches.setHorizontalHeaderItem(6, item)
        item = QtWidgets.QTableWidgetItem()
        self.lfl_matches.setHorizontalHeaderItem(7, item)
        item = QtWidgets.QTableWidgetItem()
        self.lfl_matches.setHorizontalHeaderItem(8, item)
        self.label_91 = QtWidgets.QLabel(self.page_lfl)
        self.label_91.setGeometry(QtCore.QRect(440, 20, 361, 61))
        self.label_91.setObjectName("label_91")
        self.label_92 = QtWidgets.QLabel(self.page_lfl)
        self.label_92.setGeometry(QtCore.QRect(440, 800, 61, 21))
        self.label_92.setObjectName("label_92")
        self.lfl_stats_table = QtWidgets.QTableWidget(self.page_lfl)
        self.lfl_stats_table.setGeometry(QtCore.QRect(100, 80, 831, 351))
        self.lfl_stats_table.setObjectName("lfl_stats_table")
        self.lfl_stats_table.setColumnCount(8)
        self.lfl_stats_table.setRowCount(0)
        item = QtWidgets.QTableWidgetItem()
        self.lfl_stats_table.setHorizontalHeaderItem(0, item)
        item = QtWidgets.QTableWidgetItem()
        self.lfl_stats_table.setHorizontalHeaderItem(1, item)
        item = QtWidgets.QTableWidgetItem()
        self.lfl_stats_table.setHorizontalHeaderItem(2, item)
        item = QtWidgets.QTableWidgetItem()
        self.lfl_stats_table.setHorizontalHeaderItem(3, item)
        item = QtWidgets.QTableWidgetItem()
        self.lfl_stats_table.setHorizontalHeaderItem(4, item)
        item = QtWidgets.QTableWidgetItem()
        self.lfl_stats_table.setHorizontalHeaderItem(5, item)
        item = QtWidgets.QTableWidgetItem()
        self.lfl_stats_table.setHorizontalHeaderItem(6, item)
        item = QtWidgets.QTableWidgetItem()
        self.lfl_stats_table.setHorizontalHeaderItem(7, item)
        self.lfl_line = QtWidgets.QTextEdit(self.page_lfl)
        self.lfl_line.setGeometry(QtCore.QRect(480, 800, 71, 21))
        self.lfl_line.setObjectName("lfl_line")
        self.label_93 = QtWidgets.QLabel(self.page_lfl)
        self.label_93.setGeometry(QtCore.QRect(1080, 20, 361, 61))
        self.label_93.setObjectName("label_93")
        self.ladder_lfl = QtWidgets.QTableWidget(self.page_lfl)
        self.ladder_lfl.setGeometry(QtCore.QRect(1000, 80, 231, 351))
        self.ladder_lfl.setObjectName("ladder_lfl")
        self.ladder_lfl.setColumnCount(2)
        self.ladder_lfl.setRowCount(0)
        item = QtWidgets.QTableWidgetItem()
        self.ladder_lfl.setHorizontalHeaderItem(0, item)
        item = QtWidgets.QTableWidgetItem()
        self.ladder_lfl.setHorizontalHeaderItem(1, item)
        self.lfl_odds = QtWidgets.QTextEdit(self.page_lfl)
        self.lfl_odds.setGeometry(QtCore.QRect(670, 800, 71, 21))
        self.lfl_odds.setObjectName("lfl_odds")
        self.stackedWidget.addWidget(self.page_lfl)
        self.page_lplol = QtWidgets.QWidget()
        self.page_lplol.setObjectName("page_lplol")
        self.update_lplol = QtWidgets.QPushButton(self.page_lplol)
        self.update_lplol.setGeometry(QtCore.QRect(830, 440, 101, 41))
        self.update_lplol.setObjectName("update_lplol")
        self.back_23 = QtWidgets.QPushButton(self.page_lplol)
        self.back_23.setGeometry(QtCore.QRect(10, 30, 91, 41))
        self.back_23.setObjectName("back_23")
        self.label_94 = QtWidgets.QLabel(self.page_lplol)
        self.label_94.setGeometry(QtCore.QRect(620, 800, 51, 21))
        self.label_94.setObjectName("label_94")
        self.lplol_calculate = QtWidgets.QPushButton(self.page_lplol)
        self.lplol_calculate.setGeometry(QtCore.QRect(840, 790, 141, 41))
        self.lplol_calculate.setObjectName("lplol_calculate")
        self.lplol_matches = QtWidgets.QTableWidget(self.page_lplol)
        self.lplol_matches.setGeometry(QtCore.QRect(90, 500, 1141, 281))
        self.lplol_matches.setObjectName("lplol_matches")
        self.lplol_matches.setColumnCount(9)
        self.lplol_matches.setRowCount(0)
        item = QtWidgets.QTableWidgetItem()
        self.lplol_matches.setHorizontalHeaderItem(0, item)
        item = QtWidgets.QTableWidgetItem()
        self.lplol_matches.setHorizontalHeaderItem(1, item)
        item = QtWidgets.QTableWidgetItem()
        self.lplol_matches.setHorizontalHeaderItem(2, item)
        item = QtWidgets.QTableWidgetItem()
        self.lplol_matches.setHorizontalHeaderItem(3, item)
        item = QtWidgets.QTableWidgetItem()
        self.lplol_matches.setHorizontalHeaderItem(4, item)
        item = QtWidgets.QTableWidgetItem()
        self.lplol_matches.setHorizontalHeaderItem(5, item)
        item = QtWidgets.QTableWidgetItem()
        self.lplol_matches.setHorizontalHeaderItem(6, item)
        item = QtWidgets.QTableWidgetItem()
        self.lplol_matches.setHorizontalHeaderItem(7, item)
        item = QtWidgets.QTableWidgetItem()
        self.lplol_matches.setHorizontalHeaderItem(8, item)
        self.label_95 = QtWidgets.QLabel(self.page_lplol)
        self.label_95.setGeometry(QtCore.QRect(440, 20, 361, 61))
        self.label_95.setObjectName("label_95")
        self.label_96 = QtWidgets.QLabel(self.page_lplol)
        self.label_96.setGeometry(QtCore.QRect(440, 800, 61, 21))
        self.label_96.setObjectName("label_96")
        self.lplol_stats_table = QtWidgets.QTableWidget(self.page_lplol)
        self.lplol_stats_table.setGeometry(QtCore.QRect(100, 80, 831, 351))
        self.lplol_stats_table.setObjectName("lplol_stats_table")
        self.lplol_stats_table.setColumnCount(8)
        self.lplol_stats_table.setRowCount(0)
        item = QtWidgets.QTableWidgetItem()
        self.lplol_stats_table.setHorizontalHeaderItem(0, item)
        item = QtWidgets.QTableWidgetItem()
        self.lplol_stats_table.setHorizontalHeaderItem(1, item)
        item = QtWidgets.QTableWidgetItem()
        self.lplol_stats_table.setHorizontalHeaderItem(2, item)
        item = QtWidgets.QTableWidgetItem()
        self.lplol_stats_table.setHorizontalHeaderItem(3, item)
        item = QtWidgets.QTableWidgetItem()
        self.lplol_stats_table.setHorizontalHeaderItem(4, item)
        item = QtWidgets.QTableWidgetItem()
        self.lplol_stats_table.setHorizontalHeaderItem(5, item)
        item = QtWidgets.QTableWidgetItem()
        self.lplol_stats_table.setHorizontalHeaderItem(6, item)
        item = QtWidgets.QTableWidgetItem()
        self.lplol_stats_table.setHorizontalHeaderItem(7, item)
        self.lplol_line = QtWidgets.QTextEdit(self.page_lplol)
        self.lplol_line.setGeometry(QtCore.QRect(480, 800, 71, 21))
        self.lplol_line.setObjectName("lplol_line")
        self.label_97 = QtWidgets.QLabel(self.page_lplol)
        self.label_97.setGeometry(QtCore.QRect(1080, 20, 361, 61))
        self.label_97.setObjectName("label_97")
        self.ladder_lplol = QtWidgets.QTableWidget(self.page_lplol)
        self.ladder_lplol.setGeometry(QtCore.QRect(1000, 80, 231, 351))
        self.ladder_lplol.setObjectName("ladder_lplol")
        self.ladder_lplol.setColumnCount(2)
        self.ladder_lplol.setRowCount(0)
        item = QtWidgets.QTableWidgetItem()
        self.ladder_lplol.setHorizontalHeaderItem(0, item)
        item = QtWidgets.QTableWidgetItem()
        self.ladder_lplol.setHorizontalHeaderItem(1, item)
        self.lplol_odds = QtWidgets.QTextEdit(self.page_lplol)
        self.lplol_odds.setGeometry(QtCore.QRect(670, 800, 71, 21))
        self.lplol_odds.setObjectName("lplol_odds")
        self.stackedWidget.addWidget(self.page_lplol)
        self.page_hpm = QtWidgets.QWidget()
        self.page_hpm.setObjectName("page_hpm")
        self.update_hpm = QtWidgets.QPushButton(self.page_hpm)
        self.update_hpm.setGeometry(QtCore.QRect(830, 440, 101, 41))
        self.update_hpm.setObjectName("update_hpm")
        self.back_24 = QtWidgets.QPushButton(self.page_hpm)
        self.back_24.setGeometry(QtCore.QRect(10, 30, 91, 41))
        self.back_24.setObjectName("back_24")
        self.label_98 = QtWidgets.QLabel(self.page_hpm)
        self.label_98.setGeometry(QtCore.QRect(620, 800, 51, 21))
        self.label_98.setObjectName("label_98")
        self.hpm_calculate = QtWidgets.QPushButton(self.page_hpm)
        self.hpm_calculate.setGeometry(QtCore.QRect(840, 790, 141, 41))
        self.hpm_calculate.setObjectName("hpm_calculate")
        self.hpm_matches = QtWidgets.QTableWidget(self.page_hpm)
        self.hpm_matches.setGeometry(QtCore.QRect(90, 500, 1141, 281))
        self.hpm_matches.setObjectName("hpm_matches")
        self.hpm_matches.setColumnCount(9)
        self.hpm_matches.setRowCount(0)
        item = QtWidgets.QTableWidgetItem()
        self.hpm_matches.setHorizontalHeaderItem(0, item)
        item = QtWidgets.QTableWidgetItem()
        self.hpm_matches.setHorizontalHeaderItem(1, item)
        item = QtWidgets.QTableWidgetItem()
        self.hpm_matches.setHorizontalHeaderItem(2, item)
        item = QtWidgets.QTableWidgetItem()
        self.hpm_matches.setHorizontalHeaderItem(3, item)
        item = QtWidgets.QTableWidgetItem()
        self.hpm_matches.setHorizontalHeaderItem(4, item)
        item = QtWidgets.QTableWidgetItem()
        self.hpm_matches.setHorizontalHeaderItem(5, item)
        item = QtWidgets.QTableWidgetItem()
        self.hpm_matches.setHorizontalHeaderItem(6, item)
        item = QtWidgets.QTableWidgetItem()
        self.hpm_matches.setHorizontalHeaderItem(7, item)
        item = QtWidgets.QTableWidgetItem()
        self.hpm_matches.setHorizontalHeaderItem(8, item)
        self.label_99 = QtWidgets.QLabel(self.page_hpm)
        self.label_99.setGeometry(QtCore.QRect(440, 20, 361, 61))
        self.label_99.setObjectName("label_99")
        self.label_100 = QtWidgets.QLabel(self.page_hpm)
        self.label_100.setGeometry(QtCore.QRect(440, 800, 61, 21))
        self.label_100.setObjectName("label_100")
        self.hpm_stats_table = QtWidgets.QTableWidget(self.page_hpm)
        self.hpm_stats_table.setGeometry(QtCore.QRect(100, 80, 831, 351))
        self.hpm_stats_table.setObjectName("hpm_stats_table")
        self.hpm_stats_table.setColumnCount(8)
        self.hpm_stats_table.setRowCount(0)
        item = QtWidgets.QTableWidgetItem()
        self.hpm_stats_table.setHorizontalHeaderItem(0, item)
        item = QtWidgets.QTableWidgetItem()
        self.hpm_stats_table.setHorizontalHeaderItem(1, item)
        item = QtWidgets.QTableWidgetItem()
        self.hpm_stats_table.setHorizontalHeaderItem(2, item)
        item = QtWidgets.QTableWidgetItem()
        self.hpm_stats_table.setHorizontalHeaderItem(3, item)
        item = QtWidgets.QTableWidgetItem()
        self.hpm_stats_table.setHorizontalHeaderItem(4, item)
        item = QtWidgets.QTableWidgetItem()
        self.hpm_stats_table.setHorizontalHeaderItem(5, item)
        item = QtWidgets.QTableWidgetItem()
        self.hpm_stats_table.setHorizontalHeaderItem(6, item)
        item = QtWidgets.QTableWidgetItem()
        self.hpm_stats_table.setHorizontalHeaderItem(7, item)
        self.hpm_line = QtWidgets.QTextEdit(self.page_hpm)
        self.hpm_line.setGeometry(QtCore.QRect(480, 800, 71, 21))
        self.hpm_line.setObjectName("hpm_line")
        self.label_101 = QtWidgets.QLabel(self.page_hpm)
        self.label_101.setGeometry(QtCore.QRect(1080, 20, 361, 61))
        self.label_101.setObjectName("label_101")
        self.ladder_hpm = QtWidgets.QTableWidget(self.page_hpm)
        self.ladder_hpm.setGeometry(QtCore.QRect(1000, 80, 231, 351))
        self.ladder_hpm.setObjectName("ladder_hpm")
        self.ladder_hpm.setColumnCount(2)
        self.ladder_hpm.setRowCount(0)
        item = QtWidgets.QTableWidgetItem()
        self.ladder_hpm.setHorizontalHeaderItem(0, item)
        item = QtWidgets.QTableWidgetItem()
        self.ladder_hpm.setHorizontalHeaderItem(1, item)
        self.hpm_odds = QtWidgets.QTextEdit(self.page_hpm)
        self.hpm_odds.setGeometry(QtCore.QRect(670, 800, 71, 21))
        self.hpm_odds.setObjectName("hpm_odds")
        self.stackedWidget.addWidget(self.page_hpm)
        self.page_2 = QtWidgets.QWidget()
        self.page_2.setObjectName("page_2")
        self.label_2 = QtWidgets.QLabel(self.page_2)
        self.label_2.setGeometry(QtCore.QRect(440, 20, 361, 61))
        self.label_2.setObjectName("label_2")
        self.tableWidget = QtWidgets.QTableWidget(self.page_2)
        self.tableWidget.setGeometry(QtCore.QRect(100, 80, 831, 351))
        self.tableWidget.setObjectName("tableWidget")
        self.tableWidget.setColumnCount(8)
        self.tableWidget.setRowCount(0)
        item = QtWidgets.QTableWidgetItem()
        self.tableWidget.setHorizontalHeaderItem(0, item)
        item = QtWidgets.QTableWidgetItem()
        self.tableWidget.setHorizontalHeaderItem(1, item)
        item = QtWidgets.QTableWidgetItem()
        self.tableWidget.setHorizontalHeaderItem(2, item)
        item = QtWidgets.QTableWidgetItem()
        self.tableWidget.setHorizontalHeaderItem(3, item)
        item = QtWidgets.QTableWidgetItem()
        self.tableWidget.setHorizontalHeaderItem(4, item)
        item = QtWidgets.QTableWidgetItem()
        self.tableWidget.setHorizontalHeaderItem(5, item)
        item = QtWidgets.QTableWidgetItem()
        self.tableWidget.setHorizontalHeaderItem(6, item)
        item = QtWidgets.QTableWidgetItem()
        self.tableWidget.setHorizontalHeaderItem(7, item)
        self.update_lck = QtWidgets.QPushButton(self.page_2)
        self.update_lck.setGeometry(QtCore.QRect(830, 440, 101, 41))
        self.update_lck.setObjectName("update_lck")
        self.back = QtWidgets.QPushButton(self.page_2)
        self.back.setGeometry(QtCore.QRect(10, 30, 91, 41))
        self.back.setObjectName("back")
        self.ladder_1 = QtWidgets.QTableWidget(self.page_2)
        self.ladder_1.setGeometry(QtCore.QRect(1000, 80, 231, 351))
        self.ladder_1.setObjectName("ladder_1")
        self.ladder_1.setColumnCount(2)
        self.ladder_1.setRowCount(0)
        item = QtWidgets.QTableWidgetItem()
        self.ladder_1.setHorizontalHeaderItem(0, item)
        item = QtWidgets.QTableWidgetItem()
        self.ladder_1.setHorizontalHeaderItem(1, item)
        self.label_4 = QtWidgets.QLabel(self.page_2)
        self.label_4.setGeometry(QtCore.QRect(1080, 20, 361, 61))
        self.label_4.setObjectName("label_4")
        self.lck_matches = QtWidgets.QTableWidget(self.page_2)
        self.lck_matches.setGeometry(QtCore.QRect(90, 500, 1141, 281))
        self.lck_matches.setObjectName("lck_matches")
        self.lck_matches.setColumnCount(9)
        self.lck_matches.setRowCount(0)
        item = QtWidgets.QTableWidgetItem()
        self.lck_matches.setHorizontalHeaderItem(0, item)
        item = QtWidgets.QTableWidgetItem()
        self.lck_matches.setHorizontalHeaderItem(1, item)
        item = QtWidgets.QTableWidgetItem()
        self.lck_matches.setHorizontalHeaderItem(2, item)
        item = QtWidgets.QTableWidgetItem()
        self.lck_matches.setHorizontalHeaderItem(3, item)
        item = QtWidgets.QTableWidgetItem()
        self.lck_matches.setHorizontalHeaderItem(4, item)
        item = QtWidgets.QTableWidgetItem()
        self.lck_matches.setHorizontalHeaderItem(5, item)
        item = QtWidgets.QTableWidgetItem()
        self.lck_matches.setHorizontalHeaderItem(6, item)
        item = QtWidgets.QTableWidgetItem()
        self.lck_matches.setHorizontalHeaderItem(7, item)
        item = QtWidgets.QTableWidgetItem()
        self.lck_matches.setHorizontalHeaderItem(8, item)
        self.label_8 = QtWidgets.QLabel(self.page_2)
        self.label_8.setGeometry(QtCore.QRect(440, 800, 61, 21))
        self.label_8.setObjectName("label_8")
        self.label_9 = QtWidgets.QLabel(self.page_2)
        self.label_9.setGeometry(QtCore.QRect(620, 800, 51, 21))
        self.label_9.setObjectName("label_9")
        self.lck_line = QtWidgets.QTextEdit(self.page_2)
        self.lck_line.setGeometry(QtCore.QRect(480, 800, 71, 21))
        self.lck_line.setObjectName("lck_line")
        self.lck_odds = QtWidgets.QTextEdit(self.page_2)
        self.lck_odds.setGeometry(QtCore.QRect(670, 800, 71, 21))
        self.lck_odds.setObjectName("lck_odds")
        self.lck_calculate = QtWidgets.QPushButton(self.page_2)
        self.lck_calculate.setGeometry(QtCore.QRect(840, 790, 141, 41))
        self.lck_calculate.setObjectName("lck_calculate")
        self.stackedWidget.addWidget(self.page_2)
        self.page_3 = QtWidgets.QWidget()
        self.page_3.setObjectName("page_3")
        self.tableWidget_2 = QtWidgets.QTableWidget(self.page_3)
        self.tableWidget_2.setGeometry(QtCore.QRect(100, 80, 831, 351))
        self.tableWidget_2.setObjectName("tableWidget_2")
        self.tableWidget_2.setColumnCount(8)
        self.tableWidget_2.setRowCount(0)
        item = QtWidgets.QTableWidgetItem()
        self.tableWidget_2.setHorizontalHeaderItem(0, item)
        item = QtWidgets.QTableWidgetItem()
        self.tableWidget_2.setHorizontalHeaderItem(1, item)
        item = QtWidgets.QTableWidgetItem()
        self.tableWidget_2.setHorizontalHeaderItem(2, item)
        item = QtWidgets.QTableWidgetItem()
        self.tableWidget_2.setHorizontalHeaderItem(3, item)
        item = QtWidgets.QTableWidgetItem()
        self.tableWidget_2.setHorizontalHeaderItem(4, item)
        item = QtWidgets.QTableWidgetItem()
        self.tableWidget_2.setHorizontalHeaderItem(5, item)
        item = QtWidgets.QTableWidgetItem()
        self.tableWidget_2.setHorizontalHeaderItem(6, item)
        item = QtWidgets.QTableWidgetItem()
        self.tableWidget_2.setHorizontalHeaderItem(7, item)
        self.label_3 = QtWidgets.QLabel(self.page_3)
        self.label_3.setGeometry(QtCore.QRect(440, 20, 361, 61))
        self.label_3.setObjectName("label_3")
        self.back_2 = QtWidgets.QPushButton(self.page_3)
        self.back_2.setGeometry(QtCore.QRect(10, 30, 91, 41))
        self.back_2.setObjectName("back_2")
        self.update_lcs = QtWidgets.QPushButton(self.page_3)
        self.update_lcs.setGeometry(QtCore.QRect(830, 440, 101, 41))
        self.update_lcs.setStyleSheet("")
        self.update_lcs.setObjectName("update_lcs")
        self.ladder_2 = QtWidgets.QTableWidget(self.page_3)
        self.ladder_2.setGeometry(QtCore.QRect(1000, 80, 231, 351))
        self.ladder_2.setObjectName("ladder_2")
        self.ladder_2.setColumnCount(2)
        self.ladder_2.setRowCount(0)
        item = QtWidgets.QTableWidgetItem()
        self.ladder_2.setHorizontalHeaderItem(0, item)
        item = QtWidgets.QTableWidgetItem()
        self.ladder_2.setHorizontalHeaderItem(1, item)
        self.label_5 = QtWidgets.QLabel(self.page_3)
        self.label_5.setGeometry(QtCore.QRect(1080, 20, 361, 61))
        self.label_5.setObjectName("label_5")
        self.lcs_matches = QtWidgets.QTableWidget(self.page_3)
        self.lcs_matches.setGeometry(QtCore.QRect(90, 500, 1141, 281))
        self.lcs_matches.setObjectName("lcs_matches")
        self.lcs_matches.setColumnCount(9)
        self.lcs_matches.setRowCount(0)
        item = QtWidgets.QTableWidgetItem()
        self.lcs_matches.setHorizontalHeaderItem(0, item)
        item = QtWidgets.QTableWidgetItem()
        self.lcs_matches.setHorizontalHeaderItem(1, item)
        item = QtWidgets.QTableWidgetItem()
        self.lcs_matches.setHorizontalHeaderItem(2, item)
        item = QtWidgets.QTableWidgetItem()
        self.lcs_matches.setHorizontalHeaderItem(3, item)
        item = QtWidgets.QTableWidgetItem()
        self.lcs_matches.setHorizontalHeaderItem(4, item)
        item = QtWidgets.QTableWidgetItem()
        self.lcs_matches.setHorizontalHeaderItem(5, item)
        item = QtWidgets.QTableWidgetItem()
        self.lcs_matches.setHorizontalHeaderItem(6, item)
        item = QtWidgets.QTableWidgetItem()
        self.lcs_matches.setHorizontalHeaderItem(7, item)
        item = QtWidgets.QTableWidgetItem()
        self.lcs_matches.setHorizontalHeaderItem(8, item)
        self.lcs_odds = QtWidgets.QTextEdit(self.page_3)
        self.lcs_odds.setGeometry(QtCore.QRect(650, 800, 71, 21))
        self.lcs_odds.setObjectName("lcs_odds")
        self.lcs_calculate = QtWidgets.QPushButton(self.page_3)
        self.lcs_calculate.setGeometry(QtCore.QRect(820, 790, 141, 41))
        self.lcs_calculate.setObjectName("lcs_calculate")
        self.label_10 = QtWidgets.QLabel(self.page_3)
        self.label_10.setGeometry(QtCore.QRect(600, 800, 51, 21))
        self.label_10.setObjectName("label_10")
        self.label_14 = QtWidgets.QLabel(self.page_3)
        self.label_14.setGeometry(QtCore.QRect(420, 800, 61, 21))
        self.label_14.setObjectName("label_14")
        self.lcs_line = QtWidgets.QTextEdit(self.page_3)
        self.lcs_line.setGeometry(QtCore.QRect(460, 800, 71, 21))
        self.lcs_line.setObjectName("lcs_line")
        self.stackedWidget.addWidget(self.page_3)
        self.page_4 = QtWidgets.QWidget()
        self.page_4.setObjectName("page_4")
        self.tableWidget_3 = QtWidgets.QTableWidget(self.page_4)
        self.tableWidget_3.setGeometry(QtCore.QRect(100, 80, 831, 351))
        self.tableWidget_3.setObjectName("tableWidget_3")
        self.tableWidget_3.setColumnCount(8)
        self.tableWidget_3.setRowCount(0)
        item = QtWidgets.QTableWidgetItem()
        self.tableWidget_3.setHorizontalHeaderItem(0, item)
        item = QtWidgets.QTableWidgetItem()
        self.tableWidget_3.setHorizontalHeaderItem(1, item)
        item = QtWidgets.QTableWidgetItem()
        self.tableWidget_3.setHorizontalHeaderItem(2, item)
        item = QtWidgets.QTableWidgetItem()
        self.tableWidget_3.setHorizontalHeaderItem(3, item)
        item = QtWidgets.QTableWidgetItem()
        self.tableWidget_3.setHorizontalHeaderItem(4, item)
        item = QtWidgets.QTableWidgetItem()
        self.tableWidget_3.setHorizontalHeaderItem(5, item)
        item = QtWidgets.QTableWidgetItem()
        self.tableWidget_3.setHorizontalHeaderItem(6, item)
        item = QtWidgets.QTableWidgetItem()
        self.tableWidget_3.setHorizontalHeaderItem(7, item)
        self.label_6 = QtWidgets.QLabel(self.page_4)
        self.label_6.setGeometry(QtCore.QRect(430, 20, 361, 61))
        self.label_6.setObjectName("label_6")
        self.back_3 = QtWidgets.QPushButton(self.page_4)
        self.back_3.setGeometry(QtCore.QRect(10, 30, 91, 41))
        self.back_3.setObjectName("back_3")
        self.update_lec = QtWidgets.QPushButton(self.page_4)
        self.update_lec.setGeometry(QtCore.QRect(830, 440, 101, 41))
        self.update_lec.setObjectName("update_lec")
        self.label_7 = QtWidgets.QLabel(self.page_4)
        self.label_7.setGeometry(QtCore.QRect(1080, 20, 361, 61))
        self.label_7.setObjectName("label_7")
        self.ladder_3 = QtWidgets.QTableWidget(self.page_4)
        self.ladder_3.setGeometry(QtCore.QRect(1000, 80, 231, 351))
        self.ladder_3.setObjectName("ladder_3")
        self.ladder_3.setColumnCount(2)
        self.ladder_3.setRowCount(0)
        item = QtWidgets.QTableWidgetItem()
        self.ladder_3.setHorizontalHeaderItem(0, item)
        item = QtWidgets.QTableWidgetItem()
        self.ladder_3.setHorizontalHeaderItem(1, item)
        self.lec_matches = QtWidgets.QTableWidget(self.page_4)
        self.lec_matches.setGeometry(QtCore.QRect(90, 500, 1141, 281))
        self.lec_matches.setObjectName("lec_matches")
        self.lec_matches.setColumnCount(9)
        self.lec_matches.setRowCount(0)
        item = QtWidgets.QTableWidgetItem()
        self.lec_matches.setHorizontalHeaderItem(0, item)
        item = QtWidgets.QTableWidgetItem()
        self.lec_matches.setHorizontalHeaderItem(1, item)
        item = QtWidgets.QTableWidgetItem()
        self.lec_matches.setHorizontalHeaderItem(2, item)
        item = QtWidgets.QTableWidgetItem()
        self.lec_matches.setHorizontalHeaderItem(3, item)
        item = QtWidgets.QTableWidgetItem()
        self.lec_matches.setHorizontalHeaderItem(4, item)
        item = QtWidgets.QTableWidgetItem()
        self.lec_matches.setHorizontalHeaderItem(5, item)
        item = QtWidgets.QTableWidgetItem()
        self.lec_matches.setHorizontalHeaderItem(6, item)
        item = QtWidgets.QTableWidgetItem()
        self.lec_matches.setHorizontalHeaderItem(7, item)
        item = QtWidgets.QTableWidgetItem()
        self.lec_matches.setHorizontalHeaderItem(8, item)
        self.lec_odds = QtWidgets.QTextEdit(self.page_4)
        self.lec_odds.setGeometry(QtCore.QRect(640, 800, 71, 21))
        self.lec_odds.setObjectName("lec_odds")
        self.lec_calculate = QtWidgets.QPushButton(self.page_4)
        self.lec_calculate.setGeometry(QtCore.QRect(810, 790, 141, 41))
        self.lec_calculate.setObjectName("lec_calculate")
        self.label_11 = QtWidgets.QLabel(self.page_4)
        self.label_11.setGeometry(QtCore.QRect(590, 800, 51, 21))
        self.label_11.setObjectName("label_11")
        self.label_12 = QtWidgets.QLabel(self.page_4)
        self.label_12.setGeometry(QtCore.QRect(410, 800, 61, 21))
        self.label_12.setObjectName("label_12")
        self.lec_line = QtWidgets.QTextEdit(self.page_4)
        self.lec_line.setGeometry(QtCore.QRect(450, 800, 71, 21))
        self.lec_line.setObjectName("lec_line")
        self.stackedWidget.addWidget(self.page_4)
        self.page_5 = QtWidgets.QWidget()
        self.page_5.setObjectName("page_5")
        self.label_53 = QtWidgets.QLabel(self.page_5)
        self.label_53.setGeometry(QtCore.QRect(440, 20, 361, 61))
        self.label_53.setObjectName("label_53")
        self.label_54 = QtWidgets.QLabel(self.page_5)
        self.label_54.setGeometry(QtCore.QRect(620, 800, 51, 21))
        self.label_54.setObjectName("label_54")
        self.lpl_stats_table = QtWidgets.QTableWidget(self.page_5)
        self.lpl_stats_table.setGeometry(QtCore.QRect(100, 80, 831, 351))
        self.lpl_stats_table.setObjectName("lpl_stats_table")
        self.lpl_stats_table.setColumnCount(8)
        self.lpl_stats_table.setRowCount(0)
        item = QtWidgets.QTableWidgetItem()
        self.lpl_stats_table.setHorizontalHeaderItem(0, item)
        item = QtWidgets.QTableWidgetItem()
        self.lpl_stats_table.setHorizontalHeaderItem(1, item)
        item = QtWidgets.QTableWidgetItem()
        self.lpl_stats_table.setHorizontalHeaderItem(2, item)
        item = QtWidgets.QTableWidgetItem()
        self.lpl_stats_table.setHorizontalHeaderItem(3, item)
        item = QtWidgets.QTableWidgetItem()
        self.lpl_stats_table.setHorizontalHeaderItem(4, item)
        item = QtWidgets.QTableWidgetItem()
        self.lpl_stats_table.setHorizontalHeaderItem(5, item)
        item = QtWidgets.QTableWidgetItem()
        self.lpl_stats_table.setHorizontalHeaderItem(6, item)
        item = QtWidgets.QTableWidgetItem()
        self.lpl_stats_table.setHorizontalHeaderItem(7, item)
        self.label_55 = QtWidgets.QLabel(self.page_5)
        self.label_55.setGeometry(QtCore.QRect(1080, 20, 361, 61))
        self.label_55.setObjectName("label_55")
        self.lpl_odds = QtWidgets.QTextEdit(self.page_5)
        self.lpl_odds.setGeometry(QtCore.QRect(670, 800, 71, 21))
        self.lpl_odds.setObjectName("lpl_odds")
        self.ladder_lpl = QtWidgets.QTableWidget(self.page_5)
        self.ladder_lpl.setGeometry(QtCore.QRect(1000, 80, 231, 351))
        self.ladder_lpl.setObjectName("ladder_lpl")
        self.ladder_lpl.setColumnCount(2)
        self.ladder_lpl.setRowCount(0)
        item = QtWidgets.QTableWidgetItem()
        self.ladder_lpl.setHorizontalHeaderItem(0, item)
        item = QtWidgets.QTableWidgetItem()
        self.ladder_lpl.setHorizontalHeaderItem(1, item)
        self.lpl_matches = QtWidgets.QTableWidget(self.page_5)
        self.lpl_matches.setGeometry(QtCore.QRect(90, 500, 1141, 281))
        self.lpl_matches.setObjectName("lpl_matches")
        self.lpl_matches.setColumnCount(9)
        self.lpl_matches.setRowCount(0)
        item = QtWidgets.QTableWidgetItem()
        self.lpl_matches.setHorizontalHeaderItem(0, item)
        item = QtWidgets.QTableWidgetItem()
        self.lpl_matches.setHorizontalHeaderItem(1, item)
        item = QtWidgets.QTableWidgetItem()
        self.lpl_matches.setHorizontalHeaderItem(2, item)
        item = QtWidgets.QTableWidgetItem()
        self.lpl_matches.setHorizontalHeaderItem(3, item)
        item = QtWidgets.QTableWidgetItem()
        self.lpl_matches.setHorizontalHeaderItem(4, item)
        item = QtWidgets.QTableWidgetItem()
        self.lpl_matches.setHorizontalHeaderItem(5, item)
        item = QtWidgets.QTableWidgetItem()
        self.lpl_matches.setHorizontalHeaderItem(6, item)
        item = QtWidgets.QTableWidgetItem()
        self.lpl_matches.setHorizontalHeaderItem(7, item)
        item = QtWidgets.QTableWidgetItem()
        self.lpl_matches.setHorizontalHeaderItem(8, item)
        self.lpl_calculate = QtWidgets.QPushButton(self.page_5)
        self.lpl_calculate.setGeometry(QtCore.QRect(840, 790, 141, 41))
        self.lpl_calculate.setObjectName("lpl_calculate")
        self.label_56 = QtWidgets.QLabel(self.page_5)
        self.label_56.setGeometry(QtCore.QRect(440, 800, 61, 21))
        self.label_56.setObjectName("label_56")
        self.lpl_line = QtWidgets.QTextEdit(self.page_5)
        self.lpl_line.setGeometry(QtCore.QRect(480, 800, 71, 21))
        self.lpl_line.setObjectName("lpl_line")
        self.back_13 = QtWidgets.QPushButton(self.page_5)
        self.back_13.setGeometry(QtCore.QRect(10, 30, 91, 41))
        self.back_13.setObjectName("back_13")
        self.update_lpl = QtWidgets.QPushButton(self.page_5)
        self.update_lpl.setGeometry(QtCore.QRect(830, 440, 101, 41))
        self.update_lpl.setObjectName("update_lpl")
        self.stackedWidget.addWidget(self.page_5)
        self.page_vcs = QtWidgets.QWidget()
        self.page_vcs.setObjectName("page_vcs")
        self.update_vcs = QtWidgets.QPushButton(self.page_vcs)
        self.update_vcs.setGeometry(QtCore.QRect(830, 440, 101, 41))
        self.update_vcs.setObjectName("update_vcs")
        self.vcs_calculate = QtWidgets.QPushButton(self.page_vcs)
        self.vcs_calculate.setGeometry(QtCore.QRect(840, 790, 141, 41))
        self.vcs_calculate.setObjectName("vcs_calculate")
        self.label_60 = QtWidgets.QLabel(self.page_vcs)
        self.label_60.setGeometry(QtCore.QRect(440, 800, 61, 21))
        self.label_60.setObjectName("label_60")
        self.vcs_line = QtWidgets.QTextEdit(self.page_vcs)
        self.vcs_line.setGeometry(QtCore.QRect(480, 800, 71, 21))
        self.vcs_line.setObjectName("vcs_line")
        self.back_14 = QtWidgets.QPushButton(self.page_vcs)
        self.back_14.setGeometry(QtCore.QRect(10, 30, 91, 41))
        self.back_14.setObjectName("back_14")
        self.label_57 = QtWidgets.QLabel(self.page_vcs)
        self.label_57.setGeometry(QtCore.QRect(440, 20, 361, 61))
        self.label_57.setObjectName("label_57")
        self.vcs_odds = QtWidgets.QTextEdit(self.page_vcs)
        self.vcs_odds.setGeometry(QtCore.QRect(670, 800, 71, 21))
        self.vcs_odds.setObjectName("vcs_odds")
        self.label_58 = QtWidgets.QLabel(self.page_vcs)
        self.label_58.setGeometry(QtCore.QRect(620, 800, 51, 21))
        self.label_58.setObjectName("label_58")
        self.ladder_vcs = QtWidgets.QTableWidget(self.page_vcs)
        self.ladder_vcs.setGeometry(QtCore.QRect(1000, 80, 231, 351))
        self.ladder_vcs.setObjectName("ladder_vcs")
        self.ladder_vcs.setColumnCount(2)
        self.ladder_vcs.setRowCount(0)
        item = QtWidgets.QTableWidgetItem()
        self.ladder_vcs.setHorizontalHeaderItem(0, item)
        item = QtWidgets.QTableWidgetItem()
        self.ladder_vcs.setHorizontalHeaderItem(1, item)
        self.vcs_stats_table = QtWidgets.QTableWidget(self.page_vcs)
        self.vcs_stats_table.setGeometry(QtCore.QRect(100, 80, 831, 351))
        self.vcs_stats_table.setObjectName("vcs_stats_table")
        self.vcs_stats_table.setColumnCount(8)
        self.vcs_stats_table.setRowCount(0)
        item = QtWidgets.QTableWidgetItem()
        self.vcs_stats_table.setHorizontalHeaderItem(0, item)
        item = QtWidgets.QTableWidgetItem()
        self.vcs_stats_table.setHorizontalHeaderItem(1, item)
        item = QtWidgets.QTableWidgetItem()
        self.vcs_stats_table.setHorizontalHeaderItem(2, item)
        item = QtWidgets.QTableWidgetItem()
        self.vcs_stats_table.setHorizontalHeaderItem(3, item)
        item = QtWidgets.QTableWidgetItem()
        self.vcs_stats_table.setHorizontalHeaderItem(4, item)
        item = QtWidgets.QTableWidgetItem()
        self.vcs_stats_table.setHorizontalHeaderItem(5, item)
        item = QtWidgets.QTableWidgetItem()
        self.vcs_stats_table.setHorizontalHeaderItem(6, item)
        item = QtWidgets.QTableWidgetItem()
        self.vcs_stats_table.setHorizontalHeaderItem(7, item)
        self.label_59 = QtWidgets.QLabel(self.page_vcs)
        self.label_59.setGeometry(QtCore.QRect(1080, 20, 361, 61))
        self.label_59.setObjectName("label_59")
        self.vcs_matches = QtWidgets.QTableWidget(self.page_vcs)
        self.vcs_matches.setGeometry(QtCore.QRect(90, 500, 1141, 281))
        self.vcs_matches.setObjectName("vcs_matches")
        self.vcs_matches.setColumnCount(9)
        self.vcs_matches.setRowCount(0)
        item = QtWidgets.QTableWidgetItem()
        self.vcs_matches.setHorizontalHeaderItem(0, item)
        item = QtWidgets.QTableWidgetItem()
        self.vcs_matches.setHorizontalHeaderItem(1, item)
        item = QtWidgets.QTableWidgetItem()
        self.vcs_matches.setHorizontalHeaderItem(2, item)
        item = QtWidgets.QTableWidgetItem()
        self.vcs_matches.setHorizontalHeaderItem(3, item)
        item = QtWidgets.QTableWidgetItem()
        self.vcs_matches.setHorizontalHeaderItem(4, item)
        item = QtWidgets.QTableWidgetItem()
        self.vcs_matches.setHorizontalHeaderItem(5, item)
        item = QtWidgets.QTableWidgetItem()
        self.vcs_matches.setHorizontalHeaderItem(6, item)
        item = QtWidgets.QTableWidgetItem()
        self.vcs_matches.setHorizontalHeaderItem(7, item)
        item = QtWidgets.QTableWidgetItem()
        self.vcs_matches.setHorizontalHeaderItem(8, item)
        self.stackedWidget.addWidget(self.page_vcs)
        self.page_na_acad = QtWidgets.QWidget()
        self.page_na_acad.setObjectName("page_na_acad")
        self.update_na_acad = QtWidgets.QPushButton(self.page_na_acad)
        self.update_na_acad.setGeometry(QtCore.QRect(830, 440, 101, 41))
        self.update_na_acad.setObjectName("update_na_acad")
        self.ladder_na_acad = QtWidgets.QTableWidget(self.page_na_acad)
        self.ladder_na_acad.setGeometry(QtCore.QRect(1000, 80, 231, 351))
        self.ladder_na_acad.setObjectName("ladder_na_acad")
        self.ladder_na_acad.setColumnCount(2)
        self.ladder_na_acad.setRowCount(0)
        item = QtWidgets.QTableWidgetItem()
        self.ladder_na_acad.setHorizontalHeaderItem(0, item)
        item = QtWidgets.QTableWidgetItem()
        self.ladder_na_acad.setHorizontalHeaderItem(1, item)
        self.label_62 = QtWidgets.QLabel(self.page_na_acad)
        self.label_62.setGeometry(QtCore.QRect(620, 800, 51, 21))
        self.label_62.setObjectName("label_62")
        self.na_acad_stats_table = QtWidgets.QTableWidget(self.page_na_acad)
        self.na_acad_stats_table.setGeometry(QtCore.QRect(100, 80, 831, 351))
        self.na_acad_stats_table.setObjectName("na_acad_stats_table")
        self.na_acad_stats_table.setColumnCount(8)
        self.na_acad_stats_table.setRowCount(0)
        item = QtWidgets.QTableWidgetItem()
        self.na_acad_stats_table.setHorizontalHeaderItem(0, item)
        item = QtWidgets.QTableWidgetItem()
        self.na_acad_stats_table.setHorizontalHeaderItem(1, item)
        item = QtWidgets.QTableWidgetItem()
        self.na_acad_stats_table.setHorizontalHeaderItem(2, item)
        item = QtWidgets.QTableWidgetItem()
        self.na_acad_stats_table.setHorizontalHeaderItem(3, item)
        item = QtWidgets.QTableWidgetItem()
        self.na_acad_stats_table.setHorizontalHeaderItem(4, item)
        item = QtWidgets.QTableWidgetItem()
        self.na_acad_stats_table.setHorizontalHeaderItem(5, item)
        item = QtWidgets.QTableWidgetItem()
        self.na_acad_stats_table.setHorizontalHeaderItem(6, item)
        item = QtWidgets.QTableWidgetItem()
        self.na_acad_stats_table.setHorizontalHeaderItem(7, item)
        self.na_acad_odds = QtWidgets.QTextEdit(self.page_na_acad)
        self.na_acad_odds.setGeometry(QtCore.QRect(670, 800, 71, 21))
        self.na_acad_odds.setObjectName("na_acad_odds")
        self.back_15 = QtWidgets.QPushButton(self.page_na_acad)
        self.back_15.setGeometry(QtCore.QRect(10, 30, 91, 41))
        self.back_15.setObjectName("back_15")
        self.label_61 = QtWidgets.QLabel(self.page_na_acad)
        self.label_61.setGeometry(QtCore.QRect(430, 20, 361, 61))
        self.label_61.setObjectName("label_61")
        self.na_acad_matches = QtWidgets.QTableWidget(self.page_na_acad)
        self.na_acad_matches.setGeometry(QtCore.QRect(90, 500, 1141, 281))
        self.na_acad_matches.setObjectName("na_acad_matches")
        self.na_acad_matches.setColumnCount(9)
        self.na_acad_matches.setRowCount(0)
        item = QtWidgets.QTableWidgetItem()
        self.na_acad_matches.setHorizontalHeaderItem(0, item)
        item = QtWidgets.QTableWidgetItem()
        self.na_acad_matches.setHorizontalHeaderItem(1, item)
        item = QtWidgets.QTableWidgetItem()
        self.na_acad_matches.setHorizontalHeaderItem(2, item)
        item = QtWidgets.QTableWidgetItem()
        self.na_acad_matches.setHorizontalHeaderItem(3, item)
        item = QtWidgets.QTableWidgetItem()
        self.na_acad_matches.setHorizontalHeaderItem(4, item)
        item = QtWidgets.QTableWidgetItem()
        self.na_acad_matches.setHorizontalHeaderItem(5, item)
        item = QtWidgets.QTableWidgetItem()
        self.na_acad_matches.setHorizontalHeaderItem(6, item)
        item = QtWidgets.QTableWidgetItem()
        self.na_acad_matches.setHorizontalHeaderItem(7, item)
        item = QtWidgets.QTableWidgetItem()
        self.na_acad_matches.setHorizontalHeaderItem(8, item)
        self.label_63 = QtWidgets.QLabel(self.page_na_acad)
        self.label_63.setGeometry(QtCore.QRect(1080, 20, 361, 61))
        self.label_63.setObjectName("label_63")
        self.label_64 = QtWidgets.QLabel(self.page_na_acad)
        self.label_64.setGeometry(QtCore.QRect(440, 800, 61, 21))
        self.label_64.setObjectName("label_64")
        self.na_acad_line = QtWidgets.QTextEdit(self.page_na_acad)
        self.na_acad_line.setGeometry(QtCore.QRect(480, 800, 71, 21))
        self.na_acad_line.setObjectName("na_acad_line")
        self.na_acad_calculate = QtWidgets.QPushButton(self.page_na_acad)
        self.na_acad_calculate.setGeometry(QtCore.QRect(840, 790, 141, 41))
        self.na_acad_calculate.setObjectName("na_acad_calculate")
        self.stackedWidget.addWidget(self.page_na_acad)
        self.page_cblol = QtWidgets.QWidget()
        self.page_cblol.setObjectName("page_cblol")
        self.update_cblol = QtWidgets.QPushButton(self.page_cblol)
        self.update_cblol.setGeometry(QtCore.QRect(830, 440, 101, 41))
        self.update_cblol.setObjectName("update_cblol")
        self.label_65 = QtWidgets.QLabel(self.page_cblol)
        self.label_65.setGeometry(QtCore.QRect(440, 20, 361, 61))
        self.label_65.setObjectName("label_65")
        self.label_66 = QtWidgets.QLabel(self.page_cblol)
        self.label_66.setGeometry(QtCore.QRect(620, 800, 51, 21))
        self.label_66.setObjectName("label_66")
        self.back_16 = QtWidgets.QPushButton(self.page_cblol)
        self.back_16.setGeometry(QtCore.QRect(10, 30, 91, 41))
        self.back_16.setObjectName("back_16")
        self.cblol_stats_table = QtWidgets.QTableWidget(self.page_cblol)
        self.cblol_stats_table.setGeometry(QtCore.QRect(100, 80, 831, 351))
        self.cblol_stats_table.setObjectName("cblol_stats_table")
        self.cblol_stats_table.setColumnCount(8)
        self.cblol_stats_table.setRowCount(0)
        item = QtWidgets.QTableWidgetItem()
        self.cblol_stats_table.setHorizontalHeaderItem(0, item)
        item = QtWidgets.QTableWidgetItem()
        self.cblol_stats_table.setHorizontalHeaderItem(1, item)
        item = QtWidgets.QTableWidgetItem()
        self.cblol_stats_table.setHorizontalHeaderItem(2, item)
        item = QtWidgets.QTableWidgetItem()
        self.cblol_stats_table.setHorizontalHeaderItem(3, item)
        item = QtWidgets.QTableWidgetItem()
        self.cblol_stats_table.setHorizontalHeaderItem(4, item)
        item = QtWidgets.QTableWidgetItem()
        self.cblol_stats_table.setHorizontalHeaderItem(5, item)
        item = QtWidgets.QTableWidgetItem()
        self.cblol_stats_table.setHorizontalHeaderItem(6, item)
        item = QtWidgets.QTableWidgetItem()
        self.cblol_stats_table.setHorizontalHeaderItem(7, item)
        self.cblol_calculate = QtWidgets.QPushButton(self.page_cblol)
        self.cblol_calculate.setGeometry(QtCore.QRect(840, 790, 141, 41))
        self.cblol_calculate.setObjectName("cblol_calculate")
        self.cblol_odds = QtWidgets.QTextEdit(self.page_cblol)
        self.cblol_odds.setGeometry(QtCore.QRect(670, 800, 71, 21))
        self.cblol_odds.setObjectName("cblol_odds")
        self.cblol_line = QtWidgets.QTextEdit(self.page_cblol)
        self.cblol_line.setGeometry(QtCore.QRect(480, 800, 71, 21))
        self.cblol_line.setObjectName("cblol_line")
        self.ladder_cblol = QtWidgets.QTableWidget(self.page_cblol)
        self.ladder_cblol.setGeometry(QtCore.QRect(1000, 80, 231, 351))
        self.ladder_cblol.setObjectName("ladder_cblol")
        self.ladder_cblol.setColumnCount(2)
        self.ladder_cblol.setRowCount(0)
        item = QtWidgets.QTableWidgetItem()
        self.ladder_cblol.setHorizontalHeaderItem(0, item)
        item = QtWidgets.QTableWidgetItem()
        self.ladder_cblol.setHorizontalHeaderItem(1, item)
        self.label_67 = QtWidgets.QLabel(self.page_cblol)
        self.label_67.setGeometry(QtCore.QRect(1080, 20, 361, 61))
        self.label_67.setObjectName("label_67")
        self.label_68 = QtWidgets.QLabel(self.page_cblol)
        self.label_68.setGeometry(QtCore.QRect(440, 800, 61, 21))
        self.label_68.setObjectName("label_68")
        self.cblol_matches = QtWidgets.QTableWidget(self.page_cblol)
        self.cblol_matches.setGeometry(QtCore.QRect(90, 500, 1141, 281))
        self.cblol_matches.setObjectName("cblol_matches")
        self.cblol_matches.setColumnCount(9)
        self.cblol_matches.setRowCount(0)
        item = QtWidgets.QTableWidgetItem()
        self.cblol_matches.setHorizontalHeaderItem(0, item)
        item = QtWidgets.QTableWidgetItem()
        self.cblol_matches.setHorizontalHeaderItem(1, item)
        item = QtWidgets.QTableWidgetItem()
        self.cblol_matches.setHorizontalHeaderItem(2, item)
        item = QtWidgets.QTableWidgetItem()
        self.cblol_matches.setHorizontalHeaderItem(3, item)
        item = QtWidgets.QTableWidgetItem()
        self.cblol_matches.setHorizontalHeaderItem(4, item)
        item = QtWidgets.QTableWidgetItem()
        self.cblol_matches.setHorizontalHeaderItem(5, item)
        item = QtWidgets.QTableWidgetItem()
        self.cblol_matches.setHorizontalHeaderItem(6, item)
        item = QtWidgets.QTableWidgetItem()
        self.cblol_matches.setHorizontalHeaderItem(7, item)
        item = QtWidgets.QTableWidgetItem()
        self.cblol_matches.setHorizontalHeaderItem(8, item)
        self.stackedWidget.addWidget(self.page_cblol)
        self.page_lla = QtWidgets.QWidget()
        self.page_lla.setObjectName("page_lla")
        self.update_lla = QtWidgets.QPushButton(self.page_lla)
        self.update_lla.setGeometry(QtCore.QRect(830, 440, 101, 41))
        self.update_lla.setObjectName("update_lla")
        self.ladder_lla = QtWidgets.QTableWidget(self.page_lla)
        self.ladder_lla.setGeometry(QtCore.QRect(1000, 80, 231, 351))
        self.ladder_lla.setObjectName("ladder_lla")
        self.ladder_lla.setColumnCount(2)
        self.ladder_lla.setRowCount(0)
        item = QtWidgets.QTableWidgetItem()
        self.ladder_lla.setHorizontalHeaderItem(0, item)
        item = QtWidgets.QTableWidgetItem()
        self.ladder_lla.setHorizontalHeaderItem(1, item)
        self.back_17 = QtWidgets.QPushButton(self.page_lla)
        self.back_17.setGeometry(QtCore.QRect(10, 30, 91, 41))
        self.back_17.setObjectName("back_17")
        self.lla_line = QtWidgets.QTextEdit(self.page_lla)
        self.lla_line.setGeometry(QtCore.QRect(480, 800, 71, 21))
        self.lla_line.setObjectName("lla_line")
        self.lla_odds = QtWidgets.QTextEdit(self.page_lla)
        self.lla_odds.setGeometry(QtCore.QRect(670, 800, 71, 21))
        self.lla_odds.setObjectName("lla_odds")
        self.label_69 = QtWidgets.QLabel(self.page_lla)
        self.label_69.setGeometry(QtCore.QRect(440, 20, 361, 61))
        self.label_69.setObjectName("label_69")
        self.lla_stats_table = QtWidgets.QTableWidget(self.page_lla)
        self.lla_stats_table.setGeometry(QtCore.QRect(100, 80, 831, 351))
        self.lla_stats_table.setObjectName("lla_stats_table")
        self.lla_stats_table.setColumnCount(8)
        self.lla_stats_table.setRowCount(0)
        item = QtWidgets.QTableWidgetItem()
        self.lla_stats_table.setHorizontalHeaderItem(0, item)
        item = QtWidgets.QTableWidgetItem()
        self.lla_stats_table.setHorizontalHeaderItem(1, item)
        item = QtWidgets.QTableWidgetItem()
        self.lla_stats_table.setHorizontalHeaderItem(2, item)
        item = QtWidgets.QTableWidgetItem()
        self.lla_stats_table.setHorizontalHeaderItem(3, item)
        item = QtWidgets.QTableWidgetItem()
        self.lla_stats_table.setHorizontalHeaderItem(4, item)
        item = QtWidgets.QTableWidgetItem()
        self.lla_stats_table.setHorizontalHeaderItem(5, item)
        item = QtWidgets.QTableWidgetItem()
        self.lla_stats_table.setHorizontalHeaderItem(6, item)
        item = QtWidgets.QTableWidgetItem()
        self.lla_stats_table.setHorizontalHeaderItem(7, item)
        self.label_70 = QtWidgets.QLabel(self.page_lla)
        self.label_70.setGeometry(QtCore.QRect(620, 800, 51, 21))
        self.label_70.setObjectName("label_70")
        self.label_71 = QtWidgets.QLabel(self.page_lla)
        self.label_71.setGeometry(QtCore.QRect(1080, 20, 361, 61))
        self.label_71.setObjectName("label_71")
        self.label_72 = QtWidgets.QLabel(self.page_lla)
        self.label_72.setGeometry(QtCore.QRect(440, 800, 61, 21))
        self.label_72.setObjectName("label_72")
        self.lla_matches = QtWidgets.QTableWidget(self.page_lla)
        self.lla_matches.setGeometry(QtCore.QRect(90, 500, 1141, 281))
        self.lla_matches.setObjectName("lla_matches")
        self.lla_matches.setColumnCount(9)
        self.lla_matches.setRowCount(0)
        item = QtWidgets.QTableWidgetItem()
        self.lla_matches.setHorizontalHeaderItem(0, item)
        item = QtWidgets.QTableWidgetItem()
        self.lla_matches.setHorizontalHeaderItem(1, item)
        item = QtWidgets.QTableWidgetItem()
        self.lla_matches.setHorizontalHeaderItem(2, item)
        item = QtWidgets.QTableWidgetItem()
        self.lla_matches.setHorizontalHeaderItem(3, item)
        item = QtWidgets.QTableWidgetItem()
        self.lla_matches.setHorizontalHeaderItem(4, item)
        item = QtWidgets.QTableWidgetItem()
        self.lla_matches.setHorizontalHeaderItem(5, item)
        item = QtWidgets.QTableWidgetItem()
        self.lla_matches.setHorizontalHeaderItem(6, item)
        item = QtWidgets.QTableWidgetItem()
        self.lla_matches.setHorizontalHeaderItem(7, item)
        item = QtWidgets.QTableWidgetItem()
        self.lla_matches.setHorizontalHeaderItem(8, item)
        self.lla_calculate = QtWidgets.QPushButton(self.page_lla)
        self.lla_calculate.setGeometry(QtCore.QRect(840, 790, 141, 41))
        self.lla_calculate.setObjectName("lla_calculate")
        self.stackedWidget.addWidget(self.page_lla)
        self.page_lco = QtWidgets.QWidget()
        self.page_lco.setObjectName("page_lco")
        self.update_lco = QtWidgets.QPushButton(self.page_lco)
        self.update_lco.setGeometry(QtCore.QRect(830, 440, 101, 41))
        self.update_lco.setObjectName("update_lco")
        self.label_76 = QtWidgets.QLabel(self.page_lco)
        self.label_76.setGeometry(QtCore.QRect(440, 800, 61, 21))
        self.label_76.setObjectName("label_76")
        self.ladder_lco = QtWidgets.QTableWidget(self.page_lco)
        self.ladder_lco.setGeometry(QtCore.QRect(1000, 80, 231, 351))
        self.ladder_lco.setObjectName("ladder_lco")
        self.ladder_lco.setColumnCount(2)
        self.ladder_lco.setRowCount(0)
        item = QtWidgets.QTableWidgetItem()
        self.ladder_lco.setHorizontalHeaderItem(0, item)
        item = QtWidgets.QTableWidgetItem()
        self.ladder_lco.setHorizontalHeaderItem(1, item)
        self.back_18 = QtWidgets.QPushButton(self.page_lco)
        self.back_18.setGeometry(QtCore.QRect(10, 30, 91, 41))
        self.back_18.setObjectName("back_18")
        self.label_73 = QtWidgets.QLabel(self.page_lco)
        self.label_73.setGeometry(QtCore.QRect(440, 20, 361, 61))
        self.label_73.setObjectName("label_73")
        self.lco_calculate = QtWidgets.QPushButton(self.page_lco)
        self.lco_calculate.setGeometry(QtCore.QRect(840, 790, 141, 41))
        self.lco_calculate.setObjectName("lco_calculate")
        self.lco_line = QtWidgets.QTextEdit(self.page_lco)
        self.lco_line.setGeometry(QtCore.QRect(480, 800, 71, 21))
        self.lco_line.setObjectName("lco_line")
        self.lco_stats_table = QtWidgets.QTableWidget(self.page_lco)
        self.lco_stats_table.setGeometry(QtCore.QRect(100, 80, 831, 351))
        self.lco_stats_table.setObjectName("lco_stats_table")
        self.lco_stats_table.setColumnCount(8)
        self.lco_stats_table.setRowCount(0)
        item = QtWidgets.QTableWidgetItem()
        self.lco_stats_table.setHorizontalHeaderItem(0, item)
        item = QtWidgets.QTableWidgetItem()
        self.lco_stats_table.setHorizontalHeaderItem(1, item)
        item = QtWidgets.QTableWidgetItem()
        self.lco_stats_table.setHorizontalHeaderItem(2, item)
        item = QtWidgets.QTableWidgetItem()
        self.lco_stats_table.setHorizontalHeaderItem(3, item)
        item = QtWidgets.QTableWidgetItem()
        self.lco_stats_table.setHorizontalHeaderItem(4, item)
        item = QtWidgets.QTableWidgetItem()
        self.lco_stats_table.setHorizontalHeaderItem(5, item)
        item = QtWidgets.QTableWidgetItem()
        self.lco_stats_table.setHorizontalHeaderItem(6, item)
        item = QtWidgets.QTableWidgetItem()
        self.lco_stats_table.setHorizontalHeaderItem(7, item)
        self.label_74 = QtWidgets.QLabel(self.page_lco)
        self.label_74.setGeometry(QtCore.QRect(620, 800, 51, 21))
        self.label_74.setObjectName("label_74")
        self.lco_odds = QtWidgets.QTextEdit(self.page_lco)
        self.lco_odds.setGeometry(QtCore.QRect(670, 800, 71, 21))
        self.lco_odds.setObjectName("lco_odds")
        self.lco_matches = QtWidgets.QTableWidget(self.page_lco)
        self.lco_matches.setGeometry(QtCore.QRect(90, 500, 1141, 281))
        self.lco_matches.setObjectName("lco_matches")
        self.lco_matches.setColumnCount(9)
        self.lco_matches.setRowCount(0)
        item = QtWidgets.QTableWidgetItem()
        self.lco_matches.setHorizontalHeaderItem(0, item)
        item = QtWidgets.QTableWidgetItem()
        self.lco_matches.setHorizontalHeaderItem(1, item)
        item = QtWidgets.QTableWidgetItem()
        self.lco_matches.setHorizontalHeaderItem(2, item)
        item = QtWidgets.QTableWidgetItem()
        self.lco_matches.setHorizontalHeaderItem(3, item)
        item = QtWidgets.QTableWidgetItem()
        self.lco_matches.setHorizontalHeaderItem(4, item)
        item = QtWidgets.QTableWidgetItem()
        self.lco_matches.setHorizontalHeaderItem(5, item)
        item = QtWidgets.QTableWidgetItem()
        self.lco_matches.setHorizontalHeaderItem(6, item)
        item = QtWidgets.QTableWidgetItem()
        self.lco_matches.setHorizontalHeaderItem(7, item)
        item = QtWidgets.QTableWidgetItem()
        self.lco_matches.setHorizontalHeaderItem(8, item)
        self.label_75 = QtWidgets.QLabel(self.page_lco)
        self.label_75.setGeometry(QtCore.QRect(1080, 20, 361, 61))
        self.label_75.setObjectName("label_75")
        self.stackedWidget.addWidget(self.page_lco)
        self.page_lck_chal = QtWidgets.QWidget()
        self.page_lck_chal.setObjectName("page_lck_chal")
        self.lck_chal_calculate = QtWidgets.QPushButton(self.page_lck_chal)
        self.lck_chal_calculate.setGeometry(QtCore.QRect(840, 790, 141, 41))
        self.lck_chal_calculate.setObjectName("lck_chal_calculate")
        self.lck_chal_odds = QtWidgets.QTextEdit(self.page_lck_chal)
        self.lck_chal_odds.setGeometry(QtCore.QRect(670, 800, 71, 21))
        self.lck_chal_odds.setObjectName("lck_chal_odds")
        self.back_28 = QtWidgets.QPushButton(self.page_lck_chal)
        self.back_28.setGeometry(QtCore.QRect(10, 30, 91, 41))
        self.back_28.setObjectName("back_28")
        self.lck_chal_stats_table = QtWidgets.QTableWidget(self.page_lck_chal)
        self.lck_chal_stats_table.setGeometry(QtCore.QRect(100, 80, 831, 351))
        self.lck_chal_stats_table.setObjectName("lck_chal_stats_table")
        self.lck_chal_stats_table.setColumnCount(8)
        self.lck_chal_stats_table.setRowCount(0)
        item = QtWidgets.QTableWidgetItem()
        self.lck_chal_stats_table.setHorizontalHeaderItem(0, item)
        item = QtWidgets.QTableWidgetItem()
        self.lck_chal_stats_table.setHorizontalHeaderItem(1, item)
        item = QtWidgets.QTableWidgetItem()
        self.lck_chal_stats_table.setHorizontalHeaderItem(2, item)
        item = QtWidgets.QTableWidgetItem()
        self.lck_chal_stats_table.setHorizontalHeaderItem(3, item)
        item = QtWidgets.QTableWidgetItem()
        self.lck_chal_stats_table.setHorizontalHeaderItem(4, item)
        item = QtWidgets.QTableWidgetItem()
        self.lck_chal_stats_table.setHorizontalHeaderItem(5, item)
        item = QtWidgets.QTableWidgetItem()
        self.lck_chal_stats_table.setHorizontalHeaderItem(6, item)
        item = QtWidgets.QTableWidgetItem()
        self.lck_chal_stats_table.setHorizontalHeaderItem(7, item)
        self.update_lck_chal = QtWidgets.QPushButton(self.page_lck_chal)
        self.update_lck_chal.setGeometry(QtCore.QRect(830, 440, 101, 41))
        self.update_lck_chal.setObjectName("update_lck_chal")
        self.lck_chal_matches = QtWidgets.QTableWidget(self.page_lck_chal)
        self.lck_chal_matches.setGeometry(QtCore.QRect(90, 500, 1141, 281))
        self.lck_chal_matches.setObjectName("lck_chal_matches")
        self.lck_chal_matches.setColumnCount(9)
        self.lck_chal_matches.setRowCount(0)
        item = QtWidgets.QTableWidgetItem()
        self.lck_chal_matches.setHorizontalHeaderItem(0, item)
        item = QtWidgets.QTableWidgetItem()
        self.lck_chal_matches.setHorizontalHeaderItem(1, item)
        item = QtWidgets.QTableWidgetItem()
        self.lck_chal_matches.setHorizontalHeaderItem(2, item)
        item = QtWidgets.QTableWidgetItem()
        self.lck_chal_matches.setHorizontalHeaderItem(3, item)
        item = QtWidgets.QTableWidgetItem()
        self.lck_chal_matches.setHorizontalHeaderItem(4, item)
        item = QtWidgets.QTableWidgetItem()
        self.lck_chal_matches.setHorizontalHeaderItem(5, item)
        item = QtWidgets.QTableWidgetItem()
        self.lck_chal_matches.setHorizontalHeaderItem(6, item)
        item = QtWidgets.QTableWidgetItem()
        self.lck_chal_matches.setHorizontalHeaderItem(7, item)
        item = QtWidgets.QTableWidgetItem()
        self.lck_chal_matches.setHorizontalHeaderItem(8, item)
        self.label_114 = QtWidgets.QLabel(self.page_lck_chal)
        self.label_114.setGeometry(QtCore.QRect(620, 800, 51, 21))
        self.label_114.setObjectName("label_114")
        self.label_115 = QtWidgets.QLabel(self.page_lck_chal)
        self.label_115.setGeometry(QtCore.QRect(440, 800, 61, 21))
        self.label_115.setObjectName("label_115")
        self.lck_chal_line = QtWidgets.QTextEdit(self.page_lck_chal)
        self.lck_chal_line.setGeometry(QtCore.QRect(480, 800, 71, 21))
        self.lck_chal_line.setObjectName("lck_chal_line")
        self.label_116 = QtWidgets.QLabel(self.page_lck_chal)
        self.label_116.setGeometry(QtCore.QRect(1080, 20, 361, 61))
        self.label_116.setObjectName("label_116")
        self.label_117 = QtWidgets.QLabel(self.page_lck_chal)
        self.label_117.setGeometry(QtCore.QRect(440, 20, 361, 61))
        self.label_117.setObjectName("label_117")
        self.ladder_lck_chal = QtWidgets.QTableWidget(self.page_lck_chal)
        self.ladder_lck_chal.setGeometry(QtCore.QRect(1000, 80, 231, 351))
        self.ladder_lck_chal.setObjectName("ladder_lck_chal")
        self.ladder_lck_chal.setColumnCount(2)
        self.ladder_lck_chal.setRowCount(0)
        item = QtWidgets.QTableWidgetItem()
        self.ladder_lck_chal.setHorizontalHeaderItem(0, item)
        item = QtWidgets.QTableWidgetItem()
        self.ladder_lck_chal.setHorizontalHeaderItem(1, item)
        self.stackedWidget.addWidget(self.page_lck_chal)
        self.page_tcl = QtWidgets.QWidget()
        self.page_tcl.setObjectName("page_tcl")
        self.ladder_tcl = QtWidgets.QTableWidget(self.page_tcl)
        self.ladder_tcl.setGeometry(QtCore.QRect(1000, 80, 231, 351))
        self.ladder_tcl.setObjectName("ladder_tcl")
        self.ladder_tcl.setColumnCount(2)
        self.ladder_tcl.setRowCount(0)
        item = QtWidgets.QTableWidgetItem()
        self.ladder_tcl.setHorizontalHeaderItem(0, item)
        item = QtWidgets.QTableWidgetItem()
        self.ladder_tcl.setHorizontalHeaderItem(1, item)
        self.tcl_matches = QtWidgets.QTableWidget(self.page_tcl)
        self.tcl_matches.setGeometry(QtCore.QRect(90, 500, 1141, 281))
        self.tcl_matches.setObjectName("tcl_matches")
        self.tcl_matches.setColumnCount(9)
        self.tcl_matches.setRowCount(0)
        item = QtWidgets.QTableWidgetItem()
        self.tcl_matches.setHorizontalHeaderItem(0, item)
        item = QtWidgets.QTableWidgetItem()
        self.tcl_matches.setHorizontalHeaderItem(1, item)
        item = QtWidgets.QTableWidgetItem()
        self.tcl_matches.setHorizontalHeaderItem(2, item)
        item = QtWidgets.QTableWidgetItem()
        self.tcl_matches.setHorizontalHeaderItem(3, item)
        item = QtWidgets.QTableWidgetItem()
        self.tcl_matches.setHorizontalHeaderItem(4, item)
        item = QtWidgets.QTableWidgetItem()
        self.tcl_matches.setHorizontalHeaderItem(5, item)
        item = QtWidgets.QTableWidgetItem()
        self.tcl_matches.setHorizontalHeaderItem(6, item)
        item = QtWidgets.QTableWidgetItem()
        self.tcl_matches.setHorizontalHeaderItem(7, item)
        item = QtWidgets.QTableWidgetItem()
        self.tcl_matches.setHorizontalHeaderItem(8, item)
        self.tcl_line = QtWidgets.QTextEdit(self.page_tcl)
        self.tcl_line.setGeometry(QtCore.QRect(480, 800, 71, 21))
        self.tcl_line.setObjectName("tcl_line")
        self.label_120 = QtWidgets.QLabel(self.page_tcl)
        self.label_120.setGeometry(QtCore.QRect(1080, 20, 361, 61))
        self.label_120.setObjectName("label_120")
        self.back_29 = QtWidgets.QPushButton(self.page_tcl)
        self.back_29.setGeometry(QtCore.QRect(10, 30, 91, 41))
        self.back_29.setObjectName("back_29")
        self.update_tcl = QtWidgets.QPushButton(self.page_tcl)
        self.update_tcl.setGeometry(QtCore.QRect(830, 440, 101, 41))
        self.update_tcl.setObjectName("update_tcl")
        self.tcl_stats_table = QtWidgets.QTableWidget(self.page_tcl)
        self.tcl_stats_table.setGeometry(QtCore.QRect(100, 80, 831, 351))
        self.tcl_stats_table.setObjectName("tcl_stats_table")
        self.tcl_stats_table.setColumnCount(8)
        self.tcl_stats_table.setRowCount(0)
        item = QtWidgets.QTableWidgetItem()
        self.tcl_stats_table.setHorizontalHeaderItem(0, item)
        item = QtWidgets.QTableWidgetItem()
        self.tcl_stats_table.setHorizontalHeaderItem(1, item)
        item = QtWidgets.QTableWidgetItem()
        self.tcl_stats_table.setHorizontalHeaderItem(2, item)
        item = QtWidgets.QTableWidgetItem()
        self.tcl_stats_table.setHorizontalHeaderItem(3, item)
        item = QtWidgets.QTableWidgetItem()
        self.tcl_stats_table.setHorizontalHeaderItem(4, item)
        item = QtWidgets.QTableWidgetItem()
        self.tcl_stats_table.setHorizontalHeaderItem(5, item)
        item = QtWidgets.QTableWidgetItem()
        self.tcl_stats_table.setHorizontalHeaderItem(6, item)
        item = QtWidgets.QTableWidgetItem()
        self.tcl_stats_table.setHorizontalHeaderItem(7, item)
        self.tcl_odds = QtWidgets.QTextEdit(self.page_tcl)
        self.tcl_odds.setGeometry(QtCore.QRect(670, 800, 71, 21))
        self.tcl_odds.setObjectName("tcl_odds")
        self.label_119 = QtWidgets.QLabel(self.page_tcl)
        self.label_119.setGeometry(QtCore.QRect(440, 800, 61, 21))
        self.label_119.setObjectName("label_119")
        self.label_118 = QtWidgets.QLabel(self.page_tcl)
        self.label_118.setGeometry(QtCore.QRect(620, 800, 51, 21))
        self.label_118.setObjectName("label_118")
        self.tcl_calculate = QtWidgets.QPushButton(self.page_tcl)
        self.tcl_calculate.setGeometry(QtCore.QRect(840, 790, 141, 41))
        self.tcl_calculate.setObjectName("tcl_calculate")
        self.label_121 = QtWidgets.QLabel(self.page_tcl)
        self.label_121.setGeometry(QtCore.QRect(440, 20, 361, 61))
        self.label_121.setObjectName("label_121")
        self.stackedWidget.addWidget(self.page_tcl)
        self.page_pcs = QtWidgets.QWidget()
        self.page_pcs.setObjectName("page_pcs")
        self.ladder_pcs = QtWidgets.QTableWidget(self.page_pcs)
        self.ladder_pcs.setGeometry(QtCore.QRect(1000, 80, 231, 351))
        self.ladder_pcs.setObjectName("ladder_pcs")
        self.ladder_pcs.setColumnCount(2)
        self.ladder_pcs.setRowCount(0)
        item = QtWidgets.QTableWidgetItem()
        self.ladder_pcs.setHorizontalHeaderItem(0, item)
        item = QtWidgets.QTableWidgetItem()
        self.ladder_pcs.setHorizontalHeaderItem(1, item)
        self.update_pcs = QtWidgets.QPushButton(self.page_pcs)
        self.update_pcs.setGeometry(QtCore.QRect(830, 440, 101, 41))
        self.update_pcs.setObjectName("update_pcs")
        self.pcs_line = QtWidgets.QTextEdit(self.page_pcs)
        self.pcs_line.setGeometry(QtCore.QRect(480, 800, 71, 21))
        self.pcs_line.setObjectName("pcs_line")
        self.pcs_odds = QtWidgets.QTextEdit(self.page_pcs)
        self.pcs_odds.setGeometry(QtCore.QRect(670, 800, 71, 21))
        self.pcs_odds.setObjectName("pcs_odds")
        self.back_30 = QtWidgets.QPushButton(self.page_pcs)
        self.back_30.setGeometry(QtCore.QRect(10, 30, 91, 41))
        self.back_30.setObjectName("back_30")
        self.pcs_matches = QtWidgets.QTableWidget(self.page_pcs)
        self.pcs_matches.setGeometry(QtCore.QRect(90, 500, 1141, 281))
        self.pcs_matches.setObjectName("pcs_matches")
        self.pcs_matches.setColumnCount(9)
        self.pcs_matches.setRowCount(0)
        item = QtWidgets.QTableWidgetItem()
        self.pcs_matches.setHorizontalHeaderItem(0, item)
        item = QtWidgets.QTableWidgetItem()
        self.pcs_matches.setHorizontalHeaderItem(1, item)
        item = QtWidgets.QTableWidgetItem()
        self.pcs_matches.setHorizontalHeaderItem(2, item)
        item = QtWidgets.QTableWidgetItem()
        self.pcs_matches.setHorizontalHeaderItem(3, item)
        item = QtWidgets.QTableWidgetItem()
        self.pcs_matches.setHorizontalHeaderItem(4, item)
        item = QtWidgets.QTableWidgetItem()
        self.pcs_matches.setHorizontalHeaderItem(5, item)
        item = QtWidgets.QTableWidgetItem()
        self.pcs_matches.setHorizontalHeaderItem(6, item)
        item = QtWidgets.QTableWidgetItem()
        self.pcs_matches.setHorizontalHeaderItem(7, item)
        item = QtWidgets.QTableWidgetItem()
        self.pcs_matches.setHorizontalHeaderItem(8, item)
        self.pcs_calculate = QtWidgets.QPushButton(self.page_pcs)
        self.pcs_calculate.setGeometry(QtCore.QRect(840, 790, 141, 41))
        self.pcs_calculate.setObjectName("pcs_calculate")
        self.label_124 = QtWidgets.QLabel(self.page_pcs)
        self.label_124.setGeometry(QtCore.QRect(1080, 20, 361, 61))
        self.label_124.setObjectName("label_124")
        self.label_125 = QtWidgets.QLabel(self.page_pcs)
        self.label_125.setGeometry(QtCore.QRect(440, 20, 361, 61))
        self.label_125.setObjectName("label_125")
        self.label_122 = QtWidgets.QLabel(self.page_pcs)
        self.label_122.setGeometry(QtCore.QRect(620, 800, 51, 21))
        self.label_122.setObjectName("label_122")
        self.label_123 = QtWidgets.QLabel(self.page_pcs)
        self.label_123.setGeometry(QtCore.QRect(440, 800, 61, 21))
        self.label_123.setObjectName("label_123")
        self.pcs_stats_table = QtWidgets.QTableWidget(self.page_pcs)
        self.pcs_stats_table.setGeometry(QtCore.QRect(100, 80, 831, 351))
        self.pcs_stats_table.setObjectName("pcs_stats_table")
        self.pcs_stats_table.setColumnCount(8)
        self.pcs_stats_table.setRowCount(0)
        item = QtWidgets.QTableWidgetItem()
        self.pcs_stats_table.setHorizontalHeaderItem(0, item)
        item = QtWidgets.QTableWidgetItem()
        self.pcs_stats_table.setHorizontalHeaderItem(1, item)
        item = QtWidgets.QTableWidgetItem()
        self.pcs_stats_table.setHorizontalHeaderItem(2, item)
        item = QtWidgets.QTableWidgetItem()
        self.pcs_stats_table.setHorizontalHeaderItem(3, item)
        item = QtWidgets.QTableWidgetItem()
        self.pcs_stats_table.setHorizontalHeaderItem(4, item)
        item = QtWidgets.QTableWidgetItem()
        self.pcs_stats_table.setHorizontalHeaderItem(5, item)
        item = QtWidgets.QTableWidgetItem()
        self.pcs_stats_table.setHorizontalHeaderItem(6, item)
        item = QtWidgets.QTableWidgetItem()
        self.pcs_stats_table.setHorizontalHeaderItem(7, item)
        self.stackedWidget.addWidget(self.page_pcs)
        self.page_ebl = QtWidgets.QWidget()
        self.page_ebl.setObjectName("page_ebl")
        self.ebl_calculate = QtWidgets.QPushButton(self.page_ebl)
        self.ebl_calculate.setGeometry(QtCore.QRect(840, 790, 141, 41))
        self.ebl_calculate.setObjectName("ebl_calculate")
        self.ebl_odds = QtWidgets.QTextEdit(self.page_ebl)
        self.ebl_odds.setGeometry(QtCore.QRect(670, 800, 71, 21))
        self.ebl_odds.setObjectName("ebl_odds")
        self.back_31 = QtWidgets.QPushButton(self.page_ebl)
        self.back_31.setGeometry(QtCore.QRect(10, 30, 91, 41))
        self.back_31.setObjectName("back_31")
        self.ebl_stats_table = QtWidgets.QTableWidget(self.page_ebl)
        self.ebl_stats_table.setGeometry(QtCore.QRect(100, 80, 831, 351))
        self.ebl_stats_table.setObjectName("ebl_stats_table")
        self.ebl_stats_table.setColumnCount(8)
        self.ebl_stats_table.setRowCount(0)
        item = QtWidgets.QTableWidgetItem()
        self.ebl_stats_table.setHorizontalHeaderItem(0, item)
        item = QtWidgets.QTableWidgetItem()
        self.ebl_stats_table.setHorizontalHeaderItem(1, item)
        item = QtWidgets.QTableWidgetItem()
        self.ebl_stats_table.setHorizontalHeaderItem(2, item)
        item = QtWidgets.QTableWidgetItem()
        self.ebl_stats_table.setHorizontalHeaderItem(3, item)
        item = QtWidgets.QTableWidgetItem()
        self.ebl_stats_table.setHorizontalHeaderItem(4, item)
        item = QtWidgets.QTableWidgetItem()
        self.ebl_stats_table.setHorizontalHeaderItem(5, item)
        item = QtWidgets.QTableWidgetItem()
        self.ebl_stats_table.setHorizontalHeaderItem(6, item)
        item = QtWidgets.QTableWidgetItem()
        self.ebl_stats_table.setHorizontalHeaderItem(7, item)
        self.update_ebl = QtWidgets.QPushButton(self.page_ebl)
        self.update_ebl.setGeometry(QtCore.QRect(830, 440, 101, 41))
        self.update_ebl.setObjectName("update_ebl")
        self.ebl_matches = QtWidgets.QTableWidget(self.page_ebl)
        self.ebl_matches.setGeometry(QtCore.QRect(90, 500, 1141, 281))
        self.ebl_matches.setObjectName("ebl_matches")
        self.ebl_matches.setColumnCount(9)
        self.ebl_matches.setRowCount(0)
        item = QtWidgets.QTableWidgetItem()
        self.ebl_matches.setHorizontalHeaderItem(0, item)
        item = QtWidgets.QTableWidgetItem()
        self.ebl_matches.setHorizontalHeaderItem(1, item)
        item = QtWidgets.QTableWidgetItem()
        self.ebl_matches.setHorizontalHeaderItem(2, item)
        item = QtWidgets.QTableWidgetItem()
        self.ebl_matches.setHorizontalHeaderItem(3, item)
        item = QtWidgets.QTableWidgetItem()
        self.ebl_matches.setHorizontalHeaderItem(4, item)
        item = QtWidgets.QTableWidgetItem()
        self.ebl_matches.setHorizontalHeaderItem(5, item)
        item = QtWidgets.QTableWidgetItem()
        self.ebl_matches.setHorizontalHeaderItem(6, item)
        item = QtWidgets.QTableWidgetItem()
        self.ebl_matches.setHorizontalHeaderItem(7, item)
        item = QtWidgets.QTableWidgetItem()
        self.ebl_matches.setHorizontalHeaderItem(8, item)
        self.label_126 = QtWidgets.QLabel(self.page_ebl)
        self.label_126.setGeometry(QtCore.QRect(620, 800, 51, 21))
        self.label_126.setObjectName("label_126")
        self.label_127 = QtWidgets.QLabel(self.page_ebl)
        self.label_127.setGeometry(QtCore.QRect(440, 800, 61, 21))
        self.label_127.setObjectName("label_127")
        self.ebl_line = QtWidgets.QTextEdit(self.page_ebl)
        self.ebl_line.setGeometry(QtCore.QRect(480, 800, 71, 21))
        self.ebl_line.setObjectName("ebl_line")
        self.label_128 = QtWidgets.QLabel(self.page_ebl)
        self.label_128.setGeometry(QtCore.QRect(1080, 20, 361, 61))
        self.label_128.setObjectName("label_128")
        self.label_129 = QtWidgets.QLabel(self.page_ebl)
        self.label_129.setGeometry(QtCore.QRect(440, 20, 361, 61))
        self.label_129.setObjectName("label_129")
        self.ladder_ebl = QtWidgets.QTableWidget(self.page_ebl)
        self.ladder_ebl.setGeometry(QtCore.QRect(1000, 80, 231, 351))
        self.ladder_ebl.setObjectName("ladder_ebl")
        self.ladder_ebl.setColumnCount(2)
        self.ladder_ebl.setRowCount(0)
        item = QtWidgets.QTableWidgetItem()
        self.ladder_ebl.setHorizontalHeaderItem(0, item)
        item = QtWidgets.QTableWidgetItem()
        self.ladder_ebl.setHorizontalHeaderItem(1, item)
        self.stackedWidget.addWidget(self.page_ebl)
        MainWindow.setCentralWidget(self.centralwidget)
        self.menubar = QtWidgets.QMenuBar(MainWindow)
        self.menubar.setGeometry(QtCore.QRect(0, 0, 1301, 21))
        self.menubar.setObjectName("menubar")
        MainWindow.setMenuBar(self.menubar)
        self.statusbar = QtWidgets.QStatusBar(MainWindow)
        self.statusbar.setObjectName("statusbar")
        MainWindow.setStatusBar(self.statusbar)

        # button links to created functions
        # load data associated with the selected page
        self.pushButton.clicked.connect(lambda: self.get_lck())
        self.lcs.clicked.connect(lambda: self.get_lcs())
        self.lec.clicked.connect(lambda: self.get_lec())
        self.vcs.clicked.connect(lambda: self.get_vcs())
        self.lpl.clicked.connect(lambda: self.get_lpl())
        self.na_acad.clicked.connect(lambda: self.get_na_acad())
        self.cblol.clicked.connect(lambda: self.get_cblol())
        self.lla.clicked.connect(lambda: self.get_lla())
        self.lco.clicked.connect(lambda: self.get_lco())
        self.ul.clicked.connect(lambda: self.get_ul())
        self.pgn.clicked.connect(lambda: self.get_pgn())
        self.pld.clicked.connect(lambda: self.get_pld())
        self.lvp.clicked.connect(lambda: self.get_lvp())
        self.ljl.clicked.connect(lambda: self.get_ljl())
        self.nlc.clicked.connect(lambda: self.get_nlc())
        self.lfl.clicked.connect(lambda: self.get_lfl())
        self.lplol.clicked.connect(lambda: self.get_lplol())
        self.hpm.clicked.connect(lambda: self.get_hpm())
        self.lck_chal.clicked.connect(lambda: self.get_lck_chal())
        self.tcl.clicked.connect(lambda: self.get_tcl())
        self.pcs.clicked.connect(lambda: self.get_pcs())
        self.ebl.clicked.connect(lambda: self.get_ebl())
        
        # update data in current table based on button clicked
        self.update_lck.clicked.connect(lambda: lck_data())
        self.update_lcs.clicked.connect(lambda: lcs_data())
        self.update_lec.clicked.connect(lambda: lec_data())
        self.update_vcs.clicked.connect(lambda: vcs_data())
        self.update_lpl.clicked.connect(lambda: lpl_data())
        self.update_na_acad.clicked.connect(lambda: na_acad_data())
        self.update_cblol.clicked.connect(lambda: cblol_data())
        self.update_lla.clicked.connect(lambda: lla_data())
        self.update_lco.clicked.connect(lambda: lco_data())
        self.update_ul.clicked.connect(lambda: ul_data())
        self.update_pgn.clicked.connect(lambda: pgn_data())
        self.update_pld.clicked.connect(lambda: pld_data())
        self.update_lvp.clicked.connect(lambda: lvp_data())
        self.update_ljl.clicked.connect(lambda: ljl_data())
        self.update_nlc.clicked.connect(lambda: nlc_data())
        self.update_lfl.clicked.connect(lambda: lfl_data())
        self.update_lplol.clicked.connect(lambda: lplol_data())
        self.update_hpm.clicked.connect(lambda: hpm_data())
        self.update_lck_chal.clicked.connect(lambda: lck_chal_data())
        self.update_tcl.clicked.connect(lambda: tcl_data())
        self.update_pcs.clicked.connect(lambda: pcs_data())
        self.update_ebl.clicked.connect(lambda: ebl_data())
        
        # navigate back to main page from other pages
        self.back.clicked.connect(lambda: self.stackedWidget.setCurrentWidget(self.page))
        self.back_2.clicked.connect(lambda: self.stackedWidget.setCurrentWidget(self.page))
        self.back_3.clicked.connect(lambda: self.stackedWidget.setCurrentWidget(self.page))
        self.back_13.clicked.connect(lambda: self.stackedWidget.setCurrentWidget(self.page))
        self.back_14.clicked.connect(lambda: self.stackedWidget.setCurrentWidget(self.page))
        self.back_15.clicked.connect(lambda: self.stackedWidget.setCurrentWidget(self.page))
        self.back_16.clicked.connect(lambda: self.stackedWidget.setCurrentWidget(self.page))
        self.back_17.clicked.connect(lambda: self.stackedWidget.setCurrentWidget(self.page))
        self.back_18.clicked.connect(lambda: self.stackedWidget.setCurrentWidget(self.page))
        self.back_19.clicked.connect(lambda: self.stackedWidget.setCurrentWidget(self.page))
        self.back_20.clicked.connect(lambda: self.stackedWidget.setCurrentWidget(self.page))
        self.back_21.clicked.connect(lambda: self.stackedWidget.setCurrentWidget(self.page))
        self.back_22.clicked.connect(lambda: self.stackedWidget.setCurrentWidget(self.page))
        self.back_23.clicked.connect(lambda: self.stackedWidget.setCurrentWidget(self.page))
        self.back_24.clicked.connect(lambda: self.stackedWidget.setCurrentWidget(self.page))
        self.back_25.clicked.connect(lambda: self.stackedWidget.setCurrentWidget(self.page))
        self.back_26.clicked.connect(lambda: self.stackedWidget.setCurrentWidget(self.page))
        self.back_27.clicked.connect(lambda: self.stackedWidget.setCurrentWidget(self.page))
        self.back_28.clicked.connect(lambda: self.stackedWidget.setCurrentWidget(self.page))
        self.back_29.clicked.connect(lambda: self.stackedWidget.setCurrentWidget(self.page))
        self.back_30.clicked.connect(lambda: self.stackedWidget.setCurrentWidget(self.page))
        self.back_31.clicked.connect(lambda: self.stackedWidget.setCurrentWidget(self.page))

        # alter the line/odds used in the calculation and refresh the upcoming games table
        self.lck_calculate.clicked.connect(lambda: self.calc_lck())
        self.lec_calculate.clicked.connect(lambda: self.calc_lec())
        self.lcs_calculate.clicked.connect(lambda: self.calc_lcs())
        self.vcs_calculate.clicked.connect(lambda: self.calc_vcs())
        self.lpl_calculate.clicked.connect(lambda: self.calc_lpl())
        self.na_acad_calculate.clicked.connect(lambda: self.calc_na_acad())
        self.cblol_calculate.clicked.connect(lambda: self.calc_cblol())
        self.lla_calculate.clicked.connect(lambda: self.calc_lla())
        self.lco_calculate.clicked.connect(lambda: self.calc_lco())
        self.ul_calculate.clicked.connect(lambda: self.calc_ul())
        self.pgn_calculate.clicked.connect(lambda: self.calc_pgn())
        self.pld_calculate.clicked.connect(lambda: self.calc_pld())
        self.lvp_calculate.clicked.connect(lambda: self.calc_lvp())
        self.ljl_calculate.clicked.connect(lambda: self.calc_ljl())
        self.nlc_calculate.clicked.connect(lambda: self.calc_nlc())
        self.lfl_calculate.clicked.connect(lambda: self.calc_lfl())
        self.lplol_calculate.clicked.connect(lambda: self.calc_lplol())
        self.hpm_calculate.clicked.connect(lambda: self.calc_hpm())
        self.pcs_calculate.clicked.connect(lambda: self.calc_pcs())
        self.lck_chal_calculate.clicked.connect(lambda: self.calc_lck_chal())
        self.tcl_calculate.clicked.connect(lambda: self.calc_tcl())
        self.ebl_calculate.clicked.connect(lambda: self.calc_ebl())

        # scrape the upcoming games and insert them into a table upon start up
        next_games = UpcomingGames()
        comp, match = next_games.next_games()
        self.insert_data(comp, 0, self.upcoming_table)
        self.insert_data(match, 1, self.upcoming_table)
        # resize the table for ease of use
        header = self.upcoming_table.horizontalHeader()       
        header.setSectionResizeMode(0, QtWidgets.QHeaderView.ResizeToContents)
        header.setSectionResizeMode(1, QtWidgets.QHeaderView.Stretch)

        self.retranslateUi(MainWindow)
        QtCore.QMetaObject.connectSlotsByName(MainWindow)

    # calculates the value based on the new odds/line entered for each league
    def calc_lck(self):
        line = float(self.lck_line.toPlainText())
        odds = float(self.lck_odds.toPlainText())
        self.load_data('lck', self.tableWidget, self.page_2,'https://lol.fandom.com/wiki/LCK/2022_Season/Spring_Season', self.ladder_1, 'https://www.rivalry.com/esports/league-of-legends-betting/3254-champions-korea', self.lck_matches, line, odds)

    def calc_na_acad(self):
        line = float(self.na_acad_line.toPlainText())
        odds = float(self.na_acad_odds.toPlainText())
        self.load_data('na_acad', self.na_acad_stats_table, self.page_na_acad,'https://lol.fandom.com/wiki/NA_Academy_League/2022_Season/Spring_Season', self.ladder_na_acad, 'https://www.rivalry.com/esports/league-of-legends-betting/5784-na-academy-league', self.na_acad_matches, line, odds)

    def calc_lla(self):
        line = float(self.lla_line.toPlainText())
        odds = float(self.lla_odds.toPlainText())
        self.load_data('lla', self.lla_stats_table, self.page_lla,'https://lol.fandom.com/wiki/LLA/2022_Season/Opening_Season', self.ladder_lla, 'https://www.rivalry.com/esports/league-of-legends-betting/3324-liga-latinoamerica', self.lla_matches, line, odds)

    def calc_cblol(self):
        line = float(self.cblol_line.toPlainText())
        odds = float(self.cblol_odds.toPlainText())
        self.load_data('cblol', self.cblol_stats_table, self.page_cblol,'https://lol.fandom.com/wiki/CBLOL/2022_Season/Split_1', self.ladder_cblol, 'https://www.rivalry.com/esports/league-of-legends-betting/3330-campeonato-brasileiro', self.cblol_matches, line, odds)

    def calc_lcs(self):
        line = float(self.lcs_line.toPlainText())
        odds = float(self.lcs_odds.toPlainText())
        self.load_data('lcs', self.tableWidget_2, self.page_3,'https://lol.fandom.com/wiki/LCS/2022_Season/Spring_Season', self.ladder_2, 'https://www.rivalry.com/esports/league-of-legends-betting/3713-lcs-north-america', self.lcs_matches, line, odds)
    
    def calc_lec(self):
        line = float(self.lec_line.toPlainText())
        odds = float(self.lec_odds.toPlainText())
        self.load_data('lec', self.tableWidget_3, self.page_4,'https://lol.fandom.com/wiki/LEC/2022_Season/Spring_Season', self.ladder_3, 'https://www.rivalry.com/esports/league-of-legends-betting/3282-european-championship', self.lec_matches, line, odds)

    def calc_vcs(self):
        line = float(self.vcs_line.toPlainText())
        odds = float(self.vcs_odds.toPlainText())
        self.load_data('vcs', self.vcs_stats_table, self.page_vcs,'https://lol.fandom.com/wiki/VCS/2022_Season/Spring_Season', self.ladder_vcs, 'https://www.rivalry.com/esports/league-of-legends-betting/3296-vcs', self.vcs_matches, line, odds)

    def calc_lpl(self):
        line = float(self.lpl_line.toPlainText())
        odds = float(self.lpl_odds.toPlainText())
        self.load_data('lpl', self.lpl_stats_table, self.page_5,'https://lol.fandom.com/wiki/LPL/2022_Season/Spring_Season', self.ladder_lpl, 'https://www.rivalry.com/esports/league-of-legends-betting/2762-china-lpl', self.lpl_matches, line, odds)

    def calc_lco(self):
        line = float(self.lco_line.toPlainText())
        odds = float(self.lco_odds.toPlainText())
        self.load_data('lco', self.lco_stats_table, self.page_lco,'https://lol.fandom.com/wiki/LCO/2022_Season/Split_1', self.ladder_lco, 'https://www.rivalry.com/esports/league-of-legends-betting/4139-circuit-oceania', self.lco_matches, line, odds)
    
    def calc_ul(self):
        line = float(self.ul_line.toPlainText())
        odds = float(self.ul_odds.toPlainText())
        self.load_data('ul', self.ul_stats_table, self.page_ul,'https://lol.fandom.com/wiki/Ultraliga/Season_7', self.ladder_ul, 'https://www.rivalry.com/esports/league-of-legends-betting/3314-ultraliga', self.ul_matches, line, odds)

    def calc_pgn(self):
        line = float(self.pgn_line.toPlainText())
        odds = float(self.pgn_odds.toPlainText())
        self.load_data('pgn', self.pgn_stats_table, self.page_pgn,'https://lol.fandom.com/wiki/PG_Nationals/2022_Season/Spring_Season', self.ladder_pgn, 'https://www.rivalry.com/au/esports/league-of-legends-betting/3304-pg-nationals', self.pgn_matches, line, odds)

    def calc_pld(self):
        line = float(self.pld_line.toPlainText())
        odds = float(self.pld_odds.toPlainText())
        self.load_data('pld', self.pld_stats_table, self.page_pld,'https://lol.fandom.com/wiki/Prime_League_1st_Division/2022_Season/Spring_Season', self.ladder_pld, 'https://www.rivalry.com/au/esports/league-of-legends-betting/5710-prime-league-1st-division', self.pld_matches, line, odds)

    def calc_lvp(self):
        line = float(self.lvp_line.toPlainText())
        odds = float(self.lvp_odds.toPlainText())
        self.load_data('lvp', self.lvp_stats_table, self.page_lvp,'https://lol.fandom.com/wiki/LVP_SuperLiga/2022_Season/Spring_Season', self.ladder_lvp, 'https://www.rivalry.com/au/esports/league-of-legends-betting/3301-lvp-superliga', self.lvp_matches, line, odds)

    def calc_ljl(self):
        line = float(self.ljl_line.toPlainText())
        odds = float(self.ljl_odds.toPlainText())
        self.load_data('ljl', self.ljl_stats_table, self.page_ljl,'https://lol.fandom.com/wiki/LJL/2022_Season/Spring_Season', self.ladder_ljl, 'https://www.rivalry.com/au/esports/league-of-legends-betting/3318-japan-league', self.ljl_matches, line, odds)

    def calc_nlc(self):
        line = float(self.nlc_line.toPlainText())
        odds = float(self.nlc_odds.toPlainText())
        self.load_data('nlc', self.nlc_stats_table, self.page_nlc,'https://lol.fandom.com/wiki/NLC/2022_Season/Spring_Season', self.ladder_nlc, 'https://www.rivalry.com/esports/league-of-legends-betting/2791-northern-championship', self.nlc_matches, line, odds)

    def calc_lfl(self):
        line = float(self.lfl_line.toPlainText())
        odds = float(self.lfl_odds.toPlainText())
        self.load_data('lfl', self.lfl_stats_table, self.page_lfl,'https://lol.fandom.com/wiki/LFL/2022_Season/Spring_Season', self.ladder_lfl, 'https://www.rivalry.com/au/esports/league-of-legends-betting/2842-la-ligue-francaise', self.lfl_matches, line, odds)

    def calc_lplol(self):
        line = float(self.lplol_line.toPlainText())
        odds = float(self.lplol_odds.toPlainText())
        self.load_data('lplol', self.lplol_stats_table, self.page_lplol,'https://lol.fandom.com/wiki/LPLOL/2022_Season/Spring_Season', self.ladder_lplol, 'https://www.rivalry.com/au/esports/league-of-legends-betting/2918-liga-portuguesa', self.lplol_matches, line, odds)

    def calc_hpm(self):
        line = float(self.hpm_line.toPlainText())
        odds = float(self.hpm_odds.toPlainText())
        self.load_data('hpm', self.hpm_stats_table, self.page_hpm,'https://lol.fandom.com/wiki/Hitpoint_Masters/2022_Season/Spring_Season', self.ladder_hpm, 'https://www.rivalry.com/au/esports/league-of-legends-betting/2975-hitpoint-masters', self.hpm_matches, line, odds)

    def calc_lck_chal(self):
        line = float(self.lck_chal_line.toPlainText())
        odds = float(self.lck_chal_odds.toPlainText())
        self.load_data('lck_chal', self.lck_chal_stats_table, self.page_lck_chal,'https://lol.fandom.com/wiki/LCK_CL/2022_Season/Spring_Season', self.ladder_lck_chal, 'https://www.rivalry.com/au/esports/league-of-legends-betting/3302-lck-challengers-league', self.lck_chal_matches, line, odds)

    def calc_tcl(self):
        line = float(self.tcl_line.toPlainText())
        odds = float(self.tcl_odds.toPlainText())
        self.load_data('tcl', self.tcl_stats_table, self.page_tcl,'https://lol.fandom.com/wiki/TCL/2022_Season/Winter_Season', self.ladder_tcl, 'https://www.rivalry.com/au/esports/league-of-legends-betting/2793-turkish-champions-league', self.tcl_matches, line, odds)

    def calc_pcs(self):
        line = float(self.pcs_line.toPlainText())
        odds = float(self.pcs_odds.toPlainText())
        self.load_data('pcs', self.pcs_stats_table, self.page_pcs,'https://lol.fandom.com/wiki/PCS/2022_Season/Spring_Season', self.ladder_pcs, 'https://www.rivalry.com/au/esports/league-of-legends-betting/2781-pcs', self.pcs_matches, line, odds)

    def calc_ebl(self):
        line = float(self.ebl_line.toPlainText())
        odds = float(self.ebl_odds.toPlainText())
        self.load_data('ebl', self.ebl_stats_table, self.page_ebl,'https://lol.fandom.com/wiki/Esports_Balkan_League/2022_Season/Spring_Split', self.ladder_ebl, 'https://www.rivalry.com/au/esports/league-of-legends-betting/3361-esports-balkan-league', self.ebl_matches, line, odds)

    # updates all values in the statistics table with current data scraped from the relevant websites
    def get_lck(self):
        self.load_data('lck', self.tableWidget, self.page_2,'https://lol.fandom.com/wiki/LCK/2022_Season/Spring_Season', self.ladder_1, 'https://www.rivalry.com/esports/league-of-legends-betting/3254-champions-korea', self.lck_matches, 22.5)

    def get_lcs(self):
        self.load_data('lcs', self.tableWidget_2, self.page_3, 'https://lol.fandom.com/wiki/LCS/2022_Season/Spring_Season', self.ladder_2, 'https://www.rivalry.com/esports/league-of-legends-betting/3713-lcs-north-america', self.lcs_matches, 24.5)

    def get_lec(self):
        self.load_data('lec', self.tableWidget_3, self.page_4, 'https://lol.fandom.com/wiki/LEC/2022_Season/Spring_Season', self.ladder_3, 'https://www.rivalry.com/esports/league-of-legends-betting/3282-european-championship', self.lec_matches, 25.5)

    def get_vcs(self):
        self.load_data('vcs', self.vcs_stats_table, self.page_vcs,'https://lol.fandom.com/wiki/VCS/2022_Season/Spring_Season', self.ladder_vcs, 'https://www.rivalry.com/esports/league-of-legends-betting/3296-vcs', self.vcs_matches, 31.5)

    def get_lpl(self):
        self.load_data('lpl', self.lpl_stats_table, self.page_5,'https://lol.fandom.com/wiki/LPL/2022_Season/Spring_Season', self.ladder_lpl, 'https://www.rivalry.com/esports/league-of-legends-betting/2762-china-lpl', self.lpl_matches, 25.5)

    def get_na_acad(self):
        self.load_data('na_acad', self.na_acad_stats_table, self.page_na_acad,'https://lol.fandom.com/wiki/NA_Academy_League/2022_Season/Spring_Season', self.ladder_na_acad, 'https://www.rivalry.com/esports/league-of-legends-betting/5784-na-academy-league', self.na_acad_matches, 28.5)

    def get_lla(self):
        self.load_data('lla', self.lla_stats_table, self.page_lla,'https://lol.fandom.com/wiki/LLA/2022_Season/Opening_Season', self.ladder_lla, 'https://www.rivalry.com/esports/league-of-legends-betting/3324-liga-latinoamerica', self.lla_matches, 23.5)

    def get_cblol(self):
        self.load_data('cblol', self.cblol_stats_table, self.page_cblol,'https://lol.fandom.com/wiki/CBLOL/2022_Season/Split_1', self.ladder_cblol, 'https://www.rivalry.com/esports/league-of-legends-betting/3330-campeonato-brasileiro', self.cblol_matches, 25.5)

    def get_lco(self):
        self.load_data('lco', self.lco_stats_table, self.page_lco,'https://lol.fandom.com/wiki/LCO/2022_Season/Split_1', self.ladder_lco, 'https://www.rivalry.com/esports/league-of-legends-betting/4139-circuit-oceania', self.lco_matches, 28.5)
    
    def get_ul(self):
        self.load_data('ul', self.ul_stats_table, self.page_ul,'https://lol.fandom.com/wiki/Ultraliga/Season_7', self.ladder_ul, 'https://www.rivalry.com/esports/league-of-legends-betting/3314-ultraliga', self.ul_matches, 28.5)

    def get_pgn(self):
        self.load_data('pgn', self.pgn_stats_table, self.page_pgn,'https://lol.fandom.com/wiki/PG_Nationals/2022_Season/Spring_Season', self.ladder_pgn, 'https://www.rivalry.com/au/esports/league-of-legends-betting/3304-pg-nationals', self.pgn_matches, 27.5)

    def get_pld(self):
        self.load_data('pld', self.pld_stats_table, self.page_pld,'https://lol.fandom.com/wiki/Prime_League_1st_Division/2022_Season/Spring_Season', self.ladder_pld, 'https://www.rivalry.com/au/esports/league-of-legends-betting/5710-prime-league-1st-division', self.pld_matches, 28.5)

    def get_ljl(self):
        self.load_data('ljl', self.ljl_stats_table, self.page_ljl,'https://lol.fandom.com/wiki/LJL/2022_Season/Spring_Season', self.ladder_ljl, 'https://www.rivalry.com/au/esports/league-of-legends-betting/3318-japan-league', self.ljl_matches, 24.5)

    def get_nlc(self):
        self.load_data('nlc', self.nlc_stats_table, self.page_nlc,'https://lol.fandom.com/wiki/NLC/2022_Season/Spring_Season', self.ladder_nlc, 'https://www.rivalry.com/esports/league-of-legends-betting/2791-northern-championship', self.nlc_matches, 27.5)

    def get_lfl(self):
        self.load_data('lfl', self.lfl_stats_table, self.page_lfl,'https://lol.fandom.com/wiki/LFL/2022_Season/Spring_Season', self.ladder_lfl, 'https://www.rivalry.com/au/esports/league-of-legends-betting/2842-la-ligue-francaise', self.lfl_matches, 27.5)

    def get_lplol(self):
        self.load_data('lplol', self.lplol_stats_table, self.page_lplol,'https://lol.fandom.com/wiki/LPLOL/2022_Season/Spring_Season', self.ladder_lplol, 'https://www.rivalry.com/au/esports/league-of-legends-betting/2918-liga-portuguesa', self.lplol_matches, 29.5)

    def get_hpm(self):
        self.load_data('hpm', self.hpm_stats_table, self.page_hpm,'https://lol.fandom.com/wiki/Hitpoint_Masters/2022_Season/Spring_Season', self.ladder_hpm, 'https://www.rivalry.com/au/esports/league-of-legends-betting/2975-hitpoint-masters', self.hpm_matches, 29.5)

    def get_lvp(self):
        self.load_data('lvp', self.lvp_stats_table, self.page_lvp,'https://lol.fandom.com/wiki/LVP_SuperLiga/2022_Season/Spring_Season', self.ladder_lvp, 'https://www.rivalry.com/au/esports/league-of-legends-betting/3301-lvp-superliga', self.lvp_matches, 26.5)

    def get_lck_chal(self):
        self.load_data('lck_chal', self.lck_chal_stats_table, self.page_lck_chal,'https://lol.fandom.com/wiki/LCK_CL/2022_Season/Spring_Season', self.ladder_lck_chal, 'https://www.rivalry.com/au/esports/league-of-legends-betting/3302-lck-challengers-league', self.lck_chal_matches, 26.5)

    def get_tcl(self):
        self.load_data('tcl', self.tcl_stats_table, self.page_tcl,'https://lol.fandom.com/wiki/TCL/2022_Season/Winter_Season', self.ladder_tcl, 'https://www.rivalry.com/au/esports/league-of-legends-betting/2793-turkish-champions-league', self.tcl_matches, 26.5)

    def get_pcs(self):
        self.load_data('pcs', self.pcs_stats_table, self.page_pcs,'https://lol.fandom.com/wiki/PCS/2022_Season/Spring_Season', self.ladder_pcs, 'https://www.rivalry.com/au/esports/league-of-legends-betting/2781-pcs', self.pcs_matches, 27.5)

    def get_ebl(self):
        self.load_data('ebl', self.ebl_stats_table, self.page_ebl,'https://lol.fandom.com/wiki/Esports_Balkan_League/2022_Season/Spring_Split', self.ladder_ebl, 'https://www.rivalry.com/au/esports/league-of-legends-betting/3361-esports-balkan-league', self.ebl_matches, 28.5)

    # auto generated code using PyQt5 to create the UI
    def retranslateUi(self, MainWindow):
        _translate = QtCore.QCoreApplication.translate
        MainWindow.setWindowTitle(_translate("MainWindow", "MainWindow"))
        self.label.setText(_translate("MainWindow", "<html><head/><body><p><span style=\" font-size:16pt;\">League of Legends Kill Total Gambling</span></p></body></html>"))
        self.pushButton.setText(_translate("MainWindow", "LCK"))
        self.lcs.setText(_translate("MainWindow", "LCS"))
        self.lec.setText(_translate("MainWindow", "LEC"))
        self.vcs.setText(_translate("MainWindow", "VCS"))
        self.cblol.setText(_translate("MainWindow", "CBLOL"))
        self.ul.setText(_translate("MainWindow", "UL"))
        self.lla.setText(_translate("MainWindow", "LLA"))
        self.na_acad.setText(_translate("MainWindow", "NA Acad"))
        self.lpl.setText(_translate("MainWindow", "LPL"))
        self.lco.setText(_translate("MainWindow", "LCO"))
        self.pld.setText(_translate("MainWindow", "PLD"))
        self.pgn.setText(_translate("MainWindow", "PGN"))
        item = self.upcoming_table.horizontalHeaderItem(0)
        item.setText(_translate("MainWindow", "League"))
        item = self.upcoming_table.horizontalHeaderItem(1)
        item.setText(_translate("MainWindow", "Match"))
        self.label_77.setText(_translate("MainWindow", "<html><head/><body><p><span style=\" font-size:14pt;\">Upcoming Games</span></p></body></html>"))
        self.ljl.setText(_translate("MainWindow", "LJL"))
        self.lvp.setText(_translate("MainWindow", "LVP"))
        self.nlc.setText(_translate("MainWindow", "NLC"))
        self.lfl.setText(_translate("MainWindow", "LFL"))
        self.lplol.setText(_translate("MainWindow", "LPLOL"))
        self.hpm.setText(_translate("MainWindow", "HPM"))
        self.lck_chal.setText(_translate("MainWindow", "LCK Chal"))
        self.tcl.setText(_translate("MainWindow", "TCL"))
        self.pcs.setText(_translate("MainWindow", "PCS"))
        self.ebl.setText(_translate("MainWindow", "EBL"))
        self.label_102.setText(_translate("MainWindow", "<html><head/><body><p><span style=\" font-size:12pt;\">Line:</span></p></body></html>"))
        item = self.lvp_matches.horizontalHeaderItem(0)
        item.setText(_translate("MainWindow", "Match"))
        item = self.lvp_matches.horizontalHeaderItem(1)
        item.setText(_translate("MainWindow", "Date"))
        item = self.lvp_matches.horizontalHeaderItem(2)
        item.setText(_translate("MainWindow", "Line"))
        item = self.lvp_matches.horizontalHeaderItem(3)
        item.setText(_translate("MainWindow", "G1 Avg"))
        item = self.lvp_matches.horizontalHeaderItem(4)
        item.setText(_translate("MainWindow", "G1 %"))
        item = self.lvp_matches.horizontalHeaderItem(5)
        item.setText(_translate("MainWindow", "G1 Value"))
        item = self.lvp_matches.horizontalHeaderItem(6)
        item.setText(_translate("MainWindow", "G2 Avg"))
        item = self.lvp_matches.horizontalHeaderItem(7)
        item.setText(_translate("MainWindow", "G2 %"))
        item = self.lvp_matches.horizontalHeaderItem(8)
        item.setText(_translate("MainWindow", "G2 Value"))
        self.label_103.setText(_translate("MainWindow", "<html><head/><body><p><span style=\" font-size:12pt;\">Odds:</span></p></body></html>"))
        self.lvp_calculate.setText(_translate("MainWindow", "Calculate"))
        item = self.lvp_stats_table.horizontalHeaderItem(0)
        item.setText(_translate("MainWindow", "Team Name"))
        item = self.lvp_stats_table.horizontalHeaderItem(1)
        item.setText(_translate("MainWindow", "Matches"))
        item = self.lvp_stats_table.horizontalHeaderItem(2)
        item.setText(_translate("MainWindow", "Avg Game 1"))
        item = self.lvp_stats_table.horizontalHeaderItem(3)
        item.setText(_translate("MainWindow", "% Game 1"))
        item = self.lvp_stats_table.horizontalHeaderItem(4)
        item.setText(_translate("MainWindow", "Avg Game 2"))
        item = self.lvp_stats_table.horizontalHeaderItem(5)
        item.setText(_translate("MainWindow", "% Game 2"))
        item = self.lvp_stats_table.horizontalHeaderItem(6)
        item.setText(_translate("MainWindow", "Avg Total"))
        item = self.lvp_stats_table.horizontalHeaderItem(7)
        item.setText(_translate("MainWindow", "% Total"))
        self.label_104.setText(_translate("MainWindow", "<html><head/><body><p><span style=\" font-size:16pt;\">Ladder</span></p></body></html>"))
        self.back_25.setText(_translate("MainWindow", "Back"))
        item = self.ladder_lvp.horizontalHeaderItem(0)
        item.setText(_translate("MainWindow", "Team"))
        item = self.ladder_lvp.horizontalHeaderItem(1)
        item.setText(_translate("MainWindow", "Record"))
        self.update_lvp.setText(_translate("MainWindow", "Update"))
        self.label_105.setText(_translate("MainWindow", "<html><head/><body><p><span style=\" font-size:16pt;\">LVP Statistics</span></p></body></html>"))
        self.label_108.setText(_translate("MainWindow", "<html><head/><body><p><span style=\" font-size:16pt;\">Ladder</span></p></body></html>"))
        self.label_106.setText(_translate("MainWindow", "<html><head/><body><p><span style=\" font-size:12pt;\">Line:</span></p></body></html>"))
        item = self.ljl_matches.horizontalHeaderItem(0)
        item.setText(_translate("MainWindow", "Match"))
        item = self.ljl_matches.horizontalHeaderItem(1)
        item.setText(_translate("MainWindow", "Date"))
        item = self.ljl_matches.horizontalHeaderItem(2)
        item.setText(_translate("MainWindow", "Line"))
        item = self.ljl_matches.horizontalHeaderItem(3)
        item.setText(_translate("MainWindow", "G1 Avg"))
        item = self.ljl_matches.horizontalHeaderItem(4)
        item.setText(_translate("MainWindow", "G1 %"))
        item = self.ljl_matches.horizontalHeaderItem(5)
        item.setText(_translate("MainWindow", "G1 Value"))
        item = self.ljl_matches.horizontalHeaderItem(6)
        item.setText(_translate("MainWindow", "G2 Avg"))
        item = self.ljl_matches.horizontalHeaderItem(7)
        item.setText(_translate("MainWindow", "G2 %"))
        item = self.ljl_matches.horizontalHeaderItem(8)
        item.setText(_translate("MainWindow", "G2 Value"))
        item = self.ladder_ljl.horizontalHeaderItem(0)
        item.setText(_translate("MainWindow", "Team"))
        item = self.ladder_ljl.horizontalHeaderItem(1)
        item.setText(_translate("MainWindow", "Record"))
        self.update_ljl.setText(_translate("MainWindow", "Update"))
        self.label_109.setText(_translate("MainWindow", "<html><head/><body><p><span style=\" font-size:16pt;\">LJL Statistics</span></p></body></html>"))
        self.ljl_calculate.setText(_translate("MainWindow", "Calculate"))
        self.label_107.setText(_translate("MainWindow", "<html><head/><body><p><span style=\" font-size:12pt;\">Odds:</span></p></body></html>"))
        item = self.ljl_stats_table.horizontalHeaderItem(0)
        item.setText(_translate("MainWindow", "Team Name"))
        item = self.ljl_stats_table.horizontalHeaderItem(1)
        item.setText(_translate("MainWindow", "Matches"))
        item = self.ljl_stats_table.horizontalHeaderItem(2)
        item.setText(_translate("MainWindow", "Avg Game 1"))
        item = self.ljl_stats_table.horizontalHeaderItem(3)
        item.setText(_translate("MainWindow", "% Game 1"))
        item = self.ljl_stats_table.horizontalHeaderItem(4)
        item.setText(_translate("MainWindow", "Avg Game 2"))
        item = self.ljl_stats_table.horizontalHeaderItem(5)
        item.setText(_translate("MainWindow", "% Game 2"))
        item = self.ljl_stats_table.horizontalHeaderItem(6)
        item.setText(_translate("MainWindow", "Avg Total"))
        item = self.ljl_stats_table.horizontalHeaderItem(7)
        item.setText(_translate("MainWindow", "% Total"))
        self.back_26.setText(_translate("MainWindow", "Back"))
        self.label_110.setText(_translate("MainWindow", "<html><head/><body><p><span style=\" font-size:12pt;\">Line:</span></p></body></html>"))
        item = self.nlc_matches.horizontalHeaderItem(0)
        item.setText(_translate("MainWindow", "Match"))
        item = self.nlc_matches.horizontalHeaderItem(1)
        item.setText(_translate("MainWindow", "Date"))
        item = self.nlc_matches.horizontalHeaderItem(2)
        item.setText(_translate("MainWindow", "Line"))
        item = self.nlc_matches.horizontalHeaderItem(3)
        item.setText(_translate("MainWindow", "G1 Avg"))
        item = self.nlc_matches.horizontalHeaderItem(4)
        item.setText(_translate("MainWindow", "G1 %"))
        item = self.nlc_matches.horizontalHeaderItem(5)
        item.setText(_translate("MainWindow", "G1 Value"))
        item = self.nlc_matches.horizontalHeaderItem(6)
        item.setText(_translate("MainWindow", "G2 Avg"))
        item = self.nlc_matches.horizontalHeaderItem(7)
        item.setText(_translate("MainWindow", "G2 %"))
        item = self.nlc_matches.horizontalHeaderItem(8)
        item.setText(_translate("MainWindow", "G2 Value"))
        self.label_111.setText(_translate("MainWindow", "<html><head/><body><p><span style=\" font-size:12pt;\">Odds:</span></p></body></html>"))
        self.nlc_calculate.setText(_translate("MainWindow", "Calculate"))
        item = self.nlc_stats_table.horizontalHeaderItem(0)
        item.setText(_translate("MainWindow", "Team Name"))
        item = self.nlc_stats_table.horizontalHeaderItem(1)
        item.setText(_translate("MainWindow", "Matches"))
        item = self.nlc_stats_table.horizontalHeaderItem(2)
        item.setText(_translate("MainWindow", "Avg Game 1"))
        item = self.nlc_stats_table.horizontalHeaderItem(3)
        item.setText(_translate("MainWindow", "% Game 1"))
        item = self.nlc_stats_table.horizontalHeaderItem(4)
        item.setText(_translate("MainWindow", "Avg Game 2"))
        item = self.nlc_stats_table.horizontalHeaderItem(5)
        item.setText(_translate("MainWindow", "% Game 2"))
        item = self.nlc_stats_table.horizontalHeaderItem(6)
        item.setText(_translate("MainWindow", "Avg Total"))
        item = self.nlc_stats_table.horizontalHeaderItem(7)
        item.setText(_translate("MainWindow", "% Total"))
        self.label_112.setText(_translate("MainWindow", "<html><head/><body><p><span style=\" font-size:16pt;\">Ladder</span></p></body></html>"))
        self.back_27.setText(_translate("MainWindow", "Back"))
        item = self.ladder_nlc.horizontalHeaderItem(0)
        item.setText(_translate("MainWindow", "Team"))
        item = self.ladder_nlc.horizontalHeaderItem(1)
        item.setText(_translate("MainWindow", "Record"))
        self.update_nlc.setText(_translate("MainWindow", "Update"))
        self.label_113.setText(_translate("MainWindow", "<html><head/><body><p><span style=\" font-size:16pt;\">NLC Statistics</span></p></body></html>"))
        self.update_ul.setText(_translate("MainWindow", "Update"))
        self.back_19.setText(_translate("MainWindow", "Back"))
        self.label_78.setText(_translate("MainWindow", "<html><head/><body><p><span style=\" font-size:12pt;\">Odds:</span></p></body></html>"))
        self.ul_calculate.setText(_translate("MainWindow", "Calculate"))
        item = self.ul_matches.horizontalHeaderItem(0)
        item.setText(_translate("MainWindow", "Match"))
        item = self.ul_matches.horizontalHeaderItem(1)
        item.setText(_translate("MainWindow", "Date"))
        item = self.ul_matches.horizontalHeaderItem(2)
        item.setText(_translate("MainWindow", "Line"))
        item = self.ul_matches.horizontalHeaderItem(3)
        item.setText(_translate("MainWindow", "G1 Avg"))
        item = self.ul_matches.horizontalHeaderItem(4)
        item.setText(_translate("MainWindow", "G1 %"))
        item = self.ul_matches.horizontalHeaderItem(5)
        item.setText(_translate("MainWindow", "G1 Value"))
        item = self.ul_matches.horizontalHeaderItem(6)
        item.setText(_translate("MainWindow", "G2 Avg"))
        item = self.ul_matches.horizontalHeaderItem(7)
        item.setText(_translate("MainWindow", "G2 %"))
        item = self.ul_matches.horizontalHeaderItem(8)
        item.setText(_translate("MainWindow", "G2 Value"))
        self.label_79.setText(_translate("MainWindow", "<html><head/><body><p><span style=\" font-size:16pt;\">UL Statistics</span></p></body></html>"))
        self.label_80.setText(_translate("MainWindow", "<html><head/><body><p><span style=\" font-size:12pt;\">Line:</span></p></body></html>"))
        item = self.ul_stats_table.horizontalHeaderItem(0)
        item.setText(_translate("MainWindow", "Team Name"))
        item = self.ul_stats_table.horizontalHeaderItem(1)
        item.setText(_translate("MainWindow", "Matches"))
        item = self.ul_stats_table.horizontalHeaderItem(2)
        item.setText(_translate("MainWindow", "Avg Game 1"))
        item = self.ul_stats_table.horizontalHeaderItem(3)
        item.setText(_translate("MainWindow", "% Game 1"))
        item = self.ul_stats_table.horizontalHeaderItem(4)
        item.setText(_translate("MainWindow", "Avg Game 2"))
        item = self.ul_stats_table.horizontalHeaderItem(5)
        item.setText(_translate("MainWindow", "% Game 2"))
        item = self.ul_stats_table.horizontalHeaderItem(6)
        item.setText(_translate("MainWindow", "Avg Total"))
        item = self.ul_stats_table.horizontalHeaderItem(7)
        item.setText(_translate("MainWindow", "% Total"))
        self.label_81.setText(_translate("MainWindow", "<html><head/><body><p><span style=\" font-size:16pt;\">Ladder</span></p></body></html>"))
        item = self.ladder_ul.horizontalHeaderItem(0)
        item.setText(_translate("MainWindow", "Team"))
        item = self.ladder_ul.horizontalHeaderItem(1)
        item.setText(_translate("MainWindow", "Record"))
        self.update_pld.setText(_translate("MainWindow", "Update"))
        self.back_20.setText(_translate("MainWindow", "Back"))
        self.label_82.setText(_translate("MainWindow", "<html><head/><body><p><span style=\" font-size:12pt;\">Odds:</span></p></body></html>"))
        self.pld_calculate.setText(_translate("MainWindow", "Calculate"))
        item = self.pld_matches.horizontalHeaderItem(0)
        item.setText(_translate("MainWindow", "Match"))
        item = self.pld_matches.horizontalHeaderItem(1)
        item.setText(_translate("MainWindow", "Date"))
        item = self.pld_matches.horizontalHeaderItem(2)
        item.setText(_translate("MainWindow", "Line"))
        item = self.pld_matches.horizontalHeaderItem(3)
        item.setText(_translate("MainWindow", "G1 Avg"))
        item = self.pld_matches.horizontalHeaderItem(4)
        item.setText(_translate("MainWindow", "G1 %"))
        item = self.pld_matches.horizontalHeaderItem(5)
        item.setText(_translate("MainWindow", "G1 Value"))
        item = self.pld_matches.horizontalHeaderItem(6)
        item.setText(_translate("MainWindow", "G2 Avg"))
        item = self.pld_matches.horizontalHeaderItem(7)
        item.setText(_translate("MainWindow", "G2 %"))
        item = self.pld_matches.horizontalHeaderItem(8)
        item.setText(_translate("MainWindow", "G2 Value"))
        self.label_83.setText(_translate("MainWindow", "<html><head/><body><p><span style=\" font-size:16pt;\">PLD Statistics</span></p></body></html>"))
        self.label_84.setText(_translate("MainWindow", "<html><head/><body><p><span style=\" font-size:12pt;\">Line:</span></p></body></html>"))
        item = self.pld_stats_table.horizontalHeaderItem(0)
        item.setText(_translate("MainWindow", "Team Name"))
        item = self.pld_stats_table.horizontalHeaderItem(1)
        item.setText(_translate("MainWindow", "Matches"))
        item = self.pld_stats_table.horizontalHeaderItem(2)
        item.setText(_translate("MainWindow", "Avg Game 1"))
        item = self.pld_stats_table.horizontalHeaderItem(3)
        item.setText(_translate("MainWindow", "% Game 1"))
        item = self.pld_stats_table.horizontalHeaderItem(4)
        item.setText(_translate("MainWindow", "Avg Game 2"))
        item = self.pld_stats_table.horizontalHeaderItem(5)
        item.setText(_translate("MainWindow", "% Game 2"))
        item = self.pld_stats_table.horizontalHeaderItem(6)
        item.setText(_translate("MainWindow", "Avg Total"))
        item = self.pld_stats_table.horizontalHeaderItem(7)
        item.setText(_translate("MainWindow", "% Total"))
        self.label_85.setText(_translate("MainWindow", "<html><head/><body><p><span style=\" font-size:16pt;\">Ladder</span></p></body></html>"))
        item = self.ladder_pld.horizontalHeaderItem(0)
        item.setText(_translate("MainWindow", "Team"))
        item = self.ladder_pld.horizontalHeaderItem(1)
        item.setText(_translate("MainWindow", "Record"))
        self.label_88.setText(_translate("MainWindow", "<html><head/><body><p><span style=\" font-size:12pt;\">Line:</span></p></body></html>"))
        self.label_89.setText(_translate("MainWindow", "<html><head/><body><p><span style=\" font-size:16pt;\">Ladder</span></p></body></html>"))
        self.pgn_calculate.setText(_translate("MainWindow", "Calculate"))
        self.back_21.setText(_translate("MainWindow", "Back"))
        self.label_86.setText(_translate("MainWindow", "<html><head/><body><p><span style=\" font-size:12pt;\">Odds:</span></p></body></html>"))
        item = self.pgn_stats_table.horizontalHeaderItem(0)
        item.setText(_translate("MainWindow", "Team Name"))
        item = self.pgn_stats_table.horizontalHeaderItem(1)
        item.setText(_translate("MainWindow", "Matches"))
        item = self.pgn_stats_table.horizontalHeaderItem(2)
        item.setText(_translate("MainWindow", "Avg Game 1"))
        item = self.pgn_stats_table.horizontalHeaderItem(3)
        item.setText(_translate("MainWindow", "% Game 1"))
        item = self.pgn_stats_table.horizontalHeaderItem(4)
        item.setText(_translate("MainWindow", "Avg Game 2"))
        item = self.pgn_stats_table.horizontalHeaderItem(5)
        item.setText(_translate("MainWindow", "% Game 2"))
        item = self.pgn_stats_table.horizontalHeaderItem(6)
        item.setText(_translate("MainWindow", "Avg Total"))
        item = self.pgn_stats_table.horizontalHeaderItem(7)
        item.setText(_translate("MainWindow", "% Total"))
        self.update_pgn.setText(_translate("MainWindow", "Update"))
        item = self.ladder_pgn.horizontalHeaderItem(0)
        item.setText(_translate("MainWindow", "Team"))
        item = self.ladder_pgn.horizontalHeaderItem(1)
        item.setText(_translate("MainWindow", "Record"))
        item = self.pgn_matches.horizontalHeaderItem(0)
        item.setText(_translate("MainWindow", "Match"))
        item = self.pgn_matches.horizontalHeaderItem(1)
        item.setText(_translate("MainWindow", "Date"))
        item = self.pgn_matches.horizontalHeaderItem(2)
        item.setText(_translate("MainWindow", "Line"))
        item = self.pgn_matches.horizontalHeaderItem(3)
        item.setText(_translate("MainWindow", "G1 Avg"))
        item = self.pgn_matches.horizontalHeaderItem(4)
        item.setText(_translate("MainWindow", "G1 %"))
        item = self.pgn_matches.horizontalHeaderItem(5)
        item.setText(_translate("MainWindow", "G1 Value"))
        item = self.pgn_matches.horizontalHeaderItem(6)
        item.setText(_translate("MainWindow", "G2 Avg"))
        item = self.pgn_matches.horizontalHeaderItem(7)
        item.setText(_translate("MainWindow", "G2 %"))
        item = self.pgn_matches.horizontalHeaderItem(8)
        item.setText(_translate("MainWindow", "G2 Value"))
        self.label_87.setText(_translate("MainWindow", "<html><head/><body><p><span style=\" font-size:16pt;\">PGN Statistics</span></p></body></html>"))
        self.update_lfl.setText(_translate("MainWindow", "Update"))
        self.back_22.setText(_translate("MainWindow", "Back"))
        self.label_90.setText(_translate("MainWindow", "<html><head/><body><p><span style=\" font-size:12pt;\">Odds:</span></p></body></html>"))
        self.lfl_calculate.setText(_translate("MainWindow", "Calculate"))
        item = self.lfl_matches.horizontalHeaderItem(0)
        item.setText(_translate("MainWindow", "Match"))
        item = self.lfl_matches.horizontalHeaderItem(1)
        item.setText(_translate("MainWindow", "Date"))
        item = self.lfl_matches.horizontalHeaderItem(2)
        item.setText(_translate("MainWindow", "Line"))
        item = self.lfl_matches.horizontalHeaderItem(3)
        item.setText(_translate("MainWindow", "G1 Avg"))
        item = self.lfl_matches.horizontalHeaderItem(4)
        item.setText(_translate("MainWindow", "G1 %"))
        item = self.lfl_matches.horizontalHeaderItem(5)
        item.setText(_translate("MainWindow", "G1 Value"))
        item = self.lfl_matches.horizontalHeaderItem(6)
        item.setText(_translate("MainWindow", "G2 Avg"))
        item = self.lfl_matches.horizontalHeaderItem(7)
        item.setText(_translate("MainWindow", "G2 %"))
        item = self.lfl_matches.horizontalHeaderItem(8)
        item.setText(_translate("MainWindow", "G2 Value"))
        self.label_91.setText(_translate("MainWindow", "<html><head/><body><p><span style=\" font-size:16pt;\">LFL Statistics</span></p></body></html>"))
        self.label_92.setText(_translate("MainWindow", "<html><head/><body><p><span style=\" font-size:12pt;\">Line:</span></p></body></html>"))
        item = self.lfl_stats_table.horizontalHeaderItem(0)
        item.setText(_translate("MainWindow", "Team Name"))
        item = self.lfl_stats_table.horizontalHeaderItem(1)
        item.setText(_translate("MainWindow", "Matches"))
        item = self.lfl_stats_table.horizontalHeaderItem(2)
        item.setText(_translate("MainWindow", "Avg Game 1"))
        item = self.lfl_stats_table.horizontalHeaderItem(3)
        item.setText(_translate("MainWindow", "% Game 1"))
        item = self.lfl_stats_table.horizontalHeaderItem(4)
        item.setText(_translate("MainWindow", "Avg Game 2"))
        item = self.lfl_stats_table.horizontalHeaderItem(5)
        item.setText(_translate("MainWindow", "% Game 2"))
        item = self.lfl_stats_table.horizontalHeaderItem(6)
        item.setText(_translate("MainWindow", "Avg Total"))
        item = self.lfl_stats_table.horizontalHeaderItem(7)
        item.setText(_translate("MainWindow", "% Total"))
        self.label_93.setText(_translate("MainWindow", "<html><head/><body><p><span style=\" font-size:16pt;\">Ladder</span></p></body></html>"))
        item = self.ladder_lfl.horizontalHeaderItem(0)
        item.setText(_translate("MainWindow", "Team"))
        item = self.ladder_lfl.horizontalHeaderItem(1)
        item.setText(_translate("MainWindow", "Record"))
        self.update_lplol.setText(_translate("MainWindow", "Update"))
        self.back_23.setText(_translate("MainWindow", "Back"))
        self.label_94.setText(_translate("MainWindow", "<html><head/><body><p><span style=\" font-size:12pt;\">Odds:</span></p></body></html>"))
        self.lplol_calculate.setText(_translate("MainWindow", "Calculate"))
        item = self.lplol_matches.horizontalHeaderItem(0)
        item.setText(_translate("MainWindow", "Match"))
        item = self.lplol_matches.horizontalHeaderItem(1)
        item.setText(_translate("MainWindow", "Date"))
        item = self.lplol_matches.horizontalHeaderItem(2)
        item.setText(_translate("MainWindow", "Line"))
        item = self.lplol_matches.horizontalHeaderItem(3)
        item.setText(_translate("MainWindow", "G1 Avg"))
        item = self.lplol_matches.horizontalHeaderItem(4)
        item.setText(_translate("MainWindow", "G1 %"))
        item = self.lplol_matches.horizontalHeaderItem(5)
        item.setText(_translate("MainWindow", "G1 Value"))
        item = self.lplol_matches.horizontalHeaderItem(6)
        item.setText(_translate("MainWindow", "G2 Avg"))
        item = self.lplol_matches.horizontalHeaderItem(7)
        item.setText(_translate("MainWindow", "G2 %"))
        item = self.lplol_matches.horizontalHeaderItem(8)
        item.setText(_translate("MainWindow", "G2 Value"))
        self.label_95.setText(_translate("MainWindow", "<html><head/><body><p><span style=\" font-size:16pt;\">LPLOL Statistics</span></p></body></html>"))
        self.label_96.setText(_translate("MainWindow", "<html><head/><body><p><span style=\" font-size:12pt;\">Line:</span></p></body></html>"))
        item = self.lplol_stats_table.horizontalHeaderItem(0)
        item.setText(_translate("MainWindow", "Team Name"))
        item = self.lplol_stats_table.horizontalHeaderItem(1)
        item.setText(_translate("MainWindow", "Matches"))
        item = self.lplol_stats_table.horizontalHeaderItem(2)
        item.setText(_translate("MainWindow", "Avg Game 1"))
        item = self.lplol_stats_table.horizontalHeaderItem(3)
        item.setText(_translate("MainWindow", "% Game 1"))
        item = self.lplol_stats_table.horizontalHeaderItem(4)
        item.setText(_translate("MainWindow", "Avg Game 2"))
        item = self.lplol_stats_table.horizontalHeaderItem(5)
        item.setText(_translate("MainWindow", "% Game 2"))
        item = self.lplol_stats_table.horizontalHeaderItem(6)
        item.setText(_translate("MainWindow", "Avg Total"))
        item = self.lplol_stats_table.horizontalHeaderItem(7)
        item.setText(_translate("MainWindow", "% Total"))
        self.label_97.setText(_translate("MainWindow", "<html><head/><body><p><span style=\" font-size:16pt;\">Ladder</span></p></body></html>"))
        item = self.ladder_lplol.horizontalHeaderItem(0)
        item.setText(_translate("MainWindow", "Team"))
        item = self.ladder_lplol.horizontalHeaderItem(1)
        item.setText(_translate("MainWindow", "Record"))
        self.update_hpm.setText(_translate("MainWindow", "Update"))
        self.back_24.setText(_translate("MainWindow", "Back"))
        self.label_98.setText(_translate("MainWindow", "<html><head/><body><p><span style=\" font-size:12pt;\">Odds:</span></p></body></html>"))
        self.hpm_calculate.setText(_translate("MainWindow", "Calculate"))
        item = self.hpm_matches.horizontalHeaderItem(0)
        item.setText(_translate("MainWindow", "Match"))
        item = self.hpm_matches.horizontalHeaderItem(1)
        item.setText(_translate("MainWindow", "Date"))
        item = self.hpm_matches.horizontalHeaderItem(2)
        item.setText(_translate("MainWindow", "Line"))
        item = self.hpm_matches.horizontalHeaderItem(3)
        item.setText(_translate("MainWindow", "G1 Avg"))
        item = self.hpm_matches.horizontalHeaderItem(4)
        item.setText(_translate("MainWindow", "G1 %"))
        item = self.hpm_matches.horizontalHeaderItem(5)
        item.setText(_translate("MainWindow", "G1 Value"))
        item = self.hpm_matches.horizontalHeaderItem(6)
        item.setText(_translate("MainWindow", "G2 Avg"))
        item = self.hpm_matches.horizontalHeaderItem(7)
        item.setText(_translate("MainWindow", "G2 %"))
        item = self.hpm_matches.horizontalHeaderItem(8)
        item.setText(_translate("MainWindow", "G2 Value"))
        self.label_99.setText(_translate("MainWindow", "<html><head/><body><p><span style=\" font-size:16pt;\">HPM Statistics</span></p></body></html>"))
        self.label_100.setText(_translate("MainWindow", "<html><head/><body><p><span style=\" font-size:12pt;\">Line:</span></p></body></html>"))
        item = self.hpm_stats_table.horizontalHeaderItem(0)
        item.setText(_translate("MainWindow", "Team Name"))
        item = self.hpm_stats_table.horizontalHeaderItem(1)
        item.setText(_translate("MainWindow", "Matches"))
        item = self.hpm_stats_table.horizontalHeaderItem(2)
        item.setText(_translate("MainWindow", "Avg Game 1"))
        item = self.hpm_stats_table.horizontalHeaderItem(3)
        item.setText(_translate("MainWindow", "% Game 1"))
        item = self.hpm_stats_table.horizontalHeaderItem(4)
        item.setText(_translate("MainWindow", "Avg Game 2"))
        item = self.hpm_stats_table.horizontalHeaderItem(5)
        item.setText(_translate("MainWindow", "% Game 2"))
        item = self.hpm_stats_table.horizontalHeaderItem(6)
        item.setText(_translate("MainWindow", "Avg Total"))
        item = self.hpm_stats_table.horizontalHeaderItem(7)
        item.setText(_translate("MainWindow", "% Total"))
        self.label_101.setText(_translate("MainWindow", "<html><head/><body><p><span style=\" font-size:16pt;\">Ladder</span></p></body></html>"))
        item = self.ladder_hpm.horizontalHeaderItem(0)
        item.setText(_translate("MainWindow", "Team"))
        item = self.ladder_hpm.horizontalHeaderItem(1)
        item.setText(_translate("MainWindow", "Record"))
        self.label_2.setText(_translate("MainWindow", "<html><head/><body><p><span style=\" font-size:16pt;\">LCK Statistics</span></p></body></html>"))
        item = self.tableWidget.horizontalHeaderItem(0)
        item.setText(_translate("MainWindow", "Team Name"))
        item = self.tableWidget.horizontalHeaderItem(1)
        item.setText(_translate("MainWindow", "Matches"))
        item = self.tableWidget.horizontalHeaderItem(2)
        item.setText(_translate("MainWindow", "Avg Game 1"))
        item = self.tableWidget.horizontalHeaderItem(3)
        item.setText(_translate("MainWindow", "% Game 1"))
        item = self.tableWidget.horizontalHeaderItem(4)
        item.setText(_translate("MainWindow", "Avg Game 2"))
        item = self.tableWidget.horizontalHeaderItem(5)
        item.setText(_translate("MainWindow", "% Game 2"))
        item = self.tableWidget.horizontalHeaderItem(6)
        item.setText(_translate("MainWindow", "Avg Total"))
        item = self.tableWidget.horizontalHeaderItem(7)
        item.setText(_translate("MainWindow", "% Total"))
        self.update_lck.setText(_translate("MainWindow", "Update"))
        self.back.setText(_translate("MainWindow", "Back"))
        item = self.ladder_1.horizontalHeaderItem(0)
        item.setText(_translate("MainWindow", "Team"))
        item = self.ladder_1.horizontalHeaderItem(1)
        item.setText(_translate("MainWindow", "Record"))
        self.label_4.setText(_translate("MainWindow", "<html><head/><body><p><span style=\" font-size:16pt;\">Ladder</span></p></body></html>"))
        item = self.lck_matches.horizontalHeaderItem(0)
        item.setText(_translate("MainWindow", "Match"))
        item = self.lck_matches.horizontalHeaderItem(1)
        item.setText(_translate("MainWindow", "Date"))
        item = self.lck_matches.horizontalHeaderItem(2)
        item.setText(_translate("MainWindow", "Line"))
        item = self.lck_matches.horizontalHeaderItem(3)
        item.setText(_translate("MainWindow", "G1 Avg"))
        item = self.lck_matches.horizontalHeaderItem(4)
        item.setText(_translate("MainWindow", "G1 %"))
        item = self.lck_matches.horizontalHeaderItem(5)
        item.setText(_translate("MainWindow", "G1 Value"))
        item = self.lck_matches.horizontalHeaderItem(6)
        item.setText(_translate("MainWindow", "G2 Avg"))
        item = self.lck_matches.horizontalHeaderItem(7)
        item.setText(_translate("MainWindow", "G2 %"))
        item = self.lck_matches.horizontalHeaderItem(8)
        item.setText(_translate("MainWindow", "G2 Value"))
        self.label_8.setText(_translate("MainWindow", "<html><head/><body><p><span style=\" font-size:12pt;\">Line:</span></p></body></html>"))
        self.label_9.setText(_translate("MainWindow", "<html><head/><body><p><span style=\" font-size:12pt;\">Odds:</span></p></body></html>"))
        self.lck_calculate.setText(_translate("MainWindow", "Calculate"))
        item = self.tableWidget_2.horizontalHeaderItem(0)
        item.setText(_translate("MainWindow", "Team Name"))
        item = self.tableWidget_2.horizontalHeaderItem(1)
        item.setText(_translate("MainWindow", "Matches"))
        item = self.tableWidget_2.horizontalHeaderItem(2)
        item.setText(_translate("MainWindow", "Avg Game 1"))
        item = self.tableWidget_2.horizontalHeaderItem(3)
        item.setText(_translate("MainWindow", "% Game 1"))
        item = self.tableWidget_2.horizontalHeaderItem(4)
        item.setText(_translate("MainWindow", "Avg Game 2"))
        item = self.tableWidget_2.horizontalHeaderItem(5)
        item.setText(_translate("MainWindow", "% Game 2"))
        item = self.tableWidget_2.horizontalHeaderItem(6)
        item.setText(_translate("MainWindow", "Avg Total"))
        item = self.tableWidget_2.horizontalHeaderItem(7)
        item.setText(_translate("MainWindow", "% Total"))
        self.label_3.setText(_translate("MainWindow", "<html><head/><body><p><span style=\" font-size:16pt;\">LCS Statistics</span></p></body></html>"))
        self.back_2.setText(_translate("MainWindow", "Back"))
        self.update_lcs.setText(_translate("MainWindow", "Update"))
        item = self.ladder_2.horizontalHeaderItem(0)
        item.setText(_translate("MainWindow", "Team"))
        item = self.ladder_2.horizontalHeaderItem(1)
        item.setText(_translate("MainWindow", "Record"))
        self.label_5.setText(_translate("MainWindow", "<html><head/><body><p><span style=\" font-size:16pt;\">Ladder</span></p></body></html>"))
        item = self.lcs_matches.horizontalHeaderItem(0)
        item.setText(_translate("MainWindow", "Match"))
        item = self.lcs_matches.horizontalHeaderItem(1)
        item.setText(_translate("MainWindow", "Date"))
        item = self.lcs_matches.horizontalHeaderItem(2)
        item.setText(_translate("MainWindow", "Line"))
        item = self.lcs_matches.horizontalHeaderItem(3)
        item.setText(_translate("MainWindow", "G1 Avg"))
        item = self.lcs_matches.horizontalHeaderItem(4)
        item.setText(_translate("MainWindow", "G1 %"))
        item = self.lcs_matches.horizontalHeaderItem(5)
        item.setText(_translate("MainWindow", "G1 Value"))
        item = self.lcs_matches.horizontalHeaderItem(6)
        item.setText(_translate("MainWindow", "G2 Avg"))
        item = self.lcs_matches.horizontalHeaderItem(7)
        item.setText(_translate("MainWindow", "G2 %"))
        item = self.lcs_matches.horizontalHeaderItem(8)
        item.setText(_translate("MainWindow", "G2 Value"))
        self.lcs_calculate.setText(_translate("MainWindow", "Calculate"))
        self.label_10.setText(_translate("MainWindow", "<html><head/><body><p><span style=\" font-size:12pt;\">Odds:</span></p></body></html>"))
        self.label_14.setText(_translate("MainWindow", "<html><head/><body><p><span style=\" font-size:12pt;\">Line:</span></p></body></html>"))
        item = self.tableWidget_3.horizontalHeaderItem(0)
        item.setText(_translate("MainWindow", "Team Name"))
        item = self.tableWidget_3.horizontalHeaderItem(1)
        item.setText(_translate("MainWindow", "Matches"))
        item = self.tableWidget_3.horizontalHeaderItem(2)
        item.setText(_translate("MainWindow", "Avg Game 1"))
        item = self.tableWidget_3.horizontalHeaderItem(3)
        item.setText(_translate("MainWindow", "% Game 1"))
        item = self.tableWidget_3.horizontalHeaderItem(4)
        item.setText(_translate("MainWindow", "Avg Game 2"))
        item = self.tableWidget_3.horizontalHeaderItem(5)
        item.setText(_translate("MainWindow", "% Game 2"))
        item = self.tableWidget_3.horizontalHeaderItem(6)
        item.setText(_translate("MainWindow", "Avg Total"))
        item = self.tableWidget_3.horizontalHeaderItem(7)
        item.setText(_translate("MainWindow", "% Total"))
        self.label_6.setText(_translate("MainWindow", "<html><head/><body><p><span style=\" font-size:16pt;\">LEC Statistics</span></p></body></html>"))
        self.back_3.setText(_translate("MainWindow", "Back"))
        self.update_lec.setText(_translate("MainWindow", "Update"))
        self.label_7.setText(_translate("MainWindow", "<html><head/><body><p><span style=\" font-size:16pt;\">Ladder</span></p></body></html>"))
        item = self.ladder_3.horizontalHeaderItem(0)
        item.setText(_translate("MainWindow", "Team"))
        item = self.ladder_3.horizontalHeaderItem(1)
        item.setText(_translate("MainWindow", "Record"))
        item = self.lec_matches.horizontalHeaderItem(0)
        item.setText(_translate("MainWindow", "Match"))
        item = self.lec_matches.horizontalHeaderItem(1)
        item.setText(_translate("MainWindow", "Date"))
        item = self.lec_matches.horizontalHeaderItem(2)
        item.setText(_translate("MainWindow", "Line"))
        item = self.lec_matches.horizontalHeaderItem(3)
        item.setText(_translate("MainWindow", "G1 Avg"))
        item = self.lec_matches.horizontalHeaderItem(4)
        item.setText(_translate("MainWindow", "G1 %"))
        item = self.lec_matches.horizontalHeaderItem(5)
        item.setText(_translate("MainWindow", "G1 Value"))
        item = self.lec_matches.horizontalHeaderItem(6)
        item.setText(_translate("MainWindow", "G2 Avg"))
        item = self.lec_matches.horizontalHeaderItem(7)
        item.setText(_translate("MainWindow", "G2 %"))
        item = self.lec_matches.horizontalHeaderItem(8)
        item.setText(_translate("MainWindow", "G2 Value"))
        self.lec_calculate.setText(_translate("MainWindow", "Calculate"))
        self.label_11.setText(_translate("MainWindow", "<html><head/><body><p><span style=\" font-size:12pt;\">Odds:</span></p></body></html>"))
        self.label_12.setText(_translate("MainWindow", "<html><head/><body><p><span style=\" font-size:12pt;\">Line:</span></p></body></html>"))
        self.label_53.setText(_translate("MainWindow", "<html><head/><body><p><span style=\" font-size:16pt;\">LPL Statistics</span></p></body></html>"))
        self.label_54.setText(_translate("MainWindow", "<html><head/><body><p><span style=\" font-size:12pt;\">Odds:</span></p></body></html>"))
        item = self.lpl_stats_table.horizontalHeaderItem(0)
        item.setText(_translate("MainWindow", "Team Name"))
        item = self.lpl_stats_table.horizontalHeaderItem(1)
        item.setText(_translate("MainWindow", "Matches"))
        item = self.lpl_stats_table.horizontalHeaderItem(2)
        item.setText(_translate("MainWindow", "Avg Game 1"))
        item = self.lpl_stats_table.horizontalHeaderItem(3)
        item.setText(_translate("MainWindow", "% Game 1"))
        item = self.lpl_stats_table.horizontalHeaderItem(4)
        item.setText(_translate("MainWindow", "Avg Game 2"))
        item = self.lpl_stats_table.horizontalHeaderItem(5)
        item.setText(_translate("MainWindow", "% Game 2"))
        item = self.lpl_stats_table.horizontalHeaderItem(6)
        item.setText(_translate("MainWindow", "Avg Total"))
        item = self.lpl_stats_table.horizontalHeaderItem(7)
        item.setText(_translate("MainWindow", "% Total"))
        self.label_55.setText(_translate("MainWindow", "<html><head/><body><p><span style=\" font-size:16pt;\">Ladder</span></p></body></html>"))
        item = self.ladder_lpl.horizontalHeaderItem(0)
        item.setText(_translate("MainWindow", "Team"))
        item = self.ladder_lpl.horizontalHeaderItem(1)
        item.setText(_translate("MainWindow", "Record"))
        item = self.lpl_matches.horizontalHeaderItem(0)
        item.setText(_translate("MainWindow", "Match"))
        item = self.lpl_matches.horizontalHeaderItem(1)
        item.setText(_translate("MainWindow", "Date"))
        item = self.lpl_matches.horizontalHeaderItem(2)
        item.setText(_translate("MainWindow", "Line"))
        item = self.lpl_matches.horizontalHeaderItem(3)
        item.setText(_translate("MainWindow", "G1 Avg"))
        item = self.lpl_matches.horizontalHeaderItem(4)
        item.setText(_translate("MainWindow", "G1 %"))
        item = self.lpl_matches.horizontalHeaderItem(5)
        item.setText(_translate("MainWindow", "G1 Value"))
        item = self.lpl_matches.horizontalHeaderItem(6)
        item.setText(_translate("MainWindow", "G2 Avg"))
        item = self.lpl_matches.horizontalHeaderItem(7)
        item.setText(_translate("MainWindow", "G2 %"))
        item = self.lpl_matches.horizontalHeaderItem(8)
        item.setText(_translate("MainWindow", "G2 Value"))
        self.lpl_calculate.setText(_translate("MainWindow", "Calculate"))
        self.label_56.setText(_translate("MainWindow", "<html><head/><body><p><span style=\" font-size:12pt;\">Line:</span></p></body></html>"))
        self.back_13.setText(_translate("MainWindow", "Back"))
        self.update_lpl.setText(_translate("MainWindow", "Update"))
        self.update_vcs.setText(_translate("MainWindow", "Update"))
        self.vcs_calculate.setText(_translate("MainWindow", "Calculate"))
        self.label_60.setText(_translate("MainWindow", "<html><head/><body><p><span style=\" font-size:12pt;\">Line:</span></p></body></html>"))
        self.back_14.setText(_translate("MainWindow", "Back"))
        self.label_57.setText(_translate("MainWindow", "<html><head/><body><p><span style=\" font-size:16pt;\">VCS Statistics</span></p></body></html>"))
        self.label_58.setText(_translate("MainWindow", "<html><head/><body><p><span style=\" font-size:12pt;\">Odds:</span></p></body></html>"))
        item = self.ladder_vcs.horizontalHeaderItem(0)
        item.setText(_translate("MainWindow", "Team"))
        item = self.ladder_vcs.horizontalHeaderItem(1)
        item.setText(_translate("MainWindow", "Record"))
        item = self.vcs_stats_table.horizontalHeaderItem(0)
        item.setText(_translate("MainWindow", "Team Name"))
        item = self.vcs_stats_table.horizontalHeaderItem(1)
        item.setText(_translate("MainWindow", "Matches"))
        item = self.vcs_stats_table.horizontalHeaderItem(2)
        item.setText(_translate("MainWindow", "Avg Game 1"))
        item = self.vcs_stats_table.horizontalHeaderItem(3)
        item.setText(_translate("MainWindow", "% Game 1"))
        item = self.vcs_stats_table.horizontalHeaderItem(4)
        item.setText(_translate("MainWindow", "Avg Game 2"))
        item = self.vcs_stats_table.horizontalHeaderItem(5)
        item.setText(_translate("MainWindow", "% Game 2"))
        item = self.vcs_stats_table.horizontalHeaderItem(6)
        item.setText(_translate("MainWindow", "Avg Total"))
        item = self.vcs_stats_table.horizontalHeaderItem(7)
        item.setText(_translate("MainWindow", "% Total"))
        self.label_59.setText(_translate("MainWindow", "<html><head/><body><p><span style=\" font-size:16pt;\">Ladder</span></p></body></html>"))
        item = self.vcs_matches.horizontalHeaderItem(0)
        item.setText(_translate("MainWindow", "Match"))
        item = self.vcs_matches.horizontalHeaderItem(1)
        item.setText(_translate("MainWindow", "Date"))
        item = self.vcs_matches.horizontalHeaderItem(2)
        item.setText(_translate("MainWindow", "Line"))
        item = self.vcs_matches.horizontalHeaderItem(3)
        item.setText(_translate("MainWindow", "G1 Avg"))
        item = self.vcs_matches.horizontalHeaderItem(4)
        item.setText(_translate("MainWindow", "G1 %"))
        item = self.vcs_matches.horizontalHeaderItem(5)
        item.setText(_translate("MainWindow", "G1 Value"))
        item = self.vcs_matches.horizontalHeaderItem(6)
        item.setText(_translate("MainWindow", "G2 Avg"))
        item = self.vcs_matches.horizontalHeaderItem(7)
        item.setText(_translate("MainWindow", "G2 %"))
        item = self.vcs_matches.horizontalHeaderItem(8)
        item.setText(_translate("MainWindow", "G2 Value"))
        self.update_na_acad.setText(_translate("MainWindow", "Update"))
        item = self.ladder_na_acad.horizontalHeaderItem(0)
        item.setText(_translate("MainWindow", "Team"))
        item = self.ladder_na_acad.horizontalHeaderItem(1)
        item.setText(_translate("MainWindow", "Record"))
        self.label_62.setText(_translate("MainWindow", "<html><head/><body><p><span style=\" font-size:12pt;\">Odds:</span></p></body></html>"))
        item = self.na_acad_stats_table.horizontalHeaderItem(0)
        item.setText(_translate("MainWindow", "Team Name"))
        item = self.na_acad_stats_table.horizontalHeaderItem(1)
        item.setText(_translate("MainWindow", "Matches"))
        item = self.na_acad_stats_table.horizontalHeaderItem(2)
        item.setText(_translate("MainWindow", "Avg Game 1"))
        item = self.na_acad_stats_table.horizontalHeaderItem(3)
        item.setText(_translate("MainWindow", "% Game 1"))
        item = self.na_acad_stats_table.horizontalHeaderItem(4)
        item.setText(_translate("MainWindow", "Avg Game 2"))
        item = self.na_acad_stats_table.horizontalHeaderItem(5)
        item.setText(_translate("MainWindow", "% Game 2"))
        item = self.na_acad_stats_table.horizontalHeaderItem(6)
        item.setText(_translate("MainWindow", "Avg Total"))
        item = self.na_acad_stats_table.horizontalHeaderItem(7)
        item.setText(_translate("MainWindow", "% Total"))
        self.back_15.setText(_translate("MainWindow", "Back"))
        self.label_61.setText(_translate("MainWindow", "<html><head/><body><p><span style=\" font-size:16pt;\">NA Acad Statistics</span></p></body></html>"))
        item = self.na_acad_matches.horizontalHeaderItem(0)
        item.setText(_translate("MainWindow", "Match"))
        item = self.na_acad_matches.horizontalHeaderItem(1)
        item.setText(_translate("MainWindow", "Date"))
        item = self.na_acad_matches.horizontalHeaderItem(2)
        item.setText(_translate("MainWindow", "Line"))
        item = self.na_acad_matches.horizontalHeaderItem(3)
        item.setText(_translate("MainWindow", "G1 Avg"))
        item = self.na_acad_matches.horizontalHeaderItem(4)
        item.setText(_translate("MainWindow", "G1 %"))
        item = self.na_acad_matches.horizontalHeaderItem(5)
        item.setText(_translate("MainWindow", "G1 Value"))
        item = self.na_acad_matches.horizontalHeaderItem(6)
        item.setText(_translate("MainWindow", "G2 Avg"))
        item = self.na_acad_matches.horizontalHeaderItem(7)
        item.setText(_translate("MainWindow", "G2 %"))
        item = self.na_acad_matches.horizontalHeaderItem(8)
        item.setText(_translate("MainWindow", "G2 Value"))
        self.label_63.setText(_translate("MainWindow", "<html><head/><body><p><span style=\" font-size:16pt;\">Ladder</span></p></body></html>"))
        self.label_64.setText(_translate("MainWindow", "<html><head/><body><p><span style=\" font-size:12pt;\">Line:</span></p></body></html>"))
        self.na_acad_calculate.setText(_translate("MainWindow", "Calculate"))
        self.update_cblol.setText(_translate("MainWindow", "Update"))
        self.label_65.setText(_translate("MainWindow", "<html><head/><body><p><span style=\" font-size:16pt;\">CBLOL Statistics</span></p></body></html>"))
        self.label_66.setText(_translate("MainWindow", "<html><head/><body><p><span style=\" font-size:12pt;\">Odds:</span></p></body></html>"))
        self.back_16.setText(_translate("MainWindow", "Back"))
        item = self.cblol_stats_table.horizontalHeaderItem(0)
        item.setText(_translate("MainWindow", "Team Name"))
        item = self.cblol_stats_table.horizontalHeaderItem(1)
        item.setText(_translate("MainWindow", "Matches"))
        item = self.cblol_stats_table.horizontalHeaderItem(2)
        item.setText(_translate("MainWindow", "Avg Game 1"))
        item = self.cblol_stats_table.horizontalHeaderItem(3)
        item.setText(_translate("MainWindow", "% Game 1"))
        item = self.cblol_stats_table.horizontalHeaderItem(4)
        item.setText(_translate("MainWindow", "Avg Game 2"))
        item = self.cblol_stats_table.horizontalHeaderItem(5)
        item.setText(_translate("MainWindow", "% Game 2"))
        item = self.cblol_stats_table.horizontalHeaderItem(6)
        item.setText(_translate("MainWindow", "Avg Total"))
        item = self.cblol_stats_table.horizontalHeaderItem(7)
        item.setText(_translate("MainWindow", "% Total"))
        self.cblol_calculate.setText(_translate("MainWindow", "Calculate"))
        item = self.ladder_cblol.horizontalHeaderItem(0)
        item.setText(_translate("MainWindow", "Team"))
        item = self.ladder_cblol.horizontalHeaderItem(1)
        item.setText(_translate("MainWindow", "Record"))
        self.label_67.setText(_translate("MainWindow", "<html><head/><body><p><span style=\" font-size:16pt;\">Ladder</span></p></body></html>"))
        self.label_68.setText(_translate("MainWindow", "<html><head/><body><p><span style=\" font-size:12pt;\">Line:</span></p></body></html>"))
        item = self.cblol_matches.horizontalHeaderItem(0)
        item.setText(_translate("MainWindow", "Match"))
        item = self.cblol_matches.horizontalHeaderItem(1)
        item.setText(_translate("MainWindow", "Date"))
        item = self.cblol_matches.horizontalHeaderItem(2)
        item.setText(_translate("MainWindow", "Line"))
        item = self.cblol_matches.horizontalHeaderItem(3)
        item.setText(_translate("MainWindow", "G1 Avg"))
        item = self.cblol_matches.horizontalHeaderItem(4)
        item.setText(_translate("MainWindow", "G1 %"))
        item = self.cblol_matches.horizontalHeaderItem(5)
        item.setText(_translate("MainWindow", "G1 Value"))
        item = self.cblol_matches.horizontalHeaderItem(6)
        item.setText(_translate("MainWindow", "G2 Avg"))
        item = self.cblol_matches.horizontalHeaderItem(7)
        item.setText(_translate("MainWindow", "G2 %"))
        item = self.cblol_matches.horizontalHeaderItem(8)
        item.setText(_translate("MainWindow", "G2 Value"))
        self.update_lla.setText(_translate("MainWindow", "Update"))
        item = self.ladder_lla.horizontalHeaderItem(0)
        item.setText(_translate("MainWindow", "Team"))
        item = self.ladder_lla.horizontalHeaderItem(1)
        item.setText(_translate("MainWindow", "Record"))
        self.back_17.setText(_translate("MainWindow", "Back"))
        self.label_69.setText(_translate("MainWindow", "<html><head/><body><p><span style=\" font-size:16pt;\">LLA Statistics</span></p></body></html>"))
        item = self.lla_stats_table.horizontalHeaderItem(0)
        item.setText(_translate("MainWindow", "Team Name"))
        item = self.lla_stats_table.horizontalHeaderItem(1)
        item.setText(_translate("MainWindow", "Matches"))
        item = self.lla_stats_table.horizontalHeaderItem(2)
        item.setText(_translate("MainWindow", "Avg Game 1"))
        item = self.lla_stats_table.horizontalHeaderItem(3)
        item.setText(_translate("MainWindow", "% Game 1"))
        item = self.lla_stats_table.horizontalHeaderItem(4)
        item.setText(_translate("MainWindow", "Avg Game 2"))
        item = self.lla_stats_table.horizontalHeaderItem(5)
        item.setText(_translate("MainWindow", "% Game 2"))
        item = self.lla_stats_table.horizontalHeaderItem(6)
        item.setText(_translate("MainWindow", "Avg Total"))
        item = self.lla_stats_table.horizontalHeaderItem(7)
        item.setText(_translate("MainWindow", "% Total"))
        self.label_70.setText(_translate("MainWindow", "<html><head/><body><p><span style=\" font-size:12pt;\">Odds:</span></p></body></html>"))
        self.label_71.setText(_translate("MainWindow", "<html><head/><body><p><span style=\" font-size:16pt;\">Ladder</span></p></body></html>"))
        self.label_72.setText(_translate("MainWindow", "<html><head/><body><p><span style=\" font-size:12pt;\">Line:</span></p></body></html>"))
        item = self.lla_matches.horizontalHeaderItem(0)
        item.setText(_translate("MainWindow", "Match"))
        item = self.lla_matches.horizontalHeaderItem(1)
        item.setText(_translate("MainWindow", "Date"))
        item = self.lla_matches.horizontalHeaderItem(2)
        item.setText(_translate("MainWindow", "Line"))
        item = self.lla_matches.horizontalHeaderItem(3)
        item.setText(_translate("MainWindow", "G1 Avg"))
        item = self.lla_matches.horizontalHeaderItem(4)
        item.setText(_translate("MainWindow", "G1 %"))
        item = self.lla_matches.horizontalHeaderItem(5)
        item.setText(_translate("MainWindow", "G1 Value"))
        item = self.lla_matches.horizontalHeaderItem(6)
        item.setText(_translate("MainWindow", "G2 Avg"))
        item = self.lla_matches.horizontalHeaderItem(7)
        item.setText(_translate("MainWindow", "G2 %"))
        item = self.lla_matches.horizontalHeaderItem(8)
        item.setText(_translate("MainWindow", "G2 Value"))
        self.lla_calculate.setText(_translate("MainWindow", "Calculate"))
        self.update_lco.setText(_translate("MainWindow", "Update"))
        self.label_76.setText(_translate("MainWindow", "<html><head/><body><p><span style=\" font-size:12pt;\">Line:</span></p></body></html>"))
        item = self.ladder_lco.horizontalHeaderItem(0)
        item.setText(_translate("MainWindow", "Team"))
        item = self.ladder_lco.horizontalHeaderItem(1)
        item.setText(_translate("MainWindow", "Record"))
        self.back_18.setText(_translate("MainWindow", "Back"))
        self.label_73.setText(_translate("MainWindow", "<html><head/><body><p><span style=\" font-size:16pt;\">LCO Statistics</span></p></body></html>"))
        self.lco_calculate.setText(_translate("MainWindow", "Calculate"))
        item = self.lco_stats_table.horizontalHeaderItem(0)
        item.setText(_translate("MainWindow", "Team Name"))
        item = self.lco_stats_table.horizontalHeaderItem(1)
        item.setText(_translate("MainWindow", "Matches"))
        item = self.lco_stats_table.horizontalHeaderItem(2)
        item.setText(_translate("MainWindow", "Avg Game 1"))
        item = self.lco_stats_table.horizontalHeaderItem(3)
        item.setText(_translate("MainWindow", "% Game 1"))
        item = self.lco_stats_table.horizontalHeaderItem(4)
        item.setText(_translate("MainWindow", "Avg Game 2"))
        item = self.lco_stats_table.horizontalHeaderItem(5)
        item.setText(_translate("MainWindow", "% Game 2"))
        item = self.lco_stats_table.horizontalHeaderItem(6)
        item.setText(_translate("MainWindow", "Avg Total"))
        item = self.lco_stats_table.horizontalHeaderItem(7)
        item.setText(_translate("MainWindow", "% Total"))
        self.label_74.setText(_translate("MainWindow", "<html><head/><body><p><span style=\" font-size:12pt;\">Odds:</span></p></body></html>"))
        item = self.lco_matches.horizontalHeaderItem(0)
        item.setText(_translate("MainWindow", "Match"))
        item = self.lco_matches.horizontalHeaderItem(1)
        item.setText(_translate("MainWindow", "Date"))
        item = self.lco_matches.horizontalHeaderItem(2)
        item.setText(_translate("MainWindow", "Line"))
        item = self.lco_matches.horizontalHeaderItem(3)
        item.setText(_translate("MainWindow", "G1 Avg"))
        item = self.lco_matches.horizontalHeaderItem(4)
        item.setText(_translate("MainWindow", "G1 %"))
        item = self.lco_matches.horizontalHeaderItem(5)
        item.setText(_translate("MainWindow", "G1 Value"))
        item = self.lco_matches.horizontalHeaderItem(6)
        item.setText(_translate("MainWindow", "G2 Avg"))
        item = self.lco_matches.horizontalHeaderItem(7)
        item.setText(_translate("MainWindow", "G2 %"))
        item = self.lco_matches.horizontalHeaderItem(8)
        item.setText(_translate("MainWindow", "G2 Value"))
        self.label_75.setText(_translate("MainWindow", "<html><head/><body><p><span style=\" font-size:16pt;\">Ladder</span></p></body></html>"))
        self.lck_chal_calculate.setText(_translate("MainWindow", "Calculate"))
        self.back_28.setText(_translate("MainWindow", "Back"))
        item = self.lck_chal_stats_table.horizontalHeaderItem(0)
        item.setText(_translate("MainWindow", "Team Name"))
        item = self.lck_chal_stats_table.horizontalHeaderItem(1)
        item.setText(_translate("MainWindow", "Matches"))
        item = self.lck_chal_stats_table.horizontalHeaderItem(2)
        item.setText(_translate("MainWindow", "Avg Game 1"))
        item = self.lck_chal_stats_table.horizontalHeaderItem(3)
        item.setText(_translate("MainWindow", "% Game 1"))
        item = self.lck_chal_stats_table.horizontalHeaderItem(4)
        item.setText(_translate("MainWindow", "Avg Game 2"))
        item = self.lck_chal_stats_table.horizontalHeaderItem(5)
        item.setText(_translate("MainWindow", "% Game 2"))
        item = self.lck_chal_stats_table.horizontalHeaderItem(6)
        item.setText(_translate("MainWindow", "Avg Total"))
        item = self.lck_chal_stats_table.horizontalHeaderItem(7)
        item.setText(_translate("MainWindow", "% Total"))
        self.update_lck_chal.setText(_translate("MainWindow", "Update"))
        item = self.lck_chal_matches.horizontalHeaderItem(0)
        item.setText(_translate("MainWindow", "Match"))
        item = self.lck_chal_matches.horizontalHeaderItem(1)
        item.setText(_translate("MainWindow", "Date"))
        item = self.lck_chal_matches.horizontalHeaderItem(2)
        item.setText(_translate("MainWindow", "Line"))
        item = self.lck_chal_matches.horizontalHeaderItem(3)
        item.setText(_translate("MainWindow", "G1 Avg"))
        item = self.lck_chal_matches.horizontalHeaderItem(4)
        item.setText(_translate("MainWindow", "G1 %"))
        item = self.lck_chal_matches.horizontalHeaderItem(5)
        item.setText(_translate("MainWindow", "G1 Value"))
        item = self.lck_chal_matches.horizontalHeaderItem(6)
        item.setText(_translate("MainWindow", "G2 Avg"))
        item = self.lck_chal_matches.horizontalHeaderItem(7)
        item.setText(_translate("MainWindow", "G2 %"))
        item = self.lck_chal_matches.horizontalHeaderItem(8)
        item.setText(_translate("MainWindow", "G2 Value"))
        self.label_114.setText(_translate("MainWindow", "<html><head/><body><p><span style=\" font-size:12pt;\">Odds:</span></p></body></html>"))
        self.label_115.setText(_translate("MainWindow", "<html><head/><body><p><span style=\" font-size:12pt;\">Line:</span></p></body></html>"))
        self.label_116.setText(_translate("MainWindow", "<html><head/><body><p><span style=\" font-size:16pt;\">Ladder</span></p></body></html>"))
        self.label_117.setText(_translate("MainWindow", "<html><head/><body><p><span style=\" font-size:16pt;\">LCK Chal Statistics</span></p></body></html>"))
        item = self.ladder_lck_chal.horizontalHeaderItem(0)
        item.setText(_translate("MainWindow", "Team"))
        item = self.ladder_lck_chal.horizontalHeaderItem(1)
        item.setText(_translate("MainWindow", "Record"))
        item = self.ladder_tcl.horizontalHeaderItem(0)
        item.setText(_translate("MainWindow", "Team"))
        item = self.ladder_tcl.horizontalHeaderItem(1)
        item.setText(_translate("MainWindow", "Record"))
        item = self.tcl_matches.horizontalHeaderItem(0)
        item.setText(_translate("MainWindow", "Match"))
        item = self.tcl_matches.horizontalHeaderItem(1)
        item.setText(_translate("MainWindow", "Date"))
        item = self.tcl_matches.horizontalHeaderItem(2)
        item.setText(_translate("MainWindow", "Line"))
        item = self.tcl_matches.horizontalHeaderItem(3)
        item.setText(_translate("MainWindow", "G1 Avg"))
        item = self.tcl_matches.horizontalHeaderItem(4)
        item.setText(_translate("MainWindow", "G1 %"))
        item = self.tcl_matches.horizontalHeaderItem(5)
        item.setText(_translate("MainWindow", "G1 Value"))
        item = self.tcl_matches.horizontalHeaderItem(6)
        item.setText(_translate("MainWindow", "G2 Avg"))
        item = self.tcl_matches.horizontalHeaderItem(7)
        item.setText(_translate("MainWindow", "G2 %"))
        item = self.tcl_matches.horizontalHeaderItem(8)
        item.setText(_translate("MainWindow", "G2 Value"))
        self.label_120.setText(_translate("MainWindow", "<html><head/><body><p><span style=\" font-size:16pt;\">Ladder</span></p></body></html>"))
        self.back_29.setText(_translate("MainWindow", "Back"))
        self.update_tcl.setText(_translate("MainWindow", "Update"))
        item = self.tcl_stats_table.horizontalHeaderItem(0)
        item.setText(_translate("MainWindow", "Team Name"))
        item = self.tcl_stats_table.horizontalHeaderItem(1)
        item.setText(_translate("MainWindow", "Matches"))
        item = self.tcl_stats_table.horizontalHeaderItem(2)
        item.setText(_translate("MainWindow", "Avg Game 1"))
        item = self.tcl_stats_table.horizontalHeaderItem(3)
        item.setText(_translate("MainWindow", "% Game 1"))
        item = self.tcl_stats_table.horizontalHeaderItem(4)
        item.setText(_translate("MainWindow", "Avg Game 2"))
        item = self.tcl_stats_table.horizontalHeaderItem(5)
        item.setText(_translate("MainWindow", "% Game 2"))
        item = self.tcl_stats_table.horizontalHeaderItem(6)
        item.setText(_translate("MainWindow", "Avg Total"))
        item = self.tcl_stats_table.horizontalHeaderItem(7)
        item.setText(_translate("MainWindow", "% Total"))
        self.label_119.setText(_translate("MainWindow", "<html><head/><body><p><span style=\" font-size:12pt;\">Line:</span></p></body></html>"))
        self.label_118.setText(_translate("MainWindow", "<html><head/><body><p><span style=\" font-size:12pt;\">Odds:</span></p></body></html>"))
        self.tcl_calculate.setText(_translate("MainWindow", "Calculate"))
        self.label_121.setText(_translate("MainWindow", "<html><head/><body><p><span style=\" font-size:16pt;\">TCL Statistics</span></p></body></html>"))
        item = self.ladder_pcs.horizontalHeaderItem(0)
        item.setText(_translate("MainWindow", "Team"))
        item = self.ladder_pcs.horizontalHeaderItem(1)
        item.setText(_translate("MainWindow", "Record"))
        self.update_pcs.setText(_translate("MainWindow", "Update"))
        self.back_30.setText(_translate("MainWindow", "Back"))
        item = self.pcs_matches.horizontalHeaderItem(0)
        item.setText(_translate("MainWindow", "Match"))
        item = self.pcs_matches.horizontalHeaderItem(1)
        item.setText(_translate("MainWindow", "Date"))
        item = self.pcs_matches.horizontalHeaderItem(2)
        item.setText(_translate("MainWindow", "Line"))
        item = self.pcs_matches.horizontalHeaderItem(3)
        item.setText(_translate("MainWindow", "G1 Avg"))
        item = self.pcs_matches.horizontalHeaderItem(4)
        item.setText(_translate("MainWindow", "G1 %"))
        item = self.pcs_matches.horizontalHeaderItem(5)
        item.setText(_translate("MainWindow", "G1 Value"))
        item = self.pcs_matches.horizontalHeaderItem(6)
        item.setText(_translate("MainWindow", "G2 Avg"))
        item = self.pcs_matches.horizontalHeaderItem(7)
        item.setText(_translate("MainWindow", "G2 %"))
        item = self.pcs_matches.horizontalHeaderItem(8)
        item.setText(_translate("MainWindow", "G2 Value"))
        self.pcs_calculate.setText(_translate("MainWindow", "Calculate"))
        self.label_124.setText(_translate("MainWindow", "<html><head/><body><p><span style=\" font-size:16pt;\">Ladder</span></p></body></html>"))
        self.label_125.setText(_translate("MainWindow", "<html><head/><body><p><span style=\" font-size:16pt;\">PCS Statistics</span></p></body></html>"))
        self.label_122.setText(_translate("MainWindow", "<html><head/><body><p><span style=\" font-size:12pt;\">Odds:</span></p></body></html>"))
        self.label_123.setText(_translate("MainWindow", "<html><head/><body><p><span style=\" font-size:12pt;\">Line:</span></p></body></html>"))
        item = self.pcs_stats_table.horizontalHeaderItem(0)
        item.setText(_translate("MainWindow", "Team Name"))
        item = self.pcs_stats_table.horizontalHeaderItem(1)
        item.setText(_translate("MainWindow", "Matches"))
        item = self.pcs_stats_table.horizontalHeaderItem(2)
        item.setText(_translate("MainWindow", "Avg Game 1"))
        item = self.pcs_stats_table.horizontalHeaderItem(3)
        item.setText(_translate("MainWindow", "% Game 1"))
        item = self.pcs_stats_table.horizontalHeaderItem(4)
        item.setText(_translate("MainWindow", "Avg Game 2"))
        item = self.pcs_stats_table.horizontalHeaderItem(5)
        item.setText(_translate("MainWindow", "% Game 2"))
        item = self.pcs_stats_table.horizontalHeaderItem(6)
        item.setText(_translate("MainWindow", "Avg Total"))
        item = self.pcs_stats_table.horizontalHeaderItem(7)
        item.setText(_translate("MainWindow", "% Total"))
        self.ebl_calculate.setText(_translate("MainWindow", "Calculate"))
        self.back_31.setText(_translate("MainWindow", "Back"))
        item = self.ebl_stats_table.horizontalHeaderItem(0)
        item.setText(_translate("MainWindow", "Team Name"))
        item = self.ebl_stats_table.horizontalHeaderItem(1)
        item.setText(_translate("MainWindow", "Matches"))
        item = self.ebl_stats_table.horizontalHeaderItem(2)
        item.setText(_translate("MainWindow", "Avg Game 1"))
        item = self.ebl_stats_table.horizontalHeaderItem(3)
        item.setText(_translate("MainWindow", "% Game 1"))
        item = self.ebl_stats_table.horizontalHeaderItem(4)
        item.setText(_translate("MainWindow", "Avg Game 2"))
        item = self.ebl_stats_table.horizontalHeaderItem(5)
        item.setText(_translate("MainWindow", "% Game 2"))
        item = self.ebl_stats_table.horizontalHeaderItem(6)
        item.setText(_translate("MainWindow", "Avg Total"))
        item = self.ebl_stats_table.horizontalHeaderItem(7)
        item.setText(_translate("MainWindow", "% Total"))
        self.update_ebl.setText(_translate("MainWindow", "Update"))
        item = self.ebl_matches.horizontalHeaderItem(0)
        item.setText(_translate("MainWindow", "Match"))
        item = self.ebl_matches.horizontalHeaderItem(1)
        item.setText(_translate("MainWindow", "Date"))
        item = self.ebl_matches.horizontalHeaderItem(2)
        item.setText(_translate("MainWindow", "Line"))
        item = self.ebl_matches.horizontalHeaderItem(3)
        item.setText(_translate("MainWindow", "G1 Avg"))
        item = self.ebl_matches.horizontalHeaderItem(4)
        item.setText(_translate("MainWindow", "G1 %"))
        item = self.ebl_matches.horizontalHeaderItem(5)
        item.setText(_translate("MainWindow", "G1 Value"))
        item = self.ebl_matches.horizontalHeaderItem(6)
        item.setText(_translate("MainWindow", "G2 Avg"))
        item = self.ebl_matches.horizontalHeaderItem(7)
        item.setText(_translate("MainWindow", "G2 %"))
        item = self.ebl_matches.horizontalHeaderItem(8)
        item.setText(_translate("MainWindow", "G2 Value"))
        self.label_126.setText(_translate("MainWindow", "<html><head/><body><p><span style=\" font-size:12pt;\">Odds:</span></p></body></html>"))
        self.label_127.setText(_translate("MainWindow", "<html><head/><body><p><span style=\" font-size:12pt;\">Line:</span></p></body></html>"))
        self.label_128.setText(_translate("MainWindow", "<html><head/><body><p><span style=\" font-size:16pt;\">Ladder</span></p></body></html>"))
        self.label_129.setText(_translate("MainWindow", "<html><head/><body><p><span style=\" font-size:16pt;\">EBL Statistics</span></p></body></html>"))
        item = self.ladder_ebl.horizontalHeaderItem(0)
        item.setText(_translate("MainWindow", "Team"))
        item = self.ladder_ebl.horizontalHeaderItem(1)
        item.setText(_translate("MainWindow", "Record"))

    # navigate to page and fill it with data relating to the given leagues statistics, rankings and upcomning games
    def load_data(self, competition, table, page, url, ladder_table, url_2, upcoming_table, line = 23.5, odds = 1.83):   
        # change pages if it has not been done
        try:
            self.stackedWidget.setCurrentWidget(page)
        except:
            pass

        # load the ladder upon opening/updating the page
        league = Ladder()
        ladder_details = league.ladder(url, competition)       
        self.insert_data(ladder_details[0], 0, ladder_table)
        self.insert_data(ladder_details[1], 1, ladder_table)

        # load data for the upcoming games for the specified league
        upcoming = UpcomingGames()
        games = upcoming.games(url_2)
       
        # open and retrieve all statistics data from excel sheet for relevant league
        wb = load_workbook('C:\\Users\\Legen\\Documents\\League Program\\data\\' + competition + '_data.xlsx')
        ws = wb['Sheet1']

        name = ws['A']
        matches = ws['B']
        a_g1 = ws['C']
        a_g2 = ws['E']
        a_all = ws['G']
        p_g1 = ws['D']
        p_g2 = ws['F']
        p_all = ws['H']

        name_data = [name[x].value for x in range(len(name))]
        match_data = [str(matches[x].value) for x in range(len(name))]
        a_g1_data = [a_g1[x].value for x in range(len(name))]
        a_g2_data = [a_g2[x].value for x in range(len(name))]
        a_all_data = [a_all[x].value for x in range(len(name))]
        p_g1_data = [p_g1[x].value for x in range(len(name))]
        p_g2_data = [p_g2[x].value for x in range(len(name))]
        p_all_data = [p_all[x].value for x in range(len(name))]

        wb.close()

        # use function to insert data into the statistics table
        self.insert_data(name_data, 0, table)
        self.insert_data(match_data, 1, table)
        self.insert_data(a_g1_data, 2, table)
        self.insert_data(p_g1_data, 3, table)
        self.insert_data(a_g2_data, 4, table)
        self.insert_data(p_g2_data, 5, table)
        self.insert_data(a_all_data, 6, table)
        self.insert_data(p_all_data, 7, table)

        avg_g1 = []
        avg_g2 = []
        perc_g1 = []
        perc_g2 = []

        # remove extra data pulled from scrape if it exists
        games = [x for x in games if x != 'Draw']

        # sometimes data has match line attached, remove it before comparisons
        for i in range(len(games)):
            if games[i][-5:] == ' +1.5' or games[i][-5:] == ' -1.5':
                length = len(games[i]) - 5
                games[i] = games[i][:length]
        
        # fix name discrepency in data pulled from different locations
        try:
            if competition == 'lck':
                name_data[8] = 'Nongshim RedForce'
            elif competition == 'lec':
                name_data[3] = 'G2'
                name_data[8] = 'BDS'
                name_data[5] = 'Misfits'
                name_data[9] = 'Vitality'
            elif competition == 'lcs':
                name_data[2] = 'Counter Logic'
                name_data[8] = 'Liquid'
                name_data[9] = 'Team SoloMid'
            elif competition == 'vcs':
                name_data[1] = 'Cerberus'
                name_data[2] = 'GAM'
                name_data[3] = 'Luxury'
                name_data[5] = 'SBTC'
                name_data[6] = 'Flash'
                name_data[7] = 'Secret'
            elif competition == 'lpl':
                name_data[1] = 'Bilibili'
                name_data[2] = 'Edward Gaming'
                name_data[4] = 'Invictus'
                name_data[6] = 'LGD'
                name_data[7] = 'LNG'
                name_data[10] = 'Royal Never Give Up'
                name_data[13] = 'ThunderTalk'
                name_data[16] = 'Weibo'
            elif competition == 'na_acad':
                name_data[0] = '100 Thieves Academy'
                name_data[2] = 'Cloud9 Academy'
                name_data[4] = 'Evil Geniuses Academy'
                name_data[6] = 'Golden Guardians Academy'
                name_data[7] = 'Immortals Academy'
                name_data[8] = 'Liquid Academy'
            elif competition == 'lco':
                name_data[0] = 'Chiefs'
                name_data[3] = 'Kanga'
            elif competition == 'ul':
                name_data[3] = "Gentlemen's"
                name_data[5] = 'Illuminar'
                name_data[7] = 'Komil and Friends'
                name_data[8] = 'ESCA'
            elif competition == 'pld':
                name_data[3] = 'FC Schalke 04'
                name_data[8] = 'Unicorns of Love SE'
                name_data[9] = 'WAVE'
            elif competition == 'pgn':
                name_data[1] = 'Atleta'
                name_data[5] = 'GG Esports'
                name_data[6] = 'Macko'
                name_data[7] = 'Samsung Morning Stars'
            elif competition == 'ljl':
                name_data[2] = 'Crest Act'
                name_data[3] = 'Detonation FocusMe'
                name_data[4] = 'Fukuoka Softbank Hawks'
                name_data[6] = 'Sengoku'
                name_data[7] = 'V3'
            elif competition == 'nlc':
                name_data[3] = 'JDXL'
                name_data[4] = 'MnM'
                name_data[6] = 'Riddle'
                name_data[7] = 'Singularity'
                name_data[9] = 'X7'
            elif competition == 'lfl':
                name_data[7] = 'BDS Academy'
                name_data[8] = 'Oplon'
            elif competition == 'lplol':
                name_data[1] = 'EGN'
                name_data[4] = 'Karma Clan'
                name_data[5] = 'Odivelas'
                name_data[6] = 'OFFSET'
            elif competition == 'hpm':
                name_data[0] = 'Brute'
            elif competition == 'cblol':
                name_data[0] = 'Flamengo'
                name_data[1] = 'FURIA'
                name_data[3] = 'KaBuM'
                name_data[7] = 'paiN'
                name_data[9] = 'Rensga'
            elif competition == 'lla':
                name_data[1] = 'Estral'
                name_data[3] = 'Infinity'
                name_data[6] = 'Aze'
                name_data[7] = 'XTEN'
            elif competition == 'lvp':
                name_data[0] = 'Barca'
                name_data[1] = 'BISONS'
                name_data[8] = 'Heretics'
            elif competition == 'lck_chal':
                name_data = ['DRX Chall', 'DWG KIA Chall', 'Brion Chall', 'Gen.G Chall', 'Hanwha Life Chall', 'KT Rolster Chall', 'Kwangdong Freecs Chall', 'Liiv SANDBOX Chall', 'Nongshim RedForce Chall', 'T1 Chall']
            elif competition == 'tcl':
                name_data[1] = 'Besiktas'
                name_data[3] = 'Fenerbahce'
                name_data[5] = 'Galatasaray'
                name_data[6] = 'Istanbul Wildcats'
                name_data[9] = 'AURORA'
            elif competition == 'pcs':
                name_data[0] = 'Beyond'
                name_data[2] = 'Deep Cross'
                name_data[3] = 'Frank'
                name_data[4] = 'Hurricane'
                name_data[7] = 'Meta Falcon'
            elif competition == 'ebl':
                name_data[0] = 'Auxesis'
                name_data[3] = 'Diamant'
        except:
            pass

        matches_2 = []
        combined_avg_g1 = []
        combined_avg_g2 = []
        combined_perc_g1 = []
        combined_perc_g2 = []
        g1_value = []
        g2_value = []
        line_list = []
        
        try:
            # get the stats for each team in the upcoming matches and store in a list
            for i in range(0, len(games), 1):
                no_matches = 0
                for j in range(0, len(name_data), 1):
                    if games[i] == name_data[j]:
                        avg_g1.append(float(a_g1_data[j]))
                        perc_g1.append(float(p_g1_data[j]))
                        avg_g2.append(float(a_g2_data[j]))
                        perc_g2.append(float(p_g2_data[j]))
                    else:
                        # if no matches are found there is a discrepency in the names, append a large number to avoid list index out of range 
                        # but also make it clear that the stats are not real and need to be reviewed manually
                        no_matches += 1
                        if no_matches == len(name_data):
                            avg_g1.append(999)
                            perc_g1.append(999)
                            avg_g2.append(999)
                            perc_g2.append(999)

            # turn the games into 'team 1 vs team 2' and combine data to reflect the match stats rather than individual teams
            for i in range(0, len(games) - 1, 2):
                team_1 = games[i]
                team_2 = games[i + 1]
                match = team_1 + ' vs ' + team_2
                matches_2.append(match) 
                combined_avg_g1.append(round(((avg_g1[i] + avg_g1[i + 1]) / 2), 2))
                combined_perc_g1.append(round(((perc_g1[i] + perc_g1[i + 1]) / 2), 2))
                try:    
                    combined_avg_g2.append(round(((avg_g2[i] + avg_g2[i + 1]) / 2), 2))
                    combined_perc_g2.append(round(((perc_g2[i] + perc_g2[i + 1]) / 2), 2))   
                except:
                    pass   
        except:
            print('No data in excel doc')

        # calculate the value of bets based on provided data
        g1_value = upcoming.calculate_value(combined_perc_g1, combined_avg_g1, line, odds)
        g2_value = upcoming.calculate_value(combined_perc_g2, combined_avg_g2, line, odds)

        # convert data in lists to strings for insertion into table
        a_g1 = [str(x) for x in combined_avg_g1]
        a_g2 = [str(x) for x in combined_avg_g2]
        p_g1 = [str(x) for x in combined_perc_g1]
        p_g2 = [str(x) for x in combined_perc_g2]
        v_g1 = [str(x) for x in g1_value]
        v_g2 = [str(x) for x in g2_value]
   
        # create list of given line to insert into table
        for i in range(len(matches_2)):
            line_list.append(str(line))

        # insert all data into the correct columns of the upcoming table
        self.insert_data(matches_2, 0, upcoming_table)
        self.insert_data(line_list, 2, upcoming_table)
        self.insert_data(a_g1, 3, upcoming_table)
        self.insert_data(p_g1, 4, upcoming_table)
        self.insert_data(v_g1, 5, upcoming_table)
        self.insert_data(a_g2, 6, upcoming_table)
        self.insert_data(p_g2, 7, upcoming_table)
        self.insert_data(v_g2, 8, upcoming_table)

    # insert given data into table
    def insert_data(self, data, col, table):    
        row = 0        
        table.setRowCount(len(data))       
        for i in data:
            table.setItem(row, col, QtWidgets.QTableWidgetItem(i))
            row += 1

class TableData:        
    # required initial information to establish a class
    def __init__(self, kill_average, teams, competition):
        # class variables
        self.kill_average = kill_average
        self.teams = teams
        self.competition = competition
        self.kills_g1 = []
        self.kills_g2 = []
        self.kills_all = []
        self.games_g1 = []
        self.games_g2 = []
        self.games_all = []
        self.percent_g1 = []
        self.percent_g2 = []
        self.percent_all = []
        
        # create lists of the correct length based on the amount of teams in the league
        for team in teams:
            self.kills_g1.append(0)
            self.kills_g2.append(0)
            self.kills_all.append(0)
            self.games_g1.append(0)
            self.games_g2.append(0)
            self.games_all.append(0)
            self.percent_g1.append(0)
            self.percent_g2.append(0)
            self.percent_all.append(0)

    # get team, kill and game data from the given url and store it in the appropriate list
    def kill_data(self, url):
        # get information from website
        result = requests.get(url)
        soup = BeautifulSoup(result.text,'html.parser')

        # declare variables
        max_games = 40
        team_list = []
        game_kills = []
        blue_side = []
        red_side = []
        game_total = []

        try:
            # find all games on the page to get data from
            for i in range(max_games):
                game = soup.find_all('table', class_ = 'sb')[i]

                # check number of games played on blue side
                for team in game.find_all('th', class_ = 'side-blue'):
                    game_number = team.text.strip()
                    try:
                        int(game_number)
                    except:
                        game_number = None
                    
                    if game_number != None:
                        blue_side.append(game_number)

                # check number of games played on red side
                for team in game.find_all('th', class_ = 'side-red'):
                    game_number = team.text.strip()
                    try:
                        int(game_number)
                    except:
                        game_number = None
                    
                    if game_number != None:
                        red_side.append(game_number)
                
                # get the team names associated with each game
                for team in game.find_all('span', class_ = 'teamname'):
                    team_list.append(team.text.strip())

                # get the amount of kills for game one
                for team in game.find_all('div', class_ = 'sb-header-Kills'):
                    game_kills.append(team.text.strip())
        except:
            pass

        # convert list elements from strings to ints to operate on
        blue_side = list(map(int, blue_side))
        red_side = list(map(int, red_side))
        game_kills = list(map(int, game_kills))
        # combine red and blue to determine the game number 
        game_total = [a + b for a, b in zip(red_side, blue_side)]

        # double the size of the game number list to match game kills and team names size
        for i in range(len(game_total)):
            game_total.insert((i * 2 + 1), game_total[i * 2])

        # insert kill data for each team into overall record for game 1 and 2 as well as overall
        for i in range(len(team_list)):
            for j in range(len(self.teams)):
                try:
                    if self.teams[j] == team_list[i]:              

                        # convert the team totals to match totals based on the teams position in the list, even number is team 1, odd is team 2
                        total_kills = 0
                        if i % 2 == 0:
                            total_kills = game_kills[i] + game_kills[i + 1]
                        else:
                            total_kills = game_kills[i] + game_kills[i - 1]
                        
                        self.games_all[j] += 1
                        self.kills_all[j] += total_kills

                        # use a 1 to indicate kills went over the total and 0 to indicate under
                        if total_kills > self.kill_average:
                            self.percent_all[j] += 1

                        if game_total[i] == 1:
                            self.games_g1[j] += 1
                            self.kills_g1[j] += total_kills
                            # use a 1 to indicate kills went over the total and 0 to indicate under
                            if total_kills > self.kill_average:
                                self.percent_g1[j] += 1
                        elif game_total[i] == 2:
                            self.games_g2[j] += 1
                            self.kills_g2[j] += total_kills
                            # use a 1 to indicate kills went over the total and 0 to indicate under
                            if total_kills > self.kill_average:
                                self.percent_g2[j] += 1
                except:
                    print('A game was forfeited')

    # calculate the average kills over the season
    def calculate_average(self, kills, games):
        average = []
        for i in range(len(games)):
            if games[i] != 0:
                average.append(str(round(kills[i]/games[i], 2)))
            else:
                average.append(0)
        return average

    # calculate the percentage of games that have gone over the predetermined line
    def calculate_percentage(self, percent, games):
        percentage = []
        for i in range(len(games)):
            if games[i] != 0:
                percentage.append(str(round((percent[i]/games[i]) * 100)))
            else:
                percentage.append(0)

        return percentage

    # perform all calculations on the retrieved data
    def calculate_all(self):
        average_g1 = self.calculate_average(self.kills_g1, self.games_g1)
        average_g2 = self.calculate_average(self.kills_g2, self.games_g2)
        average_all = self.calculate_average(self.kills_all, self.games_all)
        percentage_g1 = self.calculate_percentage(self.percent_g1, self.games_g1)
        percentage_g2 = self.calculate_percentage(self.percent_g2, self.games_g2)
        percentage_all = self.calculate_percentage(self.percent_all, self.games_all)

        self.save_data(average_g1, average_g2, average_all, percentage_g1, percentage_g2, percentage_all)

    # save the data to an excel doc to be later retrieved
    def save_data(self, average_g1, average_g2, average_all, percentage_g1, percentage_g2, percentage_all):
        # open correct excel doc
        wb = load_workbook('C:\\Users\\Legen\\Documents\\League Program\\data\\' + self.competition + '_data.xlsx')
        ws = wb['Sheet1']

        # save all required data to the excel doc
        for i in range(len(self.teams)):
            ws.cell(column = 1, row = i + 1, value = self.teams[i])
            ws.cell(column = 2, row = i + 1, value = self.games_g1[i])
            ws.cell(column = 3, row = i + 1, value = average_g1[i])
            ws.cell(column = 4, row = i + 1, value = percentage_g1[i])
            ws.cell(column = 5, row = i + 1, value = average_g2[i])
            ws.cell(column = 6, row = i + 1, value = percentage_g2[i])
            ws.cell(column = 7, row = i + 1, value = average_all[i])
            ws.cell(column = 8, row = i + 1, value = percentage_all[i])
                     
        # save and close to excel doc
        wb.save('C:\\Users\\Legen\\Documents\\League Program\\data\\' + self.competition + '_data.xlsx')
        wb.close()
        
# creates a ladder based on the current standings of the given league
class Ladder:
    
    # get all information associated with the ladder for the specific league
    def ladder(self, url, competition):
        result = requests.get(url)
        soup = BeautifulSoup(result.text,'html.parser')
        table = soup.find('table', class_ = 'wikitable2 standings')
        team_names = []
        
        # no class/id available so have to specify specific td in the html to retrieve teams records  
        name = table.find_all('span', class_ = 'teamname')

        for team in name:
            team_names.append(team.text)
        
        number_teams = len(team_names)
        
        # scrape data from chosen table and get information on team names for each league
        # tables are of different sizes for different leagues so size needs to be specified
        if competition == 'lck':
            team_records = self.record_position(table, 5, 8, number_teams)       
        elif competition == 'lcs':
            team_records = self.record_position(table, 5, 5, number_teams)        
        elif competition == 'lec':
            team_records = self.record_position(table, 5, 5, number_teams)
        elif competition == 'vcs':
            team_records = self.record_position(table, 5, 8, number_teams)
        elif competition == 'lpl':
            team_records = self.record_position(table, 7, 8, number_teams)
        elif competition == 'na_acad':
            team_records = self.record_position(table, 5, 6, number_teams)
        elif competition == 'cblol':
            team_records = self.record_position(table, 5, 5, number_teams)
        elif competition == 'lla':
            team_records = self.record_position(table, 5, 5, number_teams)
        elif competition == 'lco':
            team_records = self.record_position(table, 6, 5, number_teams)
        elif competition == 'ul':
            team_records = self.record_position(table, 5, 5, number_teams)
        elif competition == 'pgn':
            team_records = self.record_position(table, 5, 5, number_teams)
        elif competition == 'pld':
            team_records = self.record_position(table, 5, 5, number_teams)
        elif competition == 'ljl':
            team_records = self.record_position(table, 5, 5, number_teams)
        elif competition == 'nlc':
            team_records = self.record_position(table, 5, 5, number_teams)
        elif competition == 'lfl':
            team_records = self.record_position(table, 5, 5, number_teams)
        elif competition == 'lplol':
            team_records = self.record_position(table, 5, 5, number_teams)
        elif competition == 'hpm':
            team_records = self.record_position(table, 5, 5, number_teams)
        elif competition == 'lvp':
            team_records = self.record_position(table, 5, 5, number_teams)
        elif competition == 'lck_chal':
            team_records = self.record_position(table, 5, 5, number_teams)
        elif competition == 'tcl':
            team_records = self.record_position(table, 5, 5, number_teams)
        elif competition == 'pcs':
            team_records = self.record_position(table, 6, 5, number_teams)
        elif competition == 'ebl':
            team_records = self.record_position(table, 5, 5, number_teams)

        return team_names, team_records

    # trim off extra data gathered from scraping the team records for the correct number of teams in the league
    def record_position(self, table, starting_value, increments, number_teams):
        team_records = []
        for i in range(number_teams):
            team_records.append(self.trim_data(str(table.find_all('td')[starting_value + (increments * i)])))
        return team_records
    
    # trim extra characters from the records so they can be displayed neatly
    def trim_data(self, data):
        size = len(data)
        trimmed_data = data[4:size - 5]
        return trimmed_data

class UpcomingGames:        
    # retrieve all upcoming games for the specified league
    def games(self, url):
        result = requests.get(url)
        soup = BeautifulSoup(result.text,'lxml')
        games = []

        all_games = soup.find_all('div', class_ = 'outcome-name')
        for team in all_games:
            name = team.text
            games.append(name)        

        return games

    # apply the model calculations and return a list that represents the value in each game
    def calculate_value(self, percentage, average, line, odds):
        game_value = []
        for i in range(0, len(percentage), 1):
            system_odds_g1 = 50 + abs(((percentage[i] - 50) + (average[i] - line) * 2) / 2)
            g1_odds = (1 / odds) * 100
            game_value.append(round(system_odds_g1 - g1_odds, 2))
        return game_value

    
    def next_games(self):
        # declare variables and scrape the upcoming games
        games = self.games('https://www.rivalry.com/au/esports/league-of-legends-betting')
        matches = []

        # change the list of teams into a list of matches
        for i in range(0, len(games) - 1, 2):
            team_1 = games[i]
            team_2 = games[i + 1]
            match = team_1 + ' vs ' + team_2
            matches.append(match)

        result = requests.get('https://www.rivalry.com/au/esports/league-of-legends-betting')
        soup = BeautifulSoup(result.text,'lxml')
        league = []

        # scrape data for the competition that each match takes place in
        all_leagues = soup.find_all('div', class_ = 'text-league-of-legends-shade dark:text-league-of-legends-tint text-[11px]')
        for comp in all_leagues:
            league.append(comp.text)

        return league, matches

        
# provides necessary data used in calculation and creation of table
# yes these have a lot of links but these are the only sources for the required data, one of the reasons an edge still exists
def lck_data():
    # create object of league class for the LCK
    lck_teams = ['DRX', 'DWG KIA', 'Fredit BRION', 'Gen.G', 'Hanwha Life', 'KT Rolster', 'Kwangdong Freecs', 'Liiv SANDBOX', 'NS RedForce', 'T1']
    competition = 'lck'
    lck = TableData(22.5, lck_teams, competition)
    
    # call function to scrape data and perform calculations
    lck.kill_data('https://lol.fandom.com/wiki/LCK/2022_Season/Spring_Season/Scoreboards')
    lck.kill_data('https://lol.fandom.com/wiki/LCK/2022_Season/Spring_Season/Scoreboards/Week_1_(2)')
    lck.kill_data('https://lol.fandom.com/wiki/LCK/2022_Season/Spring_Season/Scoreboards/Week_2')
    lck.kill_data('https://lol.fandom.com/wiki/LCK/2022_Season/Spring_Season/Scoreboards/Week_2_(2)')
    lck.kill_data('https://lol.fandom.com/wiki/LCK/2022_Season/Spring_Season/Scoreboards/Week_3')
    lck.kill_data('https://lol.fandom.com/wiki/LCK/2022_Season/Spring_Season/Scoreboards/Week_3_(2)')
    lck.kill_data('https://lol.fandom.com/wiki/LCK/2022_Season/Spring_Season/Scoreboards/Week_4')
    lck.kill_data('https://lol.fandom.com/wiki/LCK/2022_Season/Spring_Season/Scoreboards/Week_4_(2)')
    lck.kill_data('https://lol.fandom.com/wiki/LCK/2022_Season/Spring_Season/Scoreboards/Week_5')
    lck.kill_data('https://lol.fandom.com/wiki/LCK/2022_Season/Spring_Season/Scoreboards/Week_5_(2)')
    lck.kill_data('https://lol.fandom.com/wiki/LCK/2022_Season/Spring_Season/Scoreboards/Week_6')
    lck.kill_data('https://lol.fandom.com/wiki/LCK/2022_Season/Spring_Season/Scoreboards/Week_6_(2)')
    lck.kill_data('https://lol.fandom.com/wiki/LCK/2022_Season/Spring_Season/Scoreboards/Week_7')
    lck.kill_data('https://lol.fandom.com/wiki/LCK/2022_Season/Spring_Season/Scoreboards/Week_7_(2)')
    lck.kill_data('https://lol.fandom.com/wiki/LCK/2022_Season/Spring_Season/Scoreboards/Week_8')
    lck.kill_data('https://lol.fandom.com/wiki/LCK/2022_Season/Spring_Season/Scoreboards/Week_8_(2)')
    lck.kill_data('https://lol.fandom.com/wiki/LCK/2022_Season/Spring_Season/Scoreboards/Week_9')
    lck.kill_data('https://lol.fandom.com/wiki/LCK/2022_Season/Spring_Season/Scoreboards/Week_9_(2)')

    # update the displayed table with the new data
    lck.calculate_all()
    ui.get_lck()

def lcs_data():
    # create object of league class for the LCS
    lcs_teams = ['100 Thieves', 'Cloud9', 'CLG', 'Dignitas', 'Evil Geniuses', 'FlyQuest', 'Golden Guardians', 'Immortals', 'Team Liquid', 'TSM']
    competition = 'lcs'
    lcs = TableData(24.5, lcs_teams, competition)
    
    # call function to scrape data and perform calculations
    lcs.kill_data('https://lol.fandom.com/wiki/LCS/2022_Season/Spring_Season/Scoreboards')
    lcs.kill_data('https://lol.fandom.com/wiki/LCS/2022_Season/Spring_Season/Scoreboards/Week_2')
    lcs.kill_data('https://lol.fandom.com/wiki/LCS/2022_Season/Spring_Season/Scoreboards/Week_3')
    lcs.kill_data('https://lol.fandom.com/wiki/LCS/2022_Season/Spring_Season/Scoreboards/Week_4')
    lcs.kill_data('https://lol.fandom.com/wiki/LCS/2022_Season/Spring_Season/Scoreboards/Week_5')
    lcs.kill_data('https://lol.fandom.com/wiki/LCS/2022_Season/Spring_Season/Scoreboards/Week_6')
    lcs.kill_data('https://lol.fandom.com/wiki/LCS/2022_Season/Spring_Season/Scoreboards/Week_7')
    lcs.kill_data('https://lol.fandom.com/wiki/LCS/2022_Season/Spring_Season/Scoreboards/Week_8')
    
    # update the displayed table with the new data
    lcs.calculate_all()
    ui.get_lcs()

# provides necessary data used in calculation and creation of table
def lec_data():
    # create object of league class for the LCS
    lec_teams = ['Astralis', 'Excel', 'Fnatic', 'G2 Esports', 'MAD Lions', 'Misfits Gaming', 'Rogue', 'SK Gaming', 'Team BDS', 'Team Vitality']
    competition = 'lec'
    lec = TableData(25.5, lec_teams, competition)
    
    # call function to scrape data and perform calculations
    lec.kill_data('https://lol.fandom.com/wiki/LEC/2022_Season/Spring_Season/Scoreboards')
    lec.kill_data('https://lol.fandom.com/wiki/LEC/2022_Season/Spring_Season/Scoreboards/Week_2')
    lec.kill_data('https://lol.fandom.com/wiki/LEC/2022_Season/Spring_Season/Scoreboards/Week_3')
    lec.kill_data('https://lol.fandom.com/wiki/LEC/2022_Season/Spring_Season/Scoreboards/Week_4')
    lec.kill_data('https://lol.fandom.com/wiki/LEC/2022_Season/Spring_Season/Scoreboards/Week_5')
    lec.kill_data('https://lol.fandom.com/wiki/LEC/2022_Season/Spring_Season/Scoreboards/Week_6')
    lec.kill_data('https://lol.fandom.com/wiki/LEC/2022_Season/Spring_Season/Scoreboards/Week_7')
    lec.kill_data('https://lol.fandom.com/wiki/LEC/2022_Season/Spring_Season/Scoreboards/Week_8')
    
    # update the displayed table with the new data
    lec.calculate_all()
    ui.get_lec()

# provides necessary data used in calculation and creation of table
def vcs_data():
    # create object of league class for the LCS
    vcs_teams = ['Burst The Sky', 'CERBERUS', 'GAM Esports', 'Luxury Esports', 'Saigon Buffalo', 'SBTC Esports', 'Team Flash', 'Team Secret']
    competition = 'vcs'
    vcs = TableData(31.5, vcs_teams, competition)
    
    # call function to scrape data and perform calculations
    vcs.kill_data('https://lol.fandom.com/wiki/VCS/2022_Season/Spring_Season/Scoreboards')
    vcs.kill_data('https://lol.fandom.com/wiki/VCS/2022_Season/Spring_Season/Scoreboards/Week_2')
    vcs.kill_data('https://lol.fandom.com/wiki/VCS/2022_Season/Spring_Season/Scoreboards/Week_3')
    vcs.kill_data('https://lol.fandom.com/wiki/VCS/2022_Season/Spring_Season/Scoreboards/Week_4')
    vcs.kill_data('https://lol.fandom.com/wiki/VCS/2022_Season/Spring_Season/Scoreboards/Week_5')
    vcs.kill_data('https://lol.fandom.com/wiki/VCS/2022_Season/Spring_Season/Scoreboards/Week_6')
    vcs.kill_data('https://lol.fandom.com/wiki/VCS/2022_Season/Spring_Season/Scoreboards/Week_7')
    vcs.kill_data('https://lol.fandom.com/wiki/VCS/2022_Season/Spring_Season/Scoreboards/Week_8')
    vcs.kill_data('https://lol.fandom.com/wiki/VCS/2022_Season/Spring_Season/Scoreboards/Week_9')
    
    # update the displayed table with the new data
    vcs.calculate_all()
    ui.get_vcs()

def lpl_data():
    # create object of league class for the LCS
    lpl_teams = ["Anyone's Legend", 'Bilibili Gaming', 'EDward Gaming', 'FunPlus Phoenix', 'Invictus Gaming', 'JD Gaming', 'LGD Gaming', 'LNG Esports', 'Oh My God', 'Rare Atom', 'RNG', 'Team WE', 'Top Esports', 'TT Gaming', 'Ultra Prime', 'Victory Five', 'Weibo Gaming']
    competition = 'lpl'
    lpl = TableData(25.5, lpl_teams, competition)
    
    # call function to scrape data and perform calculations
    lpl.kill_data('https://lol.fandom.com/wiki/LPL/2022_Season/Spring_Season/Scoreboards')
    lpl.kill_data('https://lol.fandom.com/wiki/LPL/2022_Season/Spring_Season/Scoreboards/Week_1_(2)')
    lpl.kill_data('https://lol.fandom.com/wiki/LPL/2022_Season/Spring_Season/Scoreboards/Week_1_(3)')
    lpl.kill_data('https://lol.fandom.com/wiki/LPL/2022_Season/Spring_Season/Scoreboards/Week_2')
    lpl.kill_data('https://lol.fandom.com/wiki/LPL/2022_Season/Spring_Season/Scoreboards/Week_2_(2)')
    lpl.kill_data('https://lol.fandom.com/wiki/LPL/2022_Season/Spring_Season/Scoreboards/Week_2_(3)')
    lpl.kill_data('https://lol.fandom.com/wiki/LPL/2022_Season/Spring_Season/Scoreboards/Week_3')
    lpl.kill_data('https://lol.fandom.com/wiki/LPL/2022_Season/Spring_Season/Scoreboards/Week_3_(2)')
    lpl.kill_data('https://lol.fandom.com/wiki/LPL/2022_Season/Spring_Season/Scoreboards/Week_4')
    lpl.kill_data('https://lol.fandom.com/wiki/LPL/2022_Season/Spring_Season/Scoreboards/Week_4_(2)')
    lpl.kill_data('https://lol.fandom.com/wiki/LPL/2022_Season/Spring_Season/Scoreboards/Week_5')
    lpl.kill_data('https://lol.fandom.com/wiki/LPL/2022_Season/Spring_Season/Scoreboards/Week_5_(2)')
    lpl.kill_data('https://lol.fandom.com/wiki/LPL/2022_Season/Spring_Season/Scoreboards/Week_5_(3)')
    lpl.kill_data('https://lol.fandom.com/wiki/LPL/2022_Season/Spring_Season/Scoreboards/Week_6')
    lpl.kill_data('https://lol.fandom.com/wiki/LPL/2022_Season/Spring_Season/Scoreboards/Week_6_(2)')
    lpl.kill_data('https://lol.fandom.com/wiki/LPL/2022_Season/Spring_Season/Scoreboards/Week_6_(3)')
    lpl.kill_data('https://lol.fandom.com/wiki/LPL/2022_Season/Spring_Season/Scoreboards/Week_7')
    lpl.kill_data('https://lol.fandom.com/wiki/LPL/2022_Season/Spring_Season/Scoreboards/Week_7_(2)')
    lpl.kill_data('https://lol.fandom.com/wiki/LPL/2022_Season/Spring_Season/Scoreboards/Week_7_(3)')
    lpl.kill_data('https://lol.fandom.com/wiki/LPL/2022_Season/Spring_Season/Scoreboards/Week_8')
    lpl.kill_data('https://lol.fandom.com/wiki/LPL/2022_Season/Spring_Season/Scoreboards/Week_8_(2)')
    lpl.kill_data('https://lol.fandom.com/wiki/LPL/2022_Season/Spring_Season/Scoreboards/Week_8_(3)')
    lpl.kill_data('https://lol.fandom.com/wiki/LPL/2022_Season/Spring_Season/Scoreboards/Week_9')
    lpl.kill_data('https://lol.fandom.com/wiki/LPL/2022_Season/Spring_Season/Scoreboards/Week_9_(2)')
    lpl.kill_data('https://lol.fandom.com/wiki/LPL/2022_Season/Spring_Season/Scoreboards/Week_9_(3)')
    
    # update the displayed table with the new data
    lpl.calculate_all()
    ui.get_lpl()

def na_acad_data():
    # create object of league class for the LCS
    na_acad_teams = ['100 Academy', 'CLG Academy', 'C9 Academy', 'Dignitas Academy', 'EG Academy', 'FlyQuest Academy', 'GG Academy', 'IMT Academy', 'TL Academy', 'TSM Academy']
    competition = 'na_acad'
    na_acad = TableData(28.5, na_acad_teams, competition)
    
    # call function to scrape data and perform calculations
    na_acad.kill_data('https://lol.fandom.com/wiki/NA_Academy_League/2022_Season/Spring_Season/Scoreboards')
    na_acad.kill_data('https://lol.fandom.com/wiki/NA_Academy_League/2022_Season/Spring_Season/Scoreboards/Week_2')
    na_acad.kill_data('https://lol.fandom.com/wiki/NA_Academy_League/2022_Season/Spring_Season/Scoreboards/Week_3')
    na_acad.kill_data('https://lol.fandom.com/wiki/NA_Academy_League/2022_Season/Spring_Season/Scoreboards/Week_4')
    na_acad.kill_data('https://lol.fandom.com/wiki/NA_Academy_League/2022_Season/Spring_Season/Scoreboards/Week_4_(2)')
    na_acad.kill_data('https://lol.fandom.com/wiki/NA_Academy_League/2022_Season/Spring_Season/Scoreboards/Week_5')
    na_acad.kill_data('https://lol.fandom.com/wiki/NA_Academy_League/2022_Season/Spring_Season/Scoreboards/Week_6')
    na_acad.kill_data('https://lol.fandom.com/wiki/NA_Academy_League/2022_Season/Spring_Season/Scoreboards/Week_7')
    na_acad.kill_data('https://lol.fandom.com/wiki/NA_Academy_League/2022_Season/Spring_Season/Scoreboards/Week_8')
    na_acad.kill_data('https://lol.fandom.com/wiki/NA_Academy_League/2022_Season/Spring_Season/Scoreboards/Week_8_(2)')
    
    # update the displayed table with the new data
    na_acad.calculate_all()
    ui.get_na_acad()

def cblol_data():
    # create object of league class for the LCS
    cblol_teams = ['Flamengo Esports', 'FURIA Esports', 'INTZ', 'KaBuM! e-Sports', 'Liberty', 'LOUD', 'Netshoes Miners', 'paiN Gaming', 'RED Canids', 'Rensga Esports']
    competition = 'cblol'
    cblol = TableData(25.5, cblol_teams, competition)
    
    # call function to scrape data and perform calculations
    cblol.kill_data('https://lol.fandom.com/wiki/CBLOL/2022_Season/Split_1/Scoreboards')
    cblol.kill_data('https://lol.fandom.com/wiki/CBLOL/2022_Season/Split_1/Scoreboards/Week_2')
    cblol.kill_data('https://lol.fandom.com/wiki/CBLOL/2022_Season/Split_1/Scoreboards/Week_3')
    cblol.kill_data('https://lol.fandom.com/wiki/CBLOL/2022_Season/Split_1/Scoreboards/Week_4')
    cblol.kill_data('https://lol.fandom.com/wiki/CBLOL/2022_Season/Split_1/Scoreboards/Week_5')
    cblol.kill_data('https://lol.fandom.com/wiki/CBLOL/2022_Season/Split_1/Scoreboards/Week_6')
    cblol.kill_data('https://lol.fandom.com/wiki/CBLOL/2022_Season/Split_1/Scoreboards/Week_7')
    cblol.kill_data('https://lol.fandom.com/wiki/CBLOL/2022_Season/Split_1/Scoreboards/Week_8')
    cblol.kill_data('https://lol.fandom.com/wiki/CBLOL/2022_Season/Split_1/Scoreboards/Week_9')
   
    # update the displayed table with the new data
    cblol.calculate_all()
    ui.get_cblol()

def lla_data():
    # create object of league class for the LCS
    lla_teams = ['All Knights', 'Estral Esports', 'Globant Emerald', 'INFINITY', 'Isurus', 'Rainbow7', 'Team Aze', 'XTEN Esports']
    competition = 'lla'
    lla = TableData(23.5, lla_teams, competition)
    
    # call function to scrape data and perform calculations
    lla.kill_data('https://lol.fandom.com/wiki/LLA/2022_Season/Opening_Season/Scoreboards')
    lla.kill_data('https://lol.fandom.com/wiki/LLA/2022_Season/Opening_Season/Scoreboards/Week_2')
    lla.kill_data('https://lol.fandom.com/wiki/LLA/2022_Season/Opening_Season/Scoreboards/Week_3')
    lla.kill_data('https://lol.fandom.com/wiki/LLA/2022_Season/Opening_Season/Scoreboards/Week_4')
    lla.kill_data('https://lol.fandom.com/wiki/LLA/2022_Season/Opening_Season/Scoreboards/Week_5')
    lla.kill_data('https://lol.fandom.com/wiki/LLA/2022_Season/Opening_Season/Scoreboards/Week_6')
    lla.kill_data('https://lol.fandom.com/wiki/LLA/2022_Season/Opening_Season/Scoreboards/Week_7')

    # update the displayed table with the new data
    lla.calculate_all()
    ui.get_lla()

def lco_data():
    # create object of league class for the LCS
    lco_teams = ['Chiefs Esports Club', 'Dire Wolves', 'Gravitas', 'Kanga Esports', 'MAMMOTH', 'ORDER', 'PEACE', 'Pentanet.GG']
    competition = 'lco'
    lco = TableData(28.5, lco_teams, competition)
    
    # call function to scrape data and perform calculations
    lco.kill_data('https://lol.fandom.com/wiki/LCO/2022_Season/Split_1/Scoreboards')
    lco.kill_data('https://lol.fandom.com/wiki/LCO/2022_Season/Split_1/Scoreboards/Week_2')
    lco.kill_data('https://lol.fandom.com/wiki/LCO/2022_Season/Split_1/Scoreboards/Week_3')
    lco.kill_data('https://lol.fandom.com/wiki/LCO/2022_Season/Split_1/Scoreboards/Week_4')
    lco.kill_data('https://lol.fandom.com/wiki/LCO/2022_Season/Split_1/Scoreboards/Week_5')
    lco.kill_data('https://lol.fandom.com/wiki/LCO/2022_Season/Split_1/Scoreboards/Week_6')
    lco.kill_data('https://lol.fandom.com/wiki/LCO/2022_Season/Split_1/Scoreboards/Week_7')
    lco.kill_data('https://lol.fandom.com/wiki/LCO/2022_Season/Split_1/Scoreboards/Week_8')
    lco.kill_data('https://lol.fandom.com/wiki/LCO/2022_Season/Split_1/Scoreboards/Week_9')

    # update the displayed table with the new data
    lco.calculate_all()
    ui.get_lco()

def ul_data():
    # create object of league class for the LCS
    ul_teams = ['AGO ROGUE', 'devils.one', 'Forsaken', "Gentlemen's Gaming", 'Goskilla', 'Illuminar Gaming', 'Iron Wolves', 'Komil&Friends', 'Team ESCA Gaming', 'Zero Tenacity']
    competition = 'ul'
    ul = TableData(28.5, ul_teams, competition)
    
    # call function to scrape data and perform calculations
    ul.kill_data('https://lol.fandom.com/wiki/Ultraliga/Season_7/Scoreboards')
    ul.kill_data('https://lol.fandom.com/wiki/Ultraliga/Season_7/Scoreboards/Week_2')
    ul.kill_data('https://lol.fandom.com/wiki/Ultraliga/Season_7/Scoreboards/Week_3')
    ul.kill_data('https://lol.fandom.com/wiki/Ultraliga/Season_7/Scoreboards/Week_4')
    ul.kill_data('https://lol.fandom.com/wiki/Ultraliga/Season_7/Scoreboards/Week_5')
    ul.kill_data('https://lol.fandom.com/wiki/Ultraliga/Season_7/Scoreboards/Week_6')
    ul.kill_data('https://lol.fandom.com/wiki/Ultraliga/Season_7/Scoreboards/Week_7')
    ul.kill_data('https://lol.fandom.com/wiki/Ultraliga/Season_7/Scoreboards/Week_8')
    ul.kill_data('https://lol.fandom.com/wiki/Ultraliga/Season_7/Scoreboards/Week_9')

    # update the displayed table with the new data
    ul.calculate_all()
    ui.get_ul()

def pgn_data():
    # create object of league class for the LCS
    pgn_teams = ['aNc Outplayed', 'Atleta Esport', 'Axolotl', 'Cyberground', 'Esport Empire', 'GGEsports', 'Macko Esports', 'Morning Stars']
    competition = 'pgn'
    pgn = TableData(27.5, pgn_teams, competition)
    
    # call function to scrape data and perform calculations
    pgn.kill_data('https://lol.fandom.com/wiki/PG_Nationals/2022_Season/Spring_Season/Scoreboards')
    pgn.kill_data('https://lol.fandom.com/wiki/PG_Nationals/2022_Season/Spring_Season/Scoreboards/Week_2')
    pgn.kill_data('https://lol.fandom.com/wiki/PG_Nationals/2022_Season/Spring_Season/Scoreboards/Week_3')
    pgn.kill_data('https://lol.fandom.com/wiki/PG_Nationals/2022_Season/Spring_Season/Scoreboards/Week_4')
    pgn.kill_data('https://lol.fandom.com/wiki/PG_Nationals/2022_Season/Spring_Season/Scoreboards/Week_5')
    pgn.kill_data('https://lol.fandom.com/wiki/PG_Nationals/2022_Season/Spring_Season/Scoreboards/Week_6')
    pgn.kill_data('https://lol.fandom.com/wiki/PG_Nationals/2022_Season/Spring_Season/Scoreboards/Week_7')

    # update the displayed table with the new data
    pgn.calculate_all()
    ui.get_pgn()

def pld_data():
    # create object of league class for the LCS
    pld_teams = ['BIG', 'E WIE EINFACH', 'Eintracht Spandau', 'Schalke 04', 'GamerLegion', 'MOUZ', 'PENTA 1860', 'SK Gaming Prime', 'UOL Sexy Edition', 'WAVE Esports']
    competition = 'pld'
    pld = TableData(28.5, pld_teams, competition)
    
    # call function to scrape data and perform calculations
    pld.kill_data('https://lol.fandom.com/wiki/Prime_League_1st_Division/2022_Season/Spring_Season/Scoreboards')
    pld.kill_data('https://lol.fandom.com/wiki/Prime_League_1st_Division/2022_Season/Spring_Season/Scoreboards/Week_2')
    pld.kill_data('https://lol.fandom.com/wiki/Prime_League_1st_Division/2022_Season/Spring_Season/Scoreboards/Week_3')
    pld.kill_data('https://lol.fandom.com/wiki/Prime_League_1st_Division/2022_Season/Spring_Season/Scoreboards/Week_4')
    pld.kill_data('https://lol.fandom.com/wiki/Prime_League_1st_Division/2022_Season/Spring_Season/Scoreboards/Week_5')
    pld.kill_data('https://lol.fandom.com/wiki/Prime_League_1st_Division/2022_Season/Spring_Season/Scoreboards/Week_6')
    pld.kill_data('https://lol.fandom.com/wiki/Prime_League_1st_Division/2022_Season/Spring_Season/Scoreboards/Week_7')

    # update the displayed table with the new data
    pld.calculate_all()
    ui.get_pld()

def ljl_data():
    # create object of league class for the LCS
    ljl_teams = ['AXIZ', 'Burning Core', 'Crest Gaming Act', 'DetonatioN FM', 'SoftBank HAWKS', 'Rascal Jester', 'Sengoku Gaming', 'V3 Esports']
    competition = 'ljl'
    ljl = TableData(24.5, ljl_teams, competition)
    
    # call function to scrape data and perform calculations
    ljl.kill_data('https://lol.fandom.com/wiki/LJL/2022_Season/Spring_Season/Scoreboards')
    ljl.kill_data('https://lol.fandom.com/wiki/LJL/2022_Season/Spring_Season/Scoreboards/Days_5-8')
    ljl.kill_data('https://lol.fandom.com/wiki/LJL/2022_Season/Spring_Season/Scoreboards/Days_9-12')
    ljl.kill_data('https://lol.fandom.com/wiki/LJL/2022_Season/Spring_Season/Scoreboards/Days_13-16')
    ljl.kill_data('https://lol.fandom.com/wiki/LJL/2022_Season/Spring_Season/Scoreboards/Days_17-21')

    # update the displayed table with the new data
    ljl.calculate_all()
    ui.get_ljl()

def nlc_data():
    # create object of league class for the LCS
    nlc_teams = ['Astralis Talent', 'Bifrost', 'Dusty', 'JD|XL', 'MNM Gaming', 'NYYRIKKI', 'Riddle Esports', 'Team Singularity', 'Vanir', 'X7 Esports']
    competition = 'nlc'
    nlc = TableData(27.5, nlc_teams, competition)
    
    # call function to scrape data and perform calculations
    nlc.kill_data('https://lol.fandom.com/wiki/NLC/2022_Season/Spring_Season/Scoreboards')
    nlc.kill_data('https://lol.fandom.com/wiki/NLC/2022_Season/Spring_Season/Scoreboards/Week_2')
    nlc.kill_data('https://lol.fandom.com/wiki/NLC/2022_Season/Spring_Season/Scoreboards/Week_3')
    nlc.kill_data('https://lol.fandom.com/wiki/NLC/2022_Season/Spring_Season/Scoreboards/Week_4')
    nlc.kill_data('https://lol.fandom.com/wiki/NLC/2022_Season/Spring_Season/Scoreboards/Week_5')
    nlc.kill_data('https://lol.fandom.com/wiki/NLC/2022_Season/Spring_Season/Scoreboards/Week_6')
    nlc.kill_data('https://lol.fandom.com/wiki/NLC/2022_Season/Spring_Season/Scoreboards/Week_7')
    nlc.kill_data('https://lol.fandom.com/wiki/NLC/2022_Season/Spring_Season/Scoreboards/Week_8')
    nlc.kill_data('https://lol.fandom.com/wiki/NLC/2022_Season/Spring_Season/Scoreboards/Week_9')

    # update the displayed table with the new data
    nlc.calculate_all()
    ui.get_nlc()

def lfl_data():
    # create object of league class for the LCS
    lfl_teams = ['GamersOrigin', 'GameWard', 'Karmine Corp', 'LDLC OL', 'Mirage Elyandra', 'Misfits Premier', 'Solary', 'Team BDS Academy', 'Team Oplon', 'Vitality.Bee']
    competition = 'lfl'
    lfl = TableData(27.5, lfl_teams, competition)
    
    # call function to scrape data and perform calculations
    lfl.kill_data('https://lol.fandom.com/wiki/LFL/2022_Season/Spring_Season/Scoreboards')
    lfl.kill_data('https://lol.fandom.com/wiki/LFL/2022_Season/Spring_Season/Scoreboards/Week_2')
    lfl.kill_data('https://lol.fandom.com/wiki/LFL/2022_Season/Spring_Season/Scoreboards/Week_3')
    lfl.kill_data('https://lol.fandom.com/wiki/LFL/2022_Season/Spring_Season/Scoreboards/Week_4')
    lfl.kill_data('https://lol.fandom.com/wiki/LFL/2022_Season/Spring_Season/Scoreboards/Week_5')
    lfl.kill_data('https://lol.fandom.com/wiki/LFL/2022_Season/Spring_Season/Scoreboards/Week_6')
    lfl.kill_data('https://lol.fandom.com/wiki/LFL/2022_Season/Spring_Season/Scoreboards/Week_7')
    lfl.kill_data('https://lol.fandom.com/wiki/LFL/2022_Season/Spring_Season/Scoreboards/Week_8')
    lfl.kill_data('https://lol.fandom.com/wiki/LFL/2022_Season/Spring_Season/Scoreboards/Week_9')

    # update the displayed table with the new data
    lfl.calculate_all()
    ui.get_lfl()

def lplol_data():
    # create object of league class for the LCS
    lplol_teams = ['Boavista FC', 'EGN Esports', 'For The Win', 'GTZ Bulls', 'Karma Clan Esports', 'Odivelas Sports Club', 'OFFSET Esports', 'White Dragons']
    competition = 'lplol'
    lplol = TableData(29.5, lplol_teams, competition)
    
    # call function to scrape data and perform calculations
    lplol.kill_data('https://lol.fandom.com/wiki/LPLOL/2022_Season/Spring_Season/Scoreboards')
    lplol.kill_data('https://lol.fandom.com/wiki/LPLOL/2022_Season/Spring_Season/Scoreboards/Week_2')
    lplol.kill_data('https://lol.fandom.com/wiki/LPLOL/2022_Season/Spring_Season/Scoreboards/Week_3')
    lplol.kill_data('https://lol.fandom.com/wiki/LPLOL/2022_Season/Spring_Season/Scoreboards/Week_4')
    lplol.kill_data('https://lol.fandom.com/wiki/LPLOL/2022_Season/Spring_Season/Scoreboards/Week_5')
    lplol.kill_data('https://lol.fandom.com/wiki/LPLOL/2022_Season/Spring_Season/Scoreboards/Week_6')
    lplol.kill_data('https://lol.fandom.com/wiki/LPLOL/2022_Season/Spring_Season/Scoreboards/Week_7')

    # update the displayed table with the new data
    lplol.calculate_all()
    ui.get_lplol()

def hpm_data():
    # create object of league class for the LCS
    hpm_teams = ['BRUTE', 'Cryptova', 'Dark Tigers', 'Dynamo Eclot', 'Entropiq', 'eSuba', 'Inside Games', 'SINNERS Esports']
    competition = 'hpm'
    hpm = TableData(29.5, hpm_teams, competition)
    
    # call function to scrape data and perform calculations
    hpm.kill_data('https://lol.fandom.com/wiki/Hitpoint_Masters/2022_Season/Spring_Season/Scoreboards')
    hpm.kill_data('https://lol.fandom.com/wiki/Hitpoint_Masters/2022_Season/Spring_Season/Scoreboards/Week_2')
    hpm.kill_data('https://lol.fandom.com/wiki/Hitpoint_Masters/2022_Season/Spring_Season/Scoreboards/Week_3')
    hpm.kill_data('https://lol.fandom.com/wiki/Hitpoint_Masters/2022_Season/Spring_Season/Scoreboards/Week_4')
    hpm.kill_data('https://lol.fandom.com/wiki/Hitpoint_Masters/2022_Season/Spring_Season/Scoreboards/Week_5')
    hpm.kill_data('https://lol.fandom.com/wiki/Hitpoint_Masters/2022_Season/Spring_Season/Scoreboards/Week_6')
    hpm.kill_data('https://lol.fandom.com/wiki/Hitpoint_Masters/2022_Season/Spring_Season/Scoreboards/Week_7')

    # update the displayed table with the new data
    hpm.calculate_all()
    ui.get_hpm()

def lvp_data():
    # create object of league class for the LCS
    lvp_teams = ['Barça eSports', 'BISONS ECLUB', 'Fnatic TQ', 'G2 Arctic', 'Giants', 'KOI', 'MAD Lions Madrid', 'Movistar Riders', 'Team Heretics', 'UCAM Tokiers']
    competition = 'lvp'
    lvp = TableData(26.5, lvp_teams, competition)
    
    # call function to scrape data and perform calculations
    lvp.kill_data('https://lol.fandom.com/wiki/LVP_SuperLiga/2022_Season/Spring_Season/Scoreboards')
    lvp.kill_data('https://lol.fandom.com/wiki/LVP_SuperLiga/2022_Season/Spring_Season/Scoreboards/Week_2')
    lvp.kill_data('https://lol.fandom.com/wiki/LVP_SuperLiga/2022_Season/Spring_Season/Scoreboards/Week_3')
    lvp.kill_data('https://lol.fandom.com/wiki/LVP_SuperLiga/2022_Season/Spring_Season/Scoreboards/Week_4')
    lvp.kill_data('https://lol.fandom.com/wiki/LVP_SuperLiga/2022_Season/Spring_Season/Scoreboards/Week_5')
    lvp.kill_data('https://lol.fandom.com/wiki/LVP_SuperLiga/2022_Season/Spring_Season/Scoreboards/Week_6')
    lvp.kill_data('https://lol.fandom.com/wiki/LVP_SuperLiga/2022_Season/Spring_Season/Scoreboards/Week_7')
    lvp.kill_data('https://lol.fandom.com/wiki/LVP_SuperLiga/2022_Season/Spring_Season/Scoreboards/Week_8')
    lvp.kill_data('https://lol.fandom.com/wiki/LVP_SuperLiga/2022_Season/Spring_Season/Scoreboards/Week_9')

    # update the displayed table with the new data
    lvp.calculate_all()
    ui.get_lvp()

def lck_chal_data():
    # create object of league class for the LCS
    lck_chal_teams = ['DRX Challengers', 'DWG KIA Challengers', 'BRO Challengers', 'GEN Challengers', 'HLE Challengers', 'KT Challengers', 'KDF Challengers', 'LSB Challengers', 'NS Challengers', 'T1 Challengers']
    competition = 'lck_chal'
    lck_chal = TableData(26.5, lck_chal_teams, competition)
    
    # call function to scrape data and perform calculations
    lck_chal.kill_data('https://lol.fandom.com/wiki/LCK_CL/2022_Season/Spring_Season/Scoreboards')
    lck_chal.kill_data('https://lol.fandom.com/wiki/LCK_CL/2022_Season/Spring_Season/Scoreboards/Week_2')
    lck_chal.kill_data('https://lol.fandom.com/wiki/LCK_CL/2022_Season/Spring_Season/Scoreboards/Week_3')
    lck_chal.kill_data('https://lol.fandom.com/wiki/LCK_CL/2022_Season/Spring_Season/Scoreboards/Week_4')
    lck_chal.kill_data('https://lol.fandom.com/wiki/LCK_CL/2022_Season/Spring_Season/Scoreboards/Week_5')
    lck_chal.kill_data('https://lol.fandom.com/wiki/LCK_CL/2022_Season/Spring_Season/Scoreboards/Week_6')
    lck_chal.kill_data('https://lol.fandom.com/wiki/LCK_CL/2022_Season/Spring_Season/Scoreboards/Week_7')
    lck_chal.kill_data('https://lol.fandom.com/wiki/LCK_CL/2022_Season/Spring_Season/Scoreboards/Week_8')
    lck_chal.kill_data('https://lol.fandom.com/wiki/LCK_CL/2022_Season/Spring_Season/Scoreboards/Week_9')

    # update the displayed table with the new data
    lck_chal.calculate_all()
    ui.get_lck_chal()

def tcl_data():
    # create object of league class for the LCS
    tcl_teams = ['5 Ronin', 'Beşiktaş Esports', 'Dark Passage', 'Fenerbahçe', 'Galakticos', 'Galatasaray Esports', 'İstanbul Wildcats', 'NASR Turkey', 'SuperMassive Blaze', 'Team AURORA']
    competition = 'tcl'
    tcl = TableData(26.5, tcl_teams, competition)
    
    # call function to scrape data and perform calculations
    tcl.kill_data('https://lol.fandom.com/wiki/TCL/2022_Season/Winter_Season/Scoreboards')
    tcl.kill_data('https://lol.fandom.com/wiki/TCL/2022_Season/Winter_Season/Scoreboards/Week_2')
    tcl.kill_data('https://lol.fandom.com/wiki/TCL/2022_Season/Winter_Season/Scoreboards/Week_3')
    tcl.kill_data('https://lol.fandom.com/wiki/TCL/2022_Season/Winter_Season/Scoreboards/Week_4')
    tcl.kill_data('https://lol.fandom.com/wiki/TCL/2022_Season/Winter_Season/Scoreboards/Week_5')
    tcl.kill_data('https://lol.fandom.com/wiki/TCL/2022_Season/Winter_Season/Scoreboards/Week_6')
    tcl.kill_data('https://lol.fandom.com/wiki/TCL/2022_Season/Winter_Season/Scoreboards/Week_7')
    tcl.kill_data('https://lol.fandom.com/wiki/TCL/2022_Season/Winter_Season/Scoreboards/Week_8')
    tcl.kill_data('https://lol.fandom.com/wiki/TCL/2022_Season/Winter_Season/Scoreboards/Week_9')

    # update the displayed table with the new data
    tcl.calculate_all()
    ui.get_tcl()

def pcs_data():
    # create object of league class for the LCS
    pcs_teams = ['Beyond Gaming', 'CTBC Flying Oyster', 'Deep Cross Gaming', 'Frank Esports', 'Hurricane Gaming', 'Impunity', 'J Team', 'Meta Falcon Team', 'PSG Talon', 'SEM9']
    competition = 'pcs'
    pcs = TableData(27.5, pcs_teams, competition)
    
    # call function to scrape data and perform calculations
    pcs.kill_data('https://lol.fandom.com/wiki/PCS/2022_Season/Spring_Season/Scoreboards')
    pcs.kill_data('https://lol.fandom.com/wiki/PCS/2022_Season/Spring_Season/Scoreboards/Week_2')
    pcs.kill_data('https://lol.fandom.com/wiki/PCS/2022_Season/Spring_Season/Scoreboards/Week_3')
    pcs.kill_data('https://lol.fandom.com/wiki/PCS/2022_Season/Spring_Season/Scoreboards/Week_4')
    pcs.kill_data('https://lol.fandom.com/wiki/PCS/2022_Season/Spring_Season/Scoreboards/Week_5')
    pcs.kill_data('https://lol.fandom.com/wiki/PCS/2022_Season/Spring_Season/Scoreboards/Week_6')

    # update the displayed table with the new data
    pcs.calculate_all()
    ui.get_pcs()

def ebl_data():
    # create object of league class for the LCS
    ebl_teams = ['Auxesis Esports', 'Crvena zvezda', 'Cyber Wolves', 'Diamant Esports', 'Nexus KTRL', 'SOVEJA', 'Split Raiders', 'Valiance']
    competition = 'ebl'
    ebl = TableData(28.5, ebl_teams, competition)
    
    # call function to scrape data and perform calculations
    ebl.kill_data('https://lol.fandom.com/wiki/Esports_Balkan_League/2022_Season/Spring_Split/Scoreboards')
    ebl.kill_data('https://lol.fandom.com/wiki/Esports_Balkan_League/2022_Season/Spring_Split/Scoreboards/Week_2')
    ebl.kill_data('https://lol.fandom.com/wiki/Esports_Balkan_League/2022_Season/Spring_Split/Scoreboards/Week_3')
    ebl.kill_data('https://lol.fandom.com/wiki/Esports_Balkan_League/2022_Season/Spring_Split/Scoreboards/Week_4')
    ebl.kill_data('https://lol.fandom.com/wiki/Esports_Balkan_League/2022_Season/Spring_Split/Scoreboards/Week_5')
    ebl.kill_data('https://lol.fandom.com/wiki/Esports_Balkan_League/2022_Season/Spring_Split/Scoreboards/Week_6')
    ebl.kill_data('https://lol.fandom.com/wiki/Esports_Balkan_League/2022_Season/Spring_Split/Scoreboards/Week_7')

    # update the displayed table with the new data
    ebl.calculate_all()
    ui.get_ebl()

# alter style of the UI
stylesheet = """
QToolTip
{
     border: 1px solid black;
     background-color: #ffa02f;
     padding: 1px;
     border-radius: 3px;
     opacity: 100;
}

QWidget
{
    color: #b1b1b1;
    background-color: #323232;
}

QTreeView, QListView
{
    background-color: silver;
    margin-left: 5px;
}

QWidget:item:hover
{
    background-color: QLinearGradient( x1: 0, y1: 0, x2: 0, y2: 1, stop: 0 #ffa02f, stop: 1 #ca0619);
    color: #000000;
}

QWidget:item:selected
{
    background-color: QLinearGradient( x1: 0, y1: 0, x2: 0, y2: 1, stop: 0 #ffa02f, stop: 1 #d7801a);
}

QMenuBar::item
{
    background: transparent;
}

QMenuBar::item:selected
{
    background: transparent;
    border: 1px solid #ffaa00;
}

QMenuBar::item:pressed
{
    background: #444;
    border: 1px solid #000;
    background-color: QLinearGradient(
        x1:0, y1:0,
        x2:0, y2:1,
        stop:1 #212121,
        stop:0.4 #343434/*,
        stop:0.2 #343434,
        stop:0.1 #ffaa00*/
    );
    margin-bottom:-1px;
    padding-bottom:1px;
}

QMenu
{
    border: 1px solid #000;
}

QMenu::item
{
    padding: 2px 20px 2px 20px;
}

QMenu::item:selected
{
    color: #000000;
}

QWidget:disabled
{
    color: #808080;
    background-color: #323232;
}

QAbstractItemView
{
    background-color: QLinearGradient( x1: 0, y1: 0, x2: 0, y2: 1, stop: 0 #4d4d4d, stop: 0.1 #646464, stop: 1 #5d5d5d);
}

QWidget:focus
{
    /*border: 1px solid darkgray;*/
}

QLineEdit
{
    background-color: QLinearGradient( x1: 0, y1: 0, x2: 0, y2: 1, stop: 0 #4d4d4d, stop: 0 #646464, stop: 1 #5d5d5d);
    padding: 1px;
    border-style: solid;
    border: 1px solid #1e1e1e;
    border-radius: 5;
}

QPushButton
{
    color: #b1b1b1;
    background-color: QLinearGradient( x1: 0, y1: 0, x2: 0, y2: 1, stop: 0 #565656, stop: 0.1 #525252, stop: 0.5 #4e4e4e, stop: 0.9 #4a4a4a, stop: 1 #464646);
    border-width: 1px;
    border-color: #1e1e1e;
    border-style: solid;
    border-radius: 6;
    padding: 3px;
    font-size: 12px;
    padding-left: 5px;
    padding-right: 5px;
    min-width: 40px;
}

QPushButton:pressed
{
    background-color: QLinearGradient( x1: 0, y1: 0, x2: 0, y2: 1, stop: 0 #2d2d2d, stop: 0.1 #2b2b2b, stop: 0.5 #292929, stop: 0.9 #282828, stop: 1 #252525);
}

QComboBox
{
    selection-background-color: #ffaa00;
    background-color: QLinearGradient( x1: 0, y1: 0, x2: 0, y2: 1, stop: 0 #565656, stop: 0.1 #525252, stop: 0.5 #4e4e4e, stop: 0.9 #4a4a4a, stop: 1 #464646);
    border-style: solid;
    border: 1px solid #1e1e1e;
    border-radius: 5;
}

QComboBox:hover,QPushButton:hover
{
    border: 2px solid QLinearGradient( x1: 0, y1: 0, x2: 0, y2: 1, stop: 0 #ffa02f, stop: 1 #d7801a);
}


QComboBox:on
{
    padding-top: 3px;
    padding-left: 4px;
    background-color: QLinearGradient( x1: 0, y1: 0, x2: 0, y2: 1, stop: 0 #2d2d2d, stop: 0.1 #2b2b2b, stop: 0.5 #292929, stop: 0.9 #282828, stop: 1 #252525);
    selection-background-color: #ffaa00;
}

QComboBox QAbstractItemView
{
    border: 2px solid darkgray;
    selection-background-color: QLinearGradient( x1: 0, y1: 0, x2: 0, y2: 1, stop: 0 #ffa02f, stop: 1 #d7801a);
}

QComboBox::drop-down
{
     subcontrol-origin: padding;
     subcontrol-position: top right;
     width: 15px;

     border-left-width: 0px;
     border-left-color: darkgray;
     border-left-style: solid; /* just a single line */
     border-top-right-radius: 3px; /* same radius as the QComboBox */
     border-bottom-right-radius: 3px;
 }

QComboBox::down-arrow
{
     image: url(:/dark_orange/img/down_arrow.png);
}

QGroupBox
{
    border: 1px solid darkgray;
    margin-top: 10px;
}

QGroupBox:focus
{
    border: 1px solid darkgray;
}

QTextEdit:focus
{
    border: 1px solid darkgray;
}

QScrollBar:horizontal {
     border: 1px solid #222222;
     background: QLinearGradient( x1: 0, y1: 0, x2: 0, y2: 1, stop: 0.0 #121212, stop: 0.2 #282828, stop: 1 #484848);
     height: 7px;
     margin: 0px 16px 0 16px;
}

QScrollBar::handle:horizontal
{
      background: QLinearGradient( x1: 0, y1: 0, x2: 1, y2: 0, stop: 0 #ffa02f, stop: 0.5 #d7801a, stop: 1 #ffa02f);
      min-height: 20px;
      border-radius: 2px;
}

QScrollBar::add-line:horizontal {
      border: 1px solid #1b1b19;
      border-radius: 2px;
      background: QLinearGradient( x1: 0, y1: 0, x2: 1, y2: 0, stop: 0 #ffa02f, stop: 1 #d7801a);
      width: 14px;
      subcontrol-position: right;
      subcontrol-origin: margin;
}

QScrollBar::sub-line:horizontal {
      border: 1px solid #1b1b19;
      border-radius: 2px;
      background: QLinearGradient( x1: 0, y1: 0, x2: 1, y2: 0, stop: 0 #ffa02f, stop: 1 #d7801a);
      width: 14px;
     subcontrol-position: left;
     subcontrol-origin: margin;
}

QScrollBar::right-arrow:horizontal, QScrollBar::left-arrow:horizontal
{
      border: 1px solid black;
      width: 1px;
      height: 1px;
      background: white;
}

QScrollBar::add-page:horizontal, QScrollBar::sub-page:horizontal
{
      background: none;
}

QScrollBar:vertical
{
      background: QLinearGradient( x1: 0, y1: 0, x2: 1, y2: 0, stop: 0.0 #121212, stop: 0.2 #282828, stop: 1 #484848);
      width: 7px;
      margin: 16px 0 16px 0;
      border: 1px solid #222222;
}

QScrollBar::handle:vertical
{
      background: QLinearGradient( x1: 0, y1: 0, x2: 0, y2: 1, stop: 0 #ffa02f, stop: 0.5 #d7801a, stop: 1 #ffa02f);
      min-height: 20px;
      border-radius: 2px;
}

QScrollBar::add-line:vertical
{
      border: 1px solid #1b1b19;
      border-radius: 2px;
      background: QLinearGradient( x1: 0, y1: 0, x2: 0, y2: 1, stop: 0 #ffa02f, stop: 1 #d7801a);
      height: 14px;
      subcontrol-position: bottom;
      subcontrol-origin: margin;
}

QScrollBar::sub-line:vertical
{
      border: 1px solid #1b1b19;
      border-radius: 2px;
      background: QLinearGradient( x1: 0, y1: 0, x2: 0, y2: 1, stop: 0 #d7801a, stop: 1 #ffa02f);
      height: 14px;
      subcontrol-position: top;
      subcontrol-origin: margin;
}

QScrollBar::up-arrow:vertical, QScrollBar::down-arrow:vertical
{
      border: 1px solid black;
      width: 1px;
      height: 1px;
      background: white;
}


QScrollBar::add-page:vertical, QScrollBar::sub-page:vertical
{
      background: none;
}

QTextEdit
{
    background-color: #242424;
}

QPlainTextEdit
{
    background-color: #242424;
}

QHeaderView::section
{
    background-color: QLinearGradient(x1:0, y1:0, x2:0, y2:1, stop:0 #616161, stop: 0.5 #505050, stop: 0.6 #434343, stop:1 #656565);
    color: white;
    padding-left: 4px;
    border: 1px solid #6c6c6c;
}

QCheckBox:disabled
{
color: #414141;
}

QDockWidget::title
{
    text-align: center;
    spacing: 3px; /* spacing between items in the tool bar */
    background-color: QLinearGradient(x1:0, y1:0, x2:0, y2:1, stop:0 #323232, stop: 0.5 #242424, stop:1 #323232);
}

QDockWidget::close-button, QDockWidget::float-button
{
    text-align: center;
    spacing: 1px; /* spacing between items in the tool bar */
    background-color: QLinearGradient(x1:0, y1:0, x2:0, y2:1, stop:0 #323232, stop: 0.5 #242424, stop:1 #323232);
}

QDockWidget::close-button:hover, QDockWidget::float-button:hover
{
    background: #242424;
}

QDockWidget::close-button:pressed, QDockWidget::float-button:pressed
{
    padding: 1px -1px -1px 1px;
}

QMainWindow::separator
{
    background-color: QLinearGradient(x1:0, y1:0, x2:0, y2:1, stop:0 #161616, stop: 0.5 #151515, stop: 0.6 #212121, stop:1 #343434);
    color: white;
    padding-left: 4px;
    border: 1px solid #4c4c4c;
    spacing: 3px; /* spacing between items in the tool bar */
}

QMainWindow::separator:hover
{

    background-color: QLinearGradient(x1:0, y1:0, x2:0, y2:1, stop:0 #d7801a, stop:0.5 #b56c17 stop:1 #ffa02f);
    color: white;
    padding-left: 4px;
    border: 1px solid #6c6c6c;
    spacing: 3px; /* spacing between items in the tool bar */
}

QToolBar::handle
{
     spacing: 3px; /* spacing between items in the tool bar */
     background: url(:/dark_orange/img/handle.png);
}

QMenu::separator
{
    height: 2px;
    background-color: QLinearGradient(x1:0, y1:0, x2:0, y2:1, stop:0 #161616, stop: 0.5 #151515, stop: 0.6 #212121, stop:1 #343434);
    color: white;
    padding-left: 4px;
    margin-left: 10px;
    margin-right: 5px;
}

QProgressBar
{
    border: 2px solid grey;
    border-radius: 5px;
    text-align: center;
}

QProgressBar::chunk
{
    background-color: #d7801a;
    width: 2.15px;
    margin: 0.5px;
}

QTabBar::tab {
    color: #b1b1b1;
    border: 1px solid #444;
    border-bottom-style: none;
    background-color: #323232;
    padding-left: 10px;
    padding-right: 10px;
    padding-top: 3px;
    padding-bottom: 2px;
    margin-right: -1px;
}

QTabWidget::pane {
    border: 1px solid #444;
    top: 1px;
}

QTabBar::tab:last
{
    margin-right: 0; /* the last selected tab has nothing to overlap with on the right */
    border-top-right-radius: 3px;
}

QTabBar::tab:first:!selected
{
 margin-left: 0px; /* the last selected tab has nothing to overlap with on the right */


    border-top-left-radius: 3px;
}

QTabBar::tab:!selected
{
    color: #b1b1b1;
    border-bottom-style: solid;
    margin-top: 3px;
    background-color: QLinearGradient(x1:0, y1:0, x2:0, y2:1, stop:1 #212121, stop:.4 #343434);
}

QTabBar::tab:selected
{
    border-top-left-radius: 3px;
    border-top-right-radius: 3px;
    margin-bottom: 0px;
}

QTabBar::tab:!selected:hover
{
    /*border-top: 2px solid #ffaa00;
    padding-bottom: 3px;*/
    border-top-left-radius: 3px;
    border-top-right-radius: 3px;
    background-color: QLinearGradient(x1:0, y1:0, x2:0, y2:1, stop:1 #212121, stop:0.4 #343434, stop:0.2 #343434, stop:0.1 #ffaa00);
}

QRadioButton::indicator:checked, QRadioButton::indicator:unchecked{
    color: #b1b1b1;
    background-color: #323232;
    border: 1px solid #b1b1b1;
    border-radius: 6px;
}

QRadioButton::indicator:checked
{
    background-color: qradialgradient(
        cx: 0.5, cy: 0.5,
        fx: 0.5, fy: 0.5,
        radius: 1.0,
        stop: 0.25 #ffaa00,
        stop: 0.3 #323232
    );
}

QCheckBox::indicator{
    color: #b1b1b1;
    background-color: #323232;
    border: 1px solid #b1b1b1;
    width: 9px;
    height: 9px;
}

QRadioButton::indicator
{
    border-radius: 6px;
}

QRadioButton::indicator:hover, QCheckBox::indicator:hover
{
    border: 1px solid #ffaa00;
}

QCheckBox::indicator:checked
{
    image:url(:/dark_orange/img/checkbox.png);
}

QCheckBox::indicator:disabled, QRadioButton::indicator:disabled
{
    border: 1px solid #444;
}


QSlider::groove:horizontal {
    border: 1px solid #3A3939;
    height: 8px;
    background: #201F1F;
    margin: 2px 0;
    border-radius: 2px;
}

QSlider::handle:horizontal {
    background: QLinearGradient( x1: 0, y1: 0, x2: 0, y2: 1,
      stop: 0.0 silver, stop: 0.2 #a8a8a8, stop: 1 #727272);
    border: 1px solid #3A3939;
    width: 14px;
    height: 14px;
    margin: -4px 0;
    border-radius: 2px;
}

QSlider::groove:vertical {
    border: 1px solid #3A3939;
    width: 8px;
    background: #201F1F;
    margin: 0 0px;
    border-radius: 2px;
}

QSlider::handle:vertical {
    background: QLinearGradient( x1: 0, y1: 0, x2: 0, y2: 1, stop: 0.0 silver,
      stop: 0.2 #a8a8a8, stop: 1 #727272);
    border: 1px solid #3A3939;
    width: 14px;
    height: 14px;
    margin: 0 -4px;
    border-radius: 2px;
}

QAbstractSpinBox {
    padding-top: 2px;
    padding-bottom: 2px;
    border: 1px solid darkgray;

    border-radius: 2px;
    min-width: 50px;
}
"""

if __name__ == "__main__":
    # create UI
    app = QApplication(sys.argv)
    app.setStyleSheet(stylesheet)
    MainWindow = QtWidgets.QMainWindow()
    ui = Ui_MainWindow()
    ui.setupUi(MainWindow)
    MainWindow.show()
    sys.exit(app.exec_())
